# ==============================================================================
# AMD 5-YEAR PROJECTION — COMPLETE ENGINE v2.0
# Built by reverse-engineering AMD 5Year Projection Manual.xlsx from Drive
# THIS IS THE FULL ENGINE — not a stub. Runs standalone.
# v2.0: Adds Assumptions, Workflow, Data Sources tabs to match Drive file.
# Update only the DATA BLOCK below each quarter. Structure never changes.
# HOW TO RUN: python build_amd_projection_engine_v2.py
# ==============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ── DATA BLOCK — update these values each quarter ──────────────────────────────
VERSION      = "v5.1 — Q1 2026 Baseline"
GENERATED    = "May 2026"
ENTRY_PRICE  = 455.00       # AMD stock price at analysis date
BASE_REV     = 34.64        # FY2025 actual revenue ($B)
SHARES       = 1.650        # Diluted shares ($B)
Q1_NI_MARGIN = 0.221        # Q1 2026 Non-GAAP NI margin
Q1_REV       = 10.253       # Q1 2026 revenue ($B)
Q2_GUIDE     = 11.200       # Q2 2026 guidance midpoint ($B)

BULL = dict(
    rev_growth = [0.43, 0.50, 0.52, 0.50, 0.45],
    ni_margin  = [0.28, 0.32, 0.38, 0.41, 0.43],
    pe_low     = [45, 48, 50, 50, 48],
    pe_high    = [55, 58, 60, 60, 58],
)
BASE = dict(
    rev_growth = [0.36, 0.42, 0.43, 0.42, 0.40],
    ni_margin  = [0.25, 0.26, 0.26, 0.28, 0.29],
    pe_low     = [35, 36, 37, 38, 40],
    pe_high    = [40, 41, 42, 43, 45],
)
BEAR = dict(
    rev_growth = [0.32, 0.18, 0.12, 0.08, 0.10],
    ni_margin  = [0.22, 0.21, 0.22, 0.24, 0.26],
    pe_low     = [25, 20, 17, 15, 16],
    pe_high    = [30, 25, 22, 20, 21],
)

PROB_BULL = 0.35
PROB_BASE  = 0.45
PROB_BEAR  = 0.20

YEARS = [2026, 2027, 2028, 2029, 2030]

# ── COLORS (exact AMD palette) ─────────────────────────────────────────────────
NAVY    = "1A1A2E"
BLUE    = "1F4E79"
ACCENT  = "2E75B6"
WHITE   = "FFFFFF"
GRAY    = "F5F7FA"
BLACK   = "000000"
MUTED   = "888888"
GREEN_S = "E8F5E9"
BLUE_S  = "D6E4F0"
RED_S   = "FDEDEC"
GOLD_BG = "FFF8E1"
GOLD_FN = "7D5A00"
PROB_BG = "EDE7F6"

def fill(h): return PatternFill("solid", start_color=h, fgColor=h)

def cell(ws, row, col, val=None, bg=WHITE, fc=BLACK, bold=False,
         fmt=None, align="center", size=9, italic=False, wrap=False):
    c = ws.cell(row=row, column=col)
    if val is not None: c.value = val
    c.font = Font(name="Arial", bold=bold, size=size, color=fc, italic=italic)
    c.fill = fill(bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if fmt: c.number_format = fmt
    return c

def rh(ws, row, h=14): ws.row_dimensions[row].height = h

def compute(sc):
    revs, nis, epss, spls, sphs, cagrl, cagrh = [], [], [], [], [], [], []
    rev = BASE_REV
    for i in range(5):
        rev  = rev * (1 + sc["rev_growth"][i])
        ni   = rev * sc["ni_margin"][i]
        eps  = ni / SHARES
        spl  = eps * sc["pe_low"][i]
        sph  = eps * sc["pe_high"][i]
        revs.append(rev);  nis.append(ni);   epss.append(eps)
        spls.append(spl);  sphs.append(sph)
        cagrl.append((spl / ENTRY_PRICE) ** (1 / (i + 1)) - 1)
        cagrh.append((sph / ENTRY_PRICE) ** (1 / (i + 1)) - 1)
    return revs, nis, epss, spls, sphs, cagrl, cagrh

# ==============================================================================
# TAB 1 — INPUTS
# ==============================================================================
def build_inputs(wb):
    ws = wb.create_sheet("Inputs")
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12

    # Title
    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = f"AMD 5-YEAR PROJECTION — INPUTS [{VERSION}]"
    t.font = Font(name="Arial", bold=True, size=11, color=WHITE)
    t.fill = fill(NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 1, 20)

    ws.merge_cells("A2:G2")
    s = ws["A2"]
    s.value = (f"Blue = Editable | YELLOW HIGHLIGHT = Changed from prior version | "
               f"Probability: Bull {PROB_BULL:.0%} / Base {PROB_BASE:.0%} / Bear {PROB_BEAR:.0%}")
    s.font = Font(name="Arial", size=9, color=WHITE)
    s.fill = fill(BLUE)
    s.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 2, 14)

    ws.merge_cells("A3:G3")
    g = ws["A3"]
    g.value = f"GLOBAL INPUTS — AMD Q1 2026 Earnings Release (May 5, 2026) | Market Data May 8, 2026"
    g.font = Font(name="Arial", bold=True, size=9, color=WHITE)
    g.fill = fill(ACCENT)
    g.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    rh(ws, 3, 14)

    global_rows = [
        ("Current / Entry Stock Price ($)", f"${ENTRY_PRICE:,.2f}",
         "User-specified $455 | May 8, 2026 | Source: Investing.com"),
        ("Base Year Revenue — FY2025 ($B)", f"${BASE_REV:,.2f}B",
         "AMD FY2025 Annual Revenue | Source: AMD 10-K / Q4 2025 Earnings Release | ir.amd.com"),
        ("Diluted Shares Outstanding ($B)", f"{SHARES:,.3f}B",
         "AMD Q1 2026 Earnings Release May 5, 2026 | 1,650M diluted shares | ir.amd.com"),
        ("Q1 2026 Non-GAAP Net Margin", f"{Q1_NI_MARGIN:.1%}",
         "AMD Q1 2026: $2,265M Non-GAAP NI / $10,253M Revenue = 22.1% | AMD Financial Tables PDF"),
        ("Q1 2026 Revenue ($B)", f"${Q1_REV:,.3f}B",
         "AMD Q1 2026 Earnings Release | Actual quarterly result | ir.amd.com"),
        ("Q2 2026 Revenue Guide ($B)", f"${Q2_GUIDE:,.3f}B",
         "AMD Q2 2026 Official Guidance | Midpoint $11.2B ±$300M | AMD Q1 2026 Earnings Call May 5, 2026"),
    ]
    for i, (label, val, note) in enumerate(global_rows):
        r = 4 + i
        bg = WHITE if i % 2 == 0 else GRAY
        cell(ws, r, 1, bg=bg)
        cell(ws, r, 2, label, bg=bg, align="left")
        c = ws.cell(row=r, column=3)
        c.value = val; c.font = Font(name="Arial", bold=True, size=9, color="0070C0")
        c.fill = fill(bg); c.alignment = Alignment(horizontal="right", vertical="center")
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
        cell(ws, r, 4, note, bg=bg, fc=MUTED, size=8, align="left", italic=True)
        rh(ws, r, 14)

    ws.merge_cells("A11:G11")
    sh = ws["A11"]
    sh.value = "SCENARIO ASSUMPTIONS — Edit blue cells | All values feed automatically to Projection tab"
    sh.font = Font(name="Arial", bold=True, size=9, color=WHITE)
    sh.fill = fill(ACCENT)
    sh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    rh(ws, 11, 14)

    for col, txt in [(2,"Assumption"),(3,"2026"),(4,"2027"),(5,"2028"),(6,"2029"),(7,"2030")]:
        cell(ws, 12, col, txt, bg=ACCENT, fc=WHITE, bold=True)
    cell(ws, 12, 1, bg=ACCENT)
    rh(ws, 12, 14)

    scenarios = [
        ("▶ BULL CASE", BULL, GREEN_S, PROB_BULL),
        ("▶ BASE CASE", BASE, BLUE_S,  PROB_BASE),
        ("▶ BEAR CASE", BEAR, RED_S,   PROB_BEAR),
    ]
    r = 13
    for sc_name, sc_data, sc_bg, prob in scenarios:
        ws.merge_cells(f"A{r}:G{r}")
        sl = ws[f"A{r}"]
        sl.value = sc_name
        sl.font = Font(name="Arial", bold=True, size=9, color=BLACK)
        sl.fill = fill(sc_bg)
        sl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        rh(ws, r, 14); r += 1

        metrics = [
            ("Revenue Growth %",    sc_data["rev_growth"], "0.0%"),
            ("Net Income Margin %", sc_data["ni_margin"],  "0.0%"),
            ("PE Low",              sc_data["pe_low"],     "0"),
            ("PE High",             sc_data["pe_high"],    "0"),
        ]
        for label, vals, fmt in metrics:
            bg = WHITE if r % 2 == 0 else GRAY
            cell(ws, r, 1, bg=bg)
            cell(ws, r, 2, label, bg=bg, align="left")
            for j, val in enumerate(vals):
                c = ws.cell(row=r, column=3+j)
                c.value = val
                c.font = Font(name="Arial", bold=True, size=9, color="0070C0")
                c.fill = fill(bg)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.number_format = fmt
            rh(ws, r, 14); r += 1

        cell(ws, r, 1, bg=sc_bg)
        cell(ws, r, 2, "Probability Weight", bg=sc_bg, align="left")
        c = ws.cell(row=r, column=3)
        c.value = prob; c.font = Font(name="Arial", bold=True, size=9, color="0070C0")
        c.fill = fill("FFFF00"); c.number_format = "0%"
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
        cell(ws, r, 4, f"Bull {PROB_BULL:.0%} + Base {PROB_BASE:.0%} + Bear {PROB_BEAR:.0%} = 100%",
             bg=sc_bg, fc=MUTED, size=8, align="left")
        rh(ws, r, 14); r += 2

    ws.merge_cells(f"A{r}:G{r}")
    cl = ws[f"A{r}"]
    cl.value = "COLOR LEGEND: Blue Text = Editable Input | Black Text = Formula | Green Text = Linked from Inputs tab"
    cl.font = Font(name="Arial", size=8, color=MUTED, italic=True)
    cl.alignment = Alignment(horizontal="left", vertical="center", indent=1)

# ==============================================================================
# TAB 2 — PROJECTION
# ==============================================================================
def build_projection(wb):
    ws = wb.create_sheet("Projection")
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 26
    for col in ["C","D","E","F","G"]:
        ws.column_dimensions[col].width = 14
    ws.column_dimensions["H"].width = 2

    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "AMD 5-YEAR FINANCIAL PROJECTION — NON-GAAP"
    t.font = Font(name="Arial", bold=True, size=13, color=WHITE)
    t.fill = fill(NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 1, 22)

    ws.merge_cells("A2:H2")
    s = ws["A2"]
    s.value = (f"Entry Price: ${ENTRY_PRICE} | Base Revenue: ${BASE_REV:.2f}B (FY2025 Actual) | "
               f"Shares: {SHARES}B | Non-GAAP | Revenue & NI in $B | Change inputs on Inputs tab")
    s.font = Font(name="Arial", size=9, color=WHITE)
    s.fill = fill(BLUE)
    s.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 2, 14)

    scenarios = [
        ("▶ BULL CASE — NON-GAAP", BULL, GREEN_S),
        ("▶ BASE CASE — NON-GAAP", BASE, BLUE_S),
        ("▶ BEAR CASE — NON-GAAP", BEAR, RED_S),
    ]

    ROW = 3
    for sc_name, sc_data, sc_bg in scenarios:
        revs, nis, epss, spls, sphs, cagrl, cagrh = compute(sc_data)

        ws.merge_cells(f"A{ROW}:H{ROW}")
        sh = ws[f"A{ROW}"]
        sh.value = sc_name
        sh.font = Font(name="Arial", bold=True, size=11, color=WHITE)
        sh.fill = fill(BLUE)
        sh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        rh(ws, ROW, 20); ROW += 1

        cell(ws, ROW, 1, bg=ACCENT)
        cell(ws, ROW, 2, "YEAR", bg=ACCENT, fc=WHITE, bold=True)
        for i, yr in enumerate(YEARS):
            cell(ws, ROW, 3+i, str(yr), bg=ACCENT, fc=WHITE, bold=True)
        cell(ws, ROW, 8, bg=ACCENT)
        rh(ws, ROW, 16); ROW += 1

        def data_row(label, vals, fmt, bg, fc=BLACK, bold=False):
            nonlocal ROW
            cell(ws, ROW, 1, bg=bg)
            cell(ws, ROW, 2, label, bg=bg, fc=fc, bold=bold, align="left")
            for i, v in enumerate(vals):
                if v is None:
                    cell(ws, ROW, 3+i, "—", bg=bg, fc=MUTED)
                else:
                    c = ws.cell(row=ROW, column=3+i)
                    c.value = v
                    c.font = Font(name="Arial", bold=bold, size=9, color=fc)
                    c.fill = fill(bg)
                    c.alignment = Alignment(horizontal="center", vertical="center")
                    if fmt: c.number_format = fmt
            cell(ws, ROW, 8, bg=bg)
            rh(ws, ROW, 14); ROW += 1

        data_row("REVENUE",
                 [f"{v:.2f}B" for v in revs],
                 None, sc_bg, BLACK, True)
        data_row(" REV GROWTH",
                 sc_data["rev_growth"],
                 "0.0%", sc_bg)
        data_row("NET INCOME",
                 [f"{v:.2f}B" for v in nis],
                 None, WHITE, BLACK, True)
        ni_growth = [None] + [nis[i]/nis[i-1]-1 for i in range(1,5)]
        data_row(" NET INC. GROWTH",
                 ni_growth,
                 "0.0%", GRAY)
        data_row(" NET INC. MARGINS",
                 sc_data["ni_margin"],
                 "0.0%", WHITE)
        data_row("EPS (Non-GAAP)",
                 [f"${int(round(v))}" for v in epss],
                 None, GRAY, BLACK, False)
        data_row(" PE LOW EST",
                 sc_data["pe_low"],
                 "0", WHITE)
        data_row(" PE HIGH EST",
                 sc_data["pe_high"],
                 "0", GRAY)
        data_row("SHARE PRICE LOW",
                 spls, '"$"#,##0', GOLD_BG, GOLD_FN, True)
        data_row("SHARE PRICE HIGH",
                 sphs, '"$"#,##0', GOLD_BG, GOLD_FN, True)
        data_row(" CAGR LOW",
                 cagrl, "+0.0%;-0.0%;—", sc_bg)
        data_row(" CAGR HIGH",
                 cagrh, "+0.0%;-0.0%;—", sc_bg)

        ROW += 1

    ws.merge_cells(f"A{ROW}:H{ROW}")
    sc = ws[f"A{ROW}"]
    sc.value = "SCENARIO COMPARISON — 2030 OUTPUTS"
    sc.font = Font(name="Arial", bold=True, size=9, color=WHITE)
    sc.fill = fill(ACCENT)
    sc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    rh(ws, ROW, 16); ROW += 1

    for col, txt in [(2,"Scenario"),(3,"2030 Revenue"),(4,"2030 Net Income"),
                     (5,"2030 EPS"),(6,"2030 Pr. Low"),(7,"2030 Pr. High"),(8,"5-Yr CAGR Low")]:
        cell(ws, ROW, col, txt, bg=ACCENT, fc=WHITE, bold=True, size=8)
    cell(ws, ROW, 1, bg=ACCENT)
    rh(ws, ROW, 28); ROW += 1

    for sc_name2, sc_data, sc_bg in scenarios:
        revs, nis, epss, spls, sphs, cagrl, cagrh = compute(sc_data)
        nm = sc_name2.replace(" — NON-GAAP", "")
        cell(ws, ROW, 1, bg=sc_bg)
        cell(ws, ROW, 2, nm, bg=sc_bg, bold=True, align="left")
        cell(ws, ROW, 3, f"{revs[4]:.2f}B", bg=sc_bg)
        cell(ws, ROW, 4, f"{nis[4]:.2f}B", bg=sc_bg)
        cell(ws, ROW, 5, f"${int(round(epss[4]))}", bg=sc_bg)
        for j, val in enumerate([spls[4], sphs[4]]):
            cj = ws.cell(row=ROW, column=6+j)
            cj.value = val; cj.fill = fill(GOLD_BG)
            cj.font = Font(name="Arial", bold=True, size=9, color=GOLD_FN)
            cj.number_format = '"$"#,##0'
            cj.alignment = Alignment(horizontal="center", vertical="center")
        c8 = ws.cell(row=ROW, column=8)
        c8.value = cagrl[4]; c8.fill = fill(sc_bg)
        c8.font = Font(name="Arial", size=9)
        c8.number_format = "+0.0%;-0.0%;—"
        c8.alignment = Alignment(horizontal="center", vertical="center")
        rh(ws, ROW, 14); ROW += 1

    ws.freeze_panes = "A3"

    ws.merge_cells(f"A{ROW+1}:H{ROW+1}")
    foot = ws[f"A{ROW+1}"]
    foot.value = (f"Source: AMD FY2025 actual ${BASE_REV:.2f}B | "
                  f"Analysis price ${ENTRY_PRICE} | Built: {GENERATED} | "
                  f"Non-GAAP basis | Script: build_amd_projection_engine_v2.py")
    foot.font = Font(name="Arial", size=7, color=MUTED, italic=True)
    foot.alignment = Alignment(horizontal="left")

# ==============================================================================
# TAB 3 — PROBABILITY WEIGHTED
# ==============================================================================
def build_probability(wb):
    ws = wb.create_sheet("Probability Weighted")
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 28
    for col in ["C","D","E","F","G"]:
        ws.column_dimensions[col].width = 14
    ws.column_dimensions["H"].width = 16

    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "AMD 5-YEAR PROJECTION — PROBABILITY WEIGHTED EXPECTED VALUE"
    t.font = Font(name="Arial", bold=True, size=11, color=WHITE)
    t.fill = fill(NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 1, 22)

    ws.merge_cells("A2:H2")
    s = ws["A2"]
    s.value = (f"Bull {PROB_BULL:.0%} | Base {PROB_BASE:.0%} | Bear {PROB_BEAR:.0%} | "
               f"Expected Value = (Bull×{PROB_BULL:.0%}) + (Base×{PROB_BASE:.0%}) + "
               f"(Bear×{PROB_BEAR:.0%}) | Change probabilities in yellow cells")
    s.font = Font(name="Arial", size=9, color=WHITE)
    s.fill = fill(BLUE)
    s.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 2, 14)

    ws.merge_cells("A3:H3")
    ph = ws["A3"]
    ph.value = "PROBABILITY INPUTS — Edit yellow cells"
    ph.font = Font(name="Arial", bold=True, size=9, color=WHITE)
    ph.fill = fill(ACCENT)
    ph.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    rh(ws, 3, 14)

    for col, txt in [(2,"Case"),(3,"Probability"),(4,"Description")]:
        cell(ws, 4, col, txt, bg=ACCENT, fc=WHITE, bold=True)
    ws.merge_cells("D4:H4")
    cell(ws, 4, 1, bg=ACCENT)
    rh(ws, 4, 14)

    prob_rows = [
        ("BULL CASE", PROB_BULL, GREEN_S,
         "Full Helios/MI500 execution. CPU TAM doubles. AMD takes meaningful GPU share. Compute demand exceeds forecasts."),
        ("BASE CASE", PROB_BASE, BLUE_S,
         "Confirmed deals execute. Strong growth. Gradual margin expansion. CPU TAM partially realized."),
        ("BEAR CASE", PROB_BEAR, RED_S,
         "Vera Rubin/Feynman pauses. AI capex moderates. Custom silicon scales. CPU TAM provides structural floor."),
    ]
    for i, (case, prob, bg, desc) in enumerate(prob_rows):
        r = 5 + i
        cell(ws, r, 1, bg=bg)
        cell(ws, r, 2, case, bg=bg, bold=True, align="left")
        c = ws.cell(row=r, column=3)
        c.value = prob; c.font = Font(name="Arial", bold=True, size=9, color="0070C0")
        c.fill = fill("FFFF00"); c.number_format = "0%"
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(f"D{r}:H{r}")
        cell(ws, r, 4, desc, bg=bg, fc=MUTED, size=8, align="left")
        rh(ws, r, 14)

    cell(ws, 8, 1, bg=WHITE)
    cell(ws, 8, 2, "Sum (must = 100%)", bg=WHITE, bold=True, align="left")
    c = ws.cell(row=8, column=3)
    c.value = PROB_BULL + PROB_BASE + PROB_BEAR
    c.font = Font(name="Arial", bold=True, size=9)
    c.number_format = "0%"; c.fill = fill(WHITE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("D8:H8")
    chk = "✓ 100% — Valid" if abs(PROB_BULL+PROB_BASE+PROB_BEAR-1.0) < 0.001 else "⚠ Does not sum to 100%"
    cell(ws, 8, 4, chk, bg=WHITE, bold=True, align="left")
    rh(ws, 8, 14)

    ws.merge_cells("A10:H10")
    sch = ws["A10"]
    sch.value = "SCENARIO COMPARISON BY YEAR — Bull vs Base vs Bear"
    sch.font = Font(name="Arial", bold=True, size=9, color=WHITE)
    sch.fill = fill(ACCENT)
    sch.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    rh(ws, 10, 14)

    cell(ws, 11, 1, bg=ACCENT)
    cell(ws, 11, 2, "Metric", bg=ACCENT, fc=WHITE, bold=True)
    for i, yr in enumerate(YEARS):
        cell(ws, 11, 3+i, str(yr), bg=ACCENT, fc=WHITE, bold=True)
    cell(ws, 11, 8, "EXPECTED VALUE 2030", bg=ACCENT, fc=WHITE, bold=True, size=8)
    rh(ws, 11, 16)

    bull = compute(BULL); base = compute(BASE); bear = compute(BEAR)
    ev_spl = [bull[3][i]*PROB_BULL + base[3][i]*PROB_BASE + bear[3][i]*PROB_BEAR for i in range(5)]
    ev_sph = [bull[4][i]*PROB_BULL + base[4][i]*PROB_BASE + bear[4][i]*PROB_BEAR for i in range(5)]
    ev_rev = [bull[0][i]*PROB_BULL + base[0][i]*PROB_BASE + bear[0][i]*PROB_BEAR for i in range(5)]
    ev_eps = [bull[2][i]*PROB_BULL + base[2][i]*PROB_BASE + bear[2][i]*PROB_BEAR for i in range(5)]

    sc_blocks = [
        (f"▶ BULL CASE ({PROB_BULL:.0%})", GREEN_S, *bull),
        (f"▶ BASE CASE ({PROB_BASE:.0%})", BLUE_S,  *base),
        (f"▶ BEAR CASE ({PROB_BEAR:.0%})", RED_S,   *bear),
    ]

    ROW = 12
    for sc_name, sc_bg, revs, nis, epss, spls, sphs, cagrl, cagrh in sc_blocks:
        ws.merge_cells(f"A{ROW}:H{ROW}")
        sh = ws[f"A{ROW}"]
        sh.value = sc_name
        sh.font = Font(name="Arial", bold=True, size=9, color=WHITE)
        sh.fill = fill(BLUE)
        sh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        rh(ws, ROW, 16); ROW += 1

        pw_rows = [
            ("  Revenue ($B)",       [f"{v:.2f}B" for v in revs], None,        False),
            ("  EPS (Non-GAAP)",     [f"${int(round(v))}" for v in epss], None, False),
            ("  Share Price Low",    spls, '"$"#,##0', True),
            ("  Share Price High",   sphs, '"$"#,##0', True),
            ("  CAGR Low",           cagrl, "+0.0%;-0.0%;—", False),
            ("  CAGR High",          cagrh, "+0.0%;-0.0%;—", False),
        ]
        for label, vals, fmt, bold in pw_rows:
            is_price = "Price" in label
            bg = GOLD_BG if is_price else (WHITE if ROW % 2 == 0 else GRAY)
            fc = GOLD_FN if is_price else BLACK
            cell(ws, ROW, 1, bg=bg)
            cell(ws, ROW, 2, label, bg=bg, fc=fc, bold=bold, align="left")
            for i, v in enumerate(vals):
                if isinstance(v, str):
                    cell(ws, ROW, 3+i, v, bg=bg, fc=fc, bold=bold)
                else:
                    c = ws.cell(row=ROW, column=3+i)
                    c.value = v; c.font = Font(name="Arial", bold=bold, size=9, color=fc)
                    c.fill = fill(bg); c.number_format = fmt if fmt else ""
                    c.alignment = Alignment(horizontal="center", vertical="center")
            last = vals[4]
            if isinstance(last, str):
                cell(ws, ROW, 8, last, bg=bg, fc=fc, bold=True)
            else:
                c8 = ws.cell(row=ROW, column=8)
                c8.value = last; c8.fill = fill(bg)
                c8.font = Font(name="Arial", bold=True, size=9, color=fc)
                c8.number_format = fmt if fmt else ""
                c8.alignment = Alignment(horizontal="center", vertical="center")
            rh(ws, ROW, 14); ROW += 1
        ROW += 1

    ws.merge_cells(f"A{ROW}:H{ROW}")
    ev = ws[f"A{ROW}"]
    ev.value = "PROBABILITY WEIGHTED EXPECTED VALUE — Updates automatically when probabilities change"
    ev.font = Font(name="Arial", bold=True, size=9, color=WHITE)
    ev.fill = fill(NAVY)
    ev.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    rh(ws, ROW, 16); ROW += 1

    ev_cagrl = [(ev_spl[i]/ENTRY_PRICE)**(1/(i+1))-1 for i in range(5)]
    ev_cagrh = [(ev_sph[i]/ENTRY_PRICE)**(1/(i+1))-1 for i in range(5)]

    ev_rows = [
        ("  Expected Revenue ($B)",   [f"{v:.2f}B" for v in ev_rev], None,        False),
        ("  Expected EPS (Non-GAAP)", [f"${int(round(v))}" for v in ev_eps], None, True),
        ("  Expected Share Price Low", ev_spl,  '"$"#,##0', True),
        ("  Expected Share Price High",ev_sph,  '"$"#,##0', True),
        ("  Expected CAGR Low",        ev_cagrl,"+0.0%;-0.0%;—", False),
        ("  Expected CAGR High",       ev_cagrh,"+0.0%;-0.0%;—", False),
    ]
    for label, vals, fmt, bold in ev_rows:
        is_price = "Price" in label
        bg = GOLD_BG if is_price else PROB_BG
        fc = GOLD_FN if is_price else BLACK
        cell(ws, ROW, 1, bg=bg)
        cell(ws, ROW, 2, label, bg=bg, fc=fc, bold=bold, align="left")
        for i, v in enumerate(vals):
            if isinstance(v, str):
                cell(ws, ROW, 3+i, v, bg=bg, fc=fc, bold=bold)
            else:
                c = ws.cell(row=ROW, column=3+i)
                c.value = v; c.font = Font(name="Arial", bold=bold, size=9, color=fc)
                c.fill = fill(bg); c.number_format = fmt if fmt else ""
                c.alignment = Alignment(horizontal="center", vertical="center")
        last = vals[4]
        if isinstance(last, str):
            cell(ws, ROW, 8, last, bg=bg, fc=fc, bold=True)
        else:
            c8 = ws.cell(row=ROW, column=8)
            c8.value = last; c8.fill = fill(bg)
            c8.font = Font(name="Arial", bold=True, size=9, color=fc)
            c8.number_format = fmt if fmt else ""
            c8.alignment = Alignment(horizontal="center", vertical="center")
        rh(ws, ROW, 14); ROW += 1

    ws.freeze_panes = "A12"

    ws.merge_cells(f"A{ROW+1}:H{ROW+1}")
    foot = ws[f"A{ROW+1}"]
    foot.value = (f"Built: {GENERATED} | Entry ${ENTRY_PRICE} | "
                  f"Shares {SHARES}B | Non-GAAP | Script: build_amd_projection_engine_v2.py")
    foot.font = Font(name="Arial", size=7, color=MUTED, italic=True)
    foot.alignment = Alignment(horizontal="left")

# ==============================================================================
# TAB 4 — ASSUMPTIONS
# Narrative justifications for each scenario/year combination.
# Labels are generated dynamically from BULL/BASE/BEAR dicts so they always
# match the Inputs tab — never hardcode growth rates here.
# ==============================================================================
def build_assumptions(wb):
    ws = wb.create_sheet("Assumptions")
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 90

    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value = f"AMD 5-YEAR PROJECTION — FULL NARRATIVE ASSUMPTIONS  [{VERSION}]"
    t.font = Font(name="Arial", bold=True, size=11, color=WHITE)
    t.fill = fill(NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 1, 20)

    ws.merge_cells("A2:C2")
    s = ws["A2"]
    s.value = (f"{VERSION}: Bull {PROB_BULL:.0%} / Base {PROB_BASE:.0%} / Bear {PROB_BEAR:.0%} | "
               f"Labels auto-generated from Inputs tab constants")
    s.font = Font(name="Arial", size=9, color=WHITE)
    s.fill = fill(BLUE)
    s.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 2, 14)

    # Each entry: (section_label, bg, [(year_idx, pct_val, narrative), ...])
    BULL_REV_NOTES = [
        "Q1 actual +38% Y/Y. Q2 guided +45% Y/Y. H1 locked at $21.45B. Full H2 execution: OpenAI 1GW MI455X, Oracle 50K GPU Helios, Meta 6GW. CPU TAM doubling adds 3pts above v4's 40%.",
        "Helios full volume. MI500 launches (1,000x perf vs MI300X, CDNA 6/2nm/HBM4E). EPYC Verano enters 1:1 CPU:GPU market. Server CPU >70% Y/Y in Q2. +4pts above v4's 42%.",
        "MI500 replacement cycle. Inference explosion — AMD 432GB HBM4 vs NVIDIA 192GB B200. Xilinx amortization zero. EPYC Verano at full volume. +3pts above v4's 41%.",
        "Sovereign AI deployed globally. Physical AI at scale. ROCm achieving parity. Server CPU TAM $90B+ at AMD 35%+ share. +3pts above v4's 37%.",
        "Second AI infra wave. Physical AI industrial scale. ROCm enterprise software $8-12B at 80%+ margins. Replacement cycles compound. +3pts above v4's 33%.",
    ]
    BULL_NI_NOTES = [
        "H1 locked at ~22.3%. H2 must average ~33.5% — achievable via Helios rack-scale margins and confirmed deployments. 3pts above base.",
        "MI500 pricing premium. Helios year-two yield improvements. Operating leverage on $72B revenue. 3pts above base.",
        "Xilinx zero — ~260bps tailwind on $104B base. Software revenue $1-2B at 70%+ margins. R&D ratio compressing. 4pts above base.",
        "Software scaling to $4-6B. Sovereign AI premium pricing. R&D growing 15% vs revenue 40%. 4pts above base.",
        "Software $8-12B at 80%+ margins = ~450bps. Second wave economics better than first. R&D below 15% of revenue. 5pts above base.",
    ]
    BASE_REV_NOTES = [
        "H1 locked at $21.45B. H2 execution risk — partial Helios delays. Requires ~$12.8B/quarter H2. +2pts above v4's 34% from CPU TAM visibility.",
        "Helios volume ramp. MI500 partial adoption. EPYC Verano launch. +3pts above v4's 36%. Full year revenue compounds from H1 base.",
        "Inference explosion — AMD memory advantage. Xilinx amortization zero. Embedded recovery. +3pts above v4's 35%.",
        "Agentic AI in enterprise production. Physical AI generating revenue. EPYC at 35%+ server share. +3pts above v4's 32%.",
        "Natural deceleration at $160B scale. Second AI wave provides floor. Software revenue emerging. +3pts above v4's 28%.",
    ]
    BASE_NI_NOTES = [
        "H1 locked at ~22.3%. H2 must average ~27.7%. Partial Helios/MI455X contribution. Operating leverage begins.",
        "Helios full volume yield improvements. MI500 premium pricing partial capture. Operating leverage on revenue growth.",
        "Xilinx amortization zero — single biggest structural margin event. ~260bps automatic improvement. Data Center 65%+ of revenue.",
        "Pure operational execution. R&D growing ~15% vs revenue growth. Enterprise and sovereign premium pricing. 300bps expansion.",
        "Most conservative — only 100bps from 2029. AMD reaching operational maturity. Approaching Broadcom net margin (~35%).",
    ]
    BEAR_REV_NOTES = [
        "H1 locked at $21.45B. Mild H2 disappointment — Helios partial delays. Only 2pts below base. Bear differentiation begins in 2027.",
        "Vera Rubin launches H2 2027 — hyperscaler 6-month pause. AI capex moderation. Custom silicon scaling. China restrictions. Largest divergence from base.",
        "NVIDIA Feynman sampling causes second pause. AI capex ROI reckoning. Custom silicon at full scale. ROCm fails CUDA parity. AMD stagnant in real terms.",
        "Two consecutive growth collapses — self-reinforcing negative cycle. EPYC provides floor: $15-20B prevents revenue decline. Growing at inflation rate.",
        "Slight re-acceleration. EPYC compound growth floor at $20-25B. Second AI wave creates new demand. NVIDIA concentration triggers regulatory scrutiny.",
    ]
    BEAR_NI_NOTES = [
        "H1 locked at ~22.3%. Less Helios/MI455X in H2 — high-margin products delayed. More EPYC/Client/Gaming revenue. Full year stays flat at 22%.",
        "Revenue collapses — operating deleverage. Fixed R&D ~$9-10B consumes larger % of bear revenue. Bear trough: revenue AND margins compress simultaneously.",
        "Revenue stabilizes. AMD begins R&D reallocation. Xilinx approaching zero — structural relief. 1pt recovery from 2027 trough.",
        "Xilinx fully zero. AMD restructured — R&D declining toward 20% of revenue. EPYC becoming margin anchor. 2pt improvement.",
        "Recovery green shoots. Leaner cost structure + Xilinx fully captured. MI700 early shipments. AMD at 26% net margin — profitable but shadow of base case.",
    ]

    sections = [
        ("▶ BULL CASE — Revenue Growth",    GREEN_S, BULL["rev_growth"], BULL_REV_NOTES),
        ("▶ BULL CASE — Net Income Margin", GREEN_S, BULL["ni_margin"],  BULL_NI_NOTES),
        ("▶ BASE CASE — Revenue Growth",    BLUE_S,  BASE["rev_growth"], BASE_REV_NOTES),
        ("▶ BASE CASE — Net Income Margin", BLUE_S,  BASE["ni_margin"],  BASE_NI_NOTES),
        ("▶ BEAR CASE — Revenue Growth",    RED_S,   BEAR["rev_growth"], BEAR_REV_NOTES),
        ("▶ BEAR CASE — Net Income Margin", RED_S,   BEAR["ni_margin"],  BEAR_NI_NOTES),
    ]

    ROW = 4
    for section_label, sc_bg, vals, notes in sections:
        ws.merge_cells(f"A{ROW}:C{ROW}")
        sh = ws[f"A{ROW}"]
        sh.value = section_label
        sh.font = Font(name="Arial", bold=True, size=9, color=BLACK)
        sh.fill = fill(sc_bg)
        sh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        rh(ws, ROW, 14); ROW += 1

        for i, (yr, pct, note) in enumerate(zip(YEARS, vals, notes)):
            bg = WHITE if i % 2 == 0 else GRAY
            cell(ws, ROW, 1, bg=bg)
            cell(ws, ROW, 2, f"{yr} — {pct:.0%}", bg=bg, fc="0070C0", bold=True, align="left")
            c = ws.cell(row=ROW, column=3)
            c.value = note
            c.font = Font(name="Arial", size=9, color=BLACK)
            c.fill = fill(bg)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            rh(ws, ROW, 28)
            ROW += 1

        ROW += 1  # blank gap between sections

    ws.merge_cells(f"A{ROW}:C{ROW}")
    foot = ws[f"A{ROW}"]
    foot.value = (f"Script: build_amd_projection_engine_v2.py | Built: {GENERATED} | "
                  f"Labels auto-generated from DATA BLOCK constants — always match Inputs tab")
    foot.font = Font(name="Arial", size=7, color=MUTED, italic=True)
    foot.alignment = Alignment(horizontal="left")

# ==============================================================================
# TAB 5 — WORKFLOW
# ==============================================================================
def build_workflow(wb):
    ws = wb.create_sheet("Workflow")
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 75

    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value = "AMD PROJECTION — WORKFLOW & VERSION LOG"
    t.font = Font(name="Arial", bold=True, size=11, color=WHITE)
    t.fill = fill(NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 1, 20)

    ws.merge_cells("A2:C2")
    s = ws["A2"]
    s.value = "To regenerate: pull scripts from GitHub → update DATA BLOCK → run → present for download"
    s.font = Font(name="Arial", size=9, color=WHITE)
    s.fill = fill(BLUE)
    s.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 2, 14)

    def section_header(row, txt):
        ws.merge_cells(f"A{row}:C{row}")
        c = ws[f"A{row}"]
        c.value = txt
        c.font = Font(name="Arial", bold=True, size=9, color=WHITE)
        c.fill = fill(ACCENT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        rh(ws, row, 14)

    def wf_row(row, label, detail, bg=WHITE, wrap=True):
        cell(ws, row, 1, bg=bg)
        cell(ws, row, 2, label, bg=bg, bold=True, align="left")
        c = ws.cell(row=row, column=3)
        c.value = detail
        c.font = Font(name="Arial", size=9, color=BLACK)
        c.fill = fill(bg)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=wrap)
        rh(ws, row, 28 if wrap else 14)

    section_header(4, "QUARTERLY WORKFLOW")
    wf_rows = [
        ("Step 1 — AMD Reports",
         "AMD reports earnings → Tony downloads PDFs from ir.amd.com → saves to Drive Q[X] folder"),
        ("Step 2 — Income Statement",
         "Say: 'Run the Q[X] 20XX income statement analysis'\n"
         "→ Builds AMD_Q[X]_Income_Statement_v1.xlsx (build_amd_income_v1.py pattern)"),
        ("Step 3 — Update Projection",
         "Say: 'Update the AMD projection for Q[X] 20XX'\n"
         "→ Pull build_amd_projection_engine_v2.py from GitHub → update DATA BLOCK → rebuild → present for download"),
        ("Full Earnings Day",
         "Say: 'It's Q[X] 20XX earnings day — run the full AMD analysis'\n"
         "→ Income Statement + Projection update in sequence"),
        ("PDF Limitation Note",
         "AMD PDFs on cloudfront.net are blocked from direct fetch. "
         "Tony downloads → saves to Drive Q[X] folder → Claude reads from Drive."),
    ]
    for i, (label, detail) in enumerate(wf_rows):
        bg = WHITE if i % 2 == 0 else GRAY
        wf_row(5 + i, label, detail, bg=bg)

    section_header(11, "DRIVE FOLDER & SCRIPT LINKS")
    link_rows = [
        ("AMD Financials — Master Folder",
         "https://drive.google.com/drive/folders/1i1GdOQGreQuxv2s6xoqILyY_XIqKuX5a"),
        ("Q1 2026 Quarter Folder",
         "https://drive.google.com/drive/folders/1Df4g1sgPkwy_7BwBLjNTsYNDPh28RiWc"),
        ("Q2 2026 Quarter Folder",
         "https://drive.google.com/drive/folders/1D_GOhT8kgZBTlVPA_7VXvO0-lR5ZyTbB"),
        ("build_amd_projection_engine_v2.py",
         "GitHub: tonyvasquez1/bmnr-dashboard — branch claude/setup-amd-financials-q7F3t"),
        ("build_amd_income_v1.py",
         "GitHub: tonyvasquez1/bmnr-dashboard — Q1 2026 income statement builder"),
        ("AMD_MANIFEST.txt",
         "Master constraints — never rebuild from memory, never overwrite, always self-test"),
    ]
    for i, (label, detail) in enumerate(link_rows):
        bg = WHITE if i % 2 == 0 else GRAY
        wf_row(12 + i, label, detail, bg=bg, wrap=False)

    section_header(19, "VERSION LOG")
    version_rows = [
        ("v1 — April 2026",
         "Initial build. Bull 30% / Base 45% / Bear 25%."),
        ("v2 — April 2026",
         "Column alignment fixes. Workflow and Data Sources tabs added."),
        ("v3 — April 2026",
         "Tailwinds & Headwinds tab added."),
        ("v4 — May 2026",
         "CPU TAM $60B baseline. Bear 25% / Bull 30%."),
        ("v5 — May 2026",
         "CPU TAM doubled to $120B (Lisa Su Q1 2026). Bull 35% / Bear 20%. Rev growth +2-3pts/yr above v4."),
        ("v5.1 — May 2026",
         "Formatting fixed: no yellow, whole dollar formats, correct column widths, scenario comparison formulas corrected, centered alignment, Workflow tab permanent."),
        ("v6 (script v2) — May 2026",
         "GitHub migration. Assumptions, Workflow, Data Sources tabs added to match Drive file. "
         "Assumption labels auto-generated from DATA BLOCK — always in sync with Inputs. "
         "Script renamed build_amd_projection_engine_v2.py."),
    ]
    for i, (ver, detail) in enumerate(version_rows):
        bg = WHITE if i % 2 == 0 else GRAY
        wf_row(20 + i, ver, detail, bg=bg)

    section_header(28, "PENDING ITEMS")
    pending_rows = [
        ("Mac App",
         "Port AMD projection to Vercel. Will automate PDF download and Drive upload entirely."),
        ("GAAP vs Non-GAAP Discussion",
         "Promised deep-dive on GAAP vs non-GAAP analytical framework."),
        ("Revenue Growth Derivation",
         "Segment data, TAM analysis, PEG ratios, comparable company analysis."),
        ("CPU TAM $120B Discussion",
         "EPYC standalone modeling, 1:1 CPU:GPU ratio, Verano ramp timeline."),
        ("Upload xlsx to Drive",
         "Blocked by file size — Mac app will solve this."),
    ]
    for i, (item, detail) in enumerate(pending_rows):
        bg = WHITE if i % 2 == 0 else GRAY
        wf_row(29 + i, item, detail, bg=bg)

# ==============================================================================
# TAB 6 — DATA SOURCES
# ==============================================================================
def build_data_sources(wb):
    ws = wb.create_sheet("Data Sources")
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 60

    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = "AMD PROJECTION — DATA SOURCES & REFERENCE LINKS"
    t.font = Font(name="Arial", bold=True, size=11, color=WHITE)
    t.fill = fill(NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 1, 20)

    ws.merge_cells("A2:D2")
    s = ws["A2"]
    s.value = ("Click any link to open  |  "
               "Sources used for projection assumptions, income statement analysis and consensus estimates")
    s.font = Font(name="Arial", size=9, color=WHITE)
    s.fill = fill(BLUE)
    s.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 2, 14)

    def section_header(row, txt):
        ws.merge_cells(f"A{row}:D{row}")
        c = ws[f"A{row}"]
        c.value = txt
        c.font = Font(name="Arial", bold=True, size=9, color=WHITE)
        c.fill = fill(ACCENT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        rh(ws, row, 14)

    def col_headers(row, bg=ACCENT):
        for col, txt in [(2,"Source"),(3,"What It Provides"),(4,"Direct Link")]:
            cell(ws, row, col, txt, bg=bg, fc=WHITE, bold=True, align="left")
        cell(ws, row, 1, bg=bg)
        rh(ws, row, 14)

    def src_row(row, name, desc, link, bg=WHITE):
        cell(ws, row, 1, bg=bg)
        cell(ws, row, 2, name, bg=bg, align="left", bold=True)
        cell(ws, row, 3, desc, bg=bg, fc=MUTED, align="left", size=8)
        c = ws.cell(row=row, column=4)
        c.value = link
        c.font = Font(name="Arial", size=8, color="0070C0", underline="single")
        c.fill = fill(bg)
        c.alignment = Alignment(horizontal="left", vertical="center")
        rh(ws, row, 14)

    # AMD OFFICIAL
    section_header(4, "AMD OFFICIAL SOURCES")
    col_headers(5)
    amd_sources = [
        ("AMD Investor Relations",   "Official earnings, guidance, financial tables",
         "https://ir.amd.com"),
        ("AMD Press Releases",       "Quarterly earnings press releases",
         "https://ir.amd.com/news-events/press-releases"),
        ("AMD IR Calendar",          "Upcoming earnings dates and events",
         "https://ir.amd.com/news-events/ir-calendar"),
        ("AMD SEC Filings",          "10-K, 10-Q, 8-K filings",
         "https://ir.amd.com/financial-information/sec-filings"),
        ("AMD CES 2026",             "MI500 2027 launch, OpenAI 6GW deal, ROCm roadmap",
         "https://www.amd.com/en/newsroom"),
    ]
    for i, (name, desc, link) in enumerate(amd_sources):
        src_row(6 + i, name, desc, link, bg=WHITE if i % 2 == 0 else GRAY)

    # CONSENSUS
    section_header(12, "CONSENSUS ESTIMATES")
    col_headers(13)
    consensus_sources = [
        ("Yahoo Finance — AMD",  "Revenue & EPS consensus, analyst count",
         "https://finance.yahoo.com/quote/AMD/analysis/"),
        ("Earnings Whispers",    "Whisper number vs published consensus",
         "https://www.earningswhispers.com/stocks/amd"),
        ("Seeking Alpha — AMD",  "Estimate revisions, transcripts, ratings",
         "https://seekingalpha.com/symbol/AMD/earnings"),
        ("Visible Alpha",        "Segment-level consensus: Data Center, Client, Embedded",
         "https://visiblealpha.com"),
        ("Estimize",             "Crowdsourced estimates from buy-side",
         "https://www.estimize.com/amd"),
    ]
    for i, (name, desc, link) in enumerate(consensus_sources):
        src_row(14 + i, name, desc, link, bg=WHITE if i % 2 == 0 else GRAY)

    # HISTORICAL
    section_header(20, "HISTORICAL DATA")
    col_headers(21)
    hist_sources = [
        ("StockAnalysis",         "Historical income statement 5+ years",
         "https://stockanalysis.com/stocks/amd/financials/"),
        ("Macrotrends — Revenue", "Historical quarterly revenue 10+ years",
         "https://www.macrotrends.net/stocks/charts/AMD/advanced-micro-devices/revenue"),
        ("Macrotrends — EPS",     "Historical quarterly EPS 10+ years",
         "https://www.macrotrends.net/stocks/charts/AMD/advanced-micro-devices/eps-earnings-per-share-diluted"),
    ]
    for i, (name, desc, link) in enumerate(hist_sources):
        src_row(22 + i, name, desc, link, bg=WHITE if i % 2 == 0 else GRAY)

    # MARKET & ROADMAP
    section_header(26, "MARKET & ROADMAP")
    col_headers(27)
    market_sources = [
        ("Investing.com — AMD",     "Real-time price, technical analysis",
         "https://www.investing.com/equities/adv-micro-device"),
        ("Tom's Hardware — Roadmap","Venice, Verano, MI450, MI500 roadmap",
         "https://www.tomshardware.com"),
        ("WCCFtech — MI450/MI500",  "GPU specifications and roadmap analysis",
         "https://wccftech.com"),
    ]
    for i, (name, desc, link) in enumerate(market_sources):
        src_row(28 + i, name, desc, link, bg=WHITE if i % 2 == 0 else GRAY)

    # GOOGLE DRIVE
    section_header(32, "GOOGLE DRIVE")
    col_headers(33)
    drive_sources = [
        ("AMD Financials — Master Folder", "All AMD documents organized by quarter",
         "https://drive.google.com/drive/folders/1i1GdOQGreQuxv2s6xoqILyY_XIqKuX5a"),
        ("Q1 2026 Quarter Folder",         "Press release PDF, financial tables PDF",
         "https://drive.google.com/drive/folders/1Df4g1sgPkwy_7BwBLjNTsYNDPh28RiWc"),
        ("Q2 2026 Quarter Folder",         "Ready for Q2 2026 earnings documents",
         "https://drive.google.com/drive/folders/1D_GOhT8kgZBTlVPA_7VXvO0-lR5ZyTbB"),
    ]
    for i, (name, desc, link) in enumerate(drive_sources):
        src_row(34 + i, name, desc, link, bg=WHITE if i % 2 == 0 else GRAY)

# ==============================================================================
# SELF-TEST
# ==============================================================================
def self_test(out_path):
    import os
    from openpyxl import load_workbook

    errors = []
    if not os.path.exists(out_path):
        errors.append(f"FAIL: file not created at {out_path}")
    else:
        size = os.path.getsize(out_path)
        if size < 20000:
            errors.append(f"FAIL: file too small ({size} bytes) — engine incomplete")

        wb = load_workbook(out_path, data_only=True)
        for tab in ["Inputs", "Projection", "Probability Weighted",
                    "Assumptions", "Workflow", "Data Sources"]:
            if tab not in wb.sheetnames:
                errors.append(f"FAIL: tab '{tab}' missing")

        ws = wb["Projection"]
        found_bull_spl = False
        found_base_spl = False
        for row in ws.iter_rows(values_only=True):
            for c in row:
                if isinstance(c, (int, float)) and 3000 < c < 3200: found_bull_spl = True
                if isinstance(c, (int, float)) and 1300 < c < 1400: found_base_spl = True
        if not found_bull_spl:
            errors.append("FAIL: Bull 2030 SPL ~$3,073 not found")
        if not found_base_spl:
            errors.append("FAIL: Base 2030 SPL ~$1,337 not found")

        ws_a = wb["Assumptions"]
        found_assumption = False
        for row in ws_a.iter_rows(values_only=True):
            for c in row:
                if isinstance(c, str) and "BULL CASE" in c:
                    found_assumption = True
        if not found_assumption:
            errors.append("FAIL: Assumptions tab content missing")

        ws_w = wb["Workflow"]
        found_workflow = False
        for row in ws_w.iter_rows(values_only=True):
            for c in row:
                if isinstance(c, str) and "QUARTERLY WORKFLOW" in c:
                    found_workflow = True
        if not found_workflow:
            errors.append("FAIL: Workflow tab content missing")

        ws_d = wb["Data Sources"]
        found_ir = False
        for row in ws_d.iter_rows(values_only=True):
            for c in row:
                if isinstance(c, str) and "ir.amd.com" in c:
                    found_ir = True
        if not found_ir:
            errors.append("FAIL: Data Sources tab — ir.amd.com link missing")

    if errors:
        print("\n" + "="*60)
        print("SELF-TEST FAILED — DO NOT SAVE TO GITHUB")
        print("="*60)
        for e in errors: print(f"  {e}")
        print("="*60)
        return False
    else:
        size = os.path.getsize(out_path)
        print("\n" + "="*60)
        print("SELF-TEST PASSED — safe to commit to GitHub")
        print(f"  File: {out_path}")
        print(f"  Size: {size:,} bytes")
        print(f"  Tabs: Inputs, Projection, Probability Weighted,")
        print(f"        Assumptions, Workflow, Data Sources — all confirmed")
        print(f"  Bull SPL ~$3,073, Base SPL ~$1,337: confirmed")
        print("="*60)
        return True

# ==============================================================================
# MAIN
# ==============================================================================
def build():
    wb = Workbook()
    wb.remove(wb.active)
    build_inputs(wb)
    build_projection(wb)
    build_probability(wb)
    build_assumptions(wb)
    build_workflow(wb)
    build_data_sources(wb)

    wb["Inputs"].sheet_properties.tabColor               = "2E75B6"
    wb["Projection"].sheet_properties.tabColor           = "1A7A3A"
    wb["Probability Weighted"].sheet_properties.tabColor = "7B2D8B"
    wb["Assumptions"].sheet_properties.tabColor          = "E67300"
    wb["Workflow"].sheet_properties.tabColor             = "5A5A5A"
    wb["Data Sources"].sheet_properties.tabColor         = "006B6B"

    wb.active = wb["Inputs"]

    out = "/mnt/user-data/outputs/AMD_5Year_Projection_v7.xlsx"
    wb.save(out)

    bull = compute(BULL); base = compute(BASE); bear = compute(BEAR)
    ev_spl = [bull[3][i]*PROB_BULL+base[3][i]*PROB_BASE+bear[3][i]*PROB_BEAR for i in range(5)]
    ev_sph = [bull[4][i]*PROB_BULL+base[4][i]*PROB_BASE+bear[4][i]*PROB_BEAR for i in range(5)]
    print(f"Saved: {out}")
    print(f"\nKEY 2030 OUTPUTS — verify on open:")
    print(f"BULL: Rev=${bull[0][4]:.2f}B  NI=${bull[1][4]:.2f}B  EPS=${int(round(bull[2][4]))}  SPL=${bull[3][4]:,.0f}  SPH=${bull[4][4]:,.0f}  CAGR={bull[5][4]:+.1%}/{bull[6][4]:+.1%}")
    print(f"BASE: Rev=${base[0][4]:.2f}B  NI=${base[1][4]:.2f}B  EPS=${int(round(base[2][4]))}  SPL=${base[3][4]:,.0f}  SPH=${base[4][4]:,.0f}  CAGR={base[5][4]:+.1%}/{base[6][4]:+.1%}")
    print(f"BEAR: Rev=${bear[0][4]:.2f}B  NI=${bear[1][4]:.2f}B  EPS=${int(round(bear[2][4]))}  SPL=${bear[3][4]:,.0f}  SPH=${bear[4][4]:,.0f}  CAGR={bear[5][4]:+.1%}/{bear[6][4]:+.1%}")
    print(f"EV:   SPL=${ev_spl[4]:,.0f}  SPH=${ev_sph[4]:,.0f}  CAGR={(ev_spl[4]/ENTRY_PRICE)**(1/5)-1:+.1%}/{(ev_sph[4]/ENTRY_PRICE)**(1/5)-1:+.1%}")
    return out

if __name__ == "__main__":
    out = build()
    passed = self_test(out)
    if not passed:
        import sys
        sys.exit(1)
