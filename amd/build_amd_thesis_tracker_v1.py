# ==============================================================================
# AMD THESIS TRACKER — v1.0
# Tracks whether the investment thesis is intact each quarter.
# Exit signal is driven by FUNDAMENTALS changing, not price movement.
# Fetches live data from GitHub API, Yahoo Finance, SEC EDGAR at build time.
# Falls back gracefully to DATA BLOCK manual values if fetch fails.
# HOW TO RUN: python build_amd_thesis_tracker_v1.py
# UPDATE: Edit DATA BLOCK each quarter after earnings. Re-run to rebuild.
# ==============================================================================
import urllib.request, json, math, re
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ── DATA BLOCK — update these values each quarter after earnings ───────────────
QUARTER    = "Q1 2026"
GENERATED  = "May 2026"
REPORT_DATE = "May 5, 2026"       # AMD earnings release date

# ── Financial actuals (from AMD earnings release) ─────────────────────────────
DC_REVENUE      = 5.8             # Data Center segment revenue ($B) this quarter
DC_GROWTH_YY    = 0.57            # Data Center Y/Y revenue growth (57%)
TOTAL_REVENUE   = 10.253          # Total AMD revenue ($B) this quarter
NI_MARGIN       = 0.221           # Non-GAAP net income margin (22.1%)
EPYC_SHARE      = 0.462           # Server CPU revenue share (46.2%)
GROSS_MARGIN    = 0.53            # Non-GAAP gross margin (53%)
AMD_PRICE       = 455.00          # AMD stock price at analysis date

# ── Base case targets (from build_amd_projection_engine_v7.py) ────────────────
# These are what the Base case implies for the FULL YEAR containing this quarter.
BASE_DC_REV_RUN_RATE = 5.5        # Base case quarterly DC revenue implied ($B)
BASE_NI_MARGIN_TARGET = 0.25      # Base case NI margin for FY2026
BASE_TOTAL_REV_FY     = 47.1      # Base case FY2026 total revenue ($B)

# ── Manual condition scores (1–10) — updated each quarter ─────────────────────
# 9–10: Exceeding expectations  7–8: On track  5–6: Slight concern
#  3–4: Yellow flag             1–2: Red flag
SCORE_OPENAI_DELIVERY   = 8   # 1GW H2 2026 MI450 deployment on track?
SCORE_META_DELIVERY     = 8   # 1GW H2 2026 MI450-based GPU on track?
SCORE_HUMAIN_DELIVERY   = 7   # 500MW sovereign AI — multi-exaflop by early 2026?
SCORE_NEW_DEAL_VELOCITY = 7   # New GW-scale deals announced this quarter?
SCORE_ROCM_PRODUCTION   = 6   # ROCm running actual production workloads at hyperscalers?
SCORE_COMPETITIVE       = 7   # AMD holding vs Nvidia? (10=gaining share, 5=holding, 1=losing)
SCORE_GUIDANCE_TONE     = 8   # Earnings call tone: 10=extremely bullish, 5=neutral, 1=negative

# ── Status narrative strings — updated each quarter ───────────────────────────
STATUS_OPENAI   = "1GW H2 2026 MI450 Series — on track per AMD Q1 2026 earnings call. Full 6GW ramp multi-year."
STATUS_META     = "1GW H2 2026 custom MI450-based GPU + EPYC Venice — on track. Full 6GW multi-year commitment."
STATUS_ORACLE   = "50K GPUs Q3 2026 (MI450 + EPYC Venice + Pensando Vulcano, Helios rack) — commitment confirmed Oct 14 2025."
STATUS_HUMAIN   = "500MW / $10B sovereign AI — multi-exaflop by early 2026. 5-year deployment. Full AMD stack confirmed."
STATUS_DOE      = "Lux AI (MI355X + EPYC, early 2026) + Discovery (MI430X MI400 Series + EPYC Venice, 2028). $1B combined."
STATUS_NEW_DEALS = "No new GW-scale deals announced Q1 2026. AMD investor day July 2026 — next catalyst window."
GUIDANCE_NOTE   = ("Q2 2026 guided $11.2B (+9.2% Q/Q). Non-GAAP op margin ~27%. Non-GAAP EPS ~$0.96. "
                   "Lisa Su: 'incredibly strong AI demand,' confident in H2 MI450 ramp. Language: bullish.")
ROCM_NOTE       = ("ROCm 6.x — PyTorch 2.x full support. ONNX Runtime AMD EP improving. "
                   "OpenAI/Meta production workloads not yet independently confirmed — watching. "
                   "MLCommons inference results improving Q/Q but CUDA gap remains ~15–20%.")
COMPETITIVE_NOTE = ("NVDA Q4 FY2026: DC $62.3B (+75% Y/Y). Blackwell ramp dominant. "
                    "AMD DC $5.8B vs Intel DC $5.1B — AMD now #2 in Data Center. "
                    "Dual-vendor GPU policies forming at hyperscalers — AMD beneficiary long term.")

# ── Exit / watch thresholds ───────────────────────────────────────────────────
EXIT_THRESHOLD  = 50    # Score below this for 2 consecutive quarters = review exit
WATCH_THRESHOLD = 65    # Score below this = watch closely, no new buys

# ── Score weights (must sum to 100) ───────────────────────────────────────────
W_OPENAI        = 12
W_META          = 10
W_HUMAIN_DOE    = 8
W_NEW_DEALS     = 10
W_ROCM          = 18    # highest weight — software moat is the key long-term risk
W_GITHUB        = 7     # auto-scored from fetched GitHub data
W_COMPETITIVE   = 8
W_EPYC_SHARE    = 8
W_NI_MARGIN     = 7
W_GUIDANCE      = 7
W_DC_REVENUE    = 5
# Total = 100

# ── COLORS ────────────────────────────────────────────────────────────────────
NAVY    = "1A1A2E"
BLUE    = "1F4E79"
ACCENT  = "2E75B6"
WHITE   = "FFFFFF"
GRAY    = "F5F7FA"
BLACK   = "000000"
MUTED   = "888888"
GREEN   = "1A7A3A"
GREEN_L = "E8F5E9"
AMBER   = "E67300"
AMBER_L = "FFF3E0"
RED_C   = "C0392B"
RED_L   = "FDEDEC"
TEAL    = "006B6B"
PURPLE  = "7B2D8B"

def fill(h):  return PatternFill("solid", start_color=h, fgColor=h)
def rh(ws, row, h=14): ws.row_dimensions[row].height = h

def cell(ws, row, col, val=None, bg=WHITE, fc=BLACK, bold=False,
         align="center", size=9, italic=False, wrap=False, indent=0):
    c = ws.cell(row=row, column=col)
    if val is not None: c.value = val
    c.font  = Font(name="Arial", bold=bold, size=size, color=fc, italic=italic)
    c.fill  = fill(bg)
    c.alignment = Alignment(horizontal=align, vertical="center",
                            wrap_text=wrap, indent=indent)
    return c

def score_color(score):
    if score >= 75: return GREEN,   GREEN_L
    if score >= 65: return "2E75B6", "D6E4F0"
    if score >= 50: return AMBER,   AMBER_L
    return RED_C, RED_L

def condition_color(s):
    if s >= 8: return GREEN,   GREEN_L
    if s >= 6: return "2E75B6", "D6E4F0"
    if s >= 4: return AMBER,   AMBER_L
    return RED_C, RED_L

def score_label(score):
    if score >= 80: return "STRONG HOLD — consider adding"
    if score >= 65: return "HOLD — thesis intact"
    if score >= 50: return "WATCH — do not add"
    if score >= 35: return "REVIEW — prepare exit thesis"
    return "EXIT — thesis broken"

# ==============================================================================
# LIVE DATA FETCH — runs at build time, graceful fallback on failure
# ==============================================================================
FETCH_TIMESTAMP = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
FETCHED = {}   # populated by fetch functions

def fetch_github_rocm():
    """Fetch ROCm repo stats from GitHub public API."""
    endpoints = {
        "rocm_core":    "https://api.github.com/repos/ROCm/ROCm",
        "rocm_pytorch": "https://api.github.com/repos/ROCm/pytorch",
    }
    results = {}
    for key, url in endpoints.items():
        try:
            req = urllib.request.Request(url, headers={
                "User-Agent": "AMD-thesis-tracker/1.0 (financial research)"})
            with urllib.request.urlopen(req, timeout=10) as r:
                d = json.loads(r.read())
            results[key] = {
                "stars":       d.get("stargazers_count", "N/A"),
                "forks":       d.get("forks_count", "N/A"),
                "open_issues": d.get("open_issues_count", "N/A"),
                "last_push":   d.get("pushed_at", "N/A")[:10] if d.get("pushed_at") else "N/A",
                "watchers":    d.get("watchers_count", "N/A"),
                "status":      "LIVE",
            }
        except Exception as e:
            results[key] = {"status": f"FETCH FAILED — {str(e)[:60]}. Update manually."}
    return results

def fetch_yahoo_price():
    """Fetch AMD current stock price and basic stats from Yahoo Finance."""
    try:
        url = "https://query1.finance.yahoo.com/v8/finance/chart/AMD?interval=1d&range=1d"
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=10) as r:
            d = json.loads(r.read())
        meta = d["chart"]["result"][0]["meta"]
        return {
            "price":      meta.get("regularMarketPrice", "N/A"),
            "prev_close": meta.get("previousClose", "N/A"),
            "52w_high":   meta.get("fiftyTwoWeekHigh", "N/A"),
            "52w_low":    meta.get("fiftyTwoWeekLow", "N/A"),
            "market_cap": meta.get("marketCap", "N/A"),
            "status":     "LIVE",
        }
    except Exception as e:
        return {"status": f"FETCH FAILED — {str(e)[:60]}. Update manually."}

def fetch_sec_edgar():
    """Fetch AMD's most recent SEC filing dates from EDGAR public API."""
    try:
        url = "https://data.sec.gov/submissions/CIK0000002488.json"
        req = urllib.request.Request(url, headers={
            "User-Agent": "AMD-thesis-tracker research@bmnr.local"})
        with urllib.request.urlopen(req, timeout=10) as r:
            d = json.loads(r.read())
        recent = d["filings"]["recent"]
        filings = list(zip(recent["form"], recent["filingDate"], recent["primaryDocument"]))
        key_types = ["10-K", "10-Q", "8-K"]
        latest = {}
        for form, date, doc in filings:
            if form in key_types and form not in latest:
                latest[form] = date
        return {"filings": latest, "status": "LIVE"}
    except Exception as e:
        return {"status": f"FETCH FAILED — {str(e)[:60]}. Update manually."}

def github_health_score(rocm_data):
    """Score GitHub ecosystem health 0-10 from fetched data."""
    core = rocm_data.get("rocm_core", {})
    if core.get("status", "").startswith("FETCH"):
        return 6, "Manual default — fetch failed"
    stars = core.get("stars", 0)
    issues = core.get("open_issues", 999)
    last_push = core.get("last_push", "2020-01-01")
    # Stars: >5000=3pts, >3000=2pts, >1000=1pt
    s = 3 if stars > 5000 else (2 if stars > 3000 else (1 if stars > 1000 else 0))
    # Open issues: <200=3pts, <500=2pts, <1000=1pt (lower = better maintained)
    i = 3 if issues < 200 else (2 if issues < 500 else (1 if issues < 1000 else 0))
    # Recency: pushed within 7 days = 4pts, 30 days = 3pts, 90 days = 2pts
    try:
        days = (datetime.now() - datetime.strptime(last_push, "%Y-%m-%d")).days
        p = 4 if days <= 7 else (3 if days <= 30 else (2 if days <= 90 else 1))
    except Exception:
        p = 2
    raw = s + i + p   # 0–10
    note = f"Stars {stars:,} | Open issues {issues:,} | Last push {last_push}"
    return min(10, raw), note

# ==============================================================================
# SCORING ENGINE
# ==============================================================================
def compute_score(github_score):
    # Financial auto-scores
    dc_score = min(10, max(1, round(10 * (DC_REVENUE / BASE_DC_REV_RUN_RATE))))
    ni_score  = min(10, max(1, round(10 * (NI_MARGIN  / BASE_NI_MARGIN_TARGET))))
    epyc_score = (10 if EPYC_SHARE >= 0.45 else
                   8 if EPYC_SHARE >= 0.40 else
                   5 if EPYC_SHARE >= 0.35 else 2)

    conditions = [
        ("OpenAI Delivery",          SCORE_OPENAI_DELIVERY,   W_OPENAI),
        ("Meta Delivery",            SCORE_META_DELIVERY,     W_META),
        ("HUMAIN + DoE Delivery",    SCORE_HUMAIN_DELIVERY,   W_HUMAIN_DOE),
        ("New Deal Velocity",        SCORE_NEW_DEAL_VELOCITY, W_NEW_DEALS),
        ("ROCm Production Adoption", SCORE_ROCM_PRODUCTION,   W_ROCM),
        ("GitHub Ecosystem Health",  github_score,            W_GITHUB),
        ("Competitive Position",     SCORE_COMPETITIVE,       W_COMPETITIVE),
        ("EPYC CPU Share",           epyc_score,              W_EPYC_SHARE),
        ("NI Margin Trajectory",     ni_score,                W_NI_MARGIN),
        ("Guidance Tone",            SCORE_GUIDANCE_TONE,     W_GUIDANCE),
        ("DC Revenue vs Base Case",  dc_score,                W_DC_REVENUE),
    ]
    total_w   = sum(c[2] for c in conditions)
    total_score = sum(c[1] * c[2] for c in conditions) / total_w
    return round(total_score * 10), conditions   # score 0–100

# ==============================================================================
# TAB 1 — THESIS SCORE (main dashboard)
# ==============================================================================
def build_thesis_score(wb, score, conditions, github_note):
    ws = wb.create_sheet("Thesis Score")
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 50

    sc_fg, sc_bg = score_color(score)

    # Header
    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = f"AMD INVESTMENT THESIS TRACKER — {QUARTER}"
    t.font = Font(name="Arial", bold=True, size=11, color=WHITE)
    t.fill = fill(NAVY); t.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 1, 22)

    ws.merge_cells("A2:D2")
    s = ws["A2"]
    s.value = (f"Built: {GENERATED} | Entry: ${AMD_PRICE} | "
               f"Exit threshold: {EXIT_THRESHOLD}/100 | Watch threshold: {WATCH_THRESHOLD}/100 | "
               f"Fetch timestamp: {FETCH_TIMESTAMP}")
    s.font = Font(name="Arial", size=8, color=WHITE, italic=True)
    s.fill = fill(BLUE); s.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 2, 13)

    # Score banner
    ws.merge_cells("A4:D4")
    sc = ws["A4"]
    sc.value = f"OVERALL THESIS SCORE:  {score} / 100  —  {score_label(score)}"
    sc.font = Font(name="Arial", bold=True, size=14, color=WHITE)
    sc.fill = fill(sc_fg)
    sc.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 4, 28)

    ws.merge_cells("A5:D5")
    thr = ws["A5"]
    thr.value = (f"HOLD above {WATCH_THRESHOLD}   |   "
                 f"WATCH {EXIT_THRESHOLD}–{WATCH_THRESHOLD}   |   "
                 f"REVIEW EXIT below {EXIT_THRESHOLD}   |   "
                 f"Score 2 consecutive quarters below {EXIT_THRESHOLD} = exit review")
    thr.font = Font(name="Arial", size=8, color=sc_fg, italic=True)
    thr.fill = fill(sc_bg)
    thr.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 5, 13)

    # Column headers
    for col, txt in [(2,"Condition"),(3,"Score / 10"),(4,"Weight")]:
        cell(ws, 7, col, txt, bg=ACCENT, fc=WHITE, bold=True, align="left" if col==2 else "center")
    cell(ws, 7, 1, bg=ACCENT); rh(ws, 7, 14)

    # Condition rows
    for i, (label, score_val, weight) in enumerate(conditions):
        r = 8 + i
        bg = WHITE if i % 2 == 0 else GRAY
        fg, fbg = condition_color(score_val)
        cell(ws, r, 1, bg=bg)
        cell(ws, r, 2, label, bg=bg, align="left")
        ws.merge_cells(f"C{r}:D{r}")
        bar = "█" * score_val + "░" * (10 - score_val)
        c = ws.cell(row=r, column=3)
        c.value = f"{score_val}/10  {bar}"
        c.font = Font(name="Arial", bold=True, size=9, color=fg)
        c.fill = fill(fbg)
        c.alignment = Alignment(horizontal="left", vertical="center")
        rh(ws, r, 15)

    ROW = 8 + len(conditions) + 1

    # Bucket sub-totals
    ws.merge_cells(f"A{ROW}:D{ROW}")
    bh = ws[f"A{ROW}"]
    bh.value = "BUCKET BREAKDOWN"
    bh.font = Font(name="Arial", bold=True, size=9, color=WHITE)
    bh.fill = fill(ACCENT); bh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    rh(ws, ROW, 14); ROW += 1

    deal_conds  = [c for c in conditions if c[0] in ("OpenAI Delivery","Meta Delivery","HUMAIN + DoE Delivery","New Deal Velocity")]
    moat_conds  = [c for c in conditions if c[0] in ("ROCm Production Adoption","GitHub Ecosystem Health","Competitive Position")]
    fin_conds   = [c for c in conditions if c[0] in ("EPYC CPU Share","NI Margin Trajectory","Guidance Tone","DC Revenue vs Base Case")]

    for bucket_name, conds in [("Deal Execution", deal_conds),
                                ("Software Moat",  moat_conds),
                                ("Financial Guardrails", fin_conds)]:
        tw = sum(c[2] for c in conds)
        sc_b = round(sum(c[1]*c[2] for c in conds) / tw * 10) if tw else 0
        fg, fbg = score_color(sc_b)
        cell(ws, ROW, 1, bg=fbg)
        cell(ws, ROW, 2, bucket_name, bg=fbg, fc=fg, bold=True, align="left")
        ws.merge_cells(f"C{ROW}:D{ROW}")
        bar = "█" * (sc_b//10) + "░" * (10 - sc_b//10)
        c = ws.cell(row=ROW, column=3)
        c.value = f"{sc_b}/100  {bar}  (weight {tw}%)"
        c.font = Font(name="Arial", bold=True, size=9, color=fg)
        c.fill = fill(fbg); c.alignment = Alignment(horizontal="left", vertical="center")
        rh(ws, ROW, 15); ROW += 1

    ROW += 1

    # GitHub live data note
    ws.merge_cells(f"A{ROW}:D{ROW}")
    gn = ws[f"A{ROW}"]
    gn.value = f"GitHub Ecosystem Health data ({FETCH_TIMESTAMP}): {github_note}"
    gn.font = Font(name="Arial", size=8, color=MUTED, italic=True)
    gn.alignment = Alignment(horizontal="left", vertical="center")
    rh(ws, ROW, 12); ROW += 1

    # Footer
    ws.merge_cells(f"A{ROW}:D{ROW}")
    ft = ws[f"A{ROW}"]
    ft.value = (f"Script: build_amd_thesis_tracker_v1.py | "
                f"Projection reference: build_amd_projection_engine_v7.py | "
                f"Run quarterly after earnings to refresh score")
    ft.font = Font(name="Arial", size=7, color=MUTED, italic=True)
    ft.alignment = Alignment(horizontal="left")

# ==============================================================================
# TAB 2 — DEAL EXECUTION
# ==============================================================================
def build_deal_execution(wb):
    ws = wb.create_sheet("Deal Execution")
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 65

    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = f"AMD CONFIRMED DEALS — EXECUTION TRACKER  [{QUARTER}]"
    t.font = Font(name="Arial", bold=True, size=11, color=WHITE)
    t.fill = fill(NAVY); t.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 1, 20)

    ws.merge_cells("A2:D2")
    s = ws["A2"]
    s.value = ("Score each deal 1–10 each quarter. Red = execution broken. "
               "Multiple red = demand side of thesis failing.")
    s.font = Font(name="Arial", size=9, color=WHITE)
    s.fill = fill(BLUE); s.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 2, 13)

    # Header row
    for col, txt in [(2,"Deal"),(3,"Score"),(4,"Status")]:
        cell(ws, 4, col, txt, bg=ACCENT, fc=WHITE, bold=True, align="left" if col!=3 else "center")
    cell(ws, 4, 1, bg=ACCENT); rh(ws, 4, 14)

    deals = [
        ("HUMAIN $10B / 500MW Sovereign AI",       SCORE_HUMAIN_DELIVERY,   STATUS_HUMAIN),
        ("OpenAI 6GW — MI450 Series",              SCORE_OPENAI_DELIVERY,   STATUS_OPENAI),
        ("Meta 6GW — custom MI450-based GPU",      SCORE_META_DELIVERY,     STATUS_META),
        ("Oracle 50K GPUs — MI450 + EPYC Venice",  8,                       STATUS_ORACLE),
        ("DoE/ORNL — Lux AI + Discovery",          SCORE_HUMAIN_DELIVERY,   STATUS_DOE),
        ("New Deals This Quarter",                 SCORE_NEW_DEAL_VELOCITY, STATUS_NEW_DEALS),
    ]

    for i, (name, score_val, status) in enumerate(deals):
        r = 5 + i
        bg = WHITE if i % 2 == 0 else GRAY
        fg, fbg = condition_color(score_val)
        cell(ws, r, 1, bg=bg)
        cell(ws, r, 2, name, bg=bg, bold=True, align="left")
        c3 = ws.cell(row=r, column=3)
        c3.value = f"{score_val}/10"
        c3.font = Font(name="Arial", bold=True, size=10, color=fg)
        c3.fill = fill(fbg); c3.alignment = Alignment(horizontal="center", vertical="center")
        c4 = ws.cell(row=r, column=4)
        c4.value = status
        c4.font = Font(name="Arial", size=9, color=BLACK)
        c4.fill = fill(bg); c4.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        rh(ws, r, 45)

    r = 5 + len(deals) + 1
    ws.merge_cells(f"A{r}:D{r}")
    gd = ws[f"A{r}"]
    gd.value = GUIDANCE_NOTE
    gd.font = Font(name="Arial", size=9, color=BLACK, italic=True)
    gd.fill = fill(GRAY); gd.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    rh(ws, r, 40)

    r += 2
    ws.merge_cells(f"A{r}:D{r}")
    ex = ws[f"A{r}"]
    ex.value = ("EXIT SIGNAL: If OpenAI OR Meta delivery score drops to 3 or below for 2 consecutive "
                "quarters, the primary revenue thesis is impaired. Review exit regardless of overall score.")
    ex.font = Font(name="Arial", size=9, color=RED_C, bold=True)
    ex.fill = fill(RED_L); ex.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    rh(ws, r, 30)

# ==============================================================================
# TAB 3 — SOFTWARE MOAT
# ==============================================================================
def build_software_moat(wb, github_data, github_score, github_note):
    ws = wb.create_sheet("Software Moat")
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 50

    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = f"AMD SOFTWARE MOAT TRACKER — ROCm Ecosystem  [{QUARTER}]"
    t.font = Font(name="Arial", bold=True, size=11, color=WHITE)
    t.fill = fill(NAVY); t.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 1, 20)

    ws.merge_cells("A2:D2")
    s = ws["A2"]
    s.value = ("ROCm parity with CUDA is the single biggest long-term risk to the AMD thesis. "
               "Track this more carefully than any financial metric.")
    s.font = Font(name="Arial", size=9, color=WHITE, italic=True)
    s.fill = fill(BLUE); s.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 2, 14)

    def section_hdr(row, txt):
        ws.merge_cells(f"A{row}:D{row}")
        c = ws[f"A{row}"]
        c.value = txt
        c.font = Font(name="Arial", bold=True, size=9, color=WHITE)
        c.fill = fill(ACCENT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        rh(ws, row, 14)

    # Assessment
    section_hdr(4, "MANUAL ASSESSMENT")
    for col, txt in [(2,"Metric"),(3,"Score"),(4,"Notes")]:
        cell(ws, 5, col, txt, bg=ACCENT, fc=WHITE, bold=True, align="left" if col!=3 else "center")
    cell(ws, 5, 1, bg=ACCENT); rh(ws, 5, 14)

    moat_rows = [
        ("ROCm Production Adoption", SCORE_ROCM_PRODUCTION, ROCM_NOTE),
        ("Competitive Position vs Nvidia", SCORE_COMPETITIVE, COMPETITIVE_NOTE),
        ("GitHub Ecosystem Health (auto)", github_score,
         f"Auto-scored from live GitHub data. {github_note}"),
    ]
    for i, (label, sc, note) in enumerate(moat_rows):
        r = 6 + i
        bg = WHITE if i % 2 == 0 else GRAY
        fg, fbg = condition_color(sc)
        cell(ws, r, 1, bg=bg)
        cell(ws, r, 2, label, bg=bg, bold=True, align="left")
        c3 = ws.cell(row=r, column=3)
        c3.value = f"{sc}/10"
        c3.font = Font(name="Arial", bold=True, size=10, color=fg)
        c3.fill = fill(fbg); c3.alignment = Alignment(horizontal="center", vertical="center")
        c4 = ws.cell(row=r, column=4)
        c4.value = note
        c4.font = Font(name="Arial", size=9); c4.fill = fill(bg)
        c4.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        rh(ws, r, 50)

    # GitHub live data
    section_hdr(10, f"GITHUB LIVE DATA — fetched {FETCH_TIMESTAMP}")
    for col, txt in [(2,"Repository"),(3,"Metric"),(4,"Value")]:
        cell(ws, 11, col, txt, bg=ACCENT, fc=WHITE, bold=True, align="left")
    cell(ws, 11, 1, bg=ACCENT); rh(ws, 11, 14)

    gh_display = []
    for repo_key, repo_label in [("rocm_core","ROCm/ROCm (core)"),
                                  ("rocm_pytorch","ROCm/pytorch")]:
        data = github_data.get(repo_key, {})
        if data.get("status", "").startswith("FETCH"):
            gh_display.append((repo_label, "Status", data["status"]))
        else:
            gh_display.append((repo_label, "Stars",       f"{data.get('stars','N/A'):,}" if isinstance(data.get('stars'), int) else str(data.get('stars','N/A'))))
            gh_display.append((repo_label, "Forks",       str(data.get('forks','N/A'))))
            gh_display.append((repo_label, "Open Issues", str(data.get('open_issues','N/A'))))
            gh_display.append((repo_label, "Last Push",   str(data.get('last_push','N/A'))))
            gh_display.append((repo_label, "Watchers",    str(data.get('watchers','N/A'))))

    for i, (repo, metric, val) in enumerate(gh_display):
        r = 12 + i
        bg = WHITE if i % 2 == 0 else GRAY
        cell(ws, r, 1, bg=bg)
        cell(ws, r, 2, repo, bg=bg, align="left")
        cell(ws, r, 3, metric, bg=bg, align="left")
        cell(ws, r, 4, val, bg=bg, align="left")
        rh(ws, r, 14)

    r = 12 + len(gh_display) + 2
    section_hdr(r, "WHAT TO WATCH — ROCm Signals")
    r += 1
    watch_items = [
        ("MLCommons Inference Results",
         "Published quarterly. AMD vs Nvidia on MLPerf Inference — watch the gap narrow.",
         "https://mlcommons.org/benchmarks/inference-datacenter/"),
        ("PyTorch AMD Support",
         "ROCm/pytorch GitHub — open issues, PR merge rate, release notes.",
         "https://github.com/ROCm/pytorch"),
        ("Hugging Face AMD Models",
         "Search 'device_map AMD' or 'ROCm' — count of models listing AMD as tested device.",
         "https://huggingface.co/models"),
        ("OpenAI/Meta Job Postings",
         "If they're hiring ROCm engineers, production workloads are moving. Search LinkedIn quarterly.",
         "https://www.linkedin.com/jobs/search/?keywords=ROCm"),
        ("NeurIPS / Hot Chips Conference",
         "AMD ROCm talk presence signals software ecosystem maturity.",
         "https://hotchips.org"),
    ]
    for col, txt in [(2,"Source"),(3,"What to Look For"),(4,"Link")]:
        cell(ws, r, col, txt, bg=ACCENT, fc=WHITE, bold=True, align="left")
    cell(ws, r, 1, bg=ACCENT); rh(ws, r, 14); r += 1

    for i, (src, look, link) in enumerate(watch_items):
        bg = WHITE if i % 2 == 0 else GRAY
        cell(ws, r, 1, bg=bg)
        cell(ws, r, 2, src, bg=bg, bold=True, align="left")
        c = ws.cell(row=r, column=3); c.value = look
        c.font = Font(name="Arial", size=8); c.fill = fill(bg)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        lc = ws.cell(row=r, column=4); lc.value = link
        lc.font = Font(name="Arial", size=8, color="0070C0", underline="single")
        lc.fill = fill(bg); lc.alignment = Alignment(horizontal="left", vertical="center")
        rh(ws, r, 30); r += 1

# ==============================================================================
# TAB 4 — FINANCIAL GUARDRAILS
# ==============================================================================
def build_financial_guardrails(wb):
    ws = wb.create_sheet("Financial Guardrails")
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 22

    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value = f"AMD FINANCIAL GUARDRAILS — {QUARTER} vs Projection Engine Base Case"
    t.font = Font(name="Arial", bold=True, size=11, color=WHITE)
    t.fill = fill(NAVY); t.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 1, 20)

    ws.merge_cells("A2:E2")
    s = ws["A2"]
    s.value = ("Base case targets from build_amd_projection_engine_v7.py. "
               "If actuals fall below Base for 2+ consecutive quarters, thesis is weakening.")
    s.font = Font(name="Arial", size=9, color=WHITE, italic=True)
    s.fill = fill(BLUE); s.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 2, 13)

    for col, txt in [(2,"Metric"),(3,"Base Case Target"),(4,"Actual"),(5,"Status")]:
        cell(ws, 4, col, txt, bg=ACCENT, fc=WHITE, bold=True, align="left" if col==2 else "center")
    cell(ws, 4, 1, bg=ACCENT); rh(ws, 4, 14)

    dc_vs = DC_REVENUE / BASE_DC_REV_RUN_RATE
    ni_vs = NI_MARGIN  / BASE_NI_MARGIN_TARGET
    rev_run = TOTAL_REVENUE * 4  # annualized

    def status_str(ratio):
        if ratio >= 1.05: return "✓ BEATING"
        if ratio >= 0.92: return "✓ ON TRACK"
        if ratio >= 0.80: return "⚠ WATCH"
        return "✗ BELOW TARGET"

    def status_clr(ratio):
        if ratio >= 0.92: return GREEN, GREEN_L
        if ratio >= 0.80: return AMBER, AMBER_L
        return RED_C, RED_L

    guardrails = [
        ("Data Center Revenue (quarterly)",
         f"${BASE_DC_REV_RUN_RATE:.1f}B",
         f"${DC_REVENUE:.1f}B",
         dc_vs),
        ("Non-GAAP NI Margin (full year target)",
         f"{BASE_NI_MARGIN_TARGET:.1%}",
         f"{NI_MARGIN:.1%}",
         ni_vs),
        ("EPYC Server CPU Revenue Share",
         "≥ 40%",
         f"{EPYC_SHARE:.1%}",
         EPYC_SHARE / 0.40),
        ("Non-GAAP Gross Margin",
         "≥ 52%",
         f"{GROSS_MARGIN:.1%}",
         GROSS_MARGIN / 0.52),
        ("Total Revenue annualized run rate",
         f"${BASE_TOTAL_REV_FY:.1f}B FY target",
         f"${rev_run:.1f}B annualized",
         rev_run / BASE_TOTAL_REV_FY),
        ("DC Revenue Y/Y Growth",
         "≥ 36% (Base case floor)",
         f"{DC_GROWTH_YY:.1%}",
         DC_GROWTH_YY / 0.36),
    ]

    for i, (metric, target, actual, ratio) in enumerate(guardrails):
        r = 5 + i
        bg = WHITE if i % 2 == 0 else GRAY
        fg, fbg = status_clr(ratio)
        cell(ws, r, 1, bg=bg)
        cell(ws, r, 2, metric, bg=bg, align="left")
        cell(ws, r, 3, target, bg=bg, align="center")
        cell(ws, r, 4, actual, bg=bg, align="center", bold=True)
        c5 = ws.cell(row=r, column=5)
        c5.value = status_str(ratio)
        c5.font = Font(name="Arial", bold=True, size=9, color=fg)
        c5.fill = fill(fbg); c5.alignment = Alignment(horizontal="center", vertical="center")
        rh(ws, r, 15)

    r = 5 + len(guardrails) + 2
    ws.merge_cells(f"A{r}:E{r}")
    ex = ws[f"A{r}"]
    ex.value = ("EXIT SIGNAL: If Data Center revenue falls below Base case AND NI margin compresses "
                "below 22% for 2 consecutive quarters without a clear one-time explanation, "
                "the financial thesis is impaired. Investigate before the market reacts.")
    ex.font = Font(name="Arial", size=9, color=RED_C, bold=True)
    ex.fill = fill(RED_L); ex.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    rh(ws, r, 35)

# ==============================================================================
# TAB 5 — LIVE DATA SNAPSHOT
# ==============================================================================
def build_live_snapshot(wb, github_data, yahoo_data, sec_data):
    ws = wb.create_sheet("Live Data")
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 36

    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = f"AMD LIVE DATA SNAPSHOT — fetched {FETCH_TIMESTAMP}"
    t.font = Font(name="Arial", bold=True, size=11, color=WHITE)
    t.fill = fill(NAVY); t.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 1, 20)

    ws.merge_cells("A2:D2")
    s = ws["A2"]
    s.value = ("Data fetched from public APIs at script run time. "
               "FETCH FAILED = network unavailable or source blocked; update manually.")
    s.font = Font(name="Arial", size=9, color=WHITE, italic=True)
    s.fill = fill(BLUE); s.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 2, 13)

    def section_hdr(row, txt, url=""):
        ws.merge_cells(f"A{row}:D{row}")
        c = ws[f"A{row}"]
        c.value = txt + (f"  |  {url}" if url else "")
        c.font = Font(name="Arial", bold=True, size=9, color=WHITE)
        c.fill = fill(ACCENT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        rh(ws, row, 14)

    def data_row(r, source, metric, value, bg=WHITE):
        cell(ws, r, 1, bg=bg)
        cell(ws, r, 2, source, bg=bg, align="left")
        cell(ws, r, 3, metric, bg=bg, align="left")
        c = ws.cell(row=r, column=4)
        c.value = str(value)
        c.font = Font(name="Arial", size=9)
        c.fill = fill(bg)
        c.alignment = Alignment(horizontal="left", vertical="center")
        rh(ws, r, 14)

    ROW = 4

    # Yahoo Finance
    section_hdr(ROW, "YAHOO FINANCE — AMD Stock",
                "https://finance.yahoo.com/quote/AMD"); ROW += 1
    yf_rows = [
        ("Yahoo Finance", "Current Price",    yahoo_data.get("price", "FETCH FAILED")),
        ("Yahoo Finance", "Previous Close",   yahoo_data.get("prev_close", "N/A")),
        ("Yahoo Finance", "52-Week High",     yahoo_data.get("52w_high", "N/A")),
        ("Yahoo Finance", "52-Week Low",      yahoo_data.get("52w_low", "N/A")),
        ("Yahoo Finance", "Market Cap",       yahoo_data.get("market_cap", "N/A")),
        ("Yahoo Finance", "Fetch Status",     yahoo_data.get("status", "unknown")),
    ]
    for i, (src, met, val) in enumerate(yf_rows):
        data_row(ROW, src, met, val, WHITE if i%2==0 else GRAY); ROW += 1

    ROW += 1
    # GitHub ROCm
    section_hdr(ROW, "GITHUB — ROCm Ecosystem",
                "https://github.com/ROCm"); ROW += 1
    for repo_key, repo_label in [("rocm_core","ROCm/ROCm"),("rocm_pytorch","ROCm/pytorch")]:
        data = github_data.get(repo_key, {})
        if data.get("status","").startswith("FETCH"):
            data_row(ROW, repo_label, "Status", data["status"]); ROW += 1
        else:
            for met, val in [("Stars", data.get("stars","N/A")),
                             ("Forks", data.get("forks","N/A")),
                             ("Open Issues", data.get("open_issues","N/A")),
                             ("Last Push", data.get("last_push","N/A")),
                             ("Watchers", data.get("watchers","N/A"))]:
                i_offset = list(["Stars","Forks","Open Issues","Last Push","Watchers"]).index(met)
                data_row(ROW, repo_label, met, val, WHITE if i_offset%2==0 else GRAY); ROW += 1

    ROW += 1
    # SEC EDGAR
    section_hdr(ROW, "SEC EDGAR — AMD Filings",
                "https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&CIK=0000002488"); ROW += 1
    filings = sec_data.get("filings", {})
    if filings:
        for i, (form, date) in enumerate(filings.items()):
            data_row(ROW, "SEC EDGAR", f"Latest {form}", date, WHITE if i%2==0 else GRAY); ROW += 1
    else:
        data_row(ROW, "SEC EDGAR", "Status", sec_data.get("status","FETCH FAILED")); ROW += 1

    ROW += 2
    ws.merge_cells(f"A{ROW}:D{ROW}")
    ft = ws[f"A{ROW}"]
    ft.value = ("Sources requiring manual update: SemiAnalysis (paywalled), "
                "hyperscaler earnings call transcripts (Seeking Alpha), "
                "MLCommons benchmarks (mlcommons.org/benchmarks). "
                "Check these quarterly and update DATA BLOCK scores accordingly.")
    ft.font = Font(name="Arial", size=8, color=MUTED, italic=True)
    ft.fill = fill(GRAY)
    ft.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    rh(ws, ROW, 35)

# ==============================================================================
# TAB 6 — DATA SOURCES
# ==============================================================================
def build_data_sources(wb):
    ws = wb.create_sheet("Data Sources")
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 46

    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = "AMD THESIS TRACKER — DATA SOURCES & MONITORING LINKS"
    t.font = Font(name="Arial", bold=True, size=11, color=WHITE)
    t.fill = fill(NAVY); t.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 1, 20)

    ws.merge_cells("A2:D2")
    s = ws["A2"]
    s.value = "Check each source quarterly. BOLD = highest priority signal."
    s.font = Font(name="Arial", size=9, color=WHITE)
    s.fill = fill(BLUE); s.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 2, 13)

    def section_hdr(row, txt):
        ws.merge_cells(f"A{row}:D{row}")
        c = ws[f"A{row}"]
        c.value = txt
        c.font = Font(name="Arial", bold=True, size=9, color=WHITE)
        c.fill = fill(ACCENT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        rh(ws, row, 14)

    def col_hdr(row):
        for col, txt in [(2,"Source"),(3,"Signal to Watch"),(4,"Link")]:
            cell(ws, row, col, txt, bg=ACCENT, fc=WHITE, bold=True, align="left")
        cell(ws, row, 1, bg=ACCENT); rh(ws, row, 14)

    def src_row(row, name, signal, link, bg=WHITE, bold=False):
        cell(ws, row, 1, bg=bg)
        cell(ws, row, 2, name, bg=bg, bold=bold, align="left")
        c = ws.cell(row=row, column=3); c.value = signal
        c.font = Font(name="Arial", size=8, bold=bold); c.fill = fill(bg)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        lc = ws.cell(row=row, column=4); lc.value = link
        lc.font = Font(name="Arial", size=8, color="0070C0", underline="single")
        lc.fill = fill(bg); lc.alignment = Alignment(horizontal="left", vertical="center")
        rh(ws, row, 28)

    ROW = 4

    section_hdr(ROW, "EARNINGS & OFFICIAL — Auto-fetched where possible"); ROW += 1
    col_hdr(ROW); ROW += 1
    earnings = [
        ("AMD Investor Relations", "Quarterly earnings, guidance, financial tables", "https://ir.amd.com", True),
        ("AMD Earnings Transcripts", "Lisa Su language on MI450/ROCm/deal ramp", "https://seekingalpha.com/symbol/AMD/earnings", True),
        ("SEC EDGAR — AMD", "10-K, 10-Q, 8-K filings (auto-fetched)", "https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&CIK=0000002488", False),
        ("AMD Newsroom", "New deal announcements, product launches", "https://www.amd.com/en/newsroom", True),
        ("Yahoo Finance — AMD", "Price, market cap, analyst ratings (auto-fetched)", "https://finance.yahoo.com/quote/AMD", False),
    ]
    for i, (n, s, l, b) in enumerate(earnings):
        src_row(ROW, n, s, l, WHITE if i%2==0 else GRAY, b); ROW += 1

    ROW += 1
    section_hdr(ROW, "SOFTWARE MOAT — Most Critical to Track"); ROW += 1
    col_hdr(ROW); ROW += 1
    moat = [
        ("MLCommons Inference Benchmarks", "AMD vs Nvidia gap narrowing? Published quarterly.", "https://mlcommons.org/benchmarks/inference-datacenter/", True),
        ("ROCm/ROCm GitHub", "Stars, issues, last push (auto-fetched)", "https://github.com/ROCm/ROCm", True),
        ("ROCm/pytorch GitHub", "PyTorch AMD support health (auto-fetched)", "https://github.com/ROCm/pytorch", True),
        ("Hugging Face Models", "Count models listing ROCm/AMD as tested — leading indicator", "https://huggingface.co/models?library=pytorch&sort=trending", False),
        ("SemiAnalysis", "GPU market share, deployment data — best independent source", "https://semianalysis.com", True),
        ("Hot Chips / NeurIPS", "AMD ROCm talk presence = ecosystem maturity signal", "https://hotchips.org", False),
    ]
    for i, (n, s, l, b) in enumerate(moat):
        src_row(ROW, n, s, l, WHITE if i%2==0 else GRAY, b); ROW += 1

    ROW += 1
    section_hdr(ROW, "COMPETITIVE INTELLIGENCE"); ROW += 1
    col_hdr(ROW); ROW += 1
    comp = [
        ("Nvidia Earnings", "If Nvidia DC accelerates AND AMD DC is flat = AMD losing share", "https://investor.nvidia.com/financial-information/quarterly-results/default.aspx", True),
        ("Hyperscaler Earnings Calls", "OpenAI/Meta/Oracle/Microsoft mention AMD by name = deployments live", "https://seekingalpha.com", True),
        ("Intel DCAI Earnings", "Intel recovering? AMD EPYC share reverting?", "https://www.intc.com/financial-information/quarterly-earnings", False),
        ("TSMC Earnings", "N3E packaging capacity commentary = AMD supply chain signal", "https://investor.tsmc.com/english", False),
        ("SK Hynix / Samsung HBM4", "HBM4 supply allocation = MI450 production ceiling", "https://news.samsung.com/global/category/business", False),
    ]
    for i, (n, s, l, b) in enumerate(comp):
        src_row(ROW, n, s, l, WHITE if i%2==0 else GRAY, b); ROW += 1

    ROW += 1
    section_hdr(ROW, "MACRO & POLICY"); ROW += 1
    col_hdr(ROW); ROW += 1
    macro = [
        ("Hyperscaler AI Capex Guidance", "Microsoft/Amazon/Google/Meta capex cuts = AMD demand floor weakens", "https://ir.aboutamazon.com/quarterly-results/default.aspx", True),
        ("US Export Controls", "BIS chip export rules — affects HUMAIN (Saudi) and any new sovereign deals", "https://www.bis.gov/topic/export-administration-regulations", True),
        ("Taiwan Strait / TSMC Risk", "Geopolitical escalation = supply chain risk to all AMD products", "https://www.reuters.com/technology/semiconductors/", False),
    ]
    for i, (n, s, l, b) in enumerate(macro):
        src_row(ROW, n, s, l, WHITE if i%2==0 else GRAY, b); ROW += 1

# ==============================================================================
# TAB 7 — WORKFLOW
# ==============================================================================
def build_workflow(wb, score):
    ws = wb.create_sheet("Workflow")
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 78

    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value = "AMD THESIS TRACKER — QUARTERLY WORKFLOW"
    t.font = Font(name="Arial", bold=True, size=11, color=WHITE)
    t.fill = fill(NAVY); t.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 1, 20)

    ws.merge_cells("A2:C2")
    s = ws["A2"]
    s.value = "Run this script once per quarter, within 48 hours of AMD earnings. Takes ~2 minutes."
    s.font = Font(name="Arial", size=9, color=WHITE)
    s.fill = fill(BLUE); s.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 2, 13)

    def section_hdr(row, txt):
        ws.merge_cells(f"A{row}:C{row}")
        c = ws[f"A{row}"]
        c.value = txt
        c.font = Font(name="Arial", bold=True, size=9, color=WHITE)
        c.fill = fill(ACCENT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        rh(ws, row, 14)

    def wf_row(row, label, detail, bg=WHITE):
        cell(ws, row, 1, bg=bg)
        cell(ws, row, 2, label, bg=bg, bold=True, align="left")
        c = ws.cell(row=row, column=3); c.value = detail
        c.font = Font(name="Arial", size=9); c.fill = fill(bg)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        rh(ws, row, 30)

    section_hdr(4, "QUARTERLY UPDATE STEPS")
    steps = [
        ("Step 1 — Earnings day",
         "Read AMD earnings press release (ir.amd.com). Update DATA BLOCK: DC_REVENUE, DC_GROWTH_YY, "
         "TOTAL_REVENUE, NI_MARGIN, EPYC_SHARE, GROSS_MARGIN, AMD_PRICE, QUARTER, REPORT_DATE."),
        ("Step 2 — Earnings call",
         "Read/listen to earnings call transcript. Update: SCORE_GUIDANCE_TONE, GUIDANCE_NOTE. "
         "Note any new deal announcements. Update SCORE_NEW_DEAL_VELOCITY, STATUS_NEW_DEALS."),
        ("Step 3 — Deal check",
         "Check AMD Newsroom for deployment updates. Update STATUS_OPENAI, STATUS_META, "
         "STATUS_ORACLE, STATUS_HUMAIN, STATUS_DOE. Score each delivery condition."),
        ("Step 4 — ROCm check",
         "Check MLCommons latest results. Check ROCm/pytorch GitHub for recent commits. "
         "Read SemiAnalysis for any GPU market share updates. Update SCORE_ROCM_PRODUCTION, ROCM_NOTE."),
        ("Step 5 — Competitive check",
         "Review Nvidia latest earnings. Check if hyperscalers mentioned AMD in their calls. "
         "Update SCORE_COMPETITIVE, COMPETITIVE_NOTE."),
        ("Step 6 — Run script",
         "python build_amd_thesis_tracker_v1.py — script fetches live data (GitHub, Yahoo, SEC EDGAR) "
         "and builds fresh Excel. SELF-TEST must print SELF-TEST PASSED before committing."),
        ("Step 7 — Review score",
         f"Current score: {score}/100. Compare to prior quarter. Trend matters more than one reading. "
         f"Below {EXIT_THRESHOLD} for 2 consecutive quarters = exit review. "
         f"Below {WATCH_THRESHOLD} = no new purchases."),
        ("Step 8 — Commit",
         "Commit new version to GitHub: build_amd_thesis_tracker_v2.py etc. "
         "Output: AMD_Thesis_Tracker_v2.xlsx etc. Never overwrite previous versions."),
    ]
    for i, (label, detail) in enumerate(steps):
        wf_row(5 + i, label, detail, WHITE if i%2==0 else GRAY)

    section_hdr(14, "EXIT DECISION FRAMEWORK")
    exits = [
        ("Single quarter trigger",
         "ROCm score drops to ≤3 (production workloads explicitly NOT moving to AMD per hyperscaler call). "
         "Immediate review — this is the thesis-breaking scenario."),
        ("Two quarter trigger",
         f"Overall score below {EXIT_THRESHOLD} for 2 consecutive quarters. "
         "OR: OpenAI + Meta both below 4/10 for one quarter. Deal execution broken."),
        ("NOT a trigger",
         "Stock price decline (hold on fundamentals, not price). One weak quarter without pattern. "
         "Nvidia continuing to grow (AMD can grow alongside Nvidia in a large TAM)."),
        ("Price target discipline",
         "If AMD reaches Bull 2030 SPL ($2,897) by 2028, consider trimming. "
         "The model priced in strong execution — at that point re-underwrite the next 5 years."),
    ]
    for i, (label, detail) in enumerate(exits):
        wf_row(15 + i, label, detail, WHITE if i%2==0 else GRAY)

    section_hdr(20, "VERSION LOG")
    versions = [
        ("v1 — May 2026",
         f"Initial build. Thesis Score: {score}/100. 11 conditions across Deal Execution, "
         "Software Moat, Financial Guardrails. Live fetch: GitHub ROCm, Yahoo Finance, SEC EDGAR. "
         "Script: build_amd_thesis_tracker_v1.py."),
    ]
    for i, (ver, detail) in enumerate(versions):
        wf_row(21 + i, ver, detail, WHITE if i%2==0 else GRAY)

# ==============================================================================
# SELF-TEST
# ==============================================================================
def self_test(out_path, score):
    import os
    errors = []
    if not os.path.exists(out_path):
        errors.append(f"FAIL: file not created at {out_path}")
    else:
        size = os.path.getsize(out_path)
        if size < 15000:
            errors.append(f"FAIL: file too small ({size} bytes)")

        from openpyxl import load_workbook
        wb = load_workbook(out_path, data_only=True)

        required_tabs = ["Thesis Score", "Deal Execution", "Software Moat",
                         "Financial Guardrails", "Live Data", "Data Sources", "Workflow"]
        for tab in required_tabs:
            if tab not in wb.sheetnames:
                errors.append(f"FAIL: tab '{tab}' missing")

        # Score must be numeric and in range
        if not (0 <= score <= 100):
            errors.append(f"FAIL: score {score} out of range 0-100")

        # Check Thesis Score tab has content
        ws = wb["Thesis Score"]
        found_score = False
        for row in ws.iter_rows(values_only=True):
            for c in row:
                if isinstance(c, str) and "THESIS SCORE" in c:
                    found_score = True
        if not found_score:
            errors.append("FAIL: Thesis Score tab missing THESIS SCORE header")

        # Check Deal Execution has deal names
        ws_d = wb["Deal Execution"]
        found_openai = found_humain = False
        for row in ws_d.iter_rows(values_only=True):
            for c in row:
                if isinstance(c, str):
                    if "OpenAI" in c: found_openai = True
                    if "HUMAIN" in c: found_humain = True
        if not found_openai: errors.append("FAIL: Deal Execution missing OpenAI")
        if not found_humain: errors.append("FAIL: Deal Execution missing HUMAIN")

        # Check Financial Guardrails has metrics
        ws_f = wb["Financial Guardrails"]
        found_dc = found_epyc = False
        for row in ws_f.iter_rows(values_only=True):
            for c in row:
                if isinstance(c, str):
                    if "Data Center" in c: found_dc = True
                    if "EPYC" in c: found_epyc = True
        if not found_dc:   errors.append("FAIL: Financial Guardrails missing Data Center metric")
        if not found_epyc: errors.append("FAIL: Financial Guardrails missing EPYC metric")

        # Weights must sum to 100
        weight_sum = (W_OPENAI + W_META + W_HUMAIN_DOE + W_NEW_DEALS +
                      W_ROCM + W_GITHUB + W_COMPETITIVE +
                      W_EPYC_SHARE + W_NI_MARGIN + W_GUIDANCE + W_DC_REVENUE)
        if weight_sum != 100:
            errors.append(f"FAIL: Condition weights sum to {weight_sum}, not 100")

    if errors:
        print("\n" + "="*60)
        print("SELF-TEST FAILED — DO NOT COMMIT TO GITHUB")
        print("="*60)
        for e in errors: print(f"  {e}")
        print("="*60)
        return False
    else:
        size = os.path.getsize(out_path)
        print("\n" + "="*60)
        print("SELF-TEST PASSED — safe to commit to GitHub")
        print(f"  File: {out_path}")
        print(f"  Size: {size:,} bytes")
        print(f"  Tabs: {', '.join(required_tabs)}")
        print(f"  Thesis Score: {score}/100 — {score_label(score)}")
        print(f"  Exit threshold: {EXIT_THRESHOLD} | Watch threshold: {WATCH_THRESHOLD}")
        print(f"  Fetch timestamp: {FETCH_TIMESTAMP}")
        print("="*60)
        return True

# ==============================================================================
# MAIN
# ==============================================================================
def build():
    print("Fetching live data...")
    github_data = fetch_github_rocm()
    yahoo_data  = fetch_yahoo_price()
    sec_data    = fetch_sec_edgar()

    for key, data in github_data.items():
        status = data.get("status", "LIVE")
        print(f"  GitHub {key}: {status}")
    print(f"  Yahoo Finance: {yahoo_data.get('status','unknown')}")
    print(f"  SEC EDGAR: {sec_data.get('status','unknown')}")

    github_score, github_note = github_health_score(github_data)
    print(f"  GitHub health score: {github_score}/10 — {github_note}")

    score, conditions = compute_score(github_score)
    print(f"\nThesis Score: {score}/100 — {score_label(score)}")

    wb = Workbook()
    wb.remove(wb.active)

    build_thesis_score(wb, score, conditions, github_note)
    build_deal_execution(wb)
    build_software_moat(wb, github_data, github_score, github_note)
    build_financial_guardrails(wb)
    build_live_snapshot(wb, github_data, yahoo_data, sec_data)
    build_data_sources(wb)
    build_workflow(wb, score)

    wb["Thesis Score"].sheet_properties.tabColor       = "1A7A3A"
    wb["Deal Execution"].sheet_properties.tabColor     = "2E75B6"
    wb["Software Moat"].sheet_properties.tabColor      = "7B2D8B"
    wb["Financial Guardrails"].sheet_properties.tabColor = "E67300"
    wb["Live Data"].sheet_properties.tabColor          = "006B6B"
    wb["Data Sources"].sheet_properties.tabColor       = "5A5A5A"
    wb["Workflow"].sheet_properties.tabColor           = "1A1A2E"

    wb.active = 0

    import os
    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "AMD_Thesis_Tracker_v1.xlsx")
    wb.save(out)
    print(f"Saved: {out}")
    return out, score

if __name__ == "__main__":
    out, score = build()
    self_test(out, score)
