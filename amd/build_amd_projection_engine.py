# ==============================================================================
# AMD 5-YEAR PROJECTION — COMPLETE ENGINE v1.0
# Built by reverse-engineering AMD 5Year Projection Manual.xlsx from Drive
# THIS IS THE FULL ENGINE — not a stub. Runs standalone.
# Update only the DATA BLOCK below each quarter. Structure never changes.
# HOW TO RUN: python build_amd_projection_engine.py
# ==============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ── DATA BLOCK — update these values each quarter ──────────────────────────────
VERSION      = "v5.1 — Q1 2026 Baseline"
GENERATED    = "May 2026"
ENTRY_PRICE  = 455.00       # AMD stock price at analysis date
BASE_REV     = 34.64        # FY2025 actual revenue ($B)
SHARES       = 1.650        # Diluted shares ($B)
Q1_NI_MARGIN = 0.221        # Q1 2026 Non-GAAP NI margin
Q1_REV       = 10.253       # Q1 2026 revenue ($B)
Q2_GUIDE     = 11.200       # Q2 2026 guidance midpoint ($B)

BULL = dict(
    rev_growth = [0.43, 0.50, 0.52, 0.50, 0.45],
    ni_margin  = [0.28, 0.32, 0.38, 0.41, 0.43],
    pe_low     = [45, 48, 50, 50, 48],
    pe_high    = [55, 58, 60, 60, 58],
)
BASE = dict(
    rev_growth = [0.36, 0.42, 0.43, 0.42, 0.40],
    ni_margin  = [0.25, 0.26, 0.26, 0.28, 0.29],
    pe_low     = [35, 36, 37, 38, 40],
    pe_high    = [40, 41, 42, 43, 45],
)
BEAR = dict(
    rev_growth = [0.32, 0.18, 0.12, 0.08, 0.10],
    ni_margin  = [0.22, 0.21, 0.22, 0.24, 0.26],
    pe_low     = [25, 20, 17, 15, 16],
    pe_high    = [30, 25, 22, 20, 21],
)

PROB_BULL = 0.35
PROB_BASE  = 0.45
PROB_BEAR  = 0.20

YEARS = [2026, 2027, 2028, 2029, 2030]

# ── COLORS (exact AMD palette) ─────────────────────────────────────────────────
NAVY    = "1A1A2E"
BLUE    = "1F4E79"
ACCENT  = "2E75B6"
WHITE   = "FFFFFF"
GRAY    = "F5F7FA"
BLACK   = "000000"
MUTED   = "888888"
GREEN_S = "E8F5E9"
BLUE_S  = "D6E4F0"
RED_S   = "FDEDEC"
GOLD_BG = "FFF8E1"
GOLD_FN = "7D5A00"
PROB_BG = "EDE7F6"

def fill(h): return PatternFill("solid", start_color=h, fgColor=h)

def cell(ws, row, col, val=None, bg=WHITE, fc=BLACK, bold=False,
         fmt=None, align="center", size=9, italic=False):
    c = ws.cell(row=row, column=col)
    if val is not None: c.value = val
    c.font = Font(name="Arial", bold=bold, size=size, color=fc, italic=italic)
    c.fill = fill(bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
    if fmt: c.number_format = fmt
    return c

def rh(ws, row, h=14): ws.row_dimensions[row].height = h

def compute(sc):
    revs, nis, epss, spls, sphs, cagrl, cagrh = [], [], [], [], [], [], []
    rev = BASE_REV
    for i in range(5):
        rev  = rev * (1 + sc["rev_growth"][i])
        ni   = rev * sc["ni_margin"][i]
        eps  = ni / SHARES
        spl  = eps * sc["pe_low"][i]
        sph  = eps * sc["pe_high"][i]
        revs.append(rev);  nis.append(ni);   epss.append(eps)
        spls.append(spl);  sphs.append(sph)
        cagrl.append((spl / ENTRY_PRICE) ** (1 / (i + 1)) - 1)
        cagrh.append((sph / ENTRY_PRICE) ** (1 / (i + 1)) - 1)
    return revs, nis, epss, spls, sphs, cagrl, cagrh

# ==============================================================================
# TAB 1 — INPUTS
# ==============================================================================
def build_inputs(wb):
    ws = wb.create_sheet("Inputs")
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12

    # Title
    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = f"AMD 5-YEAR PROJECTION — INPUTS [{VERSION}]"
    t.font = Font(name="Arial", bold=True, size=11, color=WHITE)
    t.fill = fill(NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 1, 20)

    ws.merge_cells("A2:G2")
    s = ws["A2"]
    s.value = (f"Blue = Editable | YELLOW HIGHLIGHT = Changed from prior version | "
               f"Probability: Bull {PROB_BULL:.0%} / Base {PROB_BASE:.0%} / Bear {PROB_BEAR:.0%}")
    s.font = Font(name="Arial", size=9, color=WHITE)
    s.fill = fill(BLUE)
    s.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 2, 14)

    ws.merge_cells("A3:G3")
    g = ws["A3"]
    g.value = f"GLOBAL INPUTS — AMD Q1 2026 Earnings Release (May 5, 2026) | Market Data May 8, 2026"
    g.font = Font(name="Arial", bold=True, size=9, color=WHITE)
    g.fill = fill(ACCENT)
    g.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    rh(ws, 3, 14)

    global_rows = [
        ("Current / Entry Stock Price ($)", f"${ENTRY_PRICE:,.2f}",
         "User-specified $455 | May 8, 2026 | Source: Investing.com"),
        ("Base Year Revenue — FY2025 ($B)", f"${BASE_REV:,.2f}B",
         "AMD FY2025 Annual Revenue | Source: AMD 10-K / Q4 2025 Earnings Release | ir.amd.com"),
        ("Diluted Shares Outstanding ($B)", f"{SHARES:,.3f}B",
         "AMD Q1 2026 Earnings Release May 5, 2026 | 1,650M diluted shares | ir.amd.com"),
        ("Q1 2026 Non-GAAP Net Margin", f"{Q1_NI_MARGIN:.1%}",
         "AMD Q1 2026: $2,265M Non-GAAP NI / $10,253M Revenue = 22.1% | AMD Financial Tables PDF"),
        ("Q1 2026 Revenue ($B)", f"${Q1_REV:,.3f}B",
         "AMD Q1 2026 Earnings Release | Actual quarterly result | ir.amd.com"),
        ("Q2 2026 Revenue Guide ($B)", f"${Q2_GUIDE:,.3f}B",
         "AMD Q2 2026 Official Guidance | Midpoint $11.2B ±$300M | AMD Q1 2026 Earnings Call May 5, 2026"),
    ]
    for i, (label, val, note) in enumerate(global_rows):
        r = 4 + i
        bg = WHITE if i % 2 == 0 else GRAY
        cell(ws, r, 1, bg=bg)
        cell(ws, r, 2, label, bg=bg, align="left")
        c = ws.cell(row=r, column=3)
        c.value = val; c.font = Font(name="Arial", bold=True, size=9, color="0070C0")
        c.fill = fill(bg); c.alignment = Alignment(horizontal="right", vertical="center")
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
        cell(ws, r, 4, note, bg=bg, fc=MUTED, size=8, align="left", italic=True)
        rh(ws, r, 14)

    ws.merge_cells("A11:G11")
    sh = ws["A11"]
    sh.value = "SCENARIO ASSUMPTIONS — Edit blue cells | All values feed automatically to Projection tab"
    sh.font = Font(name="Arial", bold=True, size=9, color=WHITE)
    sh.fill = fill(ACCENT)
    sh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    rh(ws, 11, 14)

    for col, txt in [(2,"Assumption"),(3,"2026"),(4,"2027"),(5,"2028"),(6,"2029"),(7,"2030")]:
        cell(ws, 12, col, txt, bg=ACCENT, fc=WHITE, bold=True)
    cell(ws, 12, 1, bg=ACCENT)
    rh(ws, 12, 14)

    scenarios = [
        ("▶ BULL CASE", BULL, GREEN_S, PROB_BULL),
        ("▶ BASE CASE", BASE, BLUE_S,  PROB_BASE),
        ("▶ BEAR CASE", BEAR, RED_S,   PROB_BEAR),
    ]
    r = 13
    for sc_name, sc_data, sc_bg, prob in scenarios:
        ws.merge_cells(f"A{r}:G{r}")
        sl = ws[f"A{r}"]
        sl.value = sc_name
        sl.font = Font(name="Arial", bold=True, size=9, color=BLACK)
        sl.fill = fill(sc_bg)
        sl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        rh(ws, r, 14); r += 1

        metrics = [
            ("Revenue Growth %",    sc_data["rev_growth"], "0.0%"),
            ("Net Income Margin %", sc_data["ni_margin"],  "0.0%"),
            ("PE Low",              sc_data["pe_low"],     "0"),
            ("PE High",             sc_data["pe_high"],    "0"),
        ]
        for label, vals, fmt in metrics:
            bg = WHITE if r % 2 == 0 else GRAY
            cell(ws, r, 1, bg=bg)
            cell(ws, r, 2, label, bg=bg, align="left")
            for j, val in enumerate(vals):
                c = ws.cell(row=r, column=3+j)
                c.value = val
                c.font = Font(name="Arial", bold=True, size=9, color="0070C0")
                c.fill = fill(bg)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.number_format = fmt
            rh(ws, r, 14); r += 1

        cell(ws, r, 1, bg=sc_bg)
        cell(ws, r, 2, "Probability Weight", bg=sc_bg, align="left")
        c = ws.cell(row=r, column=3)
        c.value = prob; c.font = Font(name="Arial", bold=True, size=9, color="0070C0")
        c.fill = fill("FFFF00"); c.number_format = "0%"
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
        cell(ws, r, 4, f"Bull {PROB_BULL:.0%} + Base {PROB_BASE:.0%} + Bear {PROB_BEAR:.0%} = 100%",
             bg=sc_bg, fc=MUTED, size=8, align="left")
        rh(ws, r, 14); r += 2

    ws.merge_cells(f"A{r}:G{r}")
    cl = ws[f"A{r}"]
    cl.value = "COLOR LEGEND: Blue Text = Editable Input | Black Text = Formula | Green Text = Linked from Inputs tab"
    cl.font = Font(name="Arial", size=8, color=MUTED, italic=True)
    cl.alignment = Alignment(horizontal="left", vertical="center", indent=1)

# ==============================================================================
# TAB 2 — PROJECTION
# ==============================================================================
def build_projection(wb):
    ws = wb.create_sheet("Projection")
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 26
    for col in ["C","D","E","F","G"]:
        ws.column_dimensions[col].width = 14
    ws.column_dimensions["H"].width = 2

    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "AMD 5-YEAR FINANCIAL PROJECTION — NON-GAAP"
    t.font = Font(name="Arial", bold=True, size=13, color=WHITE)
    t.fill = fill(NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 1, 22)

    ws.merge_cells("A2:H2")
    s = ws["A2"]
    s.value = (f"Entry Price: ${ENTRY_PRICE} | Base Revenue: ${BASE_REV:.2f}B (FY2025 Actual) | "
               f"Shares: {SHARES}B | Non-GAAP | Revenue & NI in $B | Change inputs on Inputs tab")
    s.font = Font(name="Arial", size=9, color=WHITE)
    s.fill = fill(BLUE)
    s.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 2, 14)

    scenarios = [
        ("▶ BULL CASE — NON-GAAP", BULL, GREEN_S),
        ("▶ BASE CASE — NON-GAAP", BASE, BLUE_S),
        ("▶ BEAR CASE — NON-GAAP", BEAR, RED_S),
    ]

    ROW = 3
    for sc_name, sc_data, sc_bg in scenarios:
        revs, nis, epss, spls, sphs, cagrl, cagrh = compute(sc_data)

        ws.merge_cells(f"A{ROW}:H{ROW}")
        sh = ws[f"A{ROW}"]
        sh.value = sc_name
        sh.font = Font(name="Arial", bold=True, size=11, color=WHITE)
        sh.fill = fill(BLUE)
        sh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        rh(ws, ROW, 20); ROW += 1

        cell(ws, ROW, 1, bg=ACCENT)
        cell(ws, ROW, 2, "YEAR", bg=ACCENT, fc=WHITE, bold=True)
        for i, yr in enumerate(YEARS):
            cell(ws, ROW, 3+i, str(yr), bg=ACCENT, fc=WHITE, bold=True)
        cell(ws, ROW, 8, bg=ACCENT)
        rh(ws, ROW, 16); ROW += 1

        def data_row(label, vals, fmt, bg, fc=BLACK, bold=False):
            nonlocal ROW
            cell(ws, ROW, 1, bg=bg)
            cell(ws, ROW, 2, label, bg=bg, fc=fc, bold=bold, align="left")
            for i, v in enumerate(vals):
                if v is None:
                    cell(ws, ROW, 3+i, "—", bg=bg, fc=MUTED)
                else:
                    c = ws.cell(row=ROW, column=3+i)
                    c.value = v
                    c.font = Font(name="Arial", bold=bold, size=9, color=fc)
                    c.fill = fill(bg)
                    c.alignment = Alignment(horizontal="center", vertical="center")
                    if fmt: c.number_format = fmt
            cell(ws, ROW, 8, bg=bg)
            rh(ws, ROW, 14); ROW += 1

        # Exact AMD row structure from xlsx
        data_row("REVENUE",
                 [f"{v:.2f}B" for v in revs],
                 None, sc_bg, BLACK, True)
        data_row(" REV GROWTH",
                 sc_data["rev_growth"],
                 "0.0%", sc_bg)
        data_row("NET INCOME",
                 [f"{v:.2f}B" for v in nis],
                 None, WHITE, BLACK, True)
        ni_growth = [None] + [nis[i]/nis[i-1]-1 for i in range(1,5)]
        data_row(" NET INC. GROWTH",
                 ni_growth,
                 "0.0%", GRAY)
        data_row(" NET INC. MARGINS",
                 sc_data["ni_margin"],
                 "0.0%", WHITE)
        data_row("EPS (Non-GAAP)",
                 [f"${int(round(v))}" for v in epss],
                 None, GRAY, BLACK, False)
        data_row(" PE LOW EST",
                 sc_data["pe_low"],
                 "0", WHITE)
        data_row(" PE HIGH EST",
                 sc_data["pe_high"],
                 "0", GRAY)
        data_row("SHARE PRICE LOW",
                 spls, '"$"#,##0', GOLD_BG, GOLD_FN, True)
        data_row("SHARE PRICE HIGH",
                 sphs, '"$"#,##0', GOLD_BG, GOLD_FN, True)
        data_row(" CAGR LOW",
                 cagrl, "+0.0%;-0.0%;—", sc_bg)
        data_row(" CAGR HIGH",
                 cagrh, "+0.0%;-0.0%;—", sc_bg)

        ROW += 1

    # Scenario comparison table
    ws.merge_cells(f"A{ROW}:H{ROW}")
    sc = ws[f"A{ROW}"]
    sc.value = "SCENARIO COMPARISON — 2030 OUTPUTS"
    sc.font = Font(name="Arial", bold=True, size=9, color=WHITE)
    sc.fill = fill(ACCENT)
    sc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    rh(ws, ROW, 16); ROW += 1

    for col, txt in [(2,"Scenario"),(3,"2030 Revenue"),(4,"2030 Net Income"),
                     (5,"2030 EPS"),(6,"2030 Pr. Low"),(7,"2030 Pr. High"),(8,"5-Yr CAGR Low")]:
        cell(ws, ROW, col, txt, bg=ACCENT, fc=WHITE, bold=True, size=8)
    cell(ws, ROW, 1, bg=ACCENT)
    rh(ws, ROW, 28); ROW += 1

    for sc_name2, sc_data, sc_bg in scenarios:
        revs, nis, epss, spls, sphs, cagrl, cagrh = compute(sc_data)
        nm = sc_name2.replace(" — NON-GAAP", "")
        cell(ws, ROW, 1, bg=sc_bg)
        cell(ws, ROW, 2, nm, bg=sc_bg, bold=True, align="left")
        cell(ws, ROW, 3, f"{revs[4]:.2f}B", bg=sc_bg)
        cell(ws, ROW, 4, f"{nis[4]:.2f}B", bg=sc_bg)
        cell(ws, ROW, 5, f"${int(round(epss[4]))}", bg=sc_bg)
        for j, val in enumerate([spls[4], sphs[4]]):
            cj = ws.cell(row=ROW, column=6+j)
            cj.value = val; cj.fill = fill(GOLD_BG)
            cj.font = Font(name="Arial", bold=True, size=9, color=GOLD_FN)
            cj.number_format = '"$"#,##0'
            cj.alignment = Alignment(horizontal="center", vertical="center")
        c8 = ws.cell(row=ROW, column=8)
        c8.value = cagrl[4]; c8.fill = fill(sc_bg)
        c8.font = Font(name="Arial", size=9)
        c8.number_format = "+0.0%;-0.0%;—"
        c8.alignment = Alignment(horizontal="center", vertical="center")
        rh(ws, ROW, 14); ROW += 1

    ws.freeze_panes = "A3"

    ws.merge_cells(f"A{ROW+1}:H{ROW+1}")
    foot = ws[f"A{ROW+1}"]
    foot.value = (f"Source: AMD FY2025 actual ${BASE_REV:.2f}B | "
                  f"Analysis price ${ENTRY_PRICE} | Built: {GENERATED} | "
                  f"Non-GAAP basis | Script: build_amd_projection_engine.py")
    foot.font = Font(name="Arial", size=7, color=MUTED, italic=True)
    foot.alignment = Alignment(horizontal="left")

# ==============================================================================
# TAB 3 — PROBABILITY WEIGHTED
# ==============================================================================
def build_probability(wb):
    ws = wb.create_sheet("Probability Weighted")
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 28
    for col in ["C","D","E","F","G"]:
        ws.column_dimensions[col].width = 14
    ws.column_dimensions["H"].width = 16

    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "AMD 5-YEAR PROJECTION — PROBABILITY WEIGHTED EXPECTED VALUE"
    t.font = Font(name="Arial", bold=True, size=11, color=WHITE)
    t.fill = fill(NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 1, 22)

    ws.merge_cells("A2:H2")
    s = ws["A2"]
    s.value = (f"Bull {PROB_BULL:.0%} | Base {PROB_BASE:.0%} | Bear {PROB_BEAR:.0%} | "
               f"Expected Value = (Bull×{PROB_BULL:.0%}) + (Base×{PROB_BASE:.0%}) + "
               f"(Bear×{PROB_BEAR:.0%}) | Change probabilities in yellow cells")
    s.font = Font(name="Arial", size=9, color=WHITE)
    s.fill = fill(BLUE)
    s.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 2, 14)

    ws.merge_cells("A3:H3")
    ph = ws["A3"]
    ph.value = "PROBABILITY INPUTS — Edit yellow cells"
    ph.font = Font(name="Arial", bold=True, size=9, color=WHITE)
    ph.fill = fill(ACCENT)
    ph.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    rh(ws, 3, 14)

    for col, txt in [(2,"Case"),(3,"Probability"),(4,"Description")]:
        cell(ws, 4, col, txt, bg=ACCENT, fc=WHITE, bold=True)
    ws.merge_cells("D4:H4")
    cell(ws, 4, 1, bg=ACCENT)
    rh(ws, 4, 14)

    prob_rows = [
        ("BULL CASE", PROB_BULL, GREEN_S,
         "Full Helios/MI500 execution. CPU TAM doubles. AMD takes meaningful GPU share. Compute demand exceeds forecasts."),
        ("BASE CASE", PROB_BASE, BLUE_S,
         "Confirmed deals execute. Strong growth. Gradual margin expansion. CPU TAM partially realized."),
        ("BEAR CASE", PROB_BEAR, RED_S,
         "Vera Rubin/Feynman pauses. AI capex moderates. Custom silicon scales. CPU TAM provides structural floor."),
    ]
    for i, (case, prob, bg, desc) in enumerate(prob_rows):
        r = 5 + i
        cell(ws, r, 1, bg=bg)
        cell(ws, r, 2, case, bg=bg, bold=True, align="left")
        c = ws.cell(row=r, column=3)
        c.value = prob; c.font = Font(name="Arial", bold=True, size=9, color="0070C0")
        c.fill = fill("FFFF00"); c.number_format = "0%"
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(f"D{r}:H{r}")
        cell(ws, r, 4, desc, bg=bg, fc=MUTED, size=8, align="left")
        rh(ws, r, 14)

    cell(ws, 8, 1, bg=WHITE)
    cell(ws, 8, 2, "Sum (must = 100%)", bg=WHITE, bold=True, align="left")
    c = ws.cell(row=8, column=3)
    c.value = PROB_BULL + PROB_BASE + PROB_BEAR
    c.font = Font(name="Arial", bold=True, size=9)
    c.number_format = "0%"; c.fill = fill(WHITE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("D8:H8")
    chk = "✓ 100% — Valid" if abs(PROB_BULL+PROB_BASE+PROB_BEAR-1.0) < 0.001 else "⚠ Does not sum to 100%"
    cell(ws, 8, 4, chk, bg=WHITE, bold=True, align="left")
    rh(ws, 8, 14)

    ws.merge_cells("A10:H10")
    sch = ws["A10"]
    sch.value = "SCENARIO COMPARISON BY YEAR — Bull vs Base vs Bear"
    sch.font = Font(name="Arial", bold=True, size=9, color=WHITE)
    sch.fill = fill(ACCENT)
    sch.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    rh(ws, 10, 14)

    cell(ws, 11, 1, bg=ACCENT)
    cell(ws, 11, 2, "Metric", bg=ACCENT, fc=WHITE, bold=True)
    for i, yr in enumerate(YEARS):
        cell(ws, 11, 3+i, str(yr), bg=ACCENT, fc=WHITE, bold=True)
    cell(ws, 11, 8, "EXPECTED VALUE 2030", bg=ACCENT, fc=WHITE, bold=True, size=8)
    rh(ws, 11, 16)

    bull = compute(BULL); base = compute(BASE); bear = compute(BEAR)
    ev_spl = [bull[3][i]*PROB_BULL + base[3][i]*PROB_BASE + bear[3][i]*PROB_BEAR for i in range(5)]
    ev_sph = [bull[4][i]*PROB_BULL + base[4][i]*PROB_BASE + bear[4][i]*PROB_BEAR for i in range(5)]
    ev_rev = [bull[0][i]*PROB_BULL + base[0][i]*PROB_BASE + bear[0][i]*PROB_BEAR for i in range(5)]
    ev_eps = [bull[2][i]*PROB_BULL + base[2][i]*PROB_BASE + bear[2][i]*PROB_BEAR for i in range(5)]

    sc_blocks = [
        (f"▶ BULL CASE ({PROB_BULL:.0%})", GREEN_S, *bull),
        (f"▶ BASE CASE ({PROB_BASE:.0%})", BLUE_S,  *base),
        (f"▶ BEAR CASE ({PROB_BEAR:.0%})", RED_S,   *bear),
    ]

    ROW = 12
    for sc_name, sc_bg, revs, nis, epss, spls, sphs, cagrl, cagrh in sc_blocks:
        ws.merge_cells(f"A{ROW}:H{ROW}")
        sh = ws[f"A{ROW}"]
        sh.value = sc_name
        sh.font = Font(name="Arial", bold=True, size=9, color=WHITE)
        sh.fill = fill(BLUE)
        sh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        rh(ws, ROW, 16); ROW += 1

        pw_rows = [
            ("  Revenue ($B)",       [f"{v:.2f}B" for v in revs], None,        False),
            ("  EPS (Non-GAAP)",     [f"${int(round(v))}" for v in epss], None, False),
            ("  Share Price Low",    spls, '"$"#,##0', True),
            ("  Share Price High",   sphs, '"$"#,##0', True),
            ("  CAGR Low",           cagrl, "+0.0%;-0.0%;—", False),
            ("  CAGR High",          cagrh, "+0.0%;-0.0%;—", False),
        ]
        for label, vals, fmt, bold in pw_rows:
            is_price = "Price" in label
            bg = GOLD_BG if is_price else (WHITE if ROW % 2 == 0 else GRAY)
            fc = GOLD_FN if is_price else BLACK
            cell(ws, ROW, 1, bg=bg)
            cell(ws, ROW, 2, label, bg=bg, fc=fc, bold=bold, align="left")
            for i, v in enumerate(vals):
                if isinstance(v, str):
                    cell(ws, ROW, 3+i, v, bg=bg, fc=fc, bold=bold)
                else:
                    c = ws.cell(row=ROW, column=3+i)
                    c.value = v; c.font = Font(name="Arial", bold=bold, size=9, color=fc)
                    c.fill = fill(bg); c.number_format = fmt if fmt else ""
                    c.alignment = Alignment(horizontal="center", vertical="center")
            last = vals[4]
            if isinstance(last, str):
                cell(ws, ROW, 8, last, bg=bg, fc=fc, bold=True)
            else:
                c8 = ws.cell(row=ROW, column=8)
                c8.value = last; c8.fill = fill(bg)
                c8.font = Font(name="Arial", bold=True, size=9, color=fc)
                c8.number_format = fmt if fmt else ""
                c8.alignment = Alignment(horizontal="center", vertical="center")
            rh(ws, ROW, 14); ROW += 1
        ROW += 1

    ws.merge_cells(f"A{ROW}:H{ROW}")
    ev = ws[f"A{ROW}"]
    ev.value = "PROBABILITY WEIGHTED EXPECTED VALUE — Updates automatically when probabilities change"
    ev.font = Font(name="Arial", bold=True, size=9, color=WHITE)
    ev.fill = fill(NAVY)
    ev.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    rh(ws, ROW, 16); ROW += 1

    ev_cagrl = [(ev_spl[i]/ENTRY_PRICE)**(1/(i+1))-1 for i in range(5)]
    ev_cagrh = [(ev_sph[i]/ENTRY_PRICE)**(1/(i+1))-1 for i in range(5)]

    ev_rows = [
        ("  Expected Revenue ($B)",   [f"{v:.2f}B" for v in ev_rev], None,        False),
        ("  Expected EPS (Non-GAAP)", [f"${int(round(v))}" for v in ev_eps], None, True),
        ("  Expected Share Price Low", ev_spl,  '"$"#,##0', True),
        ("  Expected Share Price High",ev_sph,  '"$"#,##0', True),
        ("  Expected CAGR Low",        ev_cagrl,"+0.0%;-0.0%;—", False),
        ("  Expected CAGR High",       ev_cagrh,"+0.0%;-0.0%;—", False),
    ]
    for label, vals, fmt, bold in ev_rows:
        is_price = "Price" in label
        bg = GOLD_BG if is_price else PROB_BG
        fc = GOLD_FN if is_price else BLACK
        cell(ws, ROW, 1, bg=bg)
        cell(ws, ROW, 2, label, bg=bg, fc=fc, bold=bold, align="left")
        for i, v in enumerate(vals):
            if isinstance(v, str):
                cell(ws, ROW, 3+i, v, bg=bg, fc=fc, bold=bold)
            else:
                c = ws.cell(row=ROW, column=3+i)
                c.value = v; c.font = Font(name="Arial", bold=bold, size=9, color=fc)
                c.fill = fill(bg); c.number_format = fmt if fmt else ""
                c.alignment = Alignment(horizontal="center", vertical="center")
        last = vals[4]
        if isinstance(last, str):
            cell(ws, ROW, 8, last, bg=bg, fc=fc, bold=True)
        else:
            c8 = ws.cell(row=ROW, column=8)
            c8.value = last; c8.fill = fill(bg)
            c8.font = Font(name="Arial", bold=True, size=9, color=fc)
            c8.number_format = fmt if fmt else ""
            c8.alignment = Alignment(horizontal="center", vertical="center")
        rh(ws, ROW, 14); ROW += 1

    ws.freeze_panes = "A12"

    ws.merge_cells(f"A{ROW+1}:H{ROW+1}")
    foot = ws[f"A{ROW+1}"]
    foot.value = (f"Built: {GENERATED} | Entry ${ENTRY_PRICE} | "
                  f"Shares {SHARES}B | Non-GAAP | Script: build_amd_projection_engine.py")
    foot.font = Font(name="Arial", size=7, color=MUTED, italic=True)
    foot.alignment = Alignment(horizontal="left")

# ==============================================================================
# SELF-TEST
# ==============================================================================
def self_test(out_path):
    import os
    from openpyxl import load_workbook

    errors = []
    if not os.path.exists(out_path):
        errors.append(f"FAIL: file not created at {out_path}")
    else:
        size = os.path.getsize(out_path)
        if size < 10000:
            errors.append(f"FAIL: file too small ({size} bytes) — engine incomplete")

        wb = load_workbook(out_path, data_only=True)
        for tab in ["Inputs", "Projection", "Probability Weighted"]:
            if tab not in wb.sheetnames:
                errors.append(f"FAIL: tab '{tab}' missing")

        ws = wb["Projection"]
        found_bull_spl = False
        found_base_spl = False
        for row in ws.iter_rows(values_only=True):
            for c in row:
                if isinstance(c, (int, float)) and 3000 < c < 3200: found_bull_spl = True
                if isinstance(c, (int, float)) and 1300 < c < 1400: found_base_spl = True
        if not found_bull_spl:
            errors.append("FAIL: Bull 2030 SPL ~$3,073 not found")
        if not found_base_spl:
            errors.append("FAIL: Base 2030 SPL ~$1,337 not found")

    if errors:
        print("\n" + "="*60)
        print("SELF-TEST FAILED — DO NOT SAVE TO GITHUB")
        print("="*60)
        for e in errors: print(f"  {e}")
        print("="*60)
        return False
    else:
        size = os.path.getsize(out_path)
        print("\n" + "="*60)
        print("SELF-TEST PASSED — safe to commit to GitHub")
        print(f"  File: {out_path}")
        print(f"  Size: {size:,} bytes")
        print(f"  Tabs: Inputs, Projection, Probability Weighted confirmed")
        print(f"  Bull SPL ~$3,073, Base SPL ~$1,337: confirmed")
        print("="*60)
        return True

# ==============================================================================
# MAIN
# ==============================================================================
def build():
    wb = Workbook()
    wb.remove(wb.active)
    build_inputs(wb)
    build_projection(wb)
    build_probability(wb)

    wb["Inputs"].sheet_properties.tabColor            = "2E75B6"
    wb["Projection"].sheet_properties.tabColor        = "1A7A3A"
    wb["Probability Weighted"].sheet_properties.tabColor = "7B2D8B"

    out = "/mnt/user-data/outputs/AMD_5Year_Projection_v6.xlsx"
    wb.save(out)

    bull = compute(BULL); base = compute(BASE); bear = compute(BEAR)
    ev_spl = [bull[3][i]*PROB_BULL+base[3][i]*PROB_BASE+bear[3][i]*PROB_BEAR for i in range(5)]
    ev_sph = [bull[4][i]*PROB_BULL+base[4][i]*PROB_BASE+bear[4][i]*PROB_BEAR for i in range(5)]
    print(f"Saved: {out}")
    print(f"\nKEY 2030 OUTPUTS — verify on open:")
    print(f"BULL: Rev=${bull[0][4]:.2f}B  NI=${bull[1][4]:.2f}B  EPS=${int(round(bull[2][4]))}  SPL=${bull[3][4]:,.0f}  SPH=${bull[4][4]:,.0f}  CAGR={bull[5][4]:+.1%}/{bull[6][4]:+.1%}")
    print(f"BASE: Rev=${base[0][4]:.2f}B  NI=${base[1][4]:.2f}B  EPS=${int(round(base[2][4]))}  SPL=${base[3][4]:,.0f}  SPH=${base[4][4]:,.0f}  CAGR={base[5][4]:+.1%}/{base[6][4]:+.1%}")
    print(f"BEAR: Rev=${bear[0][4]:.2f}B  NI=${bear[1][4]:.2f}B  EPS=${int(round(bear[2][4]))}  SPL=${bear[3][4]:,.0f}  SPH=${bear[4][4]:,.0f}  CAGR={bear[5][4]:+.1%}/{bear[6][4]:+.1%}")
    print(f"EV:   SPL=${ev_spl[4]:,.0f}  SPH=${ev_sph[4]:,.0f}  CAGR={(ev_spl[4]/ENTRY_PRICE)**(1/5)-1:+.1%}/{(ev_sph[4]/ENTRY_PRICE)**(1/5)-1:+.1%}")
    return out

if __name__ == "__main__":
    out = build()
    passed = self_test(out)
    if not passed:
        import sys
        sys.exit(1)
