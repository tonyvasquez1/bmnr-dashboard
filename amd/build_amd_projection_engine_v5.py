# ==============================================================================
# AMD 5-YEAR PROJECTION — COMPLETE ENGINE v5.0
# Built by reverse-engineering AMD 5Year Projection Manual.xlsx from Drive
# THIS IS THE FULL ENGINE — not a stub. Runs standalone.
# v5.0: Confirmed hyperscaler deal update — OpenAI 6GW (Oct 2025), Meta 6GW (Feb 2026),
#        Oracle 50K GPUs (Oct 2025). SHARES raised to 1.750B (320M warrant dilution from
#        OpenAI 160M + Meta 160M warrants, net of partial buyback offset).
#        Probability weights: Bull 45% / Base 40% / Bear 15%.
#        Bear narrative reframed from demand/competitive risk → execution risk.
# Update only the DATA BLOCK below each quarter. Structure never changes.
# HOW TO RUN: python build_amd_projection_engine_v5.py
# ==============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ── DATA BLOCK — update these values each quarter ──────────────────────────────
VERSION      = "v6.0 — Q1 2026 + Confirmed Deals (OpenAI/Meta/Oracle)"
GENERATED    = "May 2026"
ENTRY_PRICE  = 455.00       # AMD stock price at analysis date
BASE_REV     = 34.64        # FY2025 actual revenue ($B)
SHARES       = 1.750        # Diluted shares ($B) — 1.650B actual + 320M OpenAI/Meta warrants net
Q1_NI_MARGIN = 0.221        # Q1 2026 Non-GAAP NI margin
Q1_REV       = 10.253       # Q1 2026 revenue ($B)
Q2_GUIDE     = 11.200       # Q2 2026 guidance midpoint ($B)

BULL = dict(
    rev_growth = [0.43, 0.50, 0.52, 0.50, 0.45],
    ni_margin  = [0.28, 0.32, 0.38, 0.41, 0.43],
    pe_low     = [45, 48, 50, 50, 48],
    pe_high    = [55, 58, 60, 60, 58],
)
BASE = dict(
    rev_growth = [0.36, 0.42, 0.43, 0.42, 0.40],
    ni_margin  = [0.25, 0.26, 0.26, 0.28, 0.29],
    pe_low     = [35, 36, 37, 38, 40],
    pe_high    = [40, 41, 42, 43, 45],
)
BEAR = dict(
    rev_growth = [0.32, 0.18, 0.12, 0.08, 0.10],
    ni_margin  = [0.22, 0.21, 0.22, 0.24, 0.26],
    pe_low     = [25, 20, 17, 15, 16],
    pe_high    = [30, 25, 22, 20, 21],
)

PROB_BULL = 0.45
PROB_BASE  = 0.40
PROB_BEAR  = 0.15

YEARS = [2026, 2027, 2028, 2029, 2030]

# ── COLORS (exact AMD palette) ─────────────────────────────────────────────────
NAVY    = "1A1A2E"
BLUE    = "1F4E79"
ACCENT  = "2E75B6"
WHITE   = "FFFFFF"
GRAY    = "F5F7FA"
BLACK   = "000000"
MUTED   = "888888"
GREEN_S = "E8F5E9"
BLUE_S  = "D6E4F0"
RED_S   = "FDEDEC"
GOLD_BG = "FFF8E1"
GOLD_FN = "7D5A00"
PROB_BG = "EDE7F6"

def fill(h): return PatternFill("solid", start_color=h, fgColor=h)

def cell(ws, row, col, val=None, bg=WHITE, fc=BLACK, bold=False,
         fmt=None, align="center", size=9, italic=False, wrap=False):
    c = ws.cell(row=row, column=col)
    if val is not None: c.value = val
    c.font = Font(name="Arial", bold=bold, size=size, color=fc, italic=italic)
    c.fill = fill(bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if fmt: c.number_format = fmt
    return c

def rh(ws, row, h=14): ws.row_dimensions[row].height = h

def compute(sc):
    revs, nis, epss, spls, sphs, cagrl, cagrh = [], [], [], [], [], [], []
    rev = BASE_REV
    for i in range(5):
        rev  = rev * (1 + sc["rev_growth"][i])
        ni   = rev * sc["ni_margin"][i]
        eps  = ni / SHARES
        spl  = eps * sc["pe_low"][i]
        sph  = eps * sc["pe_high"][i]
        revs.append(rev);  nis.append(ni);   epss.append(eps)
        spls.append(spl);  sphs.append(sph)
        cagrl.append((spl / ENTRY_PRICE) ** (1 / (i + 1)) - 1)
        cagrh.append((sph / ENTRY_PRICE) ** (1 / (i + 1)) - 1)
    return revs, nis, epss, spls, sphs, cagrl, cagrh

# ==============================================================================
# TAB 1 — INPUTS
# ==============================================================================
def build_inputs(wb):
    ws = wb.create_sheet("Inputs")
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12

    # Title
    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = f"AMD 5-YEAR PROJECTION — INPUTS [{VERSION}]"
    t.font = Font(name="Arial", bold=True, size=11, color=WHITE)
    t.fill = fill(NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 1, 20)

    ws.merge_cells("A2:G2")
    s = ws["A2"]
    s.value = (f"Blue = Editable | YELLOW HIGHLIGHT = Changed from prior version | "
               f"Probability: Bull {PROB_BULL:.0%} / Base {PROB_BASE:.0%} / Bear {PROB_BEAR:.0%}")
    s.font = Font(name="Arial", size=9, color=WHITE)
    s.fill = fill(BLUE)
    s.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 2, 14)

    ws.merge_cells("A3:G3")
    g = ws["A3"]
    g.value = f"GLOBAL INPUTS — AMD Q1 2026 Earnings Release (May 5, 2026) | Market Data May 8, 2026"
    g.font = Font(name="Arial", bold=True, size=9, color=WHITE)
    g.fill = fill(ACCENT)
    g.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    rh(ws, 3, 14)

    global_rows = [
        ("Current / Entry Stock Price ($)", f"${ENTRY_PRICE:,.2f}",
         "User-specified $455 | May 8, 2026 | Source: Investing.com"),
        ("Base Year Revenue — FY2025 ($B)", f"${BASE_REV:,.2f}B",
         "AMD FY2025 Annual Revenue | Source: AMD 10-K / Q4 2025 Earnings Release | ir.amd.com"),
        ("Diluted Shares Outstanding ($B)", f"{SHARES:,.3f}B",
         "1,650M actual (AMD Q1 2026) + 320M OpenAI/Meta warrants (160M each) net of partial buyback = ~1,750M | ir.amd.com"),
        ("Q1 2026 Non-GAAP Net Margin", f"{Q1_NI_MARGIN:.1%}",
         "AMD Q1 2026: $2,265M Non-GAAP NI / $10,253M Revenue = 22.1% | AMD Financial Tables PDF"),
        ("Q1 2026 Revenue ($B)", f"${Q1_REV:,.3f}B",
         "AMD Q1 2026 Earnings Release | Actual quarterly result | ir.amd.com"),
        ("Q2 2026 Revenue Guide ($B)", f"${Q2_GUIDE:,.3f}B",
         "AMD Q2 2026 Official Guidance | Midpoint $11.2B ±$300M | AMD Q1 2026 Earnings Call May 5, 2026"),
    ]
    for i, (label, val, note) in enumerate(global_rows):
        r = 4 + i
        bg = WHITE if i % 2 == 0 else GRAY
        cell(ws, r, 1, bg=bg)
        cell(ws, r, 2, label, bg=bg, align="left")
        c = ws.cell(row=r, column=3)
        c.value = val; c.font = Font(name="Arial", bold=True, size=9, color="0070C0")
        c.fill = fill(bg); c.alignment = Alignment(horizontal="right", vertical="center")
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
        cell(ws, r, 4, note, bg=bg, fc=MUTED, size=8, align="left", italic=True)
        rh(ws, r, 14)

    ws.merge_cells("A11:G11")
    sh = ws["A11"]
    sh.value = "SCENARIO ASSUMPTIONS — Edit blue cells | All values feed automatically to Projection tab"
    sh.font = Font(name="Arial", bold=True, size=9, color=WHITE)
    sh.fill = fill(ACCENT)
    sh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    rh(ws, 11, 14)

    for col, txt in [(2,"Assumption"),(3,"2026"),(4,"2027"),(5,"2028"),(6,"2029"),(7,"2030")]:
        cell(ws, 12, col, txt, bg=ACCENT, fc=WHITE, bold=True)
    cell(ws, 12, 1, bg=ACCENT)
    rh(ws, 12, 14)

    scenarios = [
        ("▶ BULL CASE", BULL, GREEN_S, PROB_BULL),
        ("▶ BASE CASE", BASE, BLUE_S,  PROB_BASE),
        ("▶ BEAR CASE", BEAR, RED_S,   PROB_BEAR),
    ]
    r = 13
    for sc_name, sc_data, sc_bg, prob in scenarios:
        ws.merge_cells(f"A{r}:G{r}")
        sl = ws[f"A{r}"]
        sl.value = sc_name
        sl.font = Font(name="Arial", bold=True, size=9, color=BLACK)
        sl.fill = fill(sc_bg)
        sl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        rh(ws, r, 14); r += 1

        metrics = [
            ("Revenue Growth %",    sc_data["rev_growth"], "0.0%"),
            ("Net Income Margin %", sc_data["ni_margin"],  "0.0%"),
            ("PE Low",              sc_data["pe_low"],     "0"),
            ("PE High",             sc_data["pe_high"],    "0"),
        ]
        for label, vals, fmt in metrics:
            bg = WHITE if r % 2 == 0 else GRAY
            cell(ws, r, 1, bg=bg)
            cell(ws, r, 2, label, bg=bg, align="left")
            for j, val in enumerate(vals):
                c = ws.cell(row=r, column=3+j)
                c.value = val
                c.font = Font(name="Arial", bold=True, size=9, color="0070C0")
                c.fill = fill(bg)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.number_format = fmt
            rh(ws, r, 14); r += 1

        cell(ws, r, 1, bg=sc_bg)
        cell(ws, r, 2, "Probability Weight", bg=sc_bg, align="left")
        c = ws.cell(row=r, column=3)
        c.value = prob; c.font = Font(name="Arial", bold=True, size=9, color="0070C0")
        c.fill = fill("FFFF00"); c.number_format = "0%"
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
        cell(ws, r, 4, f"Bull {PROB_BULL:.0%} + Base {PROB_BASE:.0%} + Bear {PROB_BEAR:.0%} = 100%",
             bg=sc_bg, fc=MUTED, size=8, align="left")
        rh(ws, r, 14); r += 2

    ws.merge_cells(f"A{r}:G{r}")
    cl = ws[f"A{r}"]
    cl.value = "COLOR LEGEND: Blue Text = Editable Input | Black Text = Formula | Green Text = Linked from Inputs tab"
    cl.font = Font(name="Arial", size=8, color=MUTED, italic=True)
    cl.alignment = Alignment(horizontal="left", vertical="center", indent=1)

# ==============================================================================
# TAB 2 — PROJECTION
# ==============================================================================
def build_projection(wb):
    ws = wb.create_sheet("Projection")
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 26
    for col in ["C","D","E","F","G"]:
        ws.column_dimensions[col].width = 16
    ws.column_dimensions["H"].width = 14

    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "AMD 5-YEAR FINANCIAL PROJECTION — NON-GAAP"
    t.font = Font(name="Arial", bold=True, size=13, color=WHITE)
    t.fill = fill(NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 1, 22)

    ws.merge_cells("A2:H2")
    s = ws["A2"]
    s.value = (f"Entry Price: ${ENTRY_PRICE} | Base Revenue: ${BASE_REV:.2f}B (FY2025 Actual) | "
               f"Shares: {SHARES}B | Non-GAAP | Revenue & NI in $B | Change inputs on Inputs tab")
    s.font = Font(name="Arial", size=9, color=WHITE)
    s.fill = fill(BLUE)
    s.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 2, 14)

    scenarios = [
        ("▶ BULL CASE — NON-GAAP", BULL, GREEN_S),
        ("▶ BASE CASE — NON-GAAP", BASE, BLUE_S),
        ("▶ BEAR CASE — NON-GAAP", BEAR, RED_S),
    ]

    ROW = 3
    for sc_name, sc_data, sc_bg in scenarios:
        revs, nis, epss, spls, sphs, cagrl, cagrh = compute(sc_data)

        ws.merge_cells(f"A{ROW}:H{ROW}")
        sh = ws[f"A{ROW}"]
        sh.value = sc_name
        sh.font = Font(name="Arial", bold=True, size=11, color=WHITE)
        sh.fill = fill(BLUE)
        sh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        rh(ws, ROW, 20); ROW += 1

        cell(ws, ROW, 1, bg=ACCENT)
        cell(ws, ROW, 2, "YEAR", bg=ACCENT, fc=WHITE, bold=True)
        for i, yr in enumerate(YEARS):
            cell(ws, ROW, 3+i, str(yr), bg=ACCENT, fc=WHITE, bold=True)
        cell(ws, ROW, 8, bg=ACCENT)
        rh(ws, ROW, 18); ROW += 1

        def data_row(label, vals, fmt, bg, fc=BLACK, bold=False):
            nonlocal ROW
            cell(ws, ROW, 1, bg=bg)
            cell(ws, ROW, 2, label, bg=bg, fc=fc, bold=bold, align="left")
            for i, v in enumerate(vals):
                if v is None:
                    cell(ws, ROW, 3+i, "—", bg=bg, fc=MUTED)
                else:
                    c = ws.cell(row=ROW, column=3+i)
                    c.value = v
                    c.font = Font(name="Arial", bold=bold, size=9, color=fc)
                    c.fill = fill(bg)
                    c.alignment = Alignment(horizontal="center", vertical="center")
                    if fmt: c.number_format = fmt
            cell(ws, ROW, 8, bg=bg)
            rh(ws, ROW, 16); ROW += 1

        data_row("REVENUE",
                 [f"{v:.2f}B" for v in revs],
                 None, sc_bg, BLACK, True)
        data_row(" REV GROWTH",
                 sc_data["rev_growth"],
                 "0.0%", sc_bg)
        data_row("NET INCOME",
                 [f"{v:.2f}B" for v in nis],
                 None, WHITE, BLACK, True)
        ni_growth = [None] + [nis[i]/nis[i-1]-1 for i in range(1,5)]
        data_row(" NET INC. GROWTH",
                 ni_growth,
                 "0.0%", GRAY)
        data_row(" NET INC. MARGINS",
                 sc_data["ni_margin"],
                 "0.0%", WHITE)
        data_row("EPS (Non-GAAP)",
                 [f"${int(round(v))}" for v in epss],
                 None, GRAY, BLACK, False)
        data_row(" PE LOW EST",
                 sc_data["pe_low"],
                 "0", WHITE)
        data_row(" PE HIGH EST",
                 sc_data["pe_high"],
                 "0", GRAY)
        data_row("SHARE PRICE LOW",
                 spls, '"$"#,##0', GOLD_BG, GOLD_FN, True)
        data_row("SHARE PRICE HIGH",
                 sphs, '"$"#,##0', GOLD_BG, GOLD_FN, True)
        data_row(" CAGR LOW",
                 cagrl, "+0.0%;-0.0%;—", sc_bg)
        data_row(" CAGR HIGH",
                 cagrh, "+0.0%;-0.0%;—", sc_bg)

        ROW += 1

    ws.merge_cells(f"A{ROW}:H{ROW}")
    sc = ws[f"A{ROW}"]
    sc.value = "SCENARIO COMPARISON — 2030 OUTPUTS"
    sc.font = Font(name="Arial", bold=True, size=9, color=WHITE)
    sc.fill = fill(ACCENT)
    sc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    rh(ws, ROW, 16); ROW += 1

    for col, txt in [(2,"Scenario"),(3,"2030 Revenue"),(4,"2030 Net Income"),
                     (5,"2030 EPS"),(6,"2030 Pr. Low"),(7,"2030 Pr. High"),(8,"5-Yr CAGR")]:
        c = ws.cell(row=ROW, column=col)
        c.value = txt
        c.font = Font(name="Arial", bold=True, size=9, color=WHITE)
        c.fill = fill(ACCENT)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell(ws, ROW, 1, bg=ACCENT)
    rh(ws, ROW, 30); ROW += 1

    for sc_name2, sc_data, sc_bg in scenarios:
        revs, nis, epss, spls, sphs, cagrl, cagrh = compute(sc_data)
        nm = sc_name2.replace(" — NON-GAAP", "")
        cell(ws, ROW, 1, bg=sc_bg)
        cell(ws, ROW, 2, nm, bg=sc_bg, bold=True, align="left")
        cell(ws, ROW, 3, f"{revs[4]:.2f}B", bg=sc_bg)
        cell(ws, ROW, 4, f"{nis[4]:.2f}B", bg=sc_bg)
        cell(ws, ROW, 5, f"${int(round(epss[4]))}", bg=sc_bg)
        for j, val in enumerate([spls[4], sphs[4]]):
            cj = ws.cell(row=ROW, column=6+j)
            cj.value = val; cj.fill = fill(GOLD_BG)
            cj.font = Font(name="Arial", bold=True, size=9, color=GOLD_FN)
            cj.number_format = '"$"#,##0'
            cj.alignment = Alignment(horizontal="center", vertical="center")
        c8 = ws.cell(row=ROW, column=8)
        c8.value = f"{cagrl[4]:+.1%} / {cagrh[4]:+.1%}"
        c8.fill = fill(sc_bg)
        c8.font = Font(name="Arial", size=9, color=BLACK)
        c8.alignment = Alignment(horizontal="center", vertical="center")
        rh(ws, ROW, 18); ROW += 1

    ws.freeze_panes = "A3"

    ws.merge_cells(f"A{ROW+1}:H{ROW+1}")
    foot = ws[f"A{ROW+1}"]
    foot.value = (f"Source: AMD FY2025 actual ${BASE_REV:.2f}B | "
                  f"Analysis price ${ENTRY_PRICE} | Built: {GENERATED} | "
                  f"Non-GAAP basis | Script: build_amd_projection_engine_v5.py")
    foot.font = Font(name="Arial", size=7, color=MUTED, italic=True)
    foot.alignment = Alignment(horizontal="left")

# ==============================================================================
# TAB 3 — PROBABILITY WEIGHTED
# ==============================================================================
def build_probability(wb):
    ws = wb.create_sheet("Probability Weighted")
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 28
    for col in ["C","D","E","F","G"]:
        ws.column_dimensions[col].width = 14
    ws.column_dimensions["H"].width = 16

    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "AMD 5-YEAR PROJECTION — PROBABILITY WEIGHTED EXPECTED VALUE"
    t.font = Font(name="Arial", bold=True, size=11, color=WHITE)
    t.fill = fill(NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 1, 22)

    ws.merge_cells("A2:H2")
    s = ws["A2"]
    s.value = (f"Bull {PROB_BULL:.0%} | Base {PROB_BASE:.0%} | Bear {PROB_BEAR:.0%} | "
               f"Expected Value = (Bull×{PROB_BULL:.0%}) + (Base×{PROB_BASE:.0%}) + "
               f"(Bear×{PROB_BEAR:.0%}) | Change probabilities in yellow cells")
    s.font = Font(name="Arial", size=9, color=WHITE)
    s.fill = fill(BLUE)
    s.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 2, 14)

    ws.merge_cells("A3:H3")
    ph = ws["A3"]
    ph.value = "PROBABILITY INPUTS — Edit yellow cells"
    ph.font = Font(name="Arial", bold=True, size=9, color=WHITE)
    ph.fill = fill(ACCENT)
    ph.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    rh(ws, 3, 14)

    for col, txt in [(2,"Case"),(3,"Probability"),(4,"Description")]:
        cell(ws, 4, col, txt, bg=ACCENT, fc=WHITE, bold=True)
    ws.merge_cells("D4:H4")
    cell(ws, 4, 1, bg=ACCENT)
    rh(ws, 4, 14)

    prob_rows = [
        ("BULL CASE", PROB_BULL, GREEN_S,
         "Signed 12GW confirmed (OpenAI 6GW + Meta 6GW) + Oracle 50K GPU Q3 2026 removes demand risk. "
         "MI350 at Oracle scale + MI450 H2 2026 (CDNA 5, HBM4, 432GB) + MI500 2027 launch. "
         "EPYC record 46.2% server CPU revenue share — AMD Data Center $5.8B now exceeds Intel $5.1B. "
         "Hyperscaler AI capex $690-725B (+77% Y/Y) environment sustained. 45% probability: execution delivers."),
        ("BASE CASE", PROB_BASE, BLUE_S,
         "Confirmed deals execute on MI350/MI450 schedule with normal ramp delays. Strong Data Center growth. "
         "EPYC CPU share holds near record 46.2%. Gradual NI margin expansion as Xilinx amortization declines. "
         "12GW contracted pipeline provides revenue floor; 40% probability: measured execution on confirmed backlog."),
        ("BEAR CASE", PROB_BEAR, RED_S,
         "AMD execution risk: MI450 yield failures at scale, HBM4 supply constraints (SK Hynix/Samsung CoWoS "
         "capacity), ROCm software failing CUDA parity for OpenAI/Meta workloads, macro collapse causing "
         "enterprise AI capex freeze, geopolitical/TSMC risk. Demand risk is largely eliminated by signed "
         "12GW+ contracts — Bear = internal execution failure, not lost orders. 15% probability."),
    ]
    for i, (case, prob, bg, desc) in enumerate(prob_rows):
        r = 5 + i
        cell(ws, r, 1, bg=bg)
        cell(ws, r, 2, case, bg=bg, bold=True, align="left")
        c = ws.cell(row=r, column=3)
        c.value = prob; c.font = Font(name="Arial", bold=True, size=9, color="0070C0")
        c.fill = fill("FFFF00"); c.number_format = "0%"
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(f"D{r}:H{r}")
        cell(ws, r, 4, desc, bg=bg, fc=MUTED, size=8, align="left")
        rh(ws, r, 14)

    cell(ws, 8, 1, bg=WHITE)
    cell(ws, 8, 2, "Sum (must = 100%)", bg=WHITE, bold=True, align="left")
    c = ws.cell(row=8, column=3)
    c.value = PROB_BULL + PROB_BASE + PROB_BEAR
    c.font = Font(name="Arial", bold=True, size=9)
    c.number_format = "0%"; c.fill = fill(WHITE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("D8:H8")
    chk = "✓ 100% — Valid" if abs(PROB_BULL+PROB_BASE+PROB_BEAR-1.0) < 0.001 else "⚠ Does not sum to 100%"
    cell(ws, 8, 4, chk, bg=WHITE, bold=True, align="left")
    rh(ws, 8, 14)

    ws.merge_cells("A10:H10")
    sch = ws["A10"]
    sch.value = "SCENARIO COMPARISON BY YEAR — Bull vs Base vs Bear"
    sch.font = Font(name="Arial", bold=True, size=9, color=WHITE)
    sch.fill = fill(ACCENT)
    sch.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    rh(ws, 10, 14)

    cell(ws, 11, 1, bg=ACCENT)
    cell(ws, 11, 2, "Metric", bg=ACCENT, fc=WHITE, bold=True)
    for i, yr in enumerate(YEARS):
        cell(ws, 11, 3+i, str(yr), bg=ACCENT, fc=WHITE, bold=True)
    cell(ws, 11, 8, "EXPECTED VALUE 2030", bg=ACCENT, fc=WHITE, bold=True, size=8)
    rh(ws, 11, 16)

    bull = compute(BULL); base = compute(BASE); bear = compute(BEAR)
    ev_spl = [bull[3][i]*PROB_BULL + base[3][i]*PROB_BASE + bear[3][i]*PROB_BEAR for i in range(5)]
    ev_sph = [bull[4][i]*PROB_BULL + base[4][i]*PROB_BASE + bear[4][i]*PROB_BEAR for i in range(5)]
    ev_rev = [bull[0][i]*PROB_BULL + base[0][i]*PROB_BASE + bear[0][i]*PROB_BEAR for i in range(5)]
    ev_eps = [bull[2][i]*PROB_BULL + base[2][i]*PROB_BASE + bear[2][i]*PROB_BEAR for i in range(5)]

    sc_blocks = [
        (f"▶ BULL CASE ({PROB_BULL:.0%})", GREEN_S, *bull),
        (f"▶ BASE CASE ({PROB_BASE:.0%})", BLUE_S,  *base),
        (f"▶ BEAR CASE ({PROB_BEAR:.0%})", RED_S,   *bear),
    ]

    ROW = 12
    for sc_name, sc_bg, revs, nis, epss, spls, sphs, cagrl, cagrh in sc_blocks:
        ws.merge_cells(f"A{ROW}:H{ROW}")
        sh = ws[f"A{ROW}"]
        sh.value = sc_name
        sh.font = Font(name="Arial", bold=True, size=9, color=WHITE)
        sh.fill = fill(BLUE)
        sh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        rh(ws, ROW, 16); ROW += 1

        pw_rows = [
            ("  Revenue ($B)",       [f"{v:.2f}B" for v in revs], None,        False),
            ("  EPS (Non-GAAP)",     [f"${int(round(v))}" for v in epss], None, False),
            ("  Share Price Low",    spls, '"$"#,##0', True),
            ("  Share Price High",   sphs, '"$"#,##0', True),
            ("  CAGR Low",           cagrl, "+0.0%;-0.0%;—", False),
            ("  CAGR High",          cagrh, "+0.0%;-0.0%;—", False),
        ]
        for label, vals, fmt, bold in pw_rows:
            is_price = "Price" in label
            bg = GOLD_BG if is_price else (WHITE if ROW % 2 == 0 else GRAY)
            fc = GOLD_FN if is_price else BLACK
            cell(ws, ROW, 1, bg=bg)
            cell(ws, ROW, 2, label, bg=bg, fc=fc, bold=bold, align="left")
            for i, v in enumerate(vals):
                if isinstance(v, str):
                    cell(ws, ROW, 3+i, v, bg=bg, fc=fc, bold=bold)
                else:
                    c = ws.cell(row=ROW, column=3+i)
                    c.value = v; c.font = Font(name="Arial", bold=bold, size=9, color=fc)
                    c.fill = fill(bg); c.number_format = fmt if fmt else ""
                    c.alignment = Alignment(horizontal="center", vertical="center")
            last = vals[4]
            if isinstance(last, str):
                cell(ws, ROW, 8, last, bg=bg, fc=fc, bold=True)
            else:
                c8 = ws.cell(row=ROW, column=8)
                c8.value = last; c8.fill = fill(bg)
                c8.font = Font(name="Arial", bold=True, size=9, color=fc)
                c8.number_format = fmt if fmt else ""
                c8.alignment = Alignment(horizontal="center", vertical="center")
            rh(ws, ROW, 14); ROW += 1
        ROW += 1

    ws.merge_cells(f"A{ROW}:H{ROW}")
    ev = ws[f"A{ROW}"]
    ev.value = "PROBABILITY WEIGHTED EXPECTED VALUE — Updates automatically when probabilities change"
    ev.font = Font(name="Arial", bold=True, size=9, color=WHITE)
    ev.fill = fill(NAVY)
    ev.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    rh(ws, ROW, 16); ROW += 1

    ev_cagrl = [(ev_spl[i]/ENTRY_PRICE)**(1/(i+1))-1 for i in range(5)]
    ev_cagrh = [(ev_sph[i]/ENTRY_PRICE)**(1/(i+1))-1 for i in range(5)]

    ev_rows = [
        ("  Expected Revenue ($B)",   [f"{v:.2f}B" for v in ev_rev], None,        False),
        ("  Expected EPS (Non-GAAP)", [f"${int(round(v))}" for v in ev_eps], None, True),
        ("  Expected Share Price Low", ev_spl,  '"$"#,##0', True),
        ("  Expected Share Price High",ev_sph,  '"$"#,##0', True),
        ("  Expected CAGR Low",        ev_cagrl,"+0.0%;-0.0%;—", False),
        ("  Expected CAGR High",       ev_cagrh,"+0.0%;-0.0%;—", False),
    ]
    for label, vals, fmt, bold in ev_rows:
        is_price = "Price" in label
        bg = GOLD_BG if is_price else PROB_BG
        fc = GOLD_FN if is_price else BLACK
        cell(ws, ROW, 1, bg=bg)
        cell(ws, ROW, 2, label, bg=bg, fc=fc, bold=bold, align="left")
        for i, v in enumerate(vals):
            if isinstance(v, str):
                cell(ws, ROW, 3+i, v, bg=bg, fc=fc, bold=bold)
            else:
                c = ws.cell(row=ROW, column=3+i)
                c.value = v; c.font = Font(name="Arial", bold=bold, size=9, color=fc)
                c.fill = fill(bg); c.number_format = fmt if fmt else ""
                c.alignment = Alignment(horizontal="center", vertical="center")
        last = vals[4]
        if isinstance(last, str):
            cell(ws, ROW, 8, last, bg=bg, fc=fc, bold=True)
        else:
            c8 = ws.cell(row=ROW, column=8)
            c8.value = last; c8.fill = fill(bg)
            c8.font = Font(name="Arial", bold=True, size=9, color=fc)
            c8.number_format = fmt if fmt else ""
            c8.alignment = Alignment(horizontal="center", vertical="center")
        rh(ws, ROW, 14); ROW += 1

    ws.freeze_panes = "A12"

    ws.merge_cells(f"A{ROW+1}:H{ROW+1}")
    foot = ws[f"A{ROW+1}"]
    foot.value = (f"Built: {GENERATED} | Entry ${ENTRY_PRICE} | "
                  f"Shares {SHARES}B | Non-GAAP | Script: build_amd_projection_engine_v5.py")
    foot.font = Font(name="Arial", size=7, color=MUTED, italic=True)
    foot.alignment = Alignment(horizontal="left")

# ==============================================================================
# TAB 4 — ASSUMPTIONS
# Narrative justifications for each scenario/year combination.
# Labels are generated dynamically from BULL/BASE/BEAR dicts so they always
# match the Inputs tab — never hardcode growth rates here.
# ==============================================================================
def build_assumptions(wb):
    ws = wb.create_sheet("Assumptions")
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 90

    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value = f"AMD 5-YEAR PROJECTION — FULL NARRATIVE ASSUMPTIONS  [{VERSION}]"
    t.font = Font(name="Arial", bold=True, size=11, color=WHITE)
    t.fill = fill(NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 1, 20)

    ws.merge_cells("A2:C2")
    s = ws["A2"]
    s.value = (f"{VERSION}: Bull {PROB_BULL:.0%} / Base {PROB_BASE:.0%} / Bear {PROB_BEAR:.0%} | "
               f"Labels auto-generated from Inputs tab constants")
    s.font = Font(name="Arial", size=9, color=WHITE)
    s.fill = fill(BLUE)
    s.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 2, 14)

    # Each entry: (section_label, bg, [(year_idx, pct_val, narrative), ...])
    # ── BULL CASE NARRATIVES ───────────────────────────────────────────────────
    # Source: AMD Q1 2026 income statement What's New tab, May 2026.
    # Product roadmap: MI350 shipping (CDNA 4, 3nm, 288GB HBM3E) →
    #   MI450 H2 2026 production (CDNA 5, HBM4, 432GB) →
    #   MI500 2027 launch (CDNA 6, 2nm, HBM4E; MI455X training/MI430X HPC).
    # EPYC: AMD Data Center $5.8B passed Intel $5.1B in Q1 2026. Record 46.2% server CPU revenue share.
    # Macro: Hyperscaler AI capex $690-725B in 2026 (+77% Y/Y). Microsoft $190B, Amazon $200B, Google $190B, Meta $145B.
    # Confirmed deals: OpenAI 6GW total (Oct 6, 2025; 1GW initial H2 2026 MI450 Series, 160M warrants).
    #                  Meta 6GW total (Feb 24, 2026; 1GW initial H2 2026 custom MI450-based GPU + EPYC Venice, 160M warrants).
    #                  Oracle 50K GPUs Q3 2026 (Oct 14, 2025; MI450 Series + EPYC Venice + Pensando Vulcano, Helios rack).
    BULL_REV_NOTES = [
        "Q1 actual +37.9% Y/Y ($10.25B). Q2 guided $11.2B (+9.2% Q/Q). H1 locked at ~$21.45B. "
        "H2 acceleration driven by: OpenAI 6GW total agreement (1GW initial H2 2026 MI450 Series, 160M warrant shares, Oct 6 2025). "
        "Meta 6GW total (1GW initial H2 2026 custom MI450-based GPU + EPYC Venice, 160M warrant shares, Feb 24 2026). "
        "Oracle 50K GPUs Q3 2026 (MI450 Series + EPYC Venice + Pensando Vulcano, Helios rack, Oct 14 2025). "
        "12GW+ committed pipeline de-risks demand — execution is the only variable. "
        "EPYC Turin record 46.2% server CPU revenue share; AMD Data Center $5.8B now exceeds Intel $5.1B.",
        "MI450 full production volume year. OpenAI/Meta 6GW ramps begin scaling toward full commitment. "
        "MI500 launches H1 2027 (CDNA 6, 2nm, HBM4E — MI455X for training/inference, MI430X for HPC/Sovereign AI). "
        "EPYC Verano enters 1:1 CPU:GPU market. AMD Data Center at ~$8B+/quarter run rate. "
        "Server CPU TAM at 35%+ AMD share compounds with GPU wins in the same rack.",
        "MI500 replacement cycle underway. Inference explosion — MI500's HBM4E bandwidth advantage vs Nvidia. "
        "OpenAI/Meta 6GW deployments driving multi-year GPU refresh cycles at hyperscaler scale. "
        "Xilinx amortization fully zero — ~260bps structural GAAP tailwind. "
        "EPYC Verano at full volume. AMD increasingly wins full-stack (CPU+GPU) data center deployments.",
        "Sovereign AI deployed globally. Physical AI at scale. ROCm achieving enterprise software parity with CUDA. "
        "Server CPU TAM $90B+ at AMD 35%+ share. OpenAI/Meta expanding beyond initial 1GW tranches toward full 6GW. "
        "Second-source GPU adoption accelerates as hyperscalers diversify away from 100% Nvidia dependency.",
        "Second AI infrastructure wave. Physical AI industrial scale. ROCm enterprise software $8-12B at 80%+ margins. "
        "MI-series replacement cycles compound annually. OpenAI/Meta 6GW commitments driving renewal conversations. "
        "EPYC and Instinct together position AMD as the full-stack Data Center platform for the next decade.",
    ]
    BULL_NI_NOTES = [
        "H1 Non-GAAP NI margin locked at ~22.1% (Q1 actual). Q2 Non-GAAP operating margin guided ~27% → implies NI ~24%. "
        "H2 must average ~33.5% to reach 28% full-year — achievable via MI450 Helios rack-scale margins "
        "(HBM4 memory commands premium ASPs from OpenAI/Meta hyperscaler contracts) + EPYC CPU margin expansion.",
        "MI450 pricing premium (432GB HBM4 vs MI350's 288GB HBM3E commands ASP uplift on OpenAI/Meta/Oracle orders). "
        "MI500 early ramp at premium pricing. Helios year-two yield improvements reduce cost per GPU. "
        "Operating leverage on $72B revenue base. 3pts above base case.",
        "Xilinx amortization reaches zero — single largest structural margin event (~260bps automatic improvement). "
        "Software revenue $1-2B at 70%+ margins begins to scale. R&D as % of revenue compressing. "
        "Data Center 65%+ of revenue carrying the blended gross margin above 55%.",
        "Software scaling to $4-6B. Sovereign AI premium pricing — government contracts command premium margins. "
        "R&D growing ~15% vs revenue +40%. EPYC CPU margin at 50%+ contributes to NI expansion.",
        "Software $8-12B at 80%+ margins = ~450bps NI expansion. Second AI wave economics better than first — "
        "higher software attach rate, better yields, lower relative R&D. R&D below 15% of revenue.",
    ]
    # ── BASE CASE NARRATIVES ───────────────────────────────────────────────────
    BASE_REV_NOTES = [
        "H1 locked at ~$21.45B (Q1 $10.25B actual + Q2 $11.2B guided). H2 execution: MI450 production "
        "ramp partially delayed into early 2027. Oracle 50K GPUs Q3 2026 executes as scheduled. "
        "OpenAI/Meta 1GW initial tranches deploy H2 2026 but full ramp to 6GW follows normal milestone gating. "
        "EPYC CPU record 46.2% revenue share is the baseline. Hyperscaler $690-725B capex environment sustained.",
        "MI450 volume ramp year. MI500 partial launch and early adoption. EPYC Verano launch adds "
        "CPU TAM upside. AMD Data Center maintains double-digit sequential growth. "
        "OpenAI/Meta backlog provides demand floor — ramp pace is the variable.",
        "Inference explosion — AMD's HBM4 memory bandwidth advantage addresses inference workload requirements. "
        "Xilinx amortization fully zero — ~260bps structural improvement. Embedded segment recovery contributes.",
        "Agentic AI in enterprise production. EPYC at 35%+ server CPU revenue share. "
        "AMD's second-source GPU positioning compounds as hyperscalers formalize dual-vendor GPU policies.",
        "Natural deceleration at ~$160B revenue scale. Second AI infrastructure wave provides organic demand floor. "
        "Software revenue emerging at meaningful scale. EPYC and Instinct together create a durable "
        "Data Center franchise less dependent on any single product cycle.",
    ]
    BASE_NI_NOTES = [
        "H1 Non-GAAP NI locked at ~22.1% (Q1 actual). Q2 op. margin guided ~27% → NI ~23-24%. "
        "Full-year 25% requires H2 margin expansion from MI450 production and EPYC pricing power. "
        "Oracle/OpenAI/Meta initial deployments H2 contribute high-margin GPU revenue.",
        "MI450 full production yield improvements. MI500 premium pricing partially captured in H2 2027. "
        "Operating leverage on ~42% revenue growth — R&D and MG&A growing slower than revenue.",
        "Xilinx amortization zero — single biggest structural NI margin event. ~260bps automatic improvement "
        "on a $93B+ revenue base. Data Center 65%+ of revenue mix driving blended gross margin toward 55%.",
        "Pure operational execution. R&D growing ~15% vs revenue +35%. Enterprise and sovereign AI "
        "premium pricing. EPYC at 35%+ share contributes stable high-margin CPU revenue.",
        "Most conservative — only 100bps from 2029. AMD approaching operational maturity. "
        "Gross margin near 55%. Approaching Broadcom-level net margin (~35%) as software attach grows.",
    ]
    # ── BEAR CASE NARRATIVES ───────────────────────────────────────────────────
    # Bear = execution risk only. Demand risk largely eliminated by signed 12GW+ contracts.
    BEAR_REV_NOTES = [
        "H1 locked at ~$21.45B. Bear: MI450 H2 2026 yield challenges at Helios rack scale — "
        "TSMC N3E advanced packaging yields below targets; some H2 revenue pushes to early 2027. "
        "Oracle 50K GPU Q3 2026 fulfillment partially delayed. OpenAI/Meta 1GW initial tranches deploy "
        "but with hardware yield-related quality issues. Only 2pts below base in 2026; EPYC provides revenue floor.",
        "MI450 yield failure at scale: HBM4 supply constraints — SK Hynix/Samsung CoWoS/SoIC capacity "
        "consumed by competing orders (GB300 priority). ROCm fails to reach CUDA parity for transformer "
        "training workloads; OpenAI/Meta cannot migrate critical fine-tuning pipelines. "
        "The signed 1GW initial orders deploy but ramp toward full 6GW is deferred to MI500 generation. "
        "AMD Data Center GPU revenue stalls despite confirmed backlog.",
        "ROCm still failing CUDA parity in key enterprise workloads after $1B+ software investment. "
        "Macro shock — enterprise AI capex freezes as ROI scrutiny intensifies post-$690B 2026 spend. "
        "TSMC geopolitical risk (Taiwan Strait tension) creates supply chain uncertainty. "
        "HBM4 supply still constrained. AMD Instinct revenue stagnant; EPYC CPU is the only growth line.",
        "Two consecutive years of GPU growth collapse. MI450 fully obsoleted before volume ramp delivers. "
        "AMD restructuring: R&D reallocated toward EPYC leadership and inference-optimized products. "
        "EPYC CPU TAM provides the floor: $15-20B/year prevents revenue collapse. AMD growing at CPU market rates. "
        "OpenAI/Meta 6GW contracts create renewal opportunity — deferred, not cancelled.",
        "Slight re-acceleration as AMD restructures around EPYC leadership and next-gen Instinct "
        "addresses specific inference workloads. ROCm 3.0 improves software story. "
        "12GW OpenAI/Meta contracted pipeline creates renewal conversation with new product generation. "
        "Recovery is real but slow — AMD at CPU company valuations with a GPU option.",
    ]
    BEAR_NI_NOTES = [
        "H1 Non-GAAP NI locked at ~22.1% (Q1 actual). H2 MI450 yield issues mean less high-margin GPU "
        "revenue — mix shifts toward EPYC/Client/Embedded at lower blended margins. Full year stays ~22%.",
        "Revenue collapses 18% — severe operating deleverage on largely fixed R&D (~$9-10B). "
        "R&D consumes a larger % of $54B bear revenue. Worst-case: revenue AND margins compressing simultaneously.",
        "Revenue stabilizes. AMD begins R&D reallocation toward EPYC and inference-focused products. "
        "Xilinx amortization approaching zero — structural relief provides 1pt NI margin recovery. "
        "Cost base partially restructured to match the lower revenue trajectory.",
        "Xilinx fully zero — structural margin floor. AMD restructured: R&D declining toward 20% of revenue. "
        "EPYC becoming the primary margin anchor (~50% product margins at 35%+ CPU share). "
        "2pt NI margin improvement driven by cost discipline, not revenue growth.",
        "Recovery green shoots. Leaner post-restructuring cost structure + Xilinx fully amortized. "
        "Next-gen Instinct early shipments. AMD at 26% net margin — profitable and stable "
        "but well below the 38-43% Base/Bull scenarios.",
    ]

    sections = [
        ("▶ BULL CASE — Revenue Growth",    GREEN_S, BULL["rev_growth"], BULL_REV_NOTES),
        ("▶ BULL CASE — Net Income Margin", GREEN_S, BULL["ni_margin"],  BULL_NI_NOTES),
        ("▶ BASE CASE — Revenue Growth",    BLUE_S,  BASE["rev_growth"], BASE_REV_NOTES),
        ("▶ BASE CASE — Net Income Margin", BLUE_S,  BASE["ni_margin"],  BASE_NI_NOTES),
        ("▶ BEAR CASE — Revenue Growth",    RED_S,   BEAR["rev_growth"], BEAR_REV_NOTES),
        ("▶ BEAR CASE — Net Income Margin", RED_S,   BEAR["ni_margin"],  BEAR_NI_NOTES),
    ]

    ROW = 4
    for section_label, sc_bg, vals, notes in sections:
        ws.merge_cells(f"A{ROW}:C{ROW}")
        sh = ws[f"A{ROW}"]
        sh.value = section_label
        sh.font = Font(name="Arial", bold=True, size=9, color=BLACK)
        sh.fill = fill(sc_bg)
        sh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        rh(ws, ROW, 14); ROW += 1

        for i, (yr, pct, note) in enumerate(zip(YEARS, vals, notes)):
            bg = WHITE if i % 2 == 0 else GRAY
            cell(ws, ROW, 1, bg=bg)
            cell(ws, ROW, 2, f"{yr} — {pct:.0%}", bg=bg, fc="0070C0", bold=True, align="left")
            c = ws.cell(row=ROW, column=3)
            c.value = note
            c.font = Font(name="Arial", size=9, color=BLACK)
            c.fill = fill(bg)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            rh(ws, ROW, 28)
            ROW += 1

        ROW += 1  # blank gap between sections

    ws.merge_cells(f"A{ROW}:C{ROW}")
    foot = ws[f"A{ROW}"]
    foot.value = (f"Script: build_amd_projection_engine_v5.py | Built: {GENERATED} | "
                  f"Labels auto-generated from DATA BLOCK constants — always match Inputs tab")
    foot.font = Font(name="Arial", size=7, color=MUTED, italic=True)
    foot.alignment = Alignment(horizontal="left")

# ==============================================================================
# TAB 5 — WORKFLOW
# ==============================================================================
def build_workflow(wb):
    ws = wb.create_sheet("Workflow")
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 75

    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value = "AMD PROJECTION — WORKFLOW & VERSION LOG"
    t.font = Font(name="Arial", bold=True, size=11, color=WHITE)
    t.fill = fill(NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 1, 20)

    ws.merge_cells("A2:C2")
    s = ws["A2"]
    s.value = "To regenerate: pull scripts from GitHub → update DATA BLOCK → run → present for download"
    s.font = Font(name="Arial", size=9, color=WHITE)
    s.fill = fill(BLUE)
    s.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 2, 14)

    def section_header(row, txt):
        ws.merge_cells(f"A{row}:C{row}")
        c = ws[f"A{row}"]
        c.value = txt
        c.font = Font(name="Arial", bold=True, size=9, color=WHITE)
        c.fill = fill(ACCENT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        rh(ws, row, 14)

    def wf_row(row, label, detail, bg=WHITE, wrap=True):
        cell(ws, row, 1, bg=bg)
        cell(ws, row, 2, label, bg=bg, bold=True, align="left")
        c = ws.cell(row=row, column=3)
        c.value = detail
        c.font = Font(name="Arial", size=9, color=BLACK)
        c.fill = fill(bg)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=wrap)
        rh(ws, row, 28 if wrap else 14)

    section_header(4, "QUARTERLY WORKFLOW")
    wf_rows = [
        ("Step 1 — AMD Reports",
         "AMD reports earnings → Tony downloads PDFs from ir.amd.com → saves to Drive Q[X] folder"),
        ("Step 2 — Income Statement",
         "Say: 'Run the Q[X] 20XX income statement analysis'\n"
         "→ Builds AMD_Q[X]_Income_Statement_v1.xlsx (build_amd_income_v1.py pattern)"),
        ("Step 3 — Update Projection",
         "Say: 'Update the AMD projection for Q[X] 20XX'\n"
         "→ Pull build_amd_projection_engine_v5.py from GitHub → update DATA BLOCK → rebuild → present for download"),
        ("Full Earnings Day",
         "Say: 'It's Q[X] 20XX earnings day — run the full AMD analysis'\n"
         "→ Income Statement + Projection update in sequence"),
        ("PDF Limitation Note",
         "AMD PDFs on cloudfront.net are blocked from direct fetch. "
         "Tony downloads → saves to Drive Q[X] folder → Claude reads from Drive."),
    ]
    for i, (label, detail) in enumerate(wf_rows):
        bg = WHITE if i % 2 == 0 else GRAY
        wf_row(5 + i, label, detail, bg=bg)

    section_header(11, "DRIVE FOLDER & SCRIPT LINKS")
    link_rows = [
        ("AMD Financials — Master Folder",
         "https://drive.google.com/drive/folders/1i1GdOQGreQuxv2s6xoqILyY_XIqKuX5a"),
        ("Q1 2026 Quarter Folder",
         "https://drive.google.com/drive/folders/1Df4g1sgPkwy_7BwBLjNTsYNDPh28RiWc"),
        ("Q2 2026 Quarter Folder",
         "https://drive.google.com/drive/folders/1D_GOhT8kgZBTlVPA_7VXvO0-lR5ZyTbB"),
        ("build_amd_projection_engine_v5.py",
         "GitHub: tonyvasquez1/bmnr-dashboard — branch claude/setup-amd-financials-q7F3t"),
        ("build_amd_income_v2.py",
         "GitHub: tonyvasquez1/bmnr-dashboard — Q1 2026 income statement builder (confirmed deals update)"),
        ("AMD_MANIFEST.txt",
         "Master constraints — never rebuild from memory, never overwrite, always self-test"),
    ]
    for i, (label, detail) in enumerate(link_rows):
        bg = WHITE if i % 2 == 0 else GRAY
        wf_row(12 + i, label, detail, bg=bg, wrap=False)

    section_header(19, "VERSION LOG")
    version_rows = [
        ("v1 — April 2026",
         "Initial build. Bull 30% / Base 45% / Bear 25%."),
        ("v2 — April 2026",
         "Column alignment fixes. Workflow and Data Sources tabs added."),
        ("v3 — April 2026",
         "Tailwinds & Headwinds tab added."),
        ("v4 — May 2026",
         "CPU TAM $60B baseline. Bear 25% / Bull 30%."),
        ("v5 — May 2026",
         "CPU TAM doubled to $120B (Lisa Su Q1 2026). Bull 35% / Bear 20%. Rev growth +2-3pts/yr above v4."),
        ("v5.1 — May 2026",
         "Formatting fixed: no yellow, whole dollar formats, correct column widths, scenario comparison formulas corrected, centered alignment, Workflow tab permanent."),
        ("v6 (script v2) — May 2026",
         "GitHub migration. Assumptions, Workflow, Data Sources tabs added to match Drive file. "
         "Assumption labels auto-generated from DATA BLOCK — always in sync with Inputs. "
         "Script renamed build_amd_projection_engine_v2.py."),
        ("v7 (script v3) — May 2026",
         "Coherence pass: Assumptions updated to reflect Q1 2026 income statement What's New tab. "
         "MI350/MI450/MI500 product roadmap explicitly reflected per scenario year. EPYC record 46.2% "
         "server CPU revenue share incorporated. Vera Rubin / hyperscaler capex $690-725B reflected. "
         "Data Sources expanded with income statement primary sources. Script: build_amd_projection_engine_v3.py."),
        ("v8 (script v4) — May 2026",
         "Projection tab layout fixes: column H widened 2→14 (CAGR column was invisible), "
         "data columns C-G widened 14→16, row heights 14→16, comparison header wrap enabled. "
         "Script: build_amd_projection_engine_v4.py."),
        ("v10 (script v5) — May 2026",
         "Confirmed hyperscaler deal update: OpenAI 6GW total (Oct 6 2025, 1GW initial H2 2026 MI450 Series, "
         "160M warrant shares). Meta 6GW total (Feb 24 2026, 1GW initial H2 2026 custom MI450-based GPU + "
         "EPYC Venice, 160M warrant shares). Oracle 50K GPUs Q3 2026 (Oct 14 2025, MI450 Series + EPYC Venice "
         "+ Pensando Vulcano, Helios rack). SHARES raised 1.650→1.750B (320M total warrants net of partial "
         "buyback). Probabilities: Bull 35%→45%, Base 45%→40%, Bear 20%→15%. Bear narrative reframed: "
         "demand risk eliminated by signed 12GW+ contracts; Bear = execution risk only "
         "(MI450 yield failure, HBM4 supply constraints, ROCm parity, macro, geopolitical/TSMC). "
         "Script: build_amd_projection_engine_v5.py."),
    ]
    for i, (ver, detail) in enumerate(version_rows):
        bg = WHITE if i % 2 == 0 else GRAY
        wf_row(20 + i, ver, detail, bg=bg)

    section_header(30, "PENDING ITEMS")
    pending_rows = [
        ("Mac App",
         "Port AMD projection to Vercel. Will automate PDF download and Drive upload entirely."),
        ("GAAP vs Non-GAAP Discussion",
         "Promised deep-dive on GAAP vs non-GAAP analytical framework."),
        ("Revenue Growth Derivation",
         "Segment data, TAM analysis, PEG ratios, comparable company analysis."),
        ("CPU TAM $120B Discussion",
         "EPYC standalone modeling, 1:1 CPU:GPU ratio, Verano ramp timeline."),
        ("Upload xlsx to Drive",
         "Blocked by file size — Mac app will solve this."),
        ("Warrant Dilution Tracking",
         "320M total warrants (OpenAI 160M + Meta 160M). 19.4% max dilution at full vest. "
         "SHARES=1.750B reflects ~100M net after partial buyback offset. Monitor vest schedules quarterly."),
    ]
    for i, (item, detail) in enumerate(pending_rows):
        bg = WHITE if i % 2 == 0 else GRAY
        wf_row(31 + i, item, detail, bg=bg)

# ==============================================================================
# TAB 6 — DATA SOURCES
# ==============================================================================
def build_data_sources(wb):
    ws = wb.create_sheet("Data Sources")
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 60

    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = "AMD PROJECTION — DATA SOURCES & REFERENCE LINKS"
    t.font = Font(name="Arial", bold=True, size=11, color=WHITE)
    t.fill = fill(NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 1, 20)

    ws.merge_cells("A2:D2")
    s = ws["A2"]
    s.value = ("Click any link to open  |  "
               "Sources used for projection assumptions, income statement analysis and consensus estimates")
    s.font = Font(name="Arial", size=9, color=WHITE)
    s.fill = fill(BLUE)
    s.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 2, 14)

    def section_header(row, txt):
        ws.merge_cells(f"A{row}:D{row}")
        c = ws[f"A{row}"]
        c.value = txt
        c.font = Font(name="Arial", bold=True, size=9, color=WHITE)
        c.fill = fill(ACCENT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        rh(ws, row, 14)

    def col_headers(row, bg=ACCENT):
        for col, txt in [(2,"Source"),(3,"What It Provides"),(4,"Direct Link")]:
            cell(ws, row, col, txt, bg=bg, fc=WHITE, bold=True, align="left")
        cell(ws, row, 1, bg=bg)
        rh(ws, row, 14)

    def src_row(row, name, desc, link, bg=WHITE):
        cell(ws, row, 1, bg=bg)
        cell(ws, row, 2, name, bg=bg, align="left", bold=True)
        cell(ws, row, 3, desc, bg=bg, fc=MUTED, align="left", size=8)
        c = ws.cell(row=row, column=4)
        c.value = link
        c.font = Font(name="Arial", size=8, color="0070C0", underline="single")
        c.fill = fill(bg)
        c.alignment = Alignment(horizontal="left", vertical="center")
        rh(ws, row, 14)

    # CONFIRMED HYPERSCALER DEALS — NEW SECTION
    section_header(4, "CONFIRMED HYPERSCALER DEALS — PRIMARY SOURCES")
    col_headers(5)
    deal_sources = [
        ("AMD + OpenAI Strategic Partnership",
         "Oct 6, 2025. AMD-OpenAI 6GW total compute agreement. 1GW initial deployment H2 2026, "
         "MI450 Series accelerators. AMD issues 160M warrant shares to OpenAI. "
         "Source for SHARES=1.750B (+100M net of partial buyback), Bull 2026 H2 acceleration, "
         "and demand-risk elimination in probability weight rationale.",
         "AMD Newsroom — AMD and OpenAI Announce Strategic Partnership (Oct 6, 2025)"),
        ("AMD + Meta Expanded Strategic Partnership",
         "Feb 24, 2026. AMD-Meta 6GW total compute agreement. 1GW initial deployment H2 2026, "
         "custom MI450-based GPU + EPYC Venice/Verano. AMD issues 160M warrant shares to Meta. "
         "Source for SHARES=1.750B (second 160M tranche), Meta GPU + CPU cross-sell, "
         "and 12GW confirmed pipeline total in probability rationale.",
         "AMD Newsroom — AMD and Meta Announce Expanded Strategic Partnership (Feb 24, 2026)"),
        ("AMD + Oracle AI World Announcement",
         "Oct 14, 2025. Oracle commits to 50K AMD GPUs initial deployment Q3 2026. "
         "MI450 Series accelerators + EPYC Venice CPUs + Pensando Vulcano DPUs. Helios rack architecture. "
         "Expanding to 200K+ GPUs in 2027. Source for Oracle Cloud 2026 GPU revenue ramp assumption "
         "and MI350 at Oracle scale narrative.",
         "AMD Newsroom — AMD at Oracle CloudWorld (Oct 14, 2025)"),
    ]
    for i, (name, desc, link) in enumerate(deal_sources):
        src_row(6 + i, name, desc, link, bg=WHITE if i % 2 == 0 else GRAY)

    # AMD OFFICIAL
    section_header(10, "AMD OFFICIAL SOURCES")
    col_headers(11)
    amd_sources = [
        ("AMD Investor Relations",
         "Official earnings, guidance, financial tables. Primary source for all quarterly actuals.",
         "https://ir.amd.com"),
        ("AMD Q1 2026 Earnings Press Release",
         "Q1 2026 actuals: Revenue $10,253M, Gross Margin 53%, Op Income $1,476M, EPS $0.84. "
         "Q2 2026 guidance: Revenue $11.2B, Non-GAAP op margin ~27%, Non-GAAP EPS $0.96.",
         "https://ir.amd.com/news-events/press-releases"),
        ("AMD IR Calendar",
         "Q2 2026 earnings expected late July / early August 2026.",
         "https://ir.amd.com/news-events/ir-calendar"),
        ("AMD SEC Filings",
         "10-K (FY2025 actual revenue $34.64B), 10-Q, 8-K filings.",
         "https://ir.amd.com/financial-information/sec-filings"),
        ("AMD Newsroom / CES 2026",
         "MI500 2027 roadmap confirmed. OpenAI/Oracle/Meta hyperscaler deals. ROCm roadmap.",
         "https://www.amd.com/en/newsroom"),
    ]
    for i, (name, desc, link) in enumerate(amd_sources):
        src_row(12 + i, name, desc, link, bg=WHITE if i % 2 == 0 else GRAY)

    # AMD PRODUCT ROADMAP
    section_header(18, "AMD PRODUCT ROADMAP SOURCES")
    col_headers(19)
    roadmap_sources = [
        ("MI350 Launch — Data Center Dynamics",
         "MI350 now shipping. CDNA 4, 3nm, 288GB HBM3E. Deployed at scale by Oracle Cloud Infrastructure. "
         "Fastest-ramping AMD product in company history. Source for 2026 GPU revenue driver assumption.",
         "datacenterdynamics.com — AMD launches Instinct MI350"),
        ("MI450 Sampling — WCCFtech",
         "AMD sampling MI450 to lead customers for H2 2026 production deployment. "
         "CDNA 5, HBM4, 432GB memory (~50% more bandwidth vs MI350). Helios rack. "
         "Source for 2026 Bull H2 acceleration and 2027 Base/Bull GPU ramp assumptions.",
         "wccftech.com — AMD sampling MI450 GPUs"),
        ("MI500 Roadmap — WCCFtech",
         "MI500 planned for 2027. CDNA 6, 2nm process (TSMC N2), HBM4E. "
         "Two variants: MI455X (training/inference) and MI430X (HPC/Sovereign AI). "
         "Source for 2027-2028 Bull case NI margin expansion and product ASP assumptions.",
         "wccftech.com — AMD 2026-2027 AI Roadmap MI400 & MI500"),
        ("EPYC Record 46.2% Server Revenue Share — WCCFtech",
         "EPYC Turin record 46.2% server CPU revenue share in Q1 2026 (from 28.8% Q4 2025). "
         "AMD Data Center revenue $5.8B surpassed Intel Data Center $5.1B for first time. "
         "Source for CPU TAM parallel growth engine assumption across all scenarios.",
         "wccftech.com — AMD EPYC Record 46.2% Server Revenue Share"),
        ("EPYC Verano / Venice CPU Roadmap — Tom's Hardware",
         "EPYC Verano roadmap (1:1 CPU:GPU market). Venice (next-gen Zen architecture). "
         "Source for 2027+ CPU TAM contribution to Bull/Base revenue growth assumptions.",
         "https://www.tomshardware.com"),
    ]
    for i, (name, desc, link) in enumerate(roadmap_sources):
        src_row(20 + i, name, desc, link, bg=WHITE if i % 2 == 0 else GRAY)

    # COMPETITIVE INTELLIGENCE
    section_header(26, "COMPETITIVE INTELLIGENCE")
    col_headers(27)
    comp_sources = [
        ("Nvidia Q4 FY2026 Earnings",
         "Q4 FY2026: Total revenue $68.1B (+73% Y/Y), Data Center $62.3B (+75% Y/Y). "
         "91% of total revenue from Data Center. Blackwell ramp driving acceleration. "
         "Q1 FY2027 guidance: $78B (+15% Q/Q sequential). Validates hyperscaler AI GPU demand. "
         "Source for Bull/Base case TAM environment and Bear case Vera Rubin displacement risk.",
         "Nvidia Q4 FY2026 Earnings Release (period ending Jan 2026)"),
        ("Nvidia Vera Rubin (GR100) Roadmap",
         "Post-Blackwell architecture targeted 2026-2027. Expected to widen performance/efficiency lead. "
         "Source for Bear case ROCm parity risk — if ROCm cannot match CUDA for OpenAI/Meta workloads "
         "when Vera Rubin samples, ramp of signed contracts may stall.",
         "Nvidia product announcements / public roadmap"),
        ("Intel Q1 2026 DCAI Earnings",
         "Intel Data Center & AI Q1 2026: $5.1B (+22% Y/Y). AI-specific: $750M (+88% Y/Y). "
         "AMD Data Center $5.8B now exceeds Intel DCAI for first time. "
         "Source for EPYC CPU share gain validation and Bear case Intel Xeon 6 recovery risk.",
         "Intel Q1 2026 Earnings Release (period ending Mar 2026)"),
    ]
    for i, (name, desc, link) in enumerate(comp_sources):
        src_row(28 + i, name, desc, link, bg=WHITE if i % 2 == 0 else GRAY)

    # SECTOR / MACRO
    section_header(32, "SECTOR & MACRO SOURCES")
    col_headers(33)
    macro_sources = [
        ("Hyperscaler AI Capex 2026 — Public Disclosures",
         "Combined 2026 AI capex: $690-725B (+77% Y/Y vs ~$410B in 2025). "
         "Microsoft $190B, Amazon $200B, Google $190B, Meta $145B. "
         "~30% goes to AI accelerators/GPUs (~$207-217B TAM). "
         "Source for Bull/Base case revenue growth environment across 2026-2027.",
         "Microsoft/Amazon/Google/Meta Q4 2025 and Q1 2026 earnings calls"),
        ("HBM Memory Cost Inflation",
         "Microsoft attributed $25B of 2026 AI budget to rising memory/chip costs. "
         "HBM3E/HBM4 pricing elevated due to SK Hynix/Samsung supply constraints. "
         "Source for Bear case: HBM4 supply constraints throttle MI450 production volume (execution risk).",
         "Microsoft Q3 FY2026 Earnings Call (Feb 2026)"),
    ]
    for i, (name, desc, link) in enumerate(macro_sources):
        src_row(34 + i, name, desc, link, bg=WHITE if i % 2 == 0 else GRAY)

    # CONSENSUS ESTIMATES
    section_header(37, "CONSENSUS ESTIMATES")
    col_headers(38)
    consensus_sources = [
        ("Yahoo Finance — AMD",
         "Revenue & EPS consensus, analyst count, estimate revisions.",
         "https://finance.yahoo.com/quote/AMD/analysis/"),
        ("Earnings Whispers",
         "Whisper number vs published consensus — useful for beat/miss calibration.",
         "https://www.earningswhispers.com/stocks/amd"),
        ("Seeking Alpha — AMD",
         "Estimate revisions, earnings transcripts, analyst ratings.",
         "https://seekingalpha.com/symbol/AMD/earnings"),
        ("Visible Alpha",
         "Segment-level consensus: Data Center, Client, Embedded revenue breakdowns.",
         "https://visiblealpha.com"),
    ]
    for i, (name, desc, link) in enumerate(consensus_sources):
        src_row(39 + i, name, desc, link, bg=WHITE if i % 2 == 0 else GRAY)

    # HISTORICAL DATA
    section_header(44, "HISTORICAL DATA")
    col_headers(45)
    hist_sources = [
        ("StockAnalysis — AMD Financials",
         "Historical income statement 5+ years, segment breakdowns.",
         "https://stockanalysis.com/stocks/amd/financials/"),
        ("Macrotrends — Revenue",
         "Historical quarterly AMD revenue 10+ years.",
         "https://www.macrotrends.net/stocks/charts/AMD/advanced-micro-devices/revenue"),
        ("Macrotrends — EPS",
         "Historical quarterly AMD diluted EPS 10+ years.",
         "https://www.macrotrends.net/stocks/charts/AMD/advanced-micro-devices/eps-earnings-per-share-diluted"),
        ("Investing.com — AMD",
         "Real-time price, technical analysis, historical price data.",
         "https://www.investing.com/equities/adv-micro-device"),
    ]
    for i, (name, desc, link) in enumerate(hist_sources):
        src_row(46 + i, name, desc, link, bg=WHITE if i % 2 == 0 else GRAY)

    # GOOGLE DRIVE
    section_header(51, "GOOGLE DRIVE")
    col_headers(52)
    drive_sources = [
        ("AMD Financials — Master Folder",
         "All AMD documents organized by quarter.",
         "https://drive.google.com/drive/folders/1i1GdOQGreQuxv2s6xoqILyY_XIqKuX5a"),
        ("Q1 2026 Quarter Folder",
         "Press release PDF, financial tables PDF.",
         "https://drive.google.com/drive/folders/1Df4g1sgPkwy_7BwBLjNTsYNDPh28RiWc"),
        ("Q2 2026 Quarter Folder",
         "Ready for Q2 2026 earnings documents.",
         "https://drive.google.com/drive/folders/1D_GOhT8kgZBTlVPA_7VXvO0-lR5ZyTbB"),
    ]
    for i, (name, desc, link) in enumerate(drive_sources):
        src_row(53 + i, name, desc, link, bg=WHITE if i % 2 == 0 else GRAY)

# ==============================================================================
# SELF-TEST
# ==============================================================================
def self_test(out_path):
    import os
    from openpyxl import load_workbook

    errors = []
    if not os.path.exists(out_path):
        errors.append(f"FAIL: file not created at {out_path}")
    else:
        size = os.path.getsize(out_path)
        if size < 20000:
            errors.append(f"FAIL: file too small ({size} bytes) — engine incomplete")

        wb = load_workbook(out_path, data_only=True)
        for tab in ["Inputs", "Projection", "Probability Weighted",
                    "Assumptions", "Workflow", "Data Sources"]:
            if tab not in wb.sheetnames:
                errors.append(f"FAIL: tab '{tab}' missing")

        ws = wb["Projection"]
        found_bull_spl = False
        found_base_spl = False
        for row in ws.iter_rows(values_only=True):
            for c in row:
                # Bull 2030 SPL with SHARES=1.750: ~$2,897
                if isinstance(c, (int, float)) and 2700 < c < 3100: found_bull_spl = True
                # Base 2030 SPL with SHARES=1.750: ~$1,261
                if isinstance(c, (int, float)) and 1150 < c < 1350: found_base_spl = True
        if not found_bull_spl:
            errors.append("FAIL: Bull 2030 SPL ~$2,897 not found (check SHARES=1.750)")
        if not found_base_spl:
            errors.append("FAIL: Base 2030 SPL ~$1,261 not found (check SHARES=1.750)")

        ws_a = wb["Assumptions"]
        found_assumption = found_mi450_assume = found_epyc_assume = False
        found_openai_assume = found_meta_assume = False
        for row in ws_a.iter_rows(values_only=True):
            for c in row:
                if isinstance(c, str):
                    if "BULL CASE" in c:  found_assumption = True
                    if "MI450" in c:      found_mi450_assume = True
                    if "46.2" in c:       found_epyc_assume = True
                    if "OpenAI" in c:     found_openai_assume = True
                    if "Meta" in c:       found_meta_assume = True
        if not found_assumption:
            errors.append("FAIL: Assumptions tab content missing")
        if not found_mi450_assume:
            errors.append("FAIL: Assumptions tab — MI450 reference missing")
        if not found_epyc_assume:
            errors.append("FAIL: Assumptions tab — EPYC 46.2% reference missing")
        if not found_openai_assume:
            errors.append("FAIL: Assumptions tab — OpenAI deal reference missing")
        if not found_meta_assume:
            errors.append("FAIL: Assumptions tab — Meta deal reference missing")

        ws_w = wb["Workflow"]
        found_workflow = False
        for row in ws_w.iter_rows(values_only=True):
            for c in row:
                if isinstance(c, str) and "QUARTERLY WORKFLOW" in c:
                    found_workflow = True
        if not found_workflow:
            errors.append("FAIL: Workflow tab content missing")

        ws_d = wb["Data Sources"]
        found_ir = found_mi450 = found_epyc = found_openai_ds = found_oracle_ds = False
        for row in ws_d.iter_rows(values_only=True):
            for c in row:
                if isinstance(c, str):
                    if "ir.amd.com" in c:    found_ir = True
                    if "MI450" in c:          found_mi450 = True
                    if "46.2" in c:           found_epyc = True
                    if "OpenAI" in c:         found_openai_ds = True
                    if "Oracle" in c:         found_oracle_ds = True
        if not found_ir:
            errors.append("FAIL: Data Sources tab — ir.amd.com link missing")
        if not found_mi450:
            errors.append("FAIL: Data Sources tab — MI450 reference missing")
        if not found_epyc:
            errors.append("FAIL: Data Sources tab — EPYC 46.2% reference missing")
        if not found_openai_ds:
            errors.append("FAIL: Data Sources tab — OpenAI deal source missing")
        if not found_oracle_ds:
            errors.append("FAIL: Data Sources tab — Oracle deal source missing")

        # Verify probability weights sum to 1.0
        if abs(PROB_BULL + PROB_BASE + PROB_BEAR - 1.0) > 0.001:
            errors.append(f"FAIL: Probabilities do not sum to 100% ({PROB_BULL+PROB_BASE+PROB_BEAR:.0%})")

    if errors:
        print("\n" + "="*60)
        print("SELF-TEST FAILED — DO NOT SAVE TO GITHUB")
        print("="*60)
        for e in errors: print(f"  {e}")
        print("="*60)
        return False
    else:
        size = os.path.getsize(out_path)
        print("\n" + "="*60)
        print("SELF-TEST PASSED — safe to commit to GitHub")
        print(f"  File: {out_path}")
        print(f"  Size: {size:,} bytes")
        print(f"  Tabs: Inputs, Projection, Probability Weighted,")
        print(f"        Assumptions, Workflow, Data Sources — all confirmed")
        print(f"  Bull SPL ~$2,897, Base SPL ~$1,261: confirmed (SHARES=1.750B)")
        print(f"  OpenAI/Meta/Oracle deals: confirmed in Assumptions + Data Sources")
        print(f"  MI450/EPYC 46.2% coherence: confirmed in Assumptions + Data Sources")
        print(f"  Probabilities: Bull {PROB_BULL:.0%} / Base {PROB_BASE:.0%} / Bear {PROB_BEAR:.0%} = 100%")
        print("="*60)
        return True

# ==============================================================================
# MAIN
# ==============================================================================
def build():
    wb = Workbook()
    wb.remove(wb.active)
    build_inputs(wb)
    build_projection(wb)
    build_probability(wb)
    build_assumptions(wb)
    build_workflow(wb)
    build_data_sources(wb)

    wb["Inputs"].sheet_properties.tabColor               = "2E75B6"
    wb["Projection"].sheet_properties.tabColor           = "1A7A3A"
    wb["Probability Weighted"].sheet_properties.tabColor = "7B2D8B"
    wb["Assumptions"].sheet_properties.tabColor          = "E67300"
    wb["Workflow"].sheet_properties.tabColor             = "5A5A5A"
    wb["Data Sources"].sheet_properties.tabColor         = "006B6B"

    wb.active = wb["Inputs"]

    out = "/mnt/user-data/outputs/AMD_5Year_Projection_v10.xlsx"
    wb.save(out)

    bull = compute(BULL); base = compute(BASE); bear = compute(BEAR)
    ev_spl = [bull[3][i]*PROB_BULL+base[3][i]*PROB_BASE+bear[3][i]*PROB_BEAR for i in range(5)]
    ev_sph = [bull[4][i]*PROB_BULL+base[4][i]*PROB_BASE+bear[4][i]*PROB_BEAR for i in range(5)]
    print(f"Saved: {out}")
    print(f"\nKEY 2030 OUTPUTS — verify on open:")
    print(f"BULL: Rev=${bull[0][4]:.2f}B  NI=${bull[1][4]:.2f}B  EPS=${int(round(bull[2][4]))}  SPL=${bull[3][4]:,.0f}  SPH=${bull[4][4]:,.0f}  CAGR={bull[5][4]:+.1%}/{bull[6][4]:+.1%}")
    print(f"BASE: Rev=${base[0][4]:.2f}B  NI=${base[1][4]:.2f}B  EPS=${int(round(base[2][4]))}  SPL=${base[3][4]:,.0f}  SPH=${base[4][4]:,.0f}  CAGR={base[5][4]:+.1%}/{base[6][4]:+.1%}")
    print(f"BEAR: Rev=${bear[0][4]:.2f}B  NI=${bear[1][4]:.2f}B  EPS=${int(round(bear[2][4]))}  SPL=${bear[3][4]:,.0f}  SPH=${bear[4][4]:,.0f}  CAGR={bear[5][4]:+.1%}/{bear[6][4]:+.1%}")
    print(f"EV:   SPL=${ev_spl[4]:,.0f}  SPH=${ev_sph[4]:,.0f}  CAGR={(ev_spl[4]/ENTRY_PRICE)**(1/5)-1:+.1%}/{(ev_sph[4]/ENTRY_PRICE)**(1/5)-1:+.1%}")
    return out

if __name__ == "__main__":
    out = build()
    passed = self_test(out)
    if not passed:
        import sys
        sys.exit(1)
