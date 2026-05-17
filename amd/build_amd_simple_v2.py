# ══════════════════════════════════════════════════════════════════════════════
# ADVANCED MICRO DEVICES, INC. (AMD) — SIMPLE INCOME STATEMENT ANALYSIS
# build_amd_simple_v2.py
# Q1 2026 | Quarter Ended March 28, 2026 | Reported May 5, 2026
# v2: Python/openpyxl xlsx replacing the original JS/docx version
# ══════════════════════════════════════════════════════════════════════════════

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ── COLOR PALETTE ─────────────────────────────────────────────────────────────
NAVY       = "1A1A2E"
BLUE       = "1F4E79"
ACCENT     = "2E75B6"
WHITE      = "FFFFFF"
GRAY       = "F2F2F2"
GREEN      = "00B050"
AMBER      = "FF8C00"
RED        = "C00000"
GREEN_LT   = "E8F5E9"
AMBER_LT   = "FFF3E0"
RED_LT     = "FFEBEE"
MEASURE_BG = "EBF3FB"
BLACK      = "000000"
MUTED      = "666666"
SECTION_BG = "1F4E79"
STRENGTH_BG = "E8F5E9"
CONCERN_BG  = "FFEBEE"
ASSESS_BG   = "FFF8E1"

# ── PRIORITY ROWS (bold line item cell) ───────────────────────────────────────
PRIORITY = {
    "Net Revenue", "Total Cost of Sales", "Gross Profit", "Gross Margin",
    "Research and Development", "Marketing, General and Administrative",
    "Total Operating Expenses", "Operating Income", "Net Income",
    "Diluted Earnings Per Share",
}

# ── ROW DATA ──────────────────────────────────────────────────────────────────
# (label, q1_2026, q4_2025, q1_2025, yy_color, actual_color, is_pct, note,
#  what_it_measures, num_fmt_override)
#
# yy_color:        "green"|"amber"|"red"|None  → Y/Y Change + Y/Y % cells
# actual_color:    "green"|"amber"|"red"|None  → Q1 2026 actual cell only
# is_pct:          True = margin/% row (Y/Y shown in bps)
# num_fmt_override: None=auto | "$" | "#,##0" for shares | explicit string

ROWS = [
    # ── REVENUE ───────────────────────────────────────────────────────────────
    ("Net Revenue",
     10253, 10270, 7438, "green", None, False,
     "+37.9% Y/Y to $10.25B — essentially flat Q/Q (-$17M, -0.2%) despite typical Q1 "
     "seasonal weakness, confirming underlying demand strength. Data Center was the primary "
     "Y/Y driver (MI300X/MI325X GPU ramp + EPYC CPU server share gains vs Intel). "
     "Gaming declined (consumer GPU + semi-custom both down). Embedded recovering slowly. "
     "Client (Ryzen) grew modestly. The revenue story is increasingly a Data Center story — "
     "that segment is the engine and the margin expansion driver simultaneously.",
     "Total dollars sold across all products and geographies; top-line scale",
     None),

    ("Cost of Sales",
     4576, 4433, 3451, "green", None, False,
     "+32.6% Y/Y — meaningfully slower than revenue growth of +37.9%. Data Center GPU "
     "(MI300X/MI325X) carries higher gross margins than Gaming GPUs or Embedded chips. "
     "As Data Center becomes a larger share of mix, cost of sales as a % of revenue "
     "structurally improves. Q/Q increase (+$143M) reflects product mix seasonality. "
     "Green: cost growing slower than revenue is the definition of gross margin expansion.",
     "Direct production costs; growing slower than revenue = gross margin expansion",
     None),

    ("  Amortization — Acquisition Intangibles (COGS)",
     261, 260, 251, None, None, False,
     "Essentially flat Q/Q and modestly higher Y/Y ($251M → $261M). Non-cash amortization "
     "of intangibles from the 2022 Xilinx acquisition ($48.8B) — primarily Xilinx IP, "
     "technology licenses, and customer relationships allocated to cost of sales. "
     "Will decline gradually as assets fully amortize. Non-cash, directionally positive.",
     "Non-cash Xilinx acquisition intangibles in COGS; amortizes down over time",
     None),

    ("Total Cost of Sales",
     4837, 4693, 3702, "green", None, False,
     "+30.7% Y/Y — the slowest growing cost line in the income statement. Revenue grew "
     "7 percentage points faster than total COGS, directly driving +300bps gross margin "
     "expansion from 50% to 53%. When total COGS grows materially slower than revenue, "
     "gross margin expands mechanically. This is the gross-level operating leverage story.",
     "All-in production costs including non-cash amortization; key gross margin driver",
     None),

    ("Gross Profit",
     5416, 5577, 3736, "green", None, False,
     "+45.0% Y/Y — grew 700bps faster than revenue (+37.9%). $1.68B more gross profit "
     "than Q1 2025. Q/Q decline (-$161M) driven by seasonal revenue dip and product mix. "
     "The Y/Y frame is the right one: gross profit growing faster than revenue is positive "
     "operating leverage at the gross level. Dollar earning power is expanding rapidly.",
     "Absolute earning power after production costs; measures scale and efficiency",
     None),

    ("Gross Margin",
     53.0, 54.0, 50.0, "green", None, True,
     "+300bps Y/Y expansion (50% → 53%). Q/Q contracted -100bps (54% → 53%) due to "
     "seasonal product mix shift — Q4 typically carries more premium Data Center mix. "
     "Structural Y/Y driver: MI300X/MI325X GPU carries meaningfully higher gross margins "
     "than Gaming GPUs or Embedded. As Data Center becomes a larger revenue share, "
     "blended gross margin expands structurally. Management long-term target: approach "
     "Nvidia-like margins as the MI300 series matures. Current 53% vs Nvidia's ~75%+ "
     "signals significant further expansion potential if AMD continues gaining Data Center share.",
     "Production efficiency; % of each revenue dollar retained after COGS",
     None),

    # ── OPERATING EXPENSES ────────────────────────────────────────────────────
    ("Research and Development",
     2397, 2330, 1728, "amber", None, False,
     "+38.7% Y/Y — grew slightly faster than revenue (+37.9%). No operating leverage "
     "here, but the gap is thin (80bps). AMD must invest heavily in chip design to "
     "stay competitive: MI350 and MI400/CDNA4 in development, EPYC Turin (Gen 5) ramping, "
     "Zen 6 in development. R&D is the competitive lifeblood — cutting it would be fatal. "
     "The concern is trajectory: R&D needs to grow slower than revenue over time to produce "
     "operating leverage. At +38.7% vs +37.9% revenue, the gap is essentially flat. "
     "Amber: matched revenue growth — no leverage, but not a breakdown either.",
     "Product competitiveness investment; must grow slower than revenue to show leverage",
     None),

    ("Marketing, General and Administrative",
     1253, 1198, 886, "red", None, False,
     "+41.4% Y/Y — fastest growing cost line in the statement, outpacing revenue growth "
     "of +37.9% by 350bps. This is the primary operating leverage failure point this quarter. "
     "MG&A should grow significantly slower than revenue as the business scales — AMD's "
     "existing infrastructure should absorb more revenue without proportional G&A increases. "
     "Q/Q increase (+$55M) also notable. Management's path to 30%+ non-GAAP operating "
     "margins requires MG&A discipline. Red: any operating cost growing faster than revenue "
     "violates the operating leverage principle and needs to reverse direction.",
     "Overhead and sales cost discipline; must grow slower than revenue for leverage",
     None),

    ("  Amortization — Acquisition Intangibles (OpEx)",
     290, 297, 316, None, None, False,
     "Declining Y/Y ($316M → $290M, -8.2%) and Q/Q ($297M → $290M). Non-cash amortization "
     "of Xilinx intangibles in operating expenses — technology licenses, trade names, and "
     "customer relationships. This line will continue declining as assets fully amortize, "
     "providing a structural GAAP tailwind to operating income over time. "
     "No color: non-cash, declining — directionally positive.",
     "Non-cash Xilinx acquisition intangibles in OpEx; declining = GAAP operating tailwind",
     None),

    ("Total Operating Expenses",
     3940, 3825, 2930, "green", None, False,
     "+34.5% Y/Y — grew slower than revenue (+37.9%). Operating leverage confirmed in "
     "aggregate despite MG&A overshooting. The Xilinx amortization decline (-$26M Y/Y) "
     "provides a structural offset that pulls total OpEx growth below revenue growth even "
     "when MG&A disappoints. Net message: operating leverage exists in aggregate, "
     "but MG&A discipline is required to make it durable and improving.",
     "All operating costs below gross profit; slower growth than revenue = leverage",
     None),

    ("Operating Income",
     1476, 1752, 806, "green", None, False,
     "+83.1% Y/Y — more than double the rate of revenue growth (+37.9%). $670M more "
     "operating income on $2.8B more revenue = every incremental revenue dollar generates "
     "significantly more operating income than the base. GAAP operating margin: "
     "10.8% → 14.4% (+360bps Y/Y). Q/Q declined ($1,752M → $1,476M) on typical Q1 "
     "seasonality — Q4 is AMD's strongest quarter. Non-GAAP operating margin guided "
     "~27% for Q2 2026 (excludes $551M Xilinx amortization). The +83% operating income "
     "growth is the operating leverage thesis fully confirmed.",
     "Core business profitability before financing costs and taxes; leverage signal",
     None),

    # ── BELOW THE LINE ────────────────────────────────────────────────────────
    ("Interest Expense",
     -37, -36, -20, None, None, False,
     "Modest increase Q/Q (-$37M vs -$36M) and Y/Y growth from -$20M. AMD carries "
     "relatively light long-term debt (~$1.7B notes outstanding) — a significant balance "
     "sheet advantage. Interest expense is not a meaningful earnings drag at AMD's scale. "
     "The Y/Y increase from -$20M to -$37M reflects debt issued for general corporate "
     "purposes; still modest relative to $1.5B+ operating income.",
     "Cost of AMD's debt; modest relative to operating income — balance sheet strength",
     None),

    ("Other Income (Expense), net",
     165, 358, 39, None, None, False,
     "Q1: $165M. Q4 2025 was unusually elevated ($358M) — included gains on equity "
     "investments or licensing settlements not repeated in Q1. Q1 2025 was $39M, "
     "making Q1 2026 ($165M) look strong in that context. Primarily interest income "
     "on AMD's $5.1B cash balance ($100M+ annually at current rates) plus equity "
     "investment gains. Q/Q decline (-$193M) is not a concern — Q4 was the outlier.",
     "Interest income on cash + equity investment gains; Q4 2025 was the outlier quarter",
     None),

    ("Income Before Income Taxes",
     1604, 2074, 825, None, None, False,
     "$1.60B pre-tax income. Q/Q decline ($2,074M → $1,604M) driven by Q1 seasonal "
     "revenue dip and lower other income vs Q4's elevated level. Y/Y improvement "
     "($825M → $1,604M, +94%) driven entirely by operating income growth. "
     "No separate Y/Y color: derived line — the operating and other income rows above "
     "already capture the drivers.",
     "Total profitability before tax; reflects operating performance and capital structure",
     None),

    ("Income Tax Provision",
     238, 455, 123, None, None, False,
     "+$115M Y/Y in tax expense, roughly in line with income growth. Effective tax rate "
     "approximately consistent Y/Y. Q/Q large decline ($455M → $238M) mirrors Q4's "
     "elevated pre-tax income. AMD benefits from R&D tax credits, foreign income "
     "structures, and Xilinx-related deductions that keep the effective rate manageable. "
     "No color: support line.",
     "Tax obligation; R&D credits and deductions keep effective rate competitive",
     None),

    ("Equity Income in Investee",
     6, 1, 7, None, None, False,
     "Small income ($6M) from equity method investments. Essentially flat and immaterial "
     "relative to AMD's overall profitability. No color: support line.",
     "Small income from minority equity stakes; immaterial to the investment thesis",
     None),

    ("Income from Continuing Operations",
     1372, 1620, 709, None, None, False,
     "+93.5% Y/Y — nearly doubled. Q/Q decline mirrors seasonality and other income "
     "normalization. This is GAAP net income from AMD's core ongoing business, "
     "excluding discontinued operations. No separate Y/Y color: summary of above lines.",
     "Core business GAAP net income excluding discontinued operations",
     None),

    ("Income (Loss) from Discontinued Operations",
     11, -109, None, None, None, False,
     "Q1 2026: $11M gain. Q4 2025: ($109M) loss (write-down/impairment on divested assets). "
     "Q1 2025: none. These relate to Xilinx-era divested business units being wound down. "
     "The Q4 loss was a one-time impairment; Q1 gain reflects partial recovery or "
     "proceeds from asset sales. Will zero out once fully resolved. Non-recurring.",
     "Gains/losses from divested business units; non-recurring, expected to zero out",
     None),

    ("Net Income",
     1383, 1511, 709, "green", None, False,
     "+95.1% Y/Y — nearly doubled on 37.9% revenue growth. Q/Q decline ($1,511M → $1,383M) "
     "reflects Q1 seasonality plus Q4's elevated other income. GAAP net margin: "
     "9.5% → 13.5% (+400bps Y/Y). The near-doubling of net income on 38% revenue growth "
     "is textbook operating leverage. Shareholders are seeing disproportionate benefit "
     "from revenue scale — no preferred stock or debt complexity diluting the result.",
     "Total GAAP profitability; bottom-line operating leverage confirmed",
     None),

    # ── EPS ───────────────────────────────────────────────────────────────────
    ("Basic EPS — Continuing Operations",
     0.84, 1.00, 0.44, None, None, False,
     "Q1 2026: $0.84. Q4 2025: $1.00 (Q4 seasonally strongest). Q1 2025: $0.44. "
     "Reference line — diluted EPS below is the primary investor metric.",
     "Per-share profit from core operations on basic share count; reference line",
     None),

    ("Basic EPS — Discontinued Operations",
     0.01, -0.07, None, None, None, False,
     "Q1 2026: $0.01 gain. Q4 2025: ($0.07) loss from write-down. Non-recurring. "
     "Will zero out as discontinued operations are resolved.",
     "Per-share discontinued ops impact; non-recurring",
     None),

    ("Basic Earnings Per Share",
     0.85, 0.93, 0.44, None, None, False,
     "Q1 2026: $0.85. Q4: $0.93. Q1 2025: $0.44. Uses ~1,631M basic shares. "
     "Reference line — diluted EPS below is the primary metric.",
     "Total per-share profit on basic shares only; diluted EPS is the primary metric",
     None),

    ("Diluted EPS — Continuing Operations",
     0.83, 0.99, 0.44, None, None, False,
     "Q1 2026: $0.83. Diluted count: 1,650M. Basic-diluted spread of only 19M shares "
     "(1.2%) — AMD has clean, minimal dilution from stock compensation. Reference line.",
     "Per-share profit from core operations on fully diluted share count",
     None),

    ("Diluted EPS — Discontinued Operations",
     0.01, -0.07, None, None, None, False,
     "Same as basic — $0.01 gain in Q1 2026. Non-recurring.",
     "Per-share discontinued ops impact on diluted basis",
     None),

    ("Diluted Earnings Per Share",
     0.84, 0.92, 0.44, "green", None, False,
     "+90.9% Y/Y — from $0.44 to $0.84. Q/Q: $0.92 → $0.84 (Q1 seasonal decline). "
     "Non-GAAP diluted EPS for Q2 2026 guided at $0.96 midpoint by management. "
     "AMD's dilution profile is clean: 19M share basic-diluted spread (1.2%). "
     "No preferred stock complexity, no acquisition financing dilution. "
     "Green: +90.9% Y/Y EPS growth on +37.9% revenue = shareholders are seeing the "
     "full, unimpeded benefit of operating leverage. This is the cleanest EPS story "
     "in the large-cap semiconductor sector right now.",
     "Primary per-share metric; +90.9% Y/Y confirms shareholders benefit from leverage",
     None),

    # ── SHARE COUNT ───────────────────────────────────────────────────────────
    ("Basic Shares (millions)",
     1631, 1630, 1620, None, None, False,
     "Flat Q/Q (1,631M vs 1,630M) and only +0.7% Y/Y growth (+11M shares). AMD's share "
     "count is very stable — stock buybacks roughly offset SBC issuance. This is a "
     "significant advantage: EPS growth closely tracks net income growth without "
     "meaningful dilution headwinds dragging down per-share results.",
     "Basic share count; minimal growth confirms AMD is not diluting shareholders",
     "#,##0"),

    ("Diluted Shares (millions)",
     1650, 1649, 1626, None, None, False,
     "1,650M diluted vs 1,649M Q4 and 1,626M Q1 2025. Basic-diluted spread: 19M (1.2%). "
     "Very clean dilution profile. AMD's stock options and RSUs create minimal dilution "
     "relative to the share base. Compare to peers with large convertible debt or "
     "preferred stock — AMD's capital structure is straightforward and shareholder-friendly.",
     "Fully diluted count; 1.2% basic-diluted spread is minimal and clean",
     "#,##0"),
]

# ── COLUMN LAYOUT ─────────────────────────────────────────────────────────────
# A: Line Item    B: Y/Y Change    C: Y/Y %
# D: Q1 2026      E: Q4 2025       F: Q1 2025
# G: Notes        H: What This Measures

COL_WIDTHS = {
    "A": 38,   # Line Item
    "B": 14,   # Y/Y Change
    "C": 10,   # Y/Y %
    "D": 13,   # Q1 2026
    "E": 13,   # Q4 2025
    "F": 13,   # Q1 2025
    "G": 90,   # Notes
    "H": 40,   # What This Measures
}

def fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def section_header(ws, row, text, bg, cols=8):
    ws.merge_cells(f"A{row}:{chr(64+cols)}{row}")
    cell = ws[f"A{row}"]
    cell.value = text
    cell.font = Font(name="Arial", bold=True, size=10, color=WHITE)
    cell.fill = fill(bg)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 18
    return row + 1

def bullet_row(ws, row, text, bg, indent=2, font_size=9, color=BLACK):
    ws.merge_cells(f"A{row}:H{row}")
    cell = ws[f"A{row}"]
    cell.value = text
    cell.font = Font(name="Arial", size=font_size, color=color)
    cell.fill = fill(bg)
    cell.alignment = Alignment(horizontal="left", vertical="top",
                               wrap_text=True, indent=indent)
    ws.row_dimensions[row].height = max(15, 13 + len(text) // 10)
    return row + 1

def auto_fmt(val, is_pct):
    if is_pct:
        return '0.0"%"'
    if val is None:
        return "@"
    if abs(val) < 5:
        return '$#,##0.00'   # EPS
    if abs(val) < 100:
        return '$#,##0.0'
    return '$#,##0'          # large millions

def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "Q1 2026 Income Statement"

    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # ── Title ─────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "ADVANCED MICRO DEVICES, INC. (AMD) — Q1 2026 INCOME STATEMENT ANALYSIS"
    t.font = Font(name="Arial", bold=True, size=13, color=WHITE)
    t.fill = fill(NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:H2")
    s = ws["A2"]
    s.value = ("Quarter Ended March 28, 2026  |  Reported May 5, 2026  |  "
               "Source: AMD Q1 2026 Earnings Press Release  |  GAAP figures in millions except per-share amounts")
    s.font = Font(name="Arial", size=9, color=WHITE)
    s.fill = fill(BLUE)
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 15

    # ── Column headers ─────────────────────────────────────────────────────────
    headers = ["Line Item", "Y/Y Change ($M)", "Y/Y %",
               "Q1 2026 ($M)", "Q4 2025 ($M)", "Q1 2025 ($M)",
               "Notes", "What This Measures"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = h
        cell.font = Font(name="Arial", bold=True, size=9, color=WHITE)
        cell.fill = fill(ACCENT)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 28

    yy_map = {
        "green": (GREEN, GREEN_LT),
        "amber": (AMBER, AMBER_LT),
        "red":   (RED,   RED_LT),
    }

    DATA_START = 4

    for i, row_data in enumerate(ROWS):
        (label, q1, q4, q1_py,
         yy_color, actual_color, is_pct, note, measure, fmt_override) = row_data

        row = DATA_START + i
        is_priority = label.strip() in PRIORITY
        is_indent = label.startswith("  ")
        bg = WHITE if i % 2 == 0 else GRAY

        def fmt(val):
            if fmt_override:
                return fmt_override
            return auto_fmt(val, is_pct)

        # A — Line Item
        a = ws.cell(row=row, column=1)
        a.value = label
        a.font = Font(name="Arial", bold=is_priority and not is_indent,
                      size=9 if not is_indent else 8,
                      color=MUTED if is_indent else BLACK,
                      italic=is_indent)
        a.fill = fill(bg)
        a.alignment = Alignment(horizontal="left", vertical="center",
                                indent=3 if is_indent else 1)

        # B — Y/Y Change (formula)
        b = ws.cell(row=row, column=2)
        if is_indent or (q1_py is None):
            b.value = "—"
            b.font = Font(name="Arial", size=9, color=MUTED)
            b.fill = fill(bg)
        else:
            b.value = f"=D{row}-F{row}"
            if is_pct:
                b.number_format = '+0.0;-0.0;"-"'
            elif fmt_override:
                b.number_format = '+#,##0;-#,##0;"-"'
            elif q1 is not None and abs(q1) < 5:
                b.number_format = '+$#,##0.00;-$#,##0.00;"-"'
            else:
                b.number_format = '+$#,##0;-$#,##0;"-"'

        b.alignment = Alignment(horizontal="right", vertical="center")

        # C — Y/Y %
        c = ws.cell(row=row, column=3)
        if is_indent or (q1_py is None):
            c.value = "—"
            c.font = Font(name="Arial", size=9, color=MUTED)
            c.fill = fill(bg)
        else:
            if is_pct:
                c.value = f"=D{row}-F{row}"
                c.number_format = '+0.0" bps";-0.0" bps";"-"'
            else:
                c.value = f'=IF(F{row}<>0,(D{row}-F{row})/ABS(F{row}),"N/A")'
                c.number_format = '+0.0%;-0.0%;"-"'
        c.alignment = Alignment(horizontal="right", vertical="center")

        # Apply Y/Y color to B and C
        if not is_indent and q1_py is not None and yy_color in yy_map:
            ink, bg_yy = yy_map[yy_color]
            for col in [2, 3]:
                cell = ws.cell(row=row, column=col)
                cell.fill = fill(bg_yy)
                cell.font = Font(name="Arial", size=9, color=ink, bold=True)
        elif not is_indent and (b.value != "—"):
            for col in [2, 3]:
                cell = ws.cell(row=row, column=col)
                cell.fill = fill(bg)
                cell.font = Font(name="Arial", size=9, color=BLACK)

        # D — Q1 2026 actual
        d = ws.cell(row=row, column=4)
        d.value = q1
        d.number_format = fmt(q1)
        d.alignment = Alignment(horizontal="right", vertical="center")
        if actual_color in yy_map:
            ink, bg_ac = yy_map[actual_color]
            d.fill = fill(bg_ac)
            d.font = Font(name="Arial", size=9, color=ink, bold=True)
        else:
            d.fill = fill(bg)
            d.font = Font(name="Arial", size=9, bold=is_priority and not is_indent,
                          color=BLACK)

        # E — Q4 2025
        e = ws.cell(row=row, column=5)
        if q4 is None:
            e.value = "—"
            e.font = Font(name="Arial", size=9, color=MUTED)
        else:
            e.value = q4
            e.number_format = fmt(q4)
            e.font = Font(name="Arial", size=9, color=BLACK)
        e.fill = fill(bg)
        e.alignment = Alignment(horizontal="right", vertical="center")

        # F — Q1 2025
        f = ws.cell(row=row, column=6)
        if q1_py is None:
            f.value = "—"
            f.font = Font(name="Arial", size=9, color=MUTED)
        else:
            f.value = q1_py
            f.number_format = fmt(q1_py)
            f.font = Font(name="Arial", size=9, color=BLACK)
        f.fill = fill(bg)
        f.alignment = Alignment(horizontal="right", vertical="center")

        # G — Notes
        g = ws.cell(row=row, column=7)
        g.value = note
        g.font = Font(name="Arial", size=8, color="333333")
        g.fill = fill(bg)
        g.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # H — What This Measures
        h = ws.cell(row=row, column=8)
        h.value = measure
        h.font = Font(name="Arial", size=8, color="1F4E79", italic=True)
        h.fill = fill(MEASURE_BG)
        h.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True, indent=1)

        ws.row_dimensions[row].height = max(45, min(180, 12 + note.count(" ") * 1.4))

    ws.freeze_panes = "B4"

    # ── Spacer after table ────────────────────────────────────────────────────
    next_row = DATA_START + len(ROWS) + 1
    ws.row_dimensions[next_row].height = 10
    next_row += 1

    # ══════════════════════════════════════════════════════════════════════════
    # STRENGTHS
    # ══════════════════════════════════════════════════════════════════════════
    next_row = section_header(ws, next_row, "▶  STRENGTHS", "1A7A3A")
    strengths = [
        "✓  Net Revenue +37.9% Y/Y to $10.25B — essentially flat Q/Q despite typical Q1 seasonal weakness. Demand strength across Data Center and Client confirmed.",
        "✓  Gross Margin expanded +300bps Y/Y (50% → 53%) — Data Center GPU mix shift driving structural, repeatable gross margin improvement.",
        "✓  Gross Profit +45.0% Y/Y — grew 700bps faster than revenue. Dollar earning power expanding rapidly.",
        "✓  Operating Income +83.1% Y/Y — more than double the revenue growth rate. GAAP operating margin: 10.8% → 14.4% (+360bps Y/Y). Textbook operating leverage.",
        "✓  Net Income +95.1% Y/Y — nearly doubled on 37.9% revenue growth. GAAP net margin: 9.5% → 13.5% (+400bps). Shareholders seeing full benefit.",
        "✓  Diluted EPS +90.9% Y/Y ($0.44 → $0.84) — clean EPS leverage with no preferred stock dilution complexity. Basic-diluted spread only 19M shares (1.2%).",
        "✓  Total OpEx +34.5% Y/Y vs revenue +37.9% — operating leverage confirmed in aggregate across all cost lines combined.",
        "✓  Xilinx amortization declining ($316M → $290M in OpEx, -8.2% Y/Y) — structural GAAP operating income tailwind that compounds over time.",
        "✓  Balance sheet strength: ~$5.1B cash, ~$1.7B long-term debt — fortress balance sheet enabling R&D investment and share repurchases simultaneously.",
        "✓  Share count nearly stable: diluted shares +1.5% Y/Y — EPS growth closely tracks net income growth without material dilution headwinds.",
        "✓  Non-GAAP diluted EPS Q2 2026 guided at $0.96 midpoint — implies continued strong growth with management confidence in the trajectory.",
    ]
    for s in strengths:
        next_row = bullet_row(ws, next_row, s, STRENGTH_BG)

    next_row += 1

    # ══════════════════════════════════════════════════════════════════════════
    # CONCERNS
    # ══════════════════════════════════════════════════════════════════════════
    next_row = section_header(ws, next_row, "▶  CONCERNS", "8B0000")
    concerns = [
        "✗  MG&A +41.4% Y/Y — fastest growing cost line, outpacing revenue (+37.9%) by 350bps. Primary operating leverage failure point this quarter. Needs to reverse direction.",
        "✗  R&D +38.7% Y/Y — essentially matched revenue growth (+37.9%). No operating leverage in the innovation engine yet. Must grow meaningfully slower than revenue over time.",
        "✗  Q/Q Operating Income declined ($1,752M → $1,476M, -15.8%). Seasonal pattern, but Q1 to Q4 progression needs to recover strongly.",
        "✗  Gaming segment declining — consumer GPU and semi-custom both down Y/Y. No near-term catalyst for reversal; Nvidia's consumer GPU lineup is dominant.",
        "✗  Embedded recovering slowly from trough — not yet contributing meaningfully to revenue growth or margin mix improvement.",
        "✗  Nvidia competition: MI300 series is competitive but Nvidia maintains commanding Data Center GPU lead with H200/B200/B300 series. AMD must execute flawlessly to continue gaining share.",
        "✗  GAAP vs Non-GAAP gap remains large: GAAP operating margin ~14.4% vs guided non-GAAP ~27%. The $551M+ quarterly Xilinx amortization continues to suppress reported GAAP margins significantly.",
        "✗  Other income volatile: Q4 2025 $358M → Q1 2026 $165M. Some earnings contribution is dependent on equity gains and licensing settlements, not core operations.",
    ]
    for c in concerns:
        next_row = bullet_row(ws, next_row, c, CONCERN_BG)

    next_row += 1

    # ══════════════════════════════════════════════════════════════════════════
    # OVERALL ASSESSMENT
    # ══════════════════════════════════════════════════════════════════════════
    next_row = section_header(ws, next_row, "▶  OVERALL ASSESSMENT", SECTION_BG)

    grade_text = "A−"
    ws.merge_cells(f"A{next_row}:H{next_row}")
    grade_cell = ws[f"A{next_row}"]
    grade_cell.value = f"Quarter Grade:  {grade_text}"
    grade_cell.font = Font(name="Arial", bold=True, size=14, color=NAVY)
    grade_cell.fill = fill(ASSESS_BG)
    grade_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[next_row].height = 28
    next_row += 1

    assessment_text = (
        "AMD delivered a strong operating quarter with operating leverage confirmed at every level that matters. "
        "Gross Margin expanded +300bps Y/Y as the Data Center GPU mix shift drives a structural, repeatable "
        "improvement in blended margins. Operating Income grew +83% on +38% revenue — more than double the "
        "revenue growth rate. Net Income nearly doubled. Diluted EPS grew +91% with clean, minimal dilution. "
        "These are not accident — they are the direct result of the MI300X ramp and EPYC server share gains "
        "creating a higher-margin, faster-growing revenue base.\n\n"
        "The one failure this quarter is MG&A: at +41.4% vs revenue +37.9%, the G&A cost line moved in the "
        "wrong direction. Every other operating cost line grew slower than revenue. MG&A is the outlier and "
        "the one item that keeps this from a clean A. If MG&A reverts to growing slower than revenue in Q2, "
        "the operating leverage story is fully intact.\n\n"
        "The GAAP/non-GAAP gap ($551M+ quarterly Xilinx amortization) creates an optics problem: "
        "14.4% GAAP operating margin vs ~27% non-GAAP guided for Q2 is a large gap that sophisticated "
        "investors understand but retail investors often find confusing. As the Xilinx amortization "
        "gradually declines, the GAAP picture will converge with the non-GAAP picture — a multi-year tailwind.\n\n"
        "Gaming declining and Embedded recovering slowly are real headwinds, but they are overwhelmed by "
        "the Data Center engine. The concentration risk is worth monitoring: AMD is increasingly a Data Center "
        "company, and any deceleration in that segment would have outsized impact on the overall numbers."
    )
    next_row = bullet_row(ws, next_row, assessment_text, ASSESS_BG, indent=2, font_size=9)
    next_row += 1

    # ══════════════════════════════════════════════════════════════════════════
    # LETTER GRADE FRAMEWORK
    # ══════════════════════════════════════════════════════════════════════════
    next_row = section_header(ws, next_row, "▶  LETTER GRADE FRAMEWORK — AMD (Q1 2026 BASELINE)", SECTION_BG)

    grades = [
        ("A",   "MG&A < revenue growth for 2+ consecutive quarters. Non-GAAP operating margin expands to 30%+. "
                "Gross Margin reaches 55%+ as Data Center mix grows. R&D leverage achieved. Gaming stabilizes."),
        ("A−",  "Current quarter: Operating Income +83% on +38% revenue. Net Income +95%. EPS +91%. "
                "Gross Margin +300bps Y/Y. Operating leverage confirmed in aggregate. MG&A the sole exception."),
        ("B+",  "MG&A growth remains elevated above revenue growth 2+ quarters. Revenue growth decelerates "
                "below 25%. Gross margin expansion stalls at 53-54%. Data Center growth moderates."),
        ("B",   "Revenue growth decelerates to teens. Gross margin contracts from 53% peak. Data Center GPU "
                "momentum slows materially. Operating income growth falls to near revenue growth rate."),
        ("C",   "Data Center share loss to Nvidia becomes visible in segment revenue. Revenue growth below 10%. "
                "Gross margin contracts below 50%. MG&A and R&D both growing faster than revenue."),
        ("D",   "Fundamental competitive breakdown: Nvidia wins major hyperscaler contracts away from MI300. "
                "EPYC market share reversal vs Intel. Revenue decline. Operating leverage reverses."),
    ]
    for grade, desc in grades:
        next_row = bullet_row(ws, next_row,
                              f"{grade}   {desc}",
                              ASSESS_BG, indent=2, font_size=9)

    next_row += 1

    # ══════════════════════════════════════════════════════════════════════════
    # KEY METRICS TO WATCH — Q2 2026
    # ══════════════════════════════════════════════════════════════════════════
    next_row = section_header(ws, next_row,
                              "▶  KEY METRICS TO WATCH — Q2 2026 (Expected Late July / Early August 2026)",
                              SECTION_BG)

    metrics = [
        "① MG&A as % of Revenue — Q1 was elevated at 12.2% (vs revenue growth outpaced). Must show reversal. Declining MG&A % = operating leverage resuming. Most important single cost line to watch.",
        "② Non-GAAP Operating Margin — Q2 guided ~27%. Above 27.5% = outperform. Below 26.5% = concern. This is management's primary profitability KPI and the market's pricing anchor.",
        "③ Data Center Segment Revenue — must maintain sequential and Y/Y growth momentum. MI300X/MI350 ramp trajectory vs Nvidia H200/B200 competitive dynamics is the defining question.",
        "④ Gross Margin Trajectory — Q1: 53%. Q4 guidance and sequential trend matter. Can AMD reach 55%+ as Data Center mix grows? Each 100bps = ~$100M additional gross profit per quarter.",
        "⑤ Gaming Segment Stabilization — any sequential recovery in gaming GPU or semi-custom would remove a headwind. Continued decline is manageable but limits upside.",
        "⑥ R&D as % of Revenue — Q1: 23.4%. Needs to decline over time to generate R&D operating leverage. MI350 sampling and CDNA4 development keep near-term R&D spend elevated.",
        "⑦ Non-GAAP Diluted EPS vs $0.96 Guidance — beat confirms execution; miss on a guided quarter is more impactful than a miss on a non-guided quarter.",
        "⑧ Embedded Segment Recovery — Xilinx-based Embedded revenue trough may be passing. Any inflection here would be incremental to the Data Center growth story.",
        "⑨ Cash and Share Repurchase Activity — AMD generates strong free cash flow. Buyback velocity signals management's confidence in the stock at current valuations.",
    ]
    for m in metrics:
        next_row = bullet_row(ws, next_row, m, WHITE, indent=2, font_size=9)

    # ── Footer ────────────────────────────────────────────────────────────────
    next_row += 1
    ws.merge_cells(f"A{next_row}:H{next_row}")
    foot = ws[f"A{next_row}"]
    foot.value = (
        "Source: AMD Q1 2026 Earnings Press Release, May 5, 2026  |  "
        "GAAP figures only  |  Millions of dollars except per-share amounts  |  "
        "Analysis built: May 2026  |  Script: build_amd_simple_v2.py  |  "
        "Not investment advice."
    )
    foot.font = Font(name="Arial", size=7, color=MUTED, italic=True)
    foot.fill = fill(WHITE)
    foot.alignment = Alignment(horizontal="left", wrap_text=True)
    ws.row_dimensions[next_row].height = 18

    out = "/mnt/user-data/outputs/AMD_Q1_2026_Income_Statement_v2.xlsx"
    wb.save(out)
    print(f"Saved: {out}")
    return out


# ==============================================================================
# SELF-TEST
# DO NOT REMOVE — proves the script is complete and working before saving
# ==============================================================================
def self_test(out_path):
    import os
    from openpyxl import load_workbook

    errors = []

    if not os.path.exists(out_path):
        errors.append(f"FAIL: file not created at {out_path}")
    else:
        size = os.path.getsize(out_path)
        if size < 15000:
            errors.append(f"FAIL: file too small ({size} bytes) — engine may be incomplete")

        wb = load_workbook(out_path, data_only=True)
        ws = wb.active
        if ws.title != "Q1 2026 Income Statement":
            errors.append(f"FAIL: wrong sheet name '{ws.title}'")

        found_revenue = False
        found_10253 = False
        found_grade = False
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell:
                    cv = str(cell)
                    if "Net Revenue" in cv:
                        found_revenue = True
                    if "10253" in cv or "10,253" in cv:
                        found_10253 = True
                    if "A−" in cv or "A-" in cv:
                        found_grade = True

        if not found_revenue:
            errors.append("FAIL: 'Net Revenue' row not found")
        if not found_10253:
            errors.append("FAIL: Q1 2026 Revenue $10,253M not found — data block may be wrong")
        if not found_grade:
            errors.append("FAIL: Grade 'A−' not found — assessment section may be missing")

    if errors:
        print("\n" + "=" * 60)
        print("SELF-TEST FAILED — DO NOT SAVE TO DRIVE")
        print("=" * 60)
        for e in errors:
            print(f"  {e}")
        print("=" * 60)
        return False
    else:
        size = os.path.getsize(out_path)
        print("\n" + "=" * 60)
        print("SELF-TEST PASSED — safe to save to Drive")
        print(f"  File: {out_path}")
        print(f"  Size: {size:,} bytes")
        print(f"  Net Revenue $10,253M: confirmed")
        print(f"  Sheet structure: confirmed")
        print(f"  Grade A−: confirmed")
        print("=" * 60)
        return True


if __name__ == "__main__":
    out = build()
    passed = self_test(out)
    if not passed:
        import sys
        sys.exit(1)
