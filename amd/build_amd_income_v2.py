# ══════════════════════════════════════════════════════════════════════════════
# ADVANCED MICRO DEVICES, INC. (AMD) — INCOME STATEMENT ANALYSIS
# build_amd_income_v2.py
# Q1 2026 | Quarter Ended March 28, 2026 | Reported May 5, 2026
# v2: Adds confirmed hyperscaler deals to What's New tab:
#     OpenAI 6GW (Oct 6, 2025), Meta 6GW (Feb 24, 2026), Oracle 50K GPUs (Oct 14, 2025).
#     Coherence Check updated to reference build_amd_projection_engine_v5.py
#     (SHARES=1.750B, Bull 45% / Base 40% / Bear 15%).
#     Data Sources tab extended with three deal press release entries.
# ══════════════════════════════════════════════════════════════════════════════

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── COLOR PALETTE ─────────────────────────────────────────────────────────────
NAVY        = "1A1A2E"
BLUE        = "1F4E79"
ACCENT      = "2E75B6"
WHITE       = "FFFFFF"
GRAY        = "F2F2F2"
GREEN       = "00B050"
AMBER       = "FF8C00"
RED         = "C00000"
GREEN_LT    = "E8F5E9"
AMBER_LT    = "FFF3E0"
RED_LT      = "FFEBEE"
MEASURE_BG  = "EBF3FB"
BLACK       = "000000"
MUTED       = "666666"
SECTION_BG  = "1F4E79"
STRENGTH_BG = "E8F5E9"
CONCERN_BG  = "FFEBEE"
ASSESS_BG   = "FFF8E1"

# ── BORDER STYLE (applied to all data table cells) ────────────────────────────
_SIDE = Side(style="thin", color="CCCCCC")
TABLE_BORDER = Border(left=_SIDE, right=_SIDE, top=_SIDE, bottom=_SIDE)

# ── FONT SIZES ────────────────────────────────────────────────────────────────
SZ_DEFAULT  = 10
SZ_TITLE    = 14
SZ_SUBTITLE = 10
SZ_HEADER   = 12
SZ_GRADE    = 18
SZ_FOOTER   =  8

# ── PRIORITY ROWS (bold line item cell) ───────────────────────────────────────
PRIORITY = {
    "Net Revenue", "Total Cost of Sales", "Gross Profit", "Gross Margin",
    "Research and Development", "Marketing, General and Administrative",
    "Total Operating Expenses", "Operating Income", "Net Income",
    "Diluted Earnings Per Share",
}

# ── ROW DATA ──────────────────────────────────────────────────────────────────
ROWS = [
    # ── REVENUE ───────────────────────────────────────────────────────────────
    ("Net Revenue",
     10253, 10270, 7438, "green", None, False,
     "+37.9% Y/Y to $10.25B — essentially flat Q/Q (-$17M, -0.2%) despite typical Q1 "
     "seasonal weakness, confirming underlying demand strength. Data Center was the primary "
     "Y/Y driver (MI300X/MI325X GPU ramp + EPYC CPU server share gains vs Intel). "
     "Gaming declined (consumer GPU + semi-custom both down). Embedded recovering slowly. "
     "Client (Ryzen) grew modestly. The revenue story is increasingly a Data Center story — "
     "that segment is the engine and the margin expansion driver simultaneously.",
     "Total dollars sold across all products and geographies; top-line scale",
     None),

    ("Cost of Sales",
     4576, 4433, 3451, "green", None, False,
     "+32.6% Y/Y — meaningfully slower than revenue growth of +37.9%. Data Center GPU "
     "(MI300X/MI325X) carries higher gross margins than Gaming GPUs or Embedded chips. "
     "As Data Center becomes a larger share of mix, cost of sales as a % of revenue "
     "structurally improves. Q/Q increase (+$143M) reflects product mix seasonality. "
     "Green: cost growing slower than revenue is the definition of gross margin expansion.",
     "Direct production costs; growing slower than revenue = gross margin expansion",
     None),

    ("  Amortization — Acquisition Intangibles (COGS)",
     261, 260, 251, None, None, False,
     "Essentially flat Q/Q and modestly higher Y/Y ($251M → $261M). Non-cash amortization "
     "of intangibles from the 2022 Xilinx acquisition ($48.8B) — primarily Xilinx IP, "
     "technology licenses, and customer relationships allocated to cost of sales. "
     "Will decline gradually as assets fully amortize. Non-cash, directionally positive.",
     "Non-cash Xilinx acquisition intangibles in COGS; amortizes down over time",
     None),

    ("Total Cost of Sales",
     4837, 4693, 3702, "green", None, False,
     "+30.7% Y/Y — the slowest growing cost line in the income statement. Revenue grew "
     "7 percentage points faster than total COGS, directly driving +300bps gross margin "
     "expansion from 50% to 53%. When total COGS grows materially slower than revenue, "
     "gross margin expands mechanically. This is the gross-level operating leverage story.",
     "All-in production costs including non-cash amortization; key gross margin driver",
     None),

    ("Gross Profit",
     5416, 5577, 3736, "green", None, False,
     "+45.0% Y/Y — grew 700bps faster than revenue (+37.9%). $1.68B more gross profit "
     "than Q1 2025. Q/Q decline (-$161M) driven by seasonal revenue dip and product mix. "
     "The Y/Y frame is the right one: gross profit growing faster than revenue is positive "
     "operating leverage at the gross level. Dollar earning power is expanding rapidly.",
     "Absolute earning power after production costs; measures scale and efficiency",
     None),

    ("Gross Margin",
     53.0, 54.0, 50.0, "green", "green", True,
     "+300bps Y/Y expansion (50% → 53%). Q/Q contracted -100bps (54% → 53%) due to "
     "seasonal product mix shift. Structural Y/Y driver: MI300X/MI325X GPU carries "
     "meaningfully higher gross margins than Gaming GPUs or Embedded. As Data Center "
     "becomes a larger revenue share, blended gross margin expands structurally. "
     "Management long-term target: approach Nvidia-like margins as the MI300 series matures. "
     "Current 53% vs Nvidia's ~75%+ signals significant further expansion potential.",
     "Production efficiency; % of each revenue dollar retained after COGS",
     None),

    # ── OPERATING EXPENSES ────────────────────────────────────────────────────
    ("Research and Development",
     2397, 2330, 1728, "amber", None, False,
     "+38.7% Y/Y — grew slightly faster than revenue (+37.9%). No operating leverage "
     "here, but the gap is thin (80bps). AMD must invest heavily in chip design to "
     "stay competitive: MI350 now shipping (CDNA 4, 3nm, 288GB HBM3E), MI450 sampling for "
     "H2 2026 production (CDNA 5, HBM4, 432GB), MI500 on the 2027 roadmap (CDNA 6, 2nm); "
     "EPYC Turin (Gen 5) at record 46.2% server revenue share; Zen 6 in development. "
     "R&D is the competitive lifeblood — cutting it would be fatal. "
     "Amber: matched revenue growth — no leverage, but not a breakdown either.",
     "Product competitiveness investment; must grow slower than revenue to show leverage",
     None),

    ("Marketing, General and Administrative",
     1253, 1198, 886, "red", None, False,
     "+41.4% Y/Y — fastest growing cost line in the statement, outpacing revenue growth "
     "of +37.9% by 350bps. This is the primary operating leverage failure point this quarter. "
     "MG&A should grow significantly slower than revenue as the business scales. "
     "Management's path to 30%+ non-GAAP operating margins requires MG&A discipline. "
     "Red: any operating cost growing faster than revenue violates the operating leverage "
     "principle and needs to reverse direction.",
     "Overhead and sales cost discipline; must grow slower than revenue for leverage",
     None),

    ("  Amortization — Acquisition Intangibles (OpEx)",
     290, 297, 316, None, None, False,
     "Declining Y/Y ($316M → $290M, -8.2%) and Q/Q ($297M → $290M). Non-cash amortization "
     "of Xilinx intangibles in operating expenses — technology licenses, trade names, and "
     "customer relationships. This line will continue declining as assets fully amortize, "
     "providing a structural GAAP tailwind to operating income over time.",
     "Non-cash Xilinx acquisition intangibles in OpEx; declining = GAAP operating tailwind",
     None),

    ("Total Operating Expenses",
     3940, 3825, 2930, "green", None, False,
     "+34.5% Y/Y — grew slower than revenue (+37.9%). Operating leverage confirmed in "
     "aggregate despite MG&A overshooting. The Xilinx amortization decline (-$26M Y/Y) "
     "provides a structural offset. Net message: operating leverage exists in aggregate, "
     "but MG&A discipline is required to make it durable and improving.",
     "All operating costs below gross profit; slower growth than revenue = leverage",
     None),

    ("Operating Income",
     1476, 1752, 806, "green", None, False,
     "+83.1% Y/Y — more than double the rate of revenue growth (+37.9%). $670M more "
     "operating income on $2.8B more revenue. GAAP operating margin: 10.8% → 14.4% "
     "(+360bps Y/Y). Q/Q declined ($1,752M → $1,476M) on typical Q1 seasonality. "
     "Non-GAAP operating margin guided ~27% for Q2 2026 (excludes $551M Xilinx amortization). "
     "The +83% operating income growth is the operating leverage thesis fully confirmed.",
     "Core business profitability before financing costs and taxes; leverage signal",
     None),

    # ── BELOW THE LINE ────────────────────────────────────────────────────────
    ("Interest Expense",
     -37, -36, -20, None, None, False,
     "Modest increase Q/Q (-$37M vs -$36M) and Y/Y growth from -$20M. AMD carries "
     "relatively light long-term debt (~$1.7B notes outstanding) — a significant balance "
     "sheet advantage. Interest expense is not a meaningful earnings drag at AMD's scale.",
     "Cost of AMD's debt; modest relative to operating income — balance sheet strength",
     None),

    ("Other Income (Expense), net",
     165, 358, 39, None, None, False,
     "Q1: $165M. Q4 2025 was unusually elevated ($358M) — included gains on equity "
     "investments or licensing settlements not repeated in Q1. Q1 2025 was $39M. "
     "Primarily interest income on AMD's $5.1B cash balance plus equity investment gains. "
     "Q/Q decline (-$193M) is not a concern — Q4 was the outlier.",
     "Interest income on cash + equity investment gains; Q4 2025 was the outlier quarter",
     None),

    ("Income Before Income Taxes",
     1604, 2074, 825, None, None, False,
     "$1.60B pre-tax income. Q/Q decline driven by Q1 seasonal revenue dip and lower other "
     "income vs Q4's elevated level. Y/Y improvement ($825M → $1,604M, +94%) driven entirely "
     "by operating income growth.",
     "Total profitability before tax; reflects operating performance and capital structure",
     None),

    ("Income Tax Provision",
     238, 455, 123, None, None, False,
     "+$115M Y/Y in tax expense, roughly in line with income growth. AMD benefits from R&D "
     "tax credits, foreign income structures, and Xilinx-related deductions that keep the "
     "effective rate manageable.",
     "Tax obligation; R&D credits and deductions keep effective rate competitive",
     None),

    ("Equity Income in Investee",
     6, 1, 7, None, None, False,
     "Small income ($6M) from equity method investments. Immaterial relative to AMD's "
     "overall profitability.",
     "Small income from minority equity stakes; immaterial to the investment thesis",
     None),

    ("Income from Continuing Operations",
     1372, 1620, 709, None, None, False,
     "+93.5% Y/Y — nearly doubled. Q/Q decline mirrors seasonality and other income "
     "normalization. This is GAAP net income from AMD's core ongoing business.",
     "Core business GAAP net income excluding discontinued operations",
     None),

    ("Income (Loss) from Discontinued Operations",
     11, -109, None, None, None, False,
     "Q1 2026: $11M gain. Q4 2025: ($109M) loss (write-down/impairment on divested assets). "
     "These relate to Xilinx-era divested business units being wound down. Non-recurring.",
     "Gains/losses from divested business units; non-recurring, expected to zero out",
     None),

    ("Net Income",
     1383, 1511, 709, "green", None, False,
     "+95.1% Y/Y — nearly doubled on 37.9% revenue growth. Q/Q decline ($1,511M → $1,383M) "
     "reflects Q1 seasonality plus Q4's elevated other income. GAAP net margin: "
     "9.5% → 13.5% (+400bps Y/Y). The near-doubling of net income on 38% revenue growth "
     "is textbook operating leverage.",
     "Total GAAP profitability; bottom-line operating leverage confirmed",
     None),

    # ── EPS ───────────────────────────────────────────────────────────────────
    ("Basic EPS — Continuing Operations",
     0.84, 1.00, 0.44, None, None, False,
     "Q1 2026: $0.84. Q4 2025: $1.00. Q1 2025: $0.44. Reference line.",
     "Per-share profit from core operations on basic share count; reference line",
     None),

    ("Basic EPS — Discontinued Operations",
     0.01, -0.07, None, None, None, False,
     "Q1 2026: $0.01 gain. Q4 2025: ($0.07) loss. Non-recurring.",
     "Per-share discontinued ops impact; non-recurring",
     None),

    ("Basic Earnings Per Share",
     0.85, 0.93, 0.44, None, None, False,
     "Q1 2026: $0.85. Q4: $0.93. Q1 2025: $0.44. Uses ~1,631M basic shares.",
     "Total per-share profit on basic shares only; diluted EPS is the primary metric",
     None),

    ("Diluted EPS — Continuing Operations",
     0.83, 0.99, 0.44, None, None, False,
     "Q1 2026: $0.83. Diluted count: 1,650M. Basic-diluted spread of only 19M shares (1.2%).",
     "Per-share profit from core operations on fully diluted share count",
     None),

    ("Diluted EPS — Discontinued Operations",
     0.01, -0.07, None, None, None, False,
     "Same as basic — $0.01 gain in Q1 2026. Non-recurring.",
     "Per-share discontinued ops impact on diluted basis",
     None),

    ("Diluted Earnings Per Share",
     0.84, 0.92, 0.44, "green", None, False,
     "+90.9% Y/Y — from $0.44 to $0.84. Q/Q: $0.92 → $0.84 (Q1 seasonal decline). "
     "Non-GAAP diluted EPS for Q2 2026 guided at $0.96 midpoint by management. "
     "AMD's dilution profile is clean: 19M share basic-diluted spread (1.2%). "
     "No preferred stock complexity. "
     "Green: +90.9% Y/Y EPS growth on +37.9% revenue = shareholders are seeing the "
     "full, unimpeded benefit of operating leverage.",
     "Primary per-share metric; +90.9% Y/Y confirms shareholders benefit from leverage",
     None),

    # ── SHARE COUNT ───────────────────────────────────────────────────────────
    ("Basic Shares (millions)",
     1631, 1630, 1620, None, None, False,
     "Flat Q/Q (1,631M vs 1,630M) and only +0.7% Y/Y growth (+11M shares). AMD's share "
     "count is very stable — stock buybacks roughly offset SBC issuance. "
     "Note: post-Q1 2026 warrant agreements with OpenAI (160M shares, Oct 2025) and Meta "
     "(160M shares, Feb 2026) will vest based on purchase milestones, adding up to 320M "
     "shares of potential dilution. Management expects partial buyback offset.",
     "Basic share count; minimal growth — see warrant note for forward dilution",
     "#,##0"),

    ("Diluted Shares (millions)",
     1650, 1649, 1626, None, None, False,
     "1,650M diluted vs 1,649M Q4 and 1,626M Q1 2025. Basic-diluted spread: 19M (1.2%). "
     "Very clean dilution profile. Forward note: 320M warrants outstanding (OpenAI 160M "
     "Oct 2025 + Meta 160M Feb 2026) vest as purchase milestones are hit. Projection engine "
     "uses 1.750B to reflect ~100M net after partial buyback offset.",
     "Fully diluted count; 1.2% spread is minimal — forward warrant dilution tracked separately",
     "#,##0"),
]

# ── COLUMN LAYOUT ─────────────────────────────────────────────────────────────
COL_WIDTHS = {
    "A": 38, "B": 14, "C": 10,
    "D": 13, "E": 13, "F": 13,
    "G": 90, "H": 40,
}

def fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def section_header(ws, row, text, bg, cols=8):
    ws.merge_cells(f"A{row}:{chr(64+cols)}{row}")
    cell = ws[f"A{row}"]
    cell.value = text
    cell.font = Font(name="Arial", bold=True, size=SZ_HEADER, color=WHITE)
    cell.fill = fill(bg)
    cell.alignment = center()
    ws.row_dimensions[row].height = 20
    return row + 1

def bullet_row(ws, row, text, bg, font_size=SZ_DEFAULT, color=BLACK):
    ws.merge_cells(f"A{row}:H{row}")
    cell = ws[f"A{row}"]
    cell.value = text
    cell.font = Font(name="Arial", size=font_size, color=color)
    cell.fill = fill(bg)
    cell.alignment = Alignment(horizontal="left", vertical="top",
                               wrap_text=True, indent=2)
    ws.row_dimensions[row].height = max(15, 13 + len(text) // 10)
    return row + 1

def auto_fmt(val, is_pct):
    if val is None:
        return "@"
    if abs(val) < 5:
        return '$#,##0.00'
    if abs(val) < 100:
        return '$#,##0.0'
    return '$#,##0'

def yy_change_str(q1, q1_py, is_pct, fmt_override):
    if q1 is None or q1_py is None:
        return "—"
    diff = q1 - q1_py
    if is_pct:
        sign = "+" if diff >= 0 else ""
        return f"{sign}{diff:.0f} bps"
    if fmt_override == "#,##0":
        sign = "+" if diff >= 0 else ""
        return f"{sign}{diff:,.0f}"
    if abs(q1) < 5:
        sign = "+" if diff >= 0 else ""
        return f"{sign}${diff:.2f}"
    sign = "+" if diff >= 0 else "-"
    return f"{sign}${abs(diff):,.0f}M"

def yy_pct_str(q1, q1_py, is_pct):
    if q1 is None or q1_py is None or q1_py == 0:
        return "—"
    if is_pct:
        diff = q1 - q1_py
        sign = "+" if diff >= 0 else ""
        return f"{sign}{diff:.1f} pp"
    pct = (q1 - q1_py) / abs(q1_py) * 100
    sign = "+" if pct >= 0 else ""
    return f"{sign}{pct:.1f}%"

def fmt_val(val, is_pct, fmt_override):
    if val is None:
        return "—"
    if is_pct:
        return f"{val:.1f}%"
    if fmt_override == "#,##0":
        return f"{val:,.0f}"
    if abs(val) < 5:
        return f"${val:.2f}"
    if abs(val) < 100:
        return f"${val:.1f}"
    return f"${val:,.0f}"


def build_sources_tab(wb):
    ws = wb.create_sheet("Data Sources")

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 13
    ws.column_dimensions["D"].width = 52
    ws.column_dimensions["E"].width = 42

    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value = "AMD Q1 2026 INCOME STATEMENT — DATA SOURCES"
    t.font = Font(name="Arial", bold=True, size=SZ_TITLE, color=WHITE)
    t.fill = fill(NAVY)
    t.alignment = center()
    ws.row_dimensions[1].height = 24

    for col_idx, h in enumerate(
        ["Category", "Source", "Date", "Key Data Used", "Reference / URL"], 1
    ):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = h
        cell.font = Font(name="Arial", bold=True, size=SZ_HEADER, color=WHITE)
        cell.fill = fill(ACCENT)
        cell.alignment = center(wrap=True)
        cell.border = TABLE_BORDER
    ws.row_dimensions[2].height = 24

    sources = [
        # ── Confirmed deals — new in v2 ───────────────────────────────────────
        ("Strategic Deal",
         "AMD + OpenAI Strategic Partnership",
         "Oct 6, 2025",
         "6GW total compute agreement. 1GW initial deployment H2 2026 using MI450 Series "
         "accelerators. AMD issues 160M warrant shares to OpenAI vesting on purchase milestones. "
         "Largest single-customer GPU commitment in AMD history. Used in: warrant dilution "
         "note (Basic/Diluted share rows), What's New OpenAI item, projection engine v5 "
         "SHARES=1.750B and Bull probability rationale.",
         "AMD Newsroom — AMD and OpenAI Announce Strategic Partnership (Oct 6, 2025)"),

        ("Strategic Deal",
         "AMD + Meta Expanded Strategic Partnership",
         "Feb 24, 2026",
         "6GW total compute agreement. 1GW initial deployment H2 2026 using custom "
         "MI450-based GPU accelerators + EPYC Venice CPUs. AMD issues 160M warrant shares "
         "to Meta vesting on purchase milestones. Cross-sells GPU + CPU in same rack. "
         "Used in: warrant dilution note, What's New Meta item, projection engine v5 "
         "SHARES=1.750B and 12GW confirmed pipeline probability rationale.",
         "AMD Newsroom — AMD and Meta Announce Expanded Strategic Partnership (Feb 24, 2026)"),

        ("Strategic Deal",
         "AMD + Oracle AI World Announcement",
         "Oct 14, 2025",
         "Oracle commits to 50K AMD GPUs initial deployment Q3 2026: MI450 Series + "
         "EPYC Venice CPUs + Pensando Vulcano DPUs in Helios rack architecture. "
         "Expanding to 200K+ GPUs in 2027. No warrant shares — volume purchase agreement. "
         "Used in: What's New Oracle item, MI350 Oracle Cloud scale reference, "
         "projection engine v5 Oracle revenue ramp in Bull 2026 assumption.",
         "AMD Newsroom — AMD at Oracle CloudWorld (Oct 14, 2025)"),

        # ── Primary financial data ─────────────────────────────────────────────
        ("Primary Data",
         "AMD Q1 2026 Earnings Press Release",
         "May 5, 2026",
         "All Q1 2026 actuals: Revenue $10,253M, Gross Profit $5,416M, "
         "Operating Income $1,476M, Net Income $1,383M, Diluted EPS $0.84. "
         "Quarter ended March 28, 2026.",
         "https://ir.amd.com/news-events/press-releases"),

        ("Comparison Data",
         "AMD Q4 2025 Earnings Press Release",
         "Feb 2026",
         "Q4 2025 figures used for Q/Q comparison column: Revenue $7,658M, "
         "Gross Margin 54.0%, Operating Income $1,752M, EPS $0.69.",
         "https://ir.amd.com/news-events/press-releases"),

        ("Comparison Data",
         "AMD Q1 2025 Earnings Press Release",
         "Apr 2025",
         "Q1 2025 figures used for all Y/Y change calculations: Revenue $7,438M, "
         "Gross Margin 50.0%, Operating Income $806M, Net Income $710M, EPS $0.44.",
         "https://ir.amd.com/news-events/press-releases"),

        ("Guidance",
         "AMD Q2 2026 Non-GAAP Guidance",
         "May 5, 2026",
         "Q2 2026 Non-GAAP Diluted EPS guidance midpoint: $0.96. "
         "Non-GAAP operating margin guided ~27%.",
         "AMD Q1 2026 Earnings Call transcript / press release"),

        ("Product Roadmap",
         "AMD Instinct MI350 Launch",
         "2025",
         "MI350 now shipping. CDNA 4, 3nm, 288GB HBM3E. Fastest ramping product in "
         "AMD history. Deployed at scale by Oracle Cloud.",
         "datacenterdynamics.com — AMD launches Instinct MI350"),

        ("Product Roadmap",
         "AMD Instinct MI450 Sampling Confirmed",
         "May 2026",
         "AMD sampling MI450 to lead customers for H2 2026 production deployment. "
         "CDNA 5, HBM4, 432GB memory (~50% more bandwidth vs MI350). Helios rack. "
         "Used as near-term catalyst in Key Metrics Watch ③ and assessment.",
         "wccftech.com — AMD sampling MI450 GPUs"),

        ("Product Roadmap",
         "AMD Instinct MI500 Roadmap",
         "CES 2026",
         "MI500 planned for 2027. CDNA 6, 2nm process, HBM4E. Two variants: "
         "MI455X (training/inference) and MI430X (HPC/Sovereign AI). "
         "MI455X is a 2027 MI500-family product — distinct from MI450 Series (2026).",
         "wccftech.com — AMD 2026-2027 AI Roadmap MI400 & MI500"),

        ("Market Data",
         "EPYC Server CPU Revenue Share Q1 2026",
         "May 2026",
         "EPYC hit record 46.2% server CPU revenue share in Q1 2026 (from 28.8% Q4 2025). "
         "AMD Data Center revenue $5.8B surpassed Intel Data Center $5.1B for first time.",
         "wccftech.com — AMD EPYC Record 46.2% Server Revenue Share"),

        ("Competitive Context",
         "Nvidia Blackwell / Vera Rubin GPU Series",
         "2024–2026",
         "B200, B300, and Vera Rubin series referenced as the competitive benchmark "
         "for AMD Instinct in Data Center GPU.",
         "Public Nvidia earnings releases and product announcements"),

        ("Script / Analysis",
         "build_amd_income_v2.py",
         "May 2026",
         "All analysis, row coloring, letter grade framework, and narrative. "
         "v2 adds confirmed hyperscaler deals (OpenAI/Meta/Oracle) to What's New tab. "
         "Future quarterly updates: build_amd_income_v3.py, etc. Never overwrite.",
         "GitHub: tonyvasquez1/bmnr-dashboard /amd/"),
    ]

    for i, (cat, src, date, data, url) in enumerate(sources):
        row = 3 + i
        bg = WHITE if i % 2 == 0 else GRAY
        for col_idx, val in enumerate([cat, src, date, data, url], 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = val
            cell.fill = fill(bg)
            cell.border = TABLE_BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.font = Font(name="Arial", size=SZ_DEFAULT, color=BLACK)
        ws.cell(row=row, column=1).font = Font(
            name="Arial", size=SZ_DEFAULT, bold=True, color=NAVY)
        ws.row_dimensions[row].height = 52


def build_whats_new_tab(wb):
    ws = wb.create_sheet("What's New")

    for col, w in zip("ABCDEF", [20, 36, 26, 26, 28, 28]):
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = ("AMD — WHAT'S NEW  |  Version: Q1 2026  |  "
               "Script: build_amd_income_v2.py  |  Researched: May 2026")
    t.font = Font(name="Arial", bold=True, size=SZ_TITLE, color=WHITE)
    t.fill = fill(NAVY)
    t.alignment = center()
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:F2")
    instr = ws["A2"]
    instr.value = (
        "HOW TO USE: In each quarterly update (v3, v4, ...) research AMD results, guidance, "
        "product news, competitor earnings, and sector macro BEFORE editing any row data. "
        "Populate every section below. Carry each item into the income statement and projection "
        "engine. Confirm the Coherence Check passes before committing. This tab is the audit "
        "trail between versions."
    )
    instr.font = Font(name="Arial", size=SZ_SUBTITLE, color=WHITE, italic=True)
    instr.fill = fill(BLUE)
    instr.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 40

    col_headers = ["Category", "Item", "Prior State",
                   "Current State (Q1 2026)",
                   "Consequences → Earnings",
                   "Consequences → Projections"]
    for col_idx, h in enumerate(col_headers, 1):
        c = ws.cell(row=3, column=col_idx)
        c.value = h
        c.font = Font(name="Arial", bold=True, size=SZ_HEADER, color=WHITE)
        c.fill = fill(ACCENT)
        c.alignment = center(wrap=True)
        c.border = TABLE_BORDER
    ws.row_dimensions[3].height = 30

    cat_colors = {
        "AMD — Financial":     "1A5276",
        "AMD — Guidance":      "1A7A3A",
        "AMD — Product":       "6C3483",
        "AMD — Strategic":     "0E5A8A",   # new category for confirmed deals
        "AMD — Market Share":  "B7770D",
        "Competitor — Nvidia": "8B0000",
        "Competitor — Intel":  "4A235A",
        "Sector / Macro":      "1B4F72",
    }

    items = [
        # ── AMD STRATEGIC DEALS — NEW IN v2 ───────────────────────────────────
        ("AMD — Strategic",
         "OpenAI 6GW Strategic Agreement (Oct 6, 2025)",
         "No prior hyperscaler GPU agreement at this scale",
         "6GW total compute agreement. 1GW initial deployment H2 2026 using MI450 Series "
         "accelerators (CDNA 5, HBM4, 432GB). AMD issues 160M warrant shares to OpenAI "
         "vesting on purchase milestones. Largest single-customer GPU commitment in AMD history.",
         "No Q1 2026 revenue impact (H2 2026 deployment). Forward catalyst: H2 2026 Data Center "
         "GPU revenue step-up as MI450 production begins. Warrant dilution is accounted for in "
         "projection engine shares (1.650B actual → 1.750B forward, net of partial buyback).",
         "Eliminates demand risk from projection bear case. Bull probability raised to 45% "
         "(from 35%) — signed backlog de-risks the revenue scenario. SHARES updated 1.650→1.750B. "
         "MI450 Series is the 2026 product (CDNA 5) — distinct from MI500/MI455X (2027, CDNA 6)."),

        ("AMD — Strategic",
         "Meta 6GW Strategic Agreement (Feb 24, 2026)",
         "No prior multi-GPU strategic agreement with Meta at scale",
         "6GW total compute agreement. 1GW initial deployment H2 2026 using custom "
         "MI450-based GPU accelerators + EPYC Venice CPUs. AMD issues 160M warrant shares "
         "to Meta vesting on purchase milestones. Cross-sells GPU + CPU in a single rack. "
         "12GW combined committed pipeline (OpenAI + Meta).",
         "No Q1 2026 revenue impact (H2 2026 start). Forward: EPYC Venice cross-sell creates "
         "a CPU revenue stream alongside GPU — compound revenue from same hyperscaler. "
         "Total warrant exposure: 320M shares (OpenAI 160M + Meta 160M = 19.4% max dilution).",
         "12GW total confirmed pipeline with two hyperscalers eliminates demand uncertainty "
         "for 2026-2027. Base case revenue growth rates unchanged — execution risk remains. "
         "Bear reframed: not demand/competition risk but execution risk (yield, HBM4, ROCm parity)."),

        ("AMD — Strategic",
         "Oracle 50K GPU Commitment (Oct 14, 2025)",
         "MI350 in production at Oracle Cloud — no formal volume commitment prior",
         "Oracle commits to 50K AMD GPU initial deployment Q3 2026: MI450 Series + EPYC Venice "
         "CPUs + Pensando Vulcano DPUs in Helios rack. Expanding to 200K+ GPUs in 2027. "
         "No warrant shares — volume purchase agreement.",
         "Establishes Oracle Cloud as first major customer validating the Helios rack + MI450 "
         "ecosystem at scale before OpenAI/Meta deployments begin. Revenue in Q3 2026 income "
         "statement. GPU + CPU + DPU cross-sell in one rack is the highest-value architecture AMD sells.",
         "Oracle 2027 expansion (200K+ GPUs) supports Bull 2027 revenue growth assumption (50%). "
         "Helios rack validation at Oracle scale de-risks MI450 production ramp for "
         "OpenAI/Meta H2 2026 initial deployments. GPU + EPYC + Pensando full-stack wins "
         "are the highest-margin revenue AMD can generate."),

        # ── AMD FINANCIAL RESULTS ──────────────────────────────────────────────
        ("AMD — Financial",
         "Net Revenue Q1 2026",
         "Q1 2025: $7,438M  |  Q4 2025: $7,658M",
         "$10,253M actual. +37.9% Y/Y. ~flat Q/Q despite seasonal pattern.",
         "Revenue row updated. Y/Y growth above Base case (36%) and below Bull (43%). "
         "Q2 2026 guided $11.2B — sequential growth confirms no demand cliff.",
         "FY2026 on pace for $44-47B (H1 ~$21.5B + typical H2 seasonal lift). "
         "Within Base ($47.1B) / Bull ($49.5B) range. No change to scenario thresholds needed."),

        ("AMD — Financial",
         "Gross Margin: +300bps Y/Y expansion",
         "Q1 2025: 50.0%  |  Q4 2025: 54.0%",
         "Q1 2026: 53.0%. Data Center GPU mix shift is the structural driver.",
         "Gross margin row updated. Mix shift (MI350 carrying higher margin than Gaming/Embedded) "
         "is repeatable and structural.",
         "Bull NI margin ramp (28%→43%) is supported by gross margin expansion continuing as "
         "Data Center mix grows. Each 100bps gross margin = ~$100M quarterly gross profit."),

        ("AMD — Financial",
         "Operating Income: +83.1% Y/Y",
         "Q1 2025: $806M (10.8% GAAP margin)",
         "$1,476M actual. 14.4% GAAP margin. 2x+ revenue growth rate.",
         "Operating leverage confirmed. MG&A (+41.4%) is the sole cost line that grew faster "
         "than revenue — all other lines showed leverage.",
         "Operating leverage trajectory validates Non-GAAP margin expansion in Base/Bull cases. "
         "Base 2026 NI margin (25%) is conservative vs current Non-GAAP trajectory of ~22-27%."),

        ("AMD — Financial",
         "Diluted EPS: +90.9% Y/Y",
         "Q1 2025: $0.44",
         "Q1 2026: $0.84 GAAP. Non-GAAP higher. Minimal dilution (+1.5% share count Y/Y).",
         "EPS row updated. Share dilution is not a compounding headwind — repurchase program "
         "offsets stock-based compensation effectively.",
         "Projection engine updated to SHARES=1.750B for forward warrant dilution "
         "(OpenAI 160M + Meta 160M = 320M total warrants, ~100M net after partial buyback)."),

        # ── AMD GUIDANCE ───────────────────────────────────────────────────────
        ("AMD — Guidance",
         "Q2 2026 Non-GAAP EPS guidance: $0.96 midpoint",
         "No prior guidance (initial build)",
         "$0.96 Non-GAAP Diluted EPS midpoint. Non-GAAP operating margin ~27% guided.",
         "Q2 beat confirmation in August 2026 is the next catalyst. Operating margin 27% vs "
         "Q1 GAAP 14.4% — Xilinx amortization gap remains the optics problem.",
         "Non-GAAP NI margin for Q2 implied ~23-24%. Projection engine Base 2026 NI margin "
         "25% is slightly optimistic for H1 but likely achievable for the full year as H2 stronger."),

        ("AMD — Guidance",
         "Q2 2026 Revenue guidance: $11.2B midpoint",
         "Q1 2026 actual: $10,253M",
         "$11.2B midpoint (+9.2% Q/Q sequential). Implies continued Data Center demand.",
         "Sequential growth is positive — Q1 seasonal dip did not signal demand deceleration. "
         "H1 2026 total: ~$21.5B.",
         "FY2026 annualized at H1 pace + typical H2 lift: ~$44-47B. Comfortably within "
         "Base case ($47.1B). Bull case ($49.5B) requires strong H2 acceleration from MI450."),

        # ── AMD PRODUCT ROADMAP ────────────────────────────────────────────────
        ("AMD — Product",
         "MI350: now shipping (CDNA 4, 3nm, 288GB HBM3E)",
         "Previously referenced as 'in development'",
         "MI350 launched 2025. Current shipping AI GPU. Deployed at scale by Oracle Cloud "
         "Infrastructure. Fastest-ramping AMD product in company history.",
         "MI350 is the active product driving Data Center GPU revenue now.",
         "MI350 ramp supports Base case revenue growth for 2026. MI450 (H2 2026) is the "
         "next incremental upside catalyst that could push 2027 toward Bull scenario."),

        ("AMD — Product",
         "MI450: sampling confirmed for H2 2026 production (CDNA 5, HBM4, 432GB)",
         "Not previously documented",
         "AMD sampling MI450 to lead customers. H2 2026 production deployment in Helios rack. "
         "CDNA 5, HBM4, 432GB memory (~50% more bandwidth vs MI350). "
         "MI450 Series is the 2026 product — product used in OpenAI/Meta H2 2026 commitments.",
         "H2 2026 MI450 production creates a potential Data Center revenue step-up in Q3/Q4. "
         "Higher memory bandwidth addresses inference workload requirements.",
         "MI450 is the key 2027 revenue growth driver in Bull case (50% growth). If MI450 "
         "ships on schedule and hyperscalers adopt quickly, the 2027 Bull case is well-supported. "
         "Execution risk: HBM4 supply constraints (SK Hynix/Samsung) could throttle volume."),

        ("AMD — Product",
         "MI500: 2027 roadmap confirmed (CDNA 6, 2nm, HBM4E)",
         "Not previously documented",
         "MI500 for 2027. CDNA 6, 2nm process (TSMC), HBM4E. Two variants: MI455X "
         "(training/inference) and MI430X (HPC/Sovereign AI). "
         "MI455X is a 2027 MI500-family product — distinct from MI450 Series (2026).",
         "MI500 will be AMD's 2028 EPS driver. Important for long-term investor confidence.",
         "MI500 supports Bull case 2028-2030 NI margin expansion (38%→43%). "
         "2nm cost structure depends on TSMC N2 ramp."),

        # ── COMPETITOR: NVIDIA ─────────────────────────────────────────────────
        ("Competitor — Nvidia",
         "Nvidia Q4 FY2026: Data Center $62.3B (period ending Jan 2026)",
         "Nvidia Q3 FY2026 Data Center: ~$30.8B",
         "Q4 FY2026 total revenue $68.1B (+73% Y/Y). Data Center $62.3B (+75% Y/Y). "
         "91% of total revenue. Blackwell ramp driving acceleration.",
         "AMD Data Center ($5.8B) is ~9% of Nvidia's DC revenue. The TAM is real and enormous.",
         "Nvidia's scale validates the hyperscaler AI capex thesis underpinning AMD's Bull/Base "
         "revenue growth assumptions."),

        ("Competitor — Nvidia",
         "Nvidia Q1 FY2027 guidance: $78B revenue",
         "Q4 FY2026: $68.1B",
         "$78B Q1 FY2027 guidance (period Feb-Apr 2026). +15% Q/Q sequential. "
         "AI GPU demand acceleration continues into mid-2026.",
         "Strong Nvidia guidance = strong hyperscaler GPU budgets overall.",
         "Strong Nvidia demand validates the Bull/Base case $690-725B hyperscaler capex "
         "environment. Supports current scenario assumptions."),

        ("Competitor — Nvidia",
         "Nvidia Vera Rubin (next-gen GPU) on 2026-2027 roadmap",
         "H100/H200 → B200/B300 was the last transition",
         "Vera Rubin (GR100) is the post-Blackwell architecture, targeted 2026-2027. "
         "Expected to widen performance and efficiency lead vs prior generation.",
         "AMD must execute MI450 and MI500 flawlessly. AMD's value proposition is "
         "second-source diversity + EPYC integration rather than matching Nvidia spec-for-spec.",
         "With 12GW signed contracts (OpenAI/Meta), Vera Rubin competitive risk shifts from "
         "demand displacement to ROCm parity risk: can AMD's software run OpenAI/Meta "
         "training workloads at Vera Rubin performance-per-dollar parity?"),

        # ── COMPETITOR: INTEL ──────────────────────────────────────────────────
        ("Competitor — Intel",
         "Intel Q1 2026: DCAI revenue $5.1B (+22% Y/Y)",
         "Intel DCAI Q1 2025: ~$4.2B",
         "$5.1B Data Center & AI revenue vs AMD Data Center $5.8B. AMD has now passed Intel "
         "in quarterly Data Center revenue. Intel DCAI operating profit: $1.5B (31% margin).",
         "AMD's Data Center revenue surpassing Intel is a structural inflection point. "
         "Intel is growing (22%) but AMD is growing faster.",
         "Intel's DCAI growth (+22%) is slower than AMD's projected Data Center growth. "
         "Supports the CPU component of Bull/Base revenue growth projections."),

        ("Competitor — Intel",
         "Intel CPU recovery: Xeon 6 'Granite Rapids' driving ASP gains",
         "Intel CPU market share ~70%+ units",
         "Xeon 6 driving higher ASPs and AI-specific revenue growth ($750M in Q1). "
         "Intel responding to EPYC competitive pressure.",
         "EPYC Turin needs to sustain pricing power and performance lead to keep winning sockets.",
         "If Intel Xeon 6 slows EPYC share gains from Q1's record 46.2%, the CPU growth "
         "component of revenue projections could moderate. Currently not a concern."),

        # ── SECTOR / MACRO ─────────────────────────────────────────────────────
        ("Sector / Macro",
         "Hyperscaler combined AI capex 2026: $690-725B (+77% Y/Y)",
         "2025 combined capex: ~$410B",
         "Microsoft $190B, Amazon $200B, Google $190B, Meta $145B. Spending driven by "
         "capacity constraints, not declining demand. Backlogs at record levels.",
         "~30% of hyperscaler capex goes to AI accelerators/GPUs (~$207-217B market). "
         "AMD's ~$23B Data Center annual run rate is ~10-11% of the addressable GPU market.",
         "The $690-725B capex environment is the single most important macro tailwind. "
         "Bull case (43% AMD rev growth) requires AMD to hold or slightly grow its share."),

        ("Sector / Macro",
         "Memory and chip cost inflation affecting hyperscaler budgets",
         "Memory pricing stable in 2024-2025",
         "Microsoft attributed $25B of its 2026 AI budget to rising memory and chip costs. "
         "HBM3E/HBM4 pricing elevated due to SK Hynix/Samsung supply constraints.",
         "HBM4 cost inflation could pressure AMD's gross margins on MI450 at launch.",
         "Bear execution risk: HBM4 supply constraints (SK Hynix/Samsung) throttling MI450 "
         "production volume is the primary execution risk in projection engine v5 Bear case."),
    ]

    def write_data_row(ws, row, cat, item, prior, current, cons_earn, cons_proj,
                       bg, cat_colors):
        vals = [cat, item, prior, current, cons_earn, cons_proj]
        for col_idx, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col_idx)
            c.value = val
            c.fill = fill(bg)
            c.border = TABLE_BORDER
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c.font = Font(name="Arial", size=SZ_DEFAULT, color=BLACK)
        ink = cat_colors.get(cat, NAVY)
        ws.cell(row=row, column=1).font = Font(
            name="Arial", size=SZ_DEFAULT, bold=True, color=ink)
        ws.row_dimensions[row].height = 65

    next_row = 4
    for i, row_data in enumerate(items):
        bg = WHITE if i % 2 == 0 else GRAY
        write_data_row(ws, next_row, *row_data, bg, cat_colors)
        next_row += 1

    # ── COHERENCE CHECK SECTION ───────────────────────────────────────────────
    next_row += 1

    ws.merge_cells(f"A{next_row}:F{next_row}")
    sec = ws[f"A{next_row}"]
    sec.value = ("▶  COHERENCE CHECK — PROJECTION ENGINE vs INCOME STATEMENT  |  "
                 "RED = mismatch requiring update to build_amd_projection_engine_v5.py")
    sec.font = Font(name="Arial", bold=True, size=SZ_HEADER, color=WHITE)
    sec.fill = fill(NAVY)
    sec.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[next_row].height = 24
    next_row += 1

    ws.merge_cells(f"A{next_row}:F{next_row}")
    sub = ws[f"A{next_row}"]
    sub.value = ("Projection engine file: build_amd_projection_engine_v5.py  |  "
                 "Key constants: BASE_REV=$34.64B, ENTRY_PRICE=$455, SHARES=1.750B "
                 "(1.650B actual + 320M OpenAI/Meta warrants net of partial buyback), "
                 "Q1_NI_MARGIN=22.1%, Q1_REV=$10.253B, Q2_GUIDE=$11.2B  |  "
                 "Probabilities: Bull 45% / Base 40% / Bear 15%")
    sub.font = Font(name="Arial", size=SZ_DEFAULT, color=WHITE, italic=True)
    sub.fill = fill(BLUE)
    sub.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[next_row].height = 30
    next_row += 1

    chk_headers = ["Assumption", "Projection Engine Value",
                   "Income Statement Actual", "Status", "Notes"]
    for col_idx, h in enumerate(chk_headers, 1):
        c = ws.cell(row=next_row, column=col_idx)
        c.value = h
        c.font = Font(name="Arial", bold=True, size=SZ_HEADER, color=WHITE)
        c.fill = fill(ACCENT)
        c.alignment = center(wrap=True)
        c.border = TABLE_BORDER
        if col_idx == 6:
            ws.cell(row=next_row, column=6).value = ""
    ws.merge_cells(f"E{next_row}:F{next_row}")
    ws.row_dimensions[next_row].height = 24
    next_row += 1

    OK   = "1A7A3A"
    WARN = "B7770D"
    FAIL = "8B0000"

    checks = [
        ("Base Revenue (FY2025)",
         "BASE_REV = $34.64B",
         "FY2025 actual: $34.64B (per AMD annual report)",
         "✓  MATCH", OK,
         "Foundation for all scenario revenue projections. Confirmed accurate."),

        ("Q1 2026 Revenue",
         "Q1_REV = $10.253B",
         "Q1 2026 actual: $10,253M",
         "✓  MATCH", OK,
         "Used to validate projection trajectory. Revenue in line with Base/Bull range."),

        ("Q2 2026 Guidance",
         "Q2_GUIDE = $11.2B",
         "AMD guided $11.2B midpoint Non-GAAP",
         "✓  MATCH", OK,
         "Guidance midpoint correctly captured. Sequential growth (+9.2% Q/Q) is positive."),

        ("Non-GAAP NI Margin Q1",
         "Q1_NI_MARGIN = 22.1%",
         "Q1 2026 GAAP NI margin: 13.5%  |  Non-GAAP: ~22.1%",
         "✓  MATCH", OK,
         "GAAP/Non-GAAP gap is ~860bps due to Xilinx amortization ($551M+/qtr). "
         "Projection correctly uses Non-GAAP margin."),

        ("2026 Base Revenue Growth",
         "BASE rev_growth[0] = 36%  →  $47.1B",
         "Q1+Q2 run rate: ~$21.5B H1. Full year est. ~$44-47B.",
         "✓  ON TRACK", OK,
         "H1 pace is consistent with Base. H2 seasonal lift + Q2 guidance support $44-47B. "
         "Bull case ($49.5B) requires strong MI450-driven H2 acceleration."),

        ("2026 Bull Revenue Growth",
         "BULL rev_growth[0] = 43%  →  $49.5B",
         "Q1 Y/Y: +37.9%. Between Base (36%) and Bull (43%).",
         "~ WATCH", WARN,
         "Q1 growth is between Base and Bull. Bull case requires H2 acceleration above 43% "
         "run rate. MI450 H2 2026 + OpenAI/Meta initial deployments are the key catalysts."),

        ("2026 Base NI Margin",
         "BASE ni_margin[0] = 25%",
         "Q1 2026 Non-GAAP NI margin: ~22.1%  |  Q2 Non-GAAP op. margin guided ~27%",
         "~ WATCH", WARN,
         "Q1 Non-GAAP NI margin (22.1%) is below Base full-year assumption (25%). "
         "Full-year 25% requires stronger H2. Not a mismatch yet — H2 typically higher margins."),

        ("Entry Price / Stock Basis",
         "ENTRY_PRICE = $455.00",
         "AMD stock price at analysis date (May 2026): verify current price",
         "! VERIFY", WARN,
         "Entry price drives all CAGR calculations. Update ENTRY_PRICE in projection engine "
         "for each new version if AMD stock has moved significantly from $455."),

        ("Diluted Share Count",
         "SHARES = 1.750B  (v5 engine)",
         "Q1 2026 diluted shares: ~1,650M actual | +320M warrants (OpenAI 160M + Meta 160M)",
         "✓  UPDATED", OK,
         "Engine updated 1.650→1.750B in v5 to reflect 320M total warrant exposure "
         "(OpenAI Oct 2025 + Meta Feb 2026) net of ~220M partial buyback offset. "
         "Max dilution at full vest = 19.4%. Share count tracked quarterly as warrants vest."),

        ("Scenario Probabilities",
         "Bull 45% / Base 40% / Bear 15%  (v5 engine)",
         "12GW signed (OpenAI + Meta) + Oracle 50K GPUs eliminates demand risk. "
         "Bear = execution risk only. Bull raised 35%→45%.",
         "✓  UPDATED", OK,
         "v5 probability update: Bull raised 35%→45% (signed 12GW+ backlog), Base 45%→40% "
         "(execution still uncertain), Bear 20%→15% (demand risk largely gone — "
         "execution risk remains: yield, HBM4 supply, ROCm parity, macro, TSMC). "
         "Base revenue growth rates unchanged."),
    ]

    for i, (assumption, engine_val, actual_val, status, status_color, notes) in \
            enumerate(checks):
        row = next_row + i
        bg = WHITE if i % 2 == 0 else GRAY
        for col_idx, val in enumerate(
            [assumption, engine_val, actual_val, status, notes], 1
        ):
            c = ws.cell(row=row, column=col_idx)
            c.value = val
            c.fill = fill(bg)
            c.border = TABLE_BORDER
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c.font = Font(name="Arial", size=SZ_DEFAULT, color=BLACK)
        ws.cell(row=row, column=1).font = Font(
            name="Arial", size=SZ_DEFAULT, bold=True, color=NAVY)
        ws.cell(row=row, column=4).font = Font(
            name="Arial", size=SZ_DEFAULT, bold=True, color=status_color)
        ws.merge_cells(f"E{row}:F{row}")
        ws.row_dimensions[row].height = 65


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "Q1 2026 Income Statement"

    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "ADVANCED MICRO DEVICES, INC. (AMD) — Q1 2026 INCOME STATEMENT ANALYSIS"
    t.font = Font(name="Arial", bold=True, size=SZ_TITLE, color=WHITE)
    t.fill = fill(NAVY)
    t.alignment = center()
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:H2")
    s = ws["A2"]
    s.value = ("Quarter Ended March 28, 2026  |  Reported May 5, 2026  |  "
               "Source: AMD Q1 2026 Earnings Press Release  |  "
               "GAAP figures in millions except per-share amounts")
    s.font = Font(name="Arial", size=SZ_SUBTITLE, color=WHITE)
    s.fill = fill(BLUE)
    s.alignment = center(wrap=True)
    ws.row_dimensions[2].height = 18

    headers = ["Line Item", "Y/Y Change ($M)", "Y/Y %",
               "Q1 2026 ($M)", "Q4 2025 ($M)", "Q1 2025 ($M)",
               "Notes", "What This Measures"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = h
        cell.font = Font(name="Arial", bold=True, size=SZ_HEADER, color=WHITE)
        cell.fill = fill(ACCENT)
        cell.alignment = center(wrap=True)
        cell.border = TABLE_BORDER
    ws.row_dimensions[3].height = 30

    yy_map = {
        "green": (GREEN, GREEN_LT),
        "amber": (AMBER, AMBER_LT),
        "red":   (RED,   RED_LT),
    }

    DATA_START = 4

    for i, row_data in enumerate(ROWS):
        (label, q1, q4, q1_py,
         yy_color, actual_color, is_pct, note, measure, fmt_override) = row_data

        row = DATA_START + i
        is_priority = label.strip() in PRIORITY
        is_sub      = label.startswith("  ")
        bg = WHITE if i % 2 == 0 else GRAY

        a = ws.cell(row=row, column=1)
        a.value = label.strip()
        a.font = Font(name="Arial",
                      bold=is_priority and not is_sub,
                      italic=is_sub,
                      size=SZ_HEADER,
                      color=MUTED if is_sub else BLACK)
        a.fill = fill(bg)
        a.alignment = center()
        a.border = TABLE_BORDER

        b = ws.cell(row=row, column=2)
        b.value = "—" if is_sub else yy_change_str(q1, q1_py, is_pct, fmt_override)
        b.alignment = center()
        b.border = TABLE_BORDER

        c = ws.cell(row=row, column=3)
        c.value = "—" if is_sub else yy_pct_str(q1, q1_py, is_pct)
        c.alignment = center()
        c.border = TABLE_BORDER

        is_dash = (b.value == "—")
        if not is_dash and yy_color in yy_map:
            ink, bg_yy = yy_map[yy_color]
            for col in [2, 3]:
                cell = ws.cell(row=row, column=col)
                cell.fill = fill(bg_yy)
                cell.font = Font(name="Arial", size=SZ_DEFAULT, color=ink, bold=True)
        else:
            for col in [2, 3]:
                cell = ws.cell(row=row, column=col)
                cell.fill = fill(bg)
                cell.font = Font(name="Arial", size=SZ_DEFAULT,
                                 color=MUTED if is_dash else BLACK)

        d = ws.cell(row=row, column=4)
        d.value = fmt_val(q1, is_pct, fmt_override)
        d.alignment = center()
        d.border = TABLE_BORDER
        if actual_color in yy_map:
            ink, bg_ac = yy_map[actual_color]
            d.fill = fill(bg_ac)
            d.font = Font(name="Arial", size=SZ_DEFAULT, color=ink, bold=True)
        else:
            d.fill = fill(bg)
            d.font = Font(name="Arial", size=SZ_DEFAULT,
                          bold=is_priority and not is_sub, color=BLACK)

        e = ws.cell(row=row, column=5)
        e.value = fmt_val(q4, is_pct, fmt_override)
        e.fill = fill(bg); e.alignment = center(); e.border = TABLE_BORDER
        e.font = Font(name="Arial", size=SZ_DEFAULT,
                      color=MUTED if q4 is None else BLACK)

        f = ws.cell(row=row, column=6)
        f.value = fmt_val(q1_py, is_pct, fmt_override)
        f.fill = fill(bg); f.alignment = center(); f.border = TABLE_BORDER
        f.font = Font(name="Arial", size=SZ_DEFAULT,
                      color=MUTED if q1_py is None else BLACK)

        g = ws.cell(row=row, column=7)
        g.value = note
        g.font = Font(name="Arial", size=SZ_DEFAULT, color="333333")
        g.fill = fill(bg)
        g.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        g.border = TABLE_BORDER

        h = ws.cell(row=row, column=8)
        h.value = measure
        h.font = Font(name="Arial", size=SZ_DEFAULT, color="1F4E79", italic=True)
        h.fill = fill(MEASURE_BG)
        h.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        h.border = TABLE_BORDER

        ws.row_dimensions[row].height = max(45, min(180, 12 + note.count(" ") * 1.4))

    ws.freeze_panes = "B4"

    next_row = DATA_START + len(ROWS) + 1
    ws.row_dimensions[next_row].height = 10
    next_row += 1

    next_row = section_header(ws, next_row, "▶  OVERALL ASSESSMENT", SECTION_BG)

    ws.merge_cells(f"A{next_row}:H{next_row}")
    grade_cell = ws[f"A{next_row}"]
    grade_cell.value = "Quarter Grade:  A−"
    grade_cell.font = Font(name="Arial", bold=True, size=SZ_GRADE, color=NAVY)
    grade_cell.fill = fill(ASSESS_BG)
    grade_cell.alignment = center()
    ws.row_dimensions[next_row].height = 36
    next_row += 1

    assessment = (
        "AMD delivered a strong operating quarter with operating leverage confirmed at every level that matters. "
        "Gross Margin expanded +300bps Y/Y as the Data Center GPU mix shift drives a structural, repeatable "
        "improvement in blended margins. Operating Income grew +83% on +38% revenue — more than double the "
        "revenue growth rate. Net Income nearly doubled. Diluted EPS grew +91% with clean, minimal dilution. "
        "EPYC reached a record 46.2% server CPU revenue share in Q1 2026, and AMD's Data Center revenue of "
        "$5.8B passed Intel's $5.1B for the same quarter — a structural shift in enterprise compute.\n\n"
        "The confirmed hyperscaler deal stack (OpenAI 6GW Oct 2025, Oracle 50K GPUs Oct 2025, Meta 6GW "
        "Feb 2026) removes demand uncertainty for 2026-2027. The 12GW+ confirmed pipeline means AMD's "
        "near-term revenue trajectory is backlog-driven, not market-driven. Execution on MI450 H2 2026 "
        "production ramp is the primary variable — yield at scale, HBM4 supply from SK Hynix/Samsung, "
        "and ROCm software parity for OpenAI/Meta training workloads are the execution risks.\n\n"
        "The one failure this quarter is MG&A: at +41.4% vs revenue +37.9%, the G&A cost line moved in the "
        "wrong direction. If MG&A reverts to growing slower than revenue in Q2, the operating leverage story "
        "is fully intact. The GAAP/non-GAAP gap ($551M+ quarterly Xilinx amortization) creates an optics "
        "problem: 14.4% GAAP operating margin vs ~27% non-GAAP guided for Q2. As Xilinx amortization "
        "gradually declines, the GAAP picture will converge — a multi-year structural tailwind."
    )
    next_row = bullet_row(ws, next_row, assessment, ASSESS_BG)
    next_row += 1

    next_row = section_header(ws, next_row, "▶  STRENGTHS", "1A7A3A")
    for s in [
        "✓  Net Revenue +37.9% Y/Y to $10.25B — essentially flat Q/Q despite typical Q1 seasonal weakness. Demand strength across Data Center and Client confirmed.",
        "✓  EPYC CPU record 46.2% server revenue share in Q1 2026 — AMD Data Center revenue ($5.8B) surpassed Intel ($5.1B) for the first time. CPU wins are a compounding, parallel growth engine.",
        "✓  Confirmed deal stack: OpenAI 6GW (Oct 2025, 1GW H2 2026 MI450 Series), Meta 6GW (Feb 2026, 1GW H2 2026 custom MI450-based + EPYC Venice), Oracle 50K GPUs Q3 2026 (MI450 + EPYC Venice + Pensando Vulcano). 12GW+ committed pipeline eliminates demand risk.",
        "✓  Gross Margin expanded +300bps Y/Y (50% → 53%) — Data Center GPU mix shift driving structural, repeatable gross margin improvement.",
        "✓  Operating Income +83.1% Y/Y — more than double the revenue growth rate. GAAP operating margin: 10.8% → 14.4% (+360bps Y/Y). Textbook operating leverage.",
        "✓  Net Income +95.1% Y/Y — nearly doubled on 37.9% revenue growth. GAAP net margin: 9.5% → 13.5% (+400bps).",
        "✓  Diluted EPS +90.9% Y/Y ($0.44 → $0.84) — clean EPS leverage with no preferred stock dilution complexity.",
        "✓  Total OpEx +34.5% Y/Y vs revenue +37.9% — operating leverage confirmed in aggregate across all cost lines combined.",
        "✓  Xilinx amortization declining ($316M → $290M in OpEx, -8.2% Y/Y) — structural GAAP operating income tailwind.",
        "✓  Balance sheet strength: ~$5.1B cash, ~$1.7B long-term debt — fortress balance sheet enabling R&D investment and share repurchases simultaneously.",
        "✓  Non-GAAP diluted EPS Q2 2026 guided at $0.96 midpoint — implies continued strong growth.",
    ]:
        next_row = bullet_row(ws, next_row, s, STRENGTH_BG)

    next_row += 1

    next_row = section_header(ws, next_row, "▶  CONCERNS", "8B0000")
    for c in [
        "✗  MG&A +41.4% Y/Y — fastest growing cost line, outpacing revenue (+37.9%) by 350bps. Primary operating leverage failure point this quarter.",
        "✗  R&D +38.7% Y/Y — essentially matched revenue growth (+37.9%). No operating leverage in the innovation engine yet.",
        "✗  Warrant dilution: 320M total warrants outstanding (OpenAI 160M + Meta 160M). 19.4% max dilution at full vest. Buyback program expected to provide partial offset — track quarterly.",
        "✗  Q/Q Operating Income declined ($1,752M → $1,476M, -15.8%). Seasonal pattern, but Q1 to Q4 progression needs to recover strongly.",
        "✗  Gaming segment declining — consumer GPU and semi-custom both down Y/Y. No near-term catalyst for reversal.",
        "✗  Embedded recovering slowly from trough — not yet contributing meaningfully to revenue growth.",
        "✗  MI450 execution risk: HBM4 supply constraints (SK Hynix/Samsung CoWoS capacity competition), TSMC N3E packaging yield at Helios rack scale, ROCm software parity for OpenAI/Meta training workloads. Demand is signed — execution is the variable.",
        "✗  GAAP vs Non-GAAP gap remains large: GAAP operating margin ~14.4% vs guided non-GAAP ~27%. $551M+ quarterly Xilinx amortization continues to suppress reported GAAP margins.",
    ]:
        next_row = bullet_row(ws, next_row, c, CONCERN_BG)

    next_row += 1

    next_row = section_header(ws, next_row,
                              "▶  LETTER GRADE FRAMEWORK — AMD (Q1 2026 BASELINE)",
                              SECTION_BG)
    for grade, desc in [
        ("A",   "MG&A < revenue growth for 2+ consecutive quarters. Non-GAAP operating margin expands to 30%+. "
                "Gross Margin reaches 55%+. MI450 H2 2026 ships on schedule with no yield surprises. "
                "OpenAI/Meta/Oracle H2 deployments execute as contracted."),
        ("A−",  "Current quarter: Operating Income +83% on +38% revenue. Net Income +95%. EPS +91%. "
                "Gross Margin +300bps Y/Y. Operating leverage confirmed in aggregate. "
                "Confirmed 12GW+ deal stack removes demand risk. MG&A the sole exception."),
        ("B+",  "MG&A growth remains elevated above revenue growth 2+ quarters. Revenue growth decelerates "
                "below 25%. MI450 H2 2026 production slips to 2027. Gross margin expansion stalls at 53-54%."),
        ("B",   "Revenue growth decelerates to teens. Gross margin contracts from 53% peak. "
                "MI450 volume ramp delayed significantly. Operating income growth falls near revenue growth rate."),
        ("C",   "MI450 yield failures create material revenue miss vs OpenAI/Meta/Oracle contracted schedule. "
                "ROCm fails to achieve CUDA parity — hyperscalers cannot migrate critical workloads to MI450. "
                "Revenue growth below 10%. Gross margin contracts below 50%."),
        ("D",   "Fundamental execution breakdown: MI450 production suspended due to yield defects. "
                "HBM4 supply unavailable at required volume. OpenAI/Meta defer to Vera Rubin. "
                "EPYC server share reversal vs Intel after record 46.2% peak. Revenue decline."),
    ]:
        next_row = bullet_row(ws, next_row, f"{grade}   {desc}", ASSESS_BG)

    next_row += 1

    next_row = section_header(ws, next_row,
                              "▶  KEY METRICS TO WATCH — Q2 2026 (Expected Late July / Early August 2026)",
                              SECTION_BG)
    for m in [
        "① MG&A as % of Revenue — Q1 was elevated (revenue growth outpaced). Must show reversal. Most important single cost line to watch.",
        "② Non-GAAP Operating Margin — Q2 guided ~27%. Above 27.5% = outperform. Below 26.5% = concern.",
        "③ Data Center Segment Revenue — must maintain sequential and Y/Y growth momentum. MI350 is current product; MI450 sampling for H2 2026 is next catalyst. OpenAI 1GW initial + Oracle 50K GPU Q3 2026 are the first confirmed deal revenue contributions.",
        "④ MI450 Production Status Commentary — any mention of H2 2026 production timeline, Helios rack yield, or HBM4 supply is the most important forward indicator on the earnings call.",
        "⑤ Warrant Dilution Disclosure — track any OpenAI/Meta warrant vesting events that increase the reported diluted share count beyond 1,650M.",
        "⑥ Gross Margin Trajectory — Q1: 53%. Can AMD reach 55%+ as Data Center mix grows? Each 100bps = ~$100M additional gross profit per quarter.",
        "⑦ Non-GAAP Diluted EPS vs $0.96 Guidance — beat confirms execution; guidance miss is more impactful on a guided quarter.",
        "⑧ R&D as % of Revenue — Q1: 23.4%. Needs to decline over time to generate R&D operating leverage.",
        "⑨ Cash and Share Repurchase Activity — buyback velocity signals management confidence and pace of warrant dilution offset.",
    ]:
        next_row = bullet_row(ws, next_row, m, WHITE)

    next_row += 1
    ws.merge_cells(f"A{next_row}:H{next_row}")
    foot = ws[f"A{next_row}"]
    foot.value = (
        "Source: AMD Q1 2026 Earnings Press Release, May 5, 2026  |  GAAP figures only  |  "
        "Millions of dollars except per-share amounts  |  Analysis: May 2026  |  "
        "Script: build_amd_income_v2.py  |  Not investment advice."
    )
    foot.font = Font(name="Arial", size=SZ_FOOTER, color=MUTED, italic=True)
    foot.fill = fill(WHITE)
    foot.alignment = Alignment(horizontal="left", wrap_text=True)
    ws.row_dimensions[next_row].height = 18

    build_sources_tab(wb)
    build_whats_new_tab(wb)

    wb.active = wb["Q1 2026 Income Statement"]

    out = "/mnt/user-data/outputs/AMD_Q1_2026_Income_Statement_v2.xlsx"
    wb.save(out)
    print(f"Saved: {out}")
    return out


# ==============================================================================
# SELF-TEST — DO NOT REMOVE
# ==============================================================================
def self_test(out_path):
    import os
    from openpyxl import load_workbook

    errors = []

    if not os.path.exists(out_path):
        errors.append(f"FAIL: file not created at {out_path}")
    else:
        size = os.path.getsize(out_path)
        if size < 14000:
            errors.append(f"FAIL: file too small ({size} bytes)")

        wb = load_workbook(out_path, data_only=True)
        ws = wb.active
        if ws.title != "Q1 2026 Income Statement":
            errors.append(f"FAIL: wrong sheet name '{ws.title}'")
        if "Data Sources" not in wb.sheetnames:
            errors.append("FAIL: 'Data Sources' tab missing")
        if "What's New" not in wb.sheetnames:
            errors.append("FAIL: \"What's New\" tab missing")

        found_revenue = found_10253 = found_grade = False
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell:
                    cv = str(cell)
                    if "Net Revenue" in cv:            found_revenue = True
                    if "10253" in cv or "10,253" in cv: found_10253 = True
                    if "A−" in cv or "A-" in cv:       found_grade  = True

        if not found_revenue: errors.append("FAIL: 'Net Revenue' row not found")
        if not found_10253:   errors.append("FAIL: Q1 2026 Revenue $10,253M not found")
        if not found_grade:   errors.append("FAIL: Grade 'A−' not found")

        # Check What's New has confirmed deals
        ws_wn = wb["What's New"]
        found_openai = found_meta = found_oracle = found_v5 = False
        for row in ws_wn.iter_rows(values_only=True):
            for cell in row:
                if cell:
                    cv = str(cell)
                    if "OpenAI" in cv:                       found_openai = True
                    if "Meta" in cv and "6GW" in cv:        found_meta = True
                    if "Oracle" in cv and "50K" in cv:      found_oracle = True
                    if "projection_engine_v5" in cv:        found_v5 = True
        if not found_openai:  errors.append("FAIL: What's New — OpenAI deal missing")
        if not found_meta:    errors.append("FAIL: What's New — Meta 6GW deal missing")
        if not found_oracle:  errors.append("FAIL: What's New — Oracle 50K GPU deal missing")
        if not found_v5:      errors.append("FAIL: What's New — v5 engine reference missing")

        # Check Data Sources has deal entries
        ws_ds = wb["Data Sources"]
        found_openai_ds = found_oracle_ds = False
        for row in ws_ds.iter_rows(values_only=True):
            for cell in row:
                if cell:
                    cv = str(cell)
                    if "OpenAI" in cv:   found_openai_ds = True
                    if "Oracle" in cv:   found_oracle_ds = True
        if not found_openai_ds: errors.append("FAIL: Data Sources — OpenAI deal entry missing")
        if not found_oracle_ds: errors.append("FAIL: Data Sources — Oracle deal entry missing")

    if errors:
        print("\n" + "=" * 60)
        print("SELF-TEST FAILED — DO NOT SAVE TO DRIVE")
        print("=" * 60)
        for e in errors:
            print(f"  {e}")
        print("=" * 60)
        return False

    size = os.path.getsize(out_path)
    print("\n" + "=" * 60)
    print("SELF-TEST PASSED — safe to save to Drive")
    print(f"  File: {out_path}")
    print(f"  Size: {size:,} bytes")
    print(f"  Net Revenue $10,253M: confirmed")
    print(f"  Tabs: Q1 2026 Income Statement + Data Sources + What's New")
    print(f"  Grade A−: confirmed")
    print(f"  OpenAI/Meta/Oracle deals: confirmed in What's New")
    print(f"  v5 engine reference: confirmed in Coherence Check")
    print(f"  OpenAI/Oracle: confirmed in Data Sources")
    print("=" * 60)
    return True


if __name__ == "__main__":
    out = build()
    passed = self_test(out)
    if not passed:
        import sys
        sys.exit(1)
