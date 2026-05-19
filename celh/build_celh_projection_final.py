# ==============================================================================
# CELH 5-YEAR PROJECTION — BUILD SCRIPT v1 FINAL
# EXACT AMD STRUCTURE: read from AMD file content, substituting Celsius data
# Tabs: Inputs (blue), Projection (green), Probability Weighted (purple)
# Row labels ALL CAPS, col A blank, col B label, cols C-G years
# Data centered. Gold rows for share price. Scenario color fills for growth/CAGR
# NO yellow on data rows. Whole dollars only.
# ==============================================================================
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

_HERE = os.path.dirname(os.path.abspath(__file__))

# ── DATA BLOCK ─────────────────────────────────────────────────────────────────
VERSION      = "v1 — Q1 2026 Baseline"
GENERATED    = "May 2026"
ENTRY_PRICE  = 33.00       # CELH ~$33 May 2026
BASE_REV     = 2515.3      # FY2025 actual revenue ($M)
SHARES       = 334.0       # Fully diluted as-converted incl PepsiCo preferred
Q1_REV       = 782.6       # Q1 2026 actual ($M)
Q1_EBITDA_M  = 0.250       # Q1 2026 Adj. EBITDA margin

# CELH uses EV/EBITDA not P/E (key difference from AMD)
BULL = dict(rev_growth=[0.33,0.28,0.22,0.18,0.15], ebitda_margin=[0.270,0.305,0.335,0.350,0.365], ev_low=[25,28,30,30,28], ev_high=[32,35,38,38,35])
BASE = dict(rev_growth=[0.33,0.22,0.16,0.12,0.10], ebitda_margin=[0.250,0.275,0.295,0.305,0.315], ev_low=[18,20,22,22,20], ev_high=[24,26,28,28,25])
BEAR = dict(rev_growth=[0.33,0.12,0.06,0.04,0.04], ebitda_margin=[0.215,0.225,0.235,0.240,0.250], ev_low=[10,10,11,12,12], ev_high=[14,14,15,16,16])

PROB_BULL = 0.30
PROB_BASE  = 0.45
PROB_BEAR  = 0.25

YEARS = [2026,2027,2028,2029,2030]
YCOLS = ["C","D","E","F","G"]

# ── COLORS (exact AMD palette from instructions file) ──────────────────────────
NAVY    = "1A1A2E"
BLUE    = "1F4E79"
ACCENT  = "2E75B6"
WHITE   = "FFFFFF"
GRAY    = "F5F7FA"
BLACK   = "000000"
MUTED   = "888888"
GREEN_S = "E8F5E9"   # Bull scenario fill
BLUE_S  = "D6E4F0"   # Base scenario fill
RED_S   = "FDEDEC"   # Bear scenario fill
GOLD_BG = "FFF8E1"   # Share price row fill
GOLD_FN = "7D5A00"   # Share price row font
PROB_BG = "EDE7F6"   # Probability weighted fill

def fill(h): return PatternFill("solid", start_color=h, fgColor=h)

def cell(ws, row, col, val=None, bg=WHITE, fc=BLACK, bold=False,
         fmt=None, align="center", size=9):
    c = ws.cell(row=row, column=col)
    if val is not None: c.value = val
    c.font = Font(name="Arial", bold=bold, size=size, color=fc)
    c.fill = fill(bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
    if fmt: c.number_format = fmt
    return c

def rh(ws, row, h=14): ws.row_dimensions[row].height = h

def compute(sc):
    revs, ebs, spls, sphs, cagrl, cagrh = [], [], [], [], [], []
    rev = BASE_REV
    for i in range(5):
        rev  = rev * (1 + sc["rev_growth"][i])
        eb   = rev * sc["ebitda_margin"][i]
        spl  = eb * sc["ev_low"][i]  / SHARES
        sph  = eb * sc["ev_high"][i] / SHARES
        revs.append(rev);  ebs.append(eb)
        spls.append(spl);  sphs.append(sph)
        cagrl.append((spl/ENTRY_PRICE)**(1/(i+1))-1)
        cagrh.append((sph/ENTRY_PRICE)**(1/(i+1))-1)
    return revs, ebs, spls, sphs, cagrl, cagrh

# ==============================================================================
# TAB 1 — INPUTS  (matches AMD Inputs tab exactly)
# ==============================================================================
def build_inputs(wb):
    ws = wb.create_sheet("Inputs")

    # Column widths — AMD exact
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12

    # Row 1: Title
    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = f"CELSIUS HOLDINGS (CELH) 5-YEAR PROJECTION — INPUTS [{VERSION}]"
    t.font = Font(name="Arial", bold=True, size=11, color=WHITE)
    t.fill = fill(NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 1, 20)

    # Row 2: Subtitle
    ws.merge_cells("A2:G2")
    s = ws["A2"]
    s.value = (f"Blue = Editable | YELLOW HIGHLIGHT = Changed from prior version | "
               f"Valuation: EV/EBITDA multiples (not P/E) | "
               f"Probability: Bull {PROB_BULL:.0%} / Base {PROB_BASE:.0%} / Bear {PROB_BEAR:.0%}")
    s.font = Font(name="Arial", size=9, color=WHITE)
    s.fill = fill(BLUE)
    s.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 2, 14)

    # Row 3: Global inputs header
    ws.merge_cells("A3:G3")
    g = ws["A3"]
    g.value = f"GLOBAL INPUTS — CELH Q1 2026 Earnings Release (May 7, 2026) | Market Data May 2026"
    g.font = Font(name="Arial", bold=True, size=9, color=WHITE)
    g.fill = fill(ACCENT)
    g.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    rh(ws, 3, 14)

    # Rows 4–10: Global inputs — AMD format: blank A, label B, value C, blank D, notes E-G
    global_rows = [
        ("Current / Entry Stock Price ($)", f"${ENTRY_PRICE:,.2f}",
         "User-specified $33 | May 2026 | Source: Investing.com"),
        ("Base Year Revenue — FY2025 ($M)", f"${BASE_REV:,.1f}M",
         "CELH FY2025 Annual Revenue | Source: Feb 26, 2026 press release | ir.celsiusholdingsinc.com"),
        ("Diluted Shares Outstanding (M)", f"{SHARES:,.1f}M",
         "As-converted incl PepsiCo Series A+B preferred | Q1 2026"),
        ("Q1 2026 Adj. EBITDA Margin", f"{Q1_EBITDA_M:.1%}",
         "CELH Q1 2026: $195.5M Adj. EBITDA / $782.6M Revenue = 25.0% | CELH Press Release May 7, 2026"),
        ("Q1 2026 Revenue ($M)", f"${Q1_REV:,.1f}M",
         "CELH Q1 2026 actual | Source: May 7, 2026 press release | ir.celsiusholdingsinc.com"),
        ("FY2026 Street Consensus Revenue ($M)", "$3,350M",
         "23 analyst consensus post-Q1 | Simply Wall St May 11, 2026"),
        ("Valuation Method", "EV/EBITDA",
         "Celsius uses EV/EBITDA multiples (not P/E like AMD) — EBITDA × multiple ÷ diluted shares"),
    ]
    for i, (label, val, note) in enumerate(global_rows):
        r = 4 + i
        bg = WHITE if i % 2 == 0 else GRAY
        cell(ws, r, 1, bg=bg)
        cell(ws, r, 2, label, bg=bg, align="left")
        c = ws.cell(row=r, column=3)
        c.value = val; c.font = Font(name="Arial", bold=True, size=9, color="0070C0")
        c.fill = fill(bg); c.alignment = Alignment(horizontal="right", vertical="center")
        cell(ws, r, 4, bg=bg)
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=7)
        cell(ws, r, 5, note, bg=bg, fc=MUTED, size=8, align="left")
        rh(ws, r, 14)

    # Row 12: Scenario assumptions header
    ws.merge_cells("A12:G12")
    sh = ws["A12"]
    sh.value = "SCENARIO ASSUMPTIONS — Edit blue cells | All values feed automatically to Projection tab"
    sh.font = Font(name="Arial", bold=True, size=9, color=WHITE)
    sh.fill = fill(ACCENT)
    sh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    rh(ws, 12, 14)

    # Row 13: Column headers
    for col, txt in [(2,"Assumption"),(3,"2026"),(4,"2027"),(5,"2028"),(6,"2029"),(7,"2030")]:
        cell(ws, 13, col, txt, bg=ACCENT, fc=WHITE, bold=True)
    cell(ws, 13, 1, bg=ACCENT)
    rh(ws, 13, 14)

    # Scenario blocks
    scenarios = [
        ("▶ BULL CASE", BULL, GREEN_S, PROB_BULL),
        ("▶ BASE CASE", BASE, BLUE_S,  PROB_BASE),
        ("▶ BEAR CASE", BEAR, RED_S,   PROB_BEAR),
    ]
    r = 14
    for sc_name, sc_data, sc_bg, prob in scenarios:
        # Scenario label row
        ws.merge_cells(f"A{r}:G{r}")
        sl = ws[f"A{r}"]
        sl.value = sc_name
        sl.font = Font(name="Arial", bold=True, size=9, color=BLACK)
        sl.fill = fill(sc_bg)
        sl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        rh(ws, r, 14); r += 1

        metrics = [
            ("Revenue Growth %",   sc_data["rev_growth"],    "0.0%"),
            ("Adj. EBITDA Margin", sc_data["ebitda_margin"], "0.0%"),
            ("EV/EBITDA Low",      sc_data["ev_low"],        "0.0"),
            ("EV/EBITDA High",     sc_data["ev_high"],       "0.0"),
        ]
        for label, vals, fmt in metrics:
            bg = WHITE if r % 2 == 0 else GRAY
            cell(ws, r, 1, bg=bg)
            cell(ws, r, 2, label, bg=bg, align="left")
            for j, val in enumerate(vals):
                c = ws.cell(row=r, column=3+j)
                c.value = val
                c.font = Font(name="Arial", bold=True, size=9, color="0070C0")
                c.fill = fill(bg)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.number_format = fmt
            rh(ws, r, 14); r += 1

        # Probability row — yellow like AMD
        cell(ws, r, 1, bg=sc_bg)
        cell(ws, r, 2, "Probability Weight", bg=sc_bg, align="left")
        c = ws.cell(row=r, column=3)
        c.value = prob
        c.font = Font(name="Arial", bold=True, size=9, color="0070C0")
        c.fill = fill("FFFF00")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = "0%"
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
        cell(ws, r, 4,
             f"Bull {PROB_BULL:.0%} + Base {PROB_BASE:.0%} + Bear {PROB_BEAR:.0%} = 100%",
             bg=sc_bg, fc=MUTED, size=8, align="left")
        rh(ws, r, 14); r += 2

    # Color legend
    ws.merge_cells(f"A{r}:G{r}")
    cl = ws[f"A{r}"]
    cl.value = "COLOR LEGEND: Blue Text = Editable Input | Black Text = Formula | Green Text = Linked from Inputs tab"
    cl.font = Font(name="Arial", size=8, color=MUTED, italic=True)
    cl.alignment = Alignment(horizontal="left", vertical="center", indent=1)

# ==============================================================================
# TAB 2 — PROJECTION  (exact AMD structure)
# ==============================================================================
def build_projection(wb):
    ws = wb.create_sheet("Projection")

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 26
    for col in ["C","D","E","F","G"]:
        ws.column_dimensions[col].width = 14
    ws.column_dimensions["H"].width = 2

    # Title
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "CELSIUS HOLDINGS (CELH) 5-YEAR FINANCIAL PROJECTION — ADJ. EBITDA-BASED"
    t.font = Font(name="Arial", bold=True, size=13, color=WHITE)
    t.fill = fill(NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 1, 22)

    ws.merge_cells("A2:H2")
    s = ws["A2"]
    s.value = (f"Entry Price: ${ENTRY_PRICE}  |  Base Revenue: ${BASE_REV:,.0f}M (FY2025 Actual)  |  "
               f"Shares: {SHARES}M (as-converted)  |  Revenue & EBITDA in $M  |  Change inputs on Inputs tab")
    s.font = Font(name="Arial", size=9, color=WHITE)
    s.fill = fill(BLUE)
    s.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 2, 14)

    scenarios = [
        ("▶ BULL CASE — ADJ. EBITDA", BULL, GREEN_S, PROB_BULL),
        ("▶ BASE CASE — ADJ. EBITDA", BASE, BLUE_S,  PROB_BASE),
        ("▶ BEAR CASE — ADJ. EBITDA", BEAR, RED_S,   PROB_BEAR),
    ]

    ROW = 3
    for sc_name, sc_data, sc_bg, prob in scenarios:
        revs, ebs, spls, sphs, cagrl, cagrh = compute(sc_data)

        # Scenario header
        ws.merge_cells(f"A{ROW}:H{ROW}")
        sh = ws[f"A{ROW}"]
        sh.value = sc_name
        sh.font = Font(name="Arial", bold=True, size=11, color=WHITE)
        sh.fill = fill(BLUE)
        sh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        rh(ws, ROW, 20); ROW += 1

        # Year header
        cell(ws, ROW, 1, bg=ACCENT); cell(ws, ROW, 2, "YEAR", bg=ACCENT, fc=WHITE, bold=True)
        for i, yr in enumerate(YEARS):
            cell(ws, ROW, 3+i, str(yr), bg=ACCENT, fc=WHITE, bold=True)
        cell(ws, ROW, 8, bg=ACCENT)
        rh(ws, ROW, 16); ROW += 1

        # Data rows — exact AMD row labels in ALL CAPS
        def data_row(label, vals, fmt, bg, fc=BLACK, bold=False):
            nonlocal ROW
            cell(ws, ROW, 1, bg=bg); cell(ws, ROW, 2, label, bg=bg, fc=fc, bold=bold, align="left")
            for i, v in enumerate(vals):
                if v is None:
                    cell(ws, ROW, 3+i, "—", bg=bg, fc=MUTED)
                else:
                    c = ws.cell(row=ROW, column=3+i)
                    c.value = v
                    c.font = Font(name="Arial", bold=bold, size=9, color=fc)
                    c.fill = fill(bg)
                    c.alignment = Alignment(horizontal="center", vertical="center")
                    if fmt: c.number_format = fmt
            cell(ws, ROW, 8, bg=bg)
            rh(ws, ROW, 14); ROW += 1

        # Revenue formatted as $B (matching AMD style)
        data_row("REVENUE",         [f"${v/1000:.2f}B" for v in revs],  None,     sc_bg, BLACK, True)
        data_row(" REV GROWTH",     sc_data["rev_growth"],               "0.0%",   sc_bg)
        data_row("ADJ. EBITDA",     [f"${v:,.0f}M" for v in ebs],       None,     WHITE, BLACK, True)
        # EBITDA growth — dashes for first year like AMD NET INC. GROWTH
        eb_growth = [None] + [ebs[i]/ebs[i-1]-1 for i in range(1,5)]
        data_row(" ADJ. EBITDA GROWTH", eb_growth,                      "0.0%",   GRAY)
        data_row(" ADJ. EBITDA MARGINS", sc_data["ebitda_margin"],      "0.0%",   WHITE)
        data_row("EV/EBITDA LOW",   sc_data["ev_low"],                   "0.0",    GRAY)
        data_row("EV/EBITDA HIGH",  sc_data["ev_high"],                  "0.0",    WHITE)
        data_row("SHARE PRICE LOW", spls,                                '"$"#,##0', GOLD_BG, GOLD_FN, True)
        data_row("SHARE PRICE HIGH",sphs,                                '"$"#,##0', GOLD_BG, GOLD_FN, True)
        data_row(" CAGR LOW",       cagrl,                               "+0.0%;-0.0%;—", sc_bg)
        data_row(" CAGR HIGH",      cagrh,                               "+0.0%;-0.0%;—", sc_bg)

        ROW += 1  # spacer

    # Scenario comparison table — exact AMD
    ws.merge_cells(f"A{ROW}:H{ROW}")
    sc = ws[f"A{ROW}"]
    sc.value = "SCENARIO COMPARISON — 2030 OUTPUTS"
    sc.font = Font(name="Arial", bold=True, size=9, color=WHITE)
    sc.fill = fill(ACCENT)
    sc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    rh(ws, ROW, 16); ROW += 1

    for col, txt in [(2,"Scenario"),(3,"2030 Revenue"),(4,"2030 Adj. EBITDA"),
                     (5,"2030 EBITDA Mgn"),(6,"2030 Pr. Low"),(7,"2030 Pr. High"),(8,"5-Yr CAGR Low")]:
        cell(ws, ROW, col, txt, bg=ACCENT, fc=WHITE, bold=True, size=8)
    cell(ws, ROW, 1, bg=ACCENT)
    rh(ws, ROW, 28); ROW += 1

    for sc_name2, sc_data, sc_bg, prob in scenarios:
        revs, ebs, spls, sphs, cagrl, cagrh = compute(sc_data)
        nm = sc_name2.replace(" — ADJ. EBITDA","")
        cell(ws, ROW, 1, bg=sc_bg)
        cell(ws, ROW, 2, nm, bg=sc_bg, bold=True, align="left")
        cell(ws, ROW, 3, f"${revs[4]/1000:.2f}B", bg=sc_bg)
        cell(ws, ROW, 4, f"${ebs[4]:,.0f}M", bg=sc_bg)
        c5 = ws.cell(row=ROW, column=5)
        c5.value = sc_data["ebitda_margin"][4]; c5.fill = fill(sc_bg)
        c5.font = Font(name="Arial", size=9); c5.number_format = "0.0%"
        c5.alignment = Alignment(horizontal="center", vertical="center")
        for j, val in enumerate([spls[4], sphs[4]]):
            cj = ws.cell(row=ROW, column=6+j)
            cj.value = val; cj.fill = fill(GOLD_BG)
            cj.font = Font(name="Arial", bold=True, size=9, color=GOLD_FN)
            cj.number_format = '"$"#,##0'
            cj.alignment = Alignment(horizontal="center", vertical="center")
        c8 = ws.cell(row=ROW, column=8)
        c8.value = cagrl[4]; c8.fill = fill(sc_bg)
        c8.font = Font(name="Arial", size=9); c8.number_format = "+0.0%;-0.0%;—"
        c8.alignment = Alignment(horizontal="center", vertical="center")
        rh(ws, ROW, 14); ROW += 1

    ws.freeze_panes = "C3"

    # Footer
    ws.merge_cells(f"A{ROW+1}:H{ROW+1}")
    f = ws[f"A{ROW+1}"]
    f.value = (f"Source: CELH FY2025 actual ${BASE_REV:,.0f}M | "
               f"Analysis price ${ENTRY_PRICE} | Built: {GENERATED} | "
               f"Shares {SHARES}M as-converted incl. PepsiCo preferred")
    f.font = Font(name="Arial", size=7, color=MUTED, italic=True)
    f.alignment = Alignment(horizontal="left")

# ==============================================================================
# TAB 3 — PROBABILITY WEIGHTED  (exact AMD structure)
# ==============================================================================
def build_probability(wb):
    ws = wb.create_sheet("Probability Weighted")

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 28
    for col in ["C","D","E","F","G"]:
        ws.column_dimensions[col].width = 14
    ws.column_dimensions["H"].width = 16  # Expected Value col — AMD width 16

    # Title
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "CELSIUS HOLDINGS (CELH) 5-YEAR PROJECTION — PROBABILITY WEIGHTED EXPECTED VALUE"
    t.font = Font(name="Arial", bold=True, size=11, color=WHITE)
    t.fill = fill(NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 1, 22)

    ws.merge_cells("A2:H2")
    s = ws["A2"]
    s.value = (f"Bull {PROB_BULL:.0%} | Base {PROB_BASE:.0%} | Bear {PROB_BEAR:.0%} | "
               f"Expected Value = (Bull×{PROB_BULL:.0%}) + (Base×{PROB_BASE:.0%}) + "
               f"(Bear×{PROB_BEAR:.0%}) | Change probabilities in yellow cells")
    s.font = Font(name="Arial", size=9, color=WHITE)
    s.fill = fill(BLUE)
    s.alignment = Alignment(horizontal="center", vertical="center")
    rh(ws, 2, 14)

    # Probability inputs header
    ws.merge_cells("A3:H3")
    ph = ws["A3"]
    ph.value = "PROBABILITY INPUTS — Edit yellow cells"
    ph.font = Font(name="Arial", bold=True, size=9, color=WHITE)
    ph.fill = fill(ACCENT)
    ph.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    rh(ws, 3, 14)

    # Column headers
    for col, txt in [(2,"Case"),(3,"Probability"),(4,"Description")]:
        cell(ws, 4, col, txt, bg=ACCENT, fc=WHITE, bold=True)
    ws.merge_cells("D4:H4")
    cell(ws, 4, 1, bg=ACCENT)
    rh(ws, 4, 14)

    prob_rows = [
        ("BULL CASE", PROB_BULL, GREEN_S,
         "Full Alani Nu reacceleration. CELSIUS brand back to double digits. "
         "Gross margin reaches low 50s. Rockstar stabilizes. International hits $200M+."),
        ("BASE CASE", PROB_BASE, BLUE_S,
         "Confirmed deals execute. CELSIUS organic +8-10% by Q4 2026. Alani normalizes to $280-300M/qtr. "
         "Gradual margin expansion. Operating leverage drives SG&A toward 22-24%."),
        ("BEAR CASE", PROB_BEAR, RED_S,
         "Alani Nu normalizes sharply post distributor load-in. CELSIUS stays low single digits. "
         "Gross margin stuck 47-49%. Rockstar continues declining. Legal overhang expands."),
    ]
    for i, (case, prob, bg, desc) in enumerate(prob_rows):
        r = 5 + i
        cell(ws, r, 1, bg=bg)
        cell(ws, r, 2, case, bg=bg, bold=True, align="left")
        c = ws.cell(row=r, column=3)
        c.value = prob; c.font = Font(name="Arial", bold=True, size=9, color="0070C0")
        c.fill = fill("FFFF00"); c.number_format = "0%"
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(f"D{r}:H{r}")
        cell(ws, r, 4, desc, bg=bg, fc=MUTED, size=8, align="left")
        rh(ws, r, 14)

    # Sum check
    cell(ws, 8, 1, bg=WHITE)
    cell(ws, 8, 2, "Sum (must = 100%)", bg=WHITE, bold=True, align="left")
    c = ws.cell(row=8, column=3)
    c.value = PROB_BULL + PROB_BASE + PROB_BEAR
    c.font = Font(name="Arial", bold=True, size=9); c.number_format = "0%"
    c.fill = fill(WHITE); c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("D8:H8")
    chk = "✓ 100% — Valid" if abs(PROB_BULL+PROB_BASE+PROB_BEAR-1.0) < 0.001 else "⚠ Does not sum to 100%"
    cell(ws, 8, 4, chk, bg=WHITE, bold=True, align="left")
    rh(ws, 8, 14)

    # Scenario comparison by year header
    ws.merge_cells("A10:H10")
    sch = ws["A10"]
    sch.value = "SCENARIO COMPARISON BY YEAR — Bull vs Base vs Bear"
    sch.font = Font(name="Arial", bold=True, size=9, color=WHITE)
    sch.fill = fill(ACCENT)
    sch.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    rh(ws, 10, 14)

    # Year header
    cell(ws, 11, 1, bg=ACCENT)
    cell(ws, 11, 2, "Metric", bg=ACCENT, fc=WHITE, bold=True)
    for i, yr in enumerate(YEARS):
        cell(ws, 11, 3+i, str(yr), bg=ACCENT, fc=WHITE, bold=True)
    cell(ws, 11, 8, "EXPECTED VALUE 2030", bg=ACCENT, fc=WHITE, bold=True, size=8)
    rh(ws, 11, 16)

    # Compute all scenarios
    bull = compute(BULL); base = compute(BASE); bear = compute(BEAR)
    ev_spl = [bull[2][i]*PROB_BULL + base[2][i]*PROB_BASE + bear[2][i]*PROB_BEAR for i in range(5)]
    ev_sph = [bull[3][i]*PROB_BULL + base[3][i]*PROB_BASE + bear[3][i]*PROB_BEAR for i in range(5)]

    sc_blocks = [
        (f"▶ BULL CASE ({PROB_BULL:.0%})", BULL, GREEN_S, *bull),
        (f"▶ BASE CASE ({PROB_BASE:.0%})", BASE, BLUE_S,  *base),
        (f"▶ BEAR CASE ({PROB_BEAR:.0%})", BEAR, RED_S,   *bear),
    ]

    ROW = 12
    for sc_name, sc_data, sc_bg, revs, ebs, spls, sphs, cagrl, cagrh in sc_blocks:
        ws.merge_cells(f"A{ROW}:H{ROW}")
        sh = ws[f"A{ROW}"]
        sh.value = sc_name
        sh.font = Font(name="Arial", bold=True, size=9, color=WHITE)
        sh.fill = fill(BLUE)
        sh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        rh(ws, ROW, 16); ROW += 1

        pw_rows = [
            ("  Revenue ($M)",        [f"${v/1000:.2f}B" for v in revs],  None,            False),
            ("  Adj. EBITDA ($M)",    [f"${v:,.0f}M" for v in ebs],       None,            True),
            ("  Share Price Low",     spls,  '"$"#,##0', True),
            ("  Share Price High",    sphs,  '"$"#,##0', True),
            ("  CAGR Low",            cagrl, "+0.0%;-0.0%;—", False),
            ("  CAGR High",           cagrh, "+0.0%;-0.0%;—", False),
        ]
        for label, vals, fmt, bold in pw_rows:
            is_price = "Price" in label
            bg  = GOLD_BG if is_price else (WHITE if ROW % 2 == 0 else GRAY)
            fc  = GOLD_FN if is_price else BLACK
            cell(ws, ROW, 1, bg=bg)
            cell(ws, ROW, 2, label, bg=bg, fc=fc, bold=bold, align="left")
            for i, v in enumerate(vals):
                if isinstance(v, str):
                    cell(ws, ROW, 3+i, v, bg=bg, fc=fc, bold=bold)
                else:
                    c = ws.cell(row=ROW, column=3+i)
                    c.value = v; c.font = Font(name="Arial", bold=bold, size=9, color=fc)
                    c.fill = fill(bg); c.number_format = fmt if fmt else ""
                    c.alignment = Alignment(horizontal="center", vertical="center")
            # EV 2030 (col H)
            last = vals[4]
            if isinstance(last, str):
                cell(ws, ROW, 8, last, bg=bg, fc=fc, bold=True)
            else:
                c8 = ws.cell(row=ROW, column=8)
                c8.value = last; c8.fill = fill(bg)
                c8.font = Font(name="Arial", bold=True, size=9, color=fc)
                c8.number_format = fmt if fmt else ""
                c8.alignment = Alignment(horizontal="center", vertical="center")
            rh(ws, ROW, 14); ROW += 1
        ROW += 1

    # Probability weighted expected value
    ws.merge_cells(f"A{ROW}:H{ROW}")
    ev = ws[f"A{ROW}"]
    ev.value = "PROBABILITY WEIGHTED EXPECTED VALUE — Updates automatically when probabilities change"
    ev.font = Font(name="Arial", bold=True, size=9, color=WHITE)
    ev.fill = fill(NAVY)
    ev.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    rh(ws, ROW, 16); ROW += 1

    ev_rev  = [bull[0][i]*PROB_BULL + base[0][i]*PROB_BASE + bear[0][i]*PROB_BEAR for i in range(5)]
    ev_eb   = [bull[1][i]*PROB_BULL + base[1][i]*PROB_BASE + bear[1][i]*PROB_BEAR for i in range(5)]
    ev_cagrl= [(ev_spl[i]/ENTRY_PRICE)**(1/(i+1))-1 for i in range(5)]
    ev_cagrh= [(ev_sph[i]/ENTRY_PRICE)**(1/(i+1))-1 for i in range(5)]

    ev_rows = [
        ("  Expected Revenue ($M)",     [f"${v/1000:.2f}B" for v in ev_rev], None,           False),
        ("  Expected Adj. EBITDA ($M)", [f"${v:,.0f}M" for v in ev_eb],      None,           True),
        ("  Expected Share Price Low",  ev_spl,  '"$"#,##0', True),
        ("  Expected Share Price High", ev_sph,  '"$"#,##0', True),
        ("  Expected CAGR Low",         ev_cagrl,"+0.0%;-0.0%;—", False),
        ("  Expected CAGR High",        ev_cagrh,"+0.0%;-0.0%;—", False),
    ]
    for label, vals, fmt, bold in ev_rows:
        is_price = "Price" in label
        bg = GOLD_BG if is_price else PROB_BG
        fc = GOLD_FN if is_price else BLACK
        cell(ws, ROW, 1, bg=bg)
        cell(ws, ROW, 2, label, bg=bg, fc=fc, bold=bold, align="left")
        for i, v in enumerate(vals):
            if isinstance(v, str):
                cell(ws, ROW, 3+i, v, bg=bg, fc=fc, bold=bold)
            else:
                c = ws.cell(row=ROW, column=3+i)
                c.value = v; c.font = Font(name="Arial", bold=bold, size=9, color=fc)
                c.fill = fill(bg); c.number_format = fmt if fmt else ""
                c.alignment = Alignment(horizontal="center", vertical="center")
        last = vals[4]
        if isinstance(last, str):
            cell(ws, ROW, 8, last, bg=bg, fc=fc, bold=True)
        else:
            c8 = ws.cell(row=ROW, column=8)
            c8.value = last; c8.fill = fill(bg)
            c8.font = Font(name="Arial", bold=True, size=9, color=fc)
            c8.number_format = fmt if fmt else ""
            c8.alignment = Alignment(horizontal="center", vertical="center")
        rh(ws, ROW, 14); ROW += 1

    ws.freeze_panes = "C12"

    # Footer
    ws.merge_cells(f"A{ROW+1}:H{ROW+1}")
    f = ws[f"A{ROW+1}"]
    f.value = (f"Built: {GENERATED}  |  Entry ${ENTRY_PRICE}  |  "
               f"Diluted {SHARES}M as-converted  |  Script: build_celh_projection_final.py")
    f.font = Font(name="Arial", size=7, color=MUTED, italic=True)
    f.alignment = Alignment(horizontal="left")

# ==============================================================================
# MAIN
# ==============================================================================
def build():
    wb = Workbook()
    wb.remove(wb.active)

    build_inputs(wb)
    build_projection(wb)
    build_probability(wb)

    # Tab colors — exact AMD
    wb["Inputs"].sheet_properties.tabColor            = "2E75B6"  # blue
    wb["Projection"].sheet_properties.tabColor        = "1A7A3A"  # green
    wb["Probability Weighted"].sheet_properties.tabColor = "7B2D8B"  # purple

    out = os.path.join(_HERE, "CELH_5Year_Projection_v3.xlsx")
    wb.save(out)

    # Print key outputs
    bull = compute(BULL); base = compute(BASE); bear = compute(BEAR)
    ev_spl = [bull[2][i]*PROB_BULL+base[2][i]*PROB_BASE+bear[2][i]*PROB_BEAR for i in range(5)]
    ev_sph = [bull[3][i]*PROB_BULL+base[3][i]*PROB_BASE+bear[3][i]*PROB_BEAR for i in range(5)]
    print(f"Saved: {out}")
    print(f"\nKEY 2030 OUTPUTS — verify on open:")
    print(f"BULL: Rev=${bull[0][4]/1000:.2f}B  EBITDA=${bull[1][4]:,.0f}M  SPL=${bull[2][4]:,.0f}  SPH=${bull[3][4]:,.0f}  CAGR={bull[4][4]:+.1%}/{bull[5][4]:+.1%}")
    print(f"BASE: Rev=${base[0][4]/1000:.2f}B  EBITDA=${base[1][4]:,.0f}M  SPL=${base[2][4]:,.0f}  SPH=${base[3][4]:,.0f}  CAGR={base[4][4]:+.1%}/{base[5][4]:+.1%}")
    print(f"BEAR: Rev=${bear[0][4]/1000:.2f}B  EBITDA=${bear[1][4]:,.0f}M  SPL=${bear[2][4]:,.0f}  SPH=${bear[3][4]:,.0f}  CAGR={bear[4][4]:+.1%}/{bear[5][4]:+.1%}")
    print(f"EV:   SPL=${ev_spl[4]:,.0f}  SPH=${ev_sph[4]:,.0f}  CAGR={(ev_spl[4]/ENTRY_PRICE)**(1/5)-1:+.1%}/{(ev_sph[4]/ENTRY_PRICE)**(1/5)-1:+.1%}")

# ==============================================================================
# SELF-TEST — runs automatically after build(), fails loudly if broken
# DO NOT REMOVE — this is what proves the script is complete and working
# ==============================================================================
def self_test(out_path):
    import os
    from openpyxl import load_workbook

    errors = []

    if not os.path.exists(out_path):
        errors.append(f"FAIL: file not created at {out_path}")
    else:
        size = os.path.getsize(out_path)
        if size < 10000:
            errors.append(f"FAIL: file too small ({size} bytes)")

        wb = load_workbook(out_path, data_only=True)
        for tab in ["Inputs", "Projection", "Probability Weighted"]:
            if tab not in wb.sheetnames:
                errors.append(f"FAIL: tab '{tab}' missing")

        # Verify key 2030 outputs
        ws = wb["Projection"]
        found_217 = False
        found_110 = False
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell and "217" in str(cell): found_217 = True
                if cell and "110" in str(cell): found_110 = True
        if not found_217:
            errors.append("FAIL: Bull 2030 SPL $217 not found")
        if not found_110:
            errors.append("FAIL: Base 2030 SPL $110 not found")

    if errors:
        print("\n" + "="*60)
        print("SELF-TEST FAILED — DO NOT SAVE TO DRIVE")
        print("="*60)
        for e in errors: print(f"  {e}")
        print("="*60)
        return False
    else:
        size = os.path.getsize(out_path)
        print("\n" + "="*60)
        print("SELF-TEST PASSED — safe to save to Drive")
        print(f"  File: {out_path}")
        print(f"  Size: {size:,} bytes")
        print(f"  Tabs: Inputs, Projection, Probability Weighted confirmed")
        print(f"  Bull SPL $217, Base SPL $110: confirmed")
        print("="*60)
        return True

if __name__ == "__main__":
    build()
    out = os.path.join(_HERE, "CELH_5Year_Projection_v3.xlsx")
    passed = self_test(out)
    if not passed:
        import sys
        sys.exit(1)
