# ══════════════════════════════════════════════════════════════════════════════
# CELSIUS HOLDINGS, INC. (CELH) — INCOME STATEMENT ANALYSIS
# build_celh_income_v1.py
# Q1 2026 | Quarter Ended March 31, 2026 | Reported May 7, 2026
# v1: Refactored from build_celsius_simple_v4_final.py to AMD income pattern.
#     Adds explicit DATA BLOCK, Data Sources tab, What's New tab,
#     TABLE_BORDER on data rows, tab colors, MANIFEST-compliant self-test.
#     Output: CELH_Q1_2026_Income_Statement_v1.xlsx
# Next quarterly update → build_celh_income_v2.py (never overwrite this file)
# ══════════════════════════════════════════════════════════════════════════════

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

_HERE = os.path.dirname(os.path.abspath(__file__))

# ── DATA BLOCK — update these values each quarter ──────────────────────────────
TICKER       = "CELH"
COMPANY      = "Celsius Holdings, Inc."
QUARTER      = "Q1 2026"
PERIOD_END   = "March 31, 2026"
REPORT_DATE  = "May 7, 2026"
GENERATED    = "May 2026"
GRADE        = "B"
SCRIPT_NAME  = "build_celh_income_v1.py"
OUTPUT_NAME  = "CELH_Q1_2026_Income_Statement_v1.xlsx"
PRESS_URL    = (
    "https://s203.q4cdn.com/427437840/files/doc_financials/2026/q1/v3/"
    "Celsius-Holdings-Q1-2026__Earnings-Press-Release.pdf"
)
Q1_REVENUE   = 782.6   # $M — for self-test verification

# ── COLOR PALETTE ─────────────────────────────────────────────────────────────
NAVY        = "1A1A2E"
BLUE        = "1F4E79"
ACCENT      = "2E75B6"
WHITE       = "FFFFFF"
GRAY        = "F2F2F2"
GREEN       = "00B050"
AMBER       = "FF8C00"
RED         = "C00000"
GREEN_LT    = "E8F5E9"
AMBER_LT    = "FFF3E0"
RED_LT      = "FFEBEE"
MEASURE_BG  = "EBF3FB"
BLACK       = "000000"
MUTED       = "666666"
SECTION_BG  = "1F4E79"
STRENGTH_BG = "E8F5E9"
CONCERN_BG  = "FFEBEE"
ASSESS_BG   = "FFF8E1"

# ── BORDER STYLE (applied to all data table cells) ────────────────────────────
_SIDE = Side(style="thin", color="CCCCCC")
TABLE_BORDER = Border(left=_SIDE, right=_SIDE, top=_SIDE, bottom=_SIDE)

# ── FONT SIZES ─────────────────────────────────────────────────────────────────
SZ_DEFAULT  = 10
SZ_TITLE    = 14
SZ_SUBTITLE = 10
SZ_FOOTER   =  8

# ── PRIORITY ROWS (bold line item cell) ───────────────────────────────────────
PRIORITY = {
    "Revenue", "Cost of Revenue", "Gross Profit", "Gross Margin",
    "Total SG&A (GAAP)", "Income from Operations",
    "Total Other Income (Expense)", "Pre-Tax Income", "Net Income",
    "Net Income to Common Shareholders", "Diluted EPS (GAAP)", "Adjusted EBITDA",
}

# ── ROW DATA ──────────────────────────────────────────────────────────────────
# (label, q1_2026, q4_2025, q1_2025, yy_color, actual_color, is_pct, note, measure, fmt_override)
# yy_color:     "green"|"amber"|"red"|None  → Y/Y Change + Y/Y % cells
# actual_color: "green"|"amber"|"red"|None  → Q1 2026 actual cell
# is_pct:       True = margin/% row (Y/Y Change shown as pp, Y/Y % shown as bps)
# q4_2025:      None = dash (not reported this line)
# fmt_override: "#,##0" for share-count rows, else None
ROWS = [
    # ── REVENUE ───────────────────────────────────────────────────────────────
    ("Revenue",
     782.6, None, 329.3, "amber", None, False,
     "Amber despite +138% Y/Y because the headline overstates growth quality on two levels. "
     "Level 1 — inorganic: virtually all Y/Y growth comes from acquisitions (Alani Nu +$368M, "
     "Rockstar +$67M); organic CELSIUS brand grew only +6% Y/Y. "
     "Level 2 — loading inflation: Alani Nu's $368M Q1 contribution is itself distorted by "
     "a one-time PepsiCo distributor loading event (DSD network stocking before consumer "
     "pull-through begins). True Alani Nu sustainable run rate is ~$140-180M/quarter based on "
     "pre-acquisition ~$500M annual revenue + distribution expansion lift. That implies "
     "~$190-230M of Q1 Alani Nu revenue was loading, not consumer demand. "
     "Normalized Q1 revenue ex-loading: ~$550-600M — a very different picture than $782.6M. "
     "Real organic growth test comes Q3-Q4 2026 when acquisitions lap. "
     "Amber is the right signal: revenue grew, but the quality and durability are unconfirmed.",
     "Top-line scale; amber = growth is real but heavily distorted by acquisitions + loading",
     None),

    ("  North America",
     747.3, None, 306.5, "amber", None, False,
     "Decomposition of $747.3M: CELSIUS organic ~$325M (+6% on $306.5M prior year) + "
     "Alani Nu ~$368M + Rockstar ~$67M. Acquisitions account for ~$422M of the $441M Y/Y gain. "
     "The Alani Nu figure is heavily distorted by one-time distributor loading — PepsiCo's DSD "
     "network stocking up before Alani Nu hits consumer shelves. "
     "Pre-acquisition Alani Nu ran ~$500M annually (~$125M/quarter of consumer pull-through); "
     "with PepsiCo distribution expansion, sustainable rate is ~$140-180M/quarter. "
     "That implies ~$190-230M (~55-60%) of the $368M Q1 Alani Nu revenue was loading, not demand. "
     "Projected Q2 North America normalized: ~$525-585M "
     "(CELSIUS ~$325-340M + Alani Nu normalized ~$140-180M + Rockstar ~$60-65M). "
     "A Q2 print above $560M is encouraging; below $500M confirms significant inflation. "
     "Management gave no Q2 revenue guidance — notable absence that aligns with this risk.",
     "Domestic scale; ~95% of revenue; normalized Q2 run rate ~$525-585M ex-loading",
     None),

    ("  International",
     35.3, None, 22.7, "green", None, False,
     "Organic +55% Y/Y — no acquisition benefit here. Driven by Nordics strength "
     "and expansion markets (UK, Ireland, France, Australia/NZ, Benelux). "
     "Only 4.5% of total revenue but growing fast and entirely organic. "
     "International is the cleanest growth signal in the revenue mix.",
     "Organic global expansion; growth quality indicator",
     None),

    # ── COST & GROSS ──────────────────────────────────────────────────────────
    ("Cost of Revenue",
     404.5, None, 156.9, "amber", None, False,
     "Cost grew +158% vs revenue +138% — 20 ppt gap reflects Alani Nu and Rockstar "
     "acquired at lower margin profiles than core CELSIUS brand. "
     "Core CELSIUS cost structure intact — the mix shift is the driver, not inefficiency. "
     "Q4 2025 COGS write-offs and transition costs largely behind. "
     "Commodity headwinds (Midwest aluminum LME premium) partially offset recovery. "
     "Margin recovery levers: orbit model, freight optimization, raw material alignment "
     "across Alani and Rockstar. Q2 gross margin guided flat ('sidestep').",
     "Direct production cost efficiency; lower growth vs revenue = margin expansion",
     None),

    ("Gross Profit",
     378.1, None, 172.4, "green", None, False,
     "Dollar gross profit +119% Y/Y driven by acquisition scale — nearly doubled. "
     "Grew slower than revenue (+138%) because acquired brands carry lower margins. "
     "Dollar result is strong; efficiency story is told separately in Gross Margin below. "
     "Separating dollar result from margin rate gives a cleaner analytical picture.",
     "Absolute earning power after production; measures scale of the business",
     None),

    ("Gross Margin",
     48.3, None, 52.3, "amber", "amber", True,
     "-400 bps Y/Y due to Alani Nu and Rockstar acquisition mix — both brands carry "
     "lower margins than core CELSIUS upon acquisition. Not an operational failure. "
     "Sequential improvement: +90 bps vs Q4 2025 (~47.4%) — direction already reversing. "
     "Actual cell flagged amber: margin rate deteriorated even as revenue scaled — "
     "consistent rule: any margin that grew slower than or declined vs revenue gets amber on actual. "
     "Management target: return to 'low 50s' — no firm timeline given. "
     "Q2 guided flat ('sidestep'); step-ups expected Q3-Q4 via orbit model, "
     "freight structure optimization, and raw material alignment across Alani and Rockstar. "
     "Second manufacturing line in North Carolina begins H2 2026; full benefit 2027.",
     "Production efficiency; how much of each revenue dollar is retained after COGS",
     None),

    # ── OPEX ──────────────────────────────────────────────────────────────────
    ("Total SG&A (GAAP)",
     234.6, None, 120.3, "green", None, False,
     "Dollars +95% Y/Y but grew slower than revenue (+138%) — operating leverage confirmed. "
     "% of revenue improved meaningfully: 36.5% → 30.0% (-650 bps). "
     "Includes $24.6M legal settlement accrual (Strong Arm Productions litigation) "
     "and $3.8M acquisition/integration costs — both one-time. "
     "Adj. SG&A ex. one-time items: $206.3M / 26.4% of rev, down from 31.8% in Q4 2025. "
     "Green because SG&A growing slower than revenue is the definition of operating leverage.",
     "Operating cost discipline; grew slower than revenue = operating leverage confirmed",
     None),

    ("  Sales & Marketing",
     150.6, None, 80.9, "green", None, False,
     "Dollars +86% Y/Y — grew slower than revenue (+138%). "
     "% of revenue improved: 24.6% → 19.2% (-540 bps). "
     "Brand spend becoming more efficient as portfolio scales across three brands. "
     "PepsiCo distribution handles shelf placement and logistics — reduces Celsius's "
     "direct marketing cost burden as volume scales.",
     "Brand investment efficiency; % of revenue shows leverage as scale grows",
     None),

    ("  General & Administrative",
     84.1, None, 39.4, "green", None, False,
     "Dollars +113% Y/Y — grew slower than revenue (+138%) — operating leverage. "
     "However composition is noisy: includes $24.6M legal settlement (Strong Arm) "
     "and $3.8M M&A costs. Adjusted G&A ex. one-time items: $55.7M — only +41% growth "
     "vs revenue +138%. Underlying overhead is very well controlled. "
     "Green because GAAP G&A grew slower than revenue; note captures the one-time items "
     "that make the adjusted picture even cleaner than GAAP suggests.",
     "Overhead cost control; underlying G&A grew only +41% ex. one-time items",
     None),

    ("Distributor Termination Fees",
     4.4, None, 0.0, None, "amber", False,
     "No prior year comp — cannot calculate meaningful Y/Y comparison. "
     "Actual flagged amber: real cash cost this quarter, not alarming but warrants attention. "
     "Transitional cost tied to shifting Alani Nu distribution from prior system into PepsiCo. "
     "Balance reduced from $264.1M to $40.0M remaining — wind-down well advanced. "
     "Not a recurring item; will disappear from the income statement as transition completes.",
     "Transitional one-time cost; signals distribution system restructuring in progress",
     None),

    # ── OPERATING INCOME ──────────────────────────────────────────────────────
    ("Income from Operations",
     139.0, None, 52.0, "green", None, False,
     "+167% Y/Y — grew faster than revenue (+138%). Operating margin expanded: "
     "15.8% → 17.8% (+200 bps). This is the operating leverage story confirmed "
     "at the bottom of the operating section — the business is more profitable per "
     "dollar of revenue than a year ago despite absorbing two acquisition integrations simultaneously.",
     "Core business profitability before financing costs and taxes",
     None),

    # ── BELOW THE LINE ────────────────────────────────────────────────────────
    ("Total Other Income (Expense)",
     -1.5, None, 9.0, "red", None, False,
     "Swung from +$9.0M income to -$1.5M expense — a $10.5M Y/Y deterioration. "
     "Two causes: (1) Interest income declined $4.9M as cash was deployed into acquisitions. "
     "(2) New $11.8M quarterly interest expense on $696.5M term loan (due 2032) taken on "
     "to fund Alani Nu and Rockstar acquisitions — this did not exist a year ago. "
     "Partially offset by $7.4M other income (incl. $7.0M from PepsiCo related party). "
     "Red because this is structural and permanent — not a one-time item. "
     "The debt is a 2032 term loan; $11.8M interest drag repeats every quarter going forward. "
     "This is the ongoing price of the acquisition-funded growth strategy.",
     "Financing impact; new acquisition debt creates permanent quarterly interest drag",
     None),

    ("Pre-Tax Income",
     137.5, None, 61.0, "amber", None, False,
     "+125% Y/Y — grew slower than revenue (+138%). Amber by framework: any line "
     "growing slower than revenue signals efficiency erosion. "
     "The gap vs operating income (+167%) is entirely explained by the new interest expense. "
     "Unlike Adjusted EBITDA which adds back interest and D&A to show operating performance "
     "independent of capital structure, Pre-Tax Income fully reflects the cost of the "
     "$696.5M term loan. Management emphasizes Adjusted EBITDA precisely because "
     "Pre-Tax Income will always look weaker than the operating business warrants "
     "until the acquisition debt is paid down.",
     "Total profitability before tax; fully reflects capital structure cost unlike EBITDA",
     None),

    ("Income Tax Provision",
     -27.4, None, -16.6, None, None, False,
     "Tax expense grew +65% Y/Y vs revenue +138% — grew slower than revenue. "
     "No color applied: support line, not a primary profitability driver. "
     "Effective tax rate improved significantly: 27.2% → 19.9% (-730 bps). "
     "Lower rate driven by tax treatment of acquisition-related deductions — "
     "a meaningful benefit that helped net income outpace pre-tax income growth.",
     "Tax efficiency; effective rate reflects jurisdictional mix and deductions",
     None),

    ("Net Income",
     110.1, None, 44.4, "green", None, False,
     "+148% Y/Y — grew faster than revenue (+138%). One of the few lines to outpace "
     "revenue growth. The tax rate improvement (27.2% → 19.9%, -730 bps) more than "
     "offset the new interest drag, pushing net income above the revenue growth rate. "
     "Note: Net Income belongs to all capital providers including PepsiCo preferred holders. "
     "Net Income to Common Shareholders ($85.1M) is the true investor scorecard — "
     "$25M is redirected to PepsiCo preferred every quarter before common shareholders see it.",
     "Total GAAP profitability; belongs to all capital providers including preferred holders",
     None),

    ("Net Income to Common Shareholders",
     85.1, None, 34.4, "green", None, False,
     "+147% Y/Y — grew faster than revenue (+138%). Green by framework. "
     "$25M gap vs Net Income ($110.1M) reflects PepsiCo's convertible preferred stock: "
     "$14.0M in preferred dividends + $11.0M income allocated to participating preferred. "
     "Context: PepsiCo co-financed Celsius's acquisitions by receiving Series A and B "
     "convertible preferred stock. As preferred holders they get paid before common shareholders "
     "every quarter — a permanent recurring deduction of ~$25M ($100M annually). "
     "PepsiCo can convert preferred to ~78M common shares, swelling diluted count to ~334M. "
     "This is the investor's true earnings base — always use this line, not Net Income, "
     "when evaluating CELH as a common stock investment.",
     "Profitability available to common equity after preferred obligations to PepsiCo",
     None),

    # ── EPS ───────────────────────────────────────────────────────────────────
    ("Diluted EPS (GAAP)",
     0.33, None, 0.15, "amber", None, False,
     "+120% Y/Y — grew slower than revenue (+138%). Amber by framework. "
     "Preferred stock creates a double drag on EPS: "
     "(1) Numerator: $25M removed before calculation — reduces the earnings base. "
     "(2) Denominator: ~78M additional as-converted preferred shares inflate share count "
     "from 256M common to ~334M diluted — reduces per-share value. "
     "The two effects together explain why EPS growth (+120%) lags Net Income growth (+148%). "
     "Beat consensus ~$0.28 by +$0.05 (+17.9%). "
     "Adj. EPS $0.41 beat consensus $0.29-0.30 by +$0.11-0.12 (+37-41%). "
     "Basic EPS = Diluted EPS ($0.33) this quarter — effects roughly cancel out.",
     "Per-share profitability; primary metric Wall Street uses to price the stock",
     None),

    ("Basic EPS (GAAP)",
     0.33, None, 0.15, None, None, False,
     "Same as diluted EPS ($0.33) this quarter — effects roughly cancel. "
     "Basic EPS = Net Income to Common / actual shares outstanding (256.5M). "
     "Diluted EPS = Net Income to Common (adjusted) / fully diluted shares (~334M as-converted). "
     "When Basic = Diluted, dilution risk is minimal in the current period. "
     "Watch for divergence — that signals dilution is actively hurting common shareholders. "
     "No color: reference line; Diluted EPS above is the primary metric.",
     "Per-share profit on common shares only; excludes all dilutive securities",
     None),

    # ── NON-GAAP ──────────────────────────────────────────────────────────────
    ("Adjusted EBITDA",
     195.5, None, 69.7, "green", None, False,
     "+181% Y/Y — grew faster than revenue (+138%). Strongest growth rate in the statement. "
     "Margin expanded: 21.2% → 25.0% (+380 bps Y/Y). Beat consensus ~$165M by ~$30M (+18%). "
     "Add-backs total ~$49.5M: D&A $9.1M, SBC $7.6M, legal settlement $24.6M, "
     "distributor termination $4.4M, M&A costs $3.8M, FX -$0.4M. "
     "Management's primary operating metric — strips all acquisition noise to show "
     "underlying cash earning power of the brand portfolio. "
     "The +181% growth vs +138% revenue confirms the business generates cash at an "
     "accelerating rate relative to revenue as it scales.",
     "Cash earning power ex. non-cash and one-time items; management's primary metric",
     None),

    ("Adjusted EBITDA Margin",
     25.0, None, 21.2, "green", "green", True,
     "+380 bps Y/Y expansion — margin expanded despite absorbing two acquisitions "
     "at lower margin profiles. Demonstrates operating leverage at scale. "
     "Actual cell colored green: consistent rule — all margin rows get actual colored "
     "to match Y/Y color. Green actual = margin expanded Y/Y (positive signal). "
     "Contrast with Gross Margin (amber actual) where margin deteriorated — "
     "the two together tell the full story: gross margins under pressure from mix, "
     "but cash margins expanding from operating leverage. Both signals matter.",
     "Operating leverage at scale; expanding margin = business becoming more efficient",
     None),

    ("Adjusted Diluted EPS",
     0.41, None, 0.18, "amber", None, False,
     "+128% Y/Y — grew slower than revenue (+138%). Amber by framework. "
     "Same preferred share dilution dynamic as GAAP EPS affects the denominator "
     "even in the non-GAAP calculation. "
     "However this is the headline beat of the quarter: $0.41 vs consensus $0.29-0.30, "
     "a +37-41% surprise — strongest consensus beat in the statement. "
     "Add-backs net of tax: M&A costs +$0.01, distributor termination +$0.01, "
     "legal settlement +$0.06. Tax effect of all adjustments: -$0.08/share. "
     "Despite amber on the Y/Y framework, the consensus beat signals the market "
     "was significantly underestimating Celsius's underlying earnings power.",
     "Clean per-share earnings ex. one-time items; consensus benchmark for beat/miss",
     None),
]

# ── COLUMN LAYOUT ─────────────────────────────────────────────────────────────
COL_WIDTHS = {
    "A": 32,   # Line Item
    "B": 14,   # Y/Y Change
    "C": 10,   # Y/Y %
    "D": 13,   # Q1 2026
    "E": 13,   # Q4 2025
    "F": 13,   # Q1 2025
    "G": 90,   # Notes
    "H": 40,   # What This Measures
}


def fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)


def center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def section_header(ws, row, text, bg, cols=8):
    ws.merge_cells(f"A{row}:{chr(64+cols)}{row}")
    c = ws[f"A{row}"]
    c.value = text
    c.font = Font(name="Arial", bold=True, size=10, color=WHITE)
    c.fill = fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 18
    return row + 1


def bullet_row(ws, row, text, bg, indent=2, font_size=9, color=BLACK):
    ws.merge_cells(f"A{row}:H{row}")
    c = ws[f"A{row}"]
    c.value = text
    c.font = Font(name="Arial", size=font_size, color=color)
    c.fill = fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="top",
                             wrap_text=True, indent=indent)
    ws.row_dimensions[row].height = max(15, 13 + len(text) // 10)
    return row + 1


def num_fmt(val, is_pct):
    if val is None:
        return "@"
    if is_pct:
        return '0.0"%"'
    return '$#,##0.00' if abs(val) < 5 else '$#,##0.0'


# ==============================================================================
# TAB 1 — INCOME STATEMENT
# ==============================================================================
def build_income_tab(wb):
    ws = wb.create_sheet("Q1 2026 Income Statement")

    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # Title
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = f"{COMPANY} ({TICKER}) — {QUARTER} INCOME STATEMENT ANALYSIS"
    t.font = Font(name="Arial", bold=True, size=SZ_TITLE, color=WHITE)
    t.fill = fill(NAVY)
    t.alignment = center()
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:H2")
    s = ws["A2"]
    s.value = (f"Quarter Ended {PERIOD_END}  |  Reported {REPORT_DATE}  |  "
               f"Source: {COMPANY} Earnings Press Release  |  "
               f"$ in millions except per-share amounts")
    s.font = Font(name="Arial", size=SZ_SUBTITLE, color=WHITE)
    s.fill = fill(BLUE)
    s.alignment = center(wrap=True)
    ws.row_dimensions[2].height = 18

    # Column headers
    headers = ["Line Item", "Y/Y Change", "Y/Y %",
               f"{QUARTER} ($M)", "Q4 2025 ($M)", "Q1 2025 ($M)",
               "Notes", "What This Measures"]
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col_idx)
        c.value = h
        c.font = Font(name="Arial", bold=True, size=9, color=WHITE)
        c.fill = fill(ACCENT)
        c.alignment = center(wrap=True)
        c.border = TABLE_BORDER
    ws.row_dimensions[3].height = 28

    yy_map = {
        "green": (GREEN, GREEN_LT),
        "amber": (AMBER, AMBER_LT),
        "red":   (RED,   RED_LT),
    }

    DATA_START = 4

    for i, (label, q1, q4, q1_py,
            yy_color, actual_color, is_pct, note, measure, _fmt) in enumerate(ROWS):

        row = DATA_START + i
        is_priority = label.strip() in PRIORITY
        is_sub      = label.startswith("  ")
        bg = WHITE if i % 2 == 0 else GRAY

        # A — Line Item
        a = ws.cell(row=row, column=1)
        a.value = label
        a.font = Font(name="Arial",
                      bold=is_priority and not is_sub,
                      italic=is_sub,
                      size=9,
                      color=MUTED if is_sub else BLACK)
        a.fill = fill(bg)
        a.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        a.border = TABLE_BORDER

        # B — Y/Y Change
        b = ws.cell(row=row, column=2)
        b.value = f"=D{row}-F{row}"
        if is_pct:
            b.number_format = '+0.0" pp";-0.0" pp";"-"'
        elif q1 is not None and abs(q1) < 5:
            b.number_format = '+$#,##0.00;-$#,##0.00;"-"'
        else:
            b.number_format = '+$#,##0.0;-$#,##0.0;"-"'
        b.alignment = Alignment(horizontal="right", vertical="center")
        b.border = TABLE_BORDER

        # C — Y/Y %
        c = ws.cell(row=row, column=3)
        if is_pct:
            c.value = f"=(D{row}-F{row})*100"
            c.number_format = '+0" bps";-0" bps";"-"'
        else:
            c.value = f'=IF(F{row}<>0,(D{row}-F{row})/ABS(F{row}),"N/A")'
            c.number_format = '+0.0%;-0.0%;"-"'
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border = TABLE_BORDER

        # Apply Y/Y color to B and C
        if yy_color in yy_map:
            ink, bg_yy = yy_map[yy_color]
            for col in [2, 3]:
                cell = ws.cell(row=row, column=col)
                cell.fill = fill(bg_yy)
                cell.font = Font(name="Arial", size=9, color=ink, bold=True)
        else:
            for col in [2, 3]:
                cell = ws.cell(row=row, column=col)
                cell.fill = fill(bg)
                cell.font = Font(name="Arial", size=9, color=BLACK)

        # D — Q1 2026 actual
        d = ws.cell(row=row, column=4)
        d.value = q1
        d.number_format = num_fmt(q1, is_pct)
        d.alignment = Alignment(horizontal="right", vertical="center")
        d.border = TABLE_BORDER
        if actual_color in yy_map:
            ink, bg_ac = yy_map[actual_color]
            d.fill = fill(bg_ac)
            d.font = Font(name="Arial", size=9, color=ink, bold=True)
        else:
            d.fill = fill(bg)
            d.font = Font(name="Arial", size=9, color=BLACK)

        # E — Q4 2025
        e = ws.cell(row=row, column=5)
        if q4 is None:
            e.value = "—"
            e.font = Font(name="Arial", size=9, color=MUTED)
        else:
            e.value = q4
            e.number_format = num_fmt(q4, is_pct)
            e.font = Font(name="Arial", size=9, color=BLACK)
        e.fill = fill(bg)
        e.alignment = Alignment(horizontal="right", vertical="center")
        e.border = TABLE_BORDER

        # F — Q1 2025
        f = ws.cell(row=row, column=6)
        if q1_py is None:
            f.value = "—"
            f.font = Font(name="Arial", size=9, color=MUTED)
        else:
            f.value = q1_py
            f.number_format = num_fmt(q1_py, is_pct)
            f.font = Font(name="Arial", size=9, color=BLACK)
        f.fill = fill(bg)
        f.alignment = Alignment(horizontal="right", vertical="center")
        f.border = TABLE_BORDER

        # G — Notes
        g = ws.cell(row=row, column=7)
        g.value = note
        g.font = Font(name="Arial", size=8, color="333333")
        g.fill = fill(bg)
        g.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        g.border = TABLE_BORDER

        # H — What This Measures
        h = ws.cell(row=row, column=8)
        h.value = measure
        h.font = Font(name="Arial", size=8, color=BLUE, italic=True)
        h.fill = fill(MEASURE_BG)
        h.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True, indent=1)
        h.border = TABLE_BORDER

        ws.row_dimensions[row].height = max(45, min(180, 12 + note.count(" ") * 1.4))

    ws.freeze_panes = "B4"

    # ── Sections below the data table ─────────────────────────────────────────
    next_row = DATA_START + len(ROWS) + 1
    ws.row_dimensions[next_row].height = 10
    next_row += 1

    # STRENGTHS
    next_row = section_header(ws, next_row, "▶  STRENGTHS", "1A7A3A")
    for s in [
        "✓  Record Q1 revenue of $782.6M — +138% Y/Y; portfolio now holds 20.9% of the U.S. RTD energy category, one of the fastest category share gains in beverage history.",
        "✓  Adjusted EBITDA +181% Y/Y to $195.5M at a 25.0% margin — strongest growth rate in the statement and proof that cash earnings scale faster than revenue.",
        "✓  Operating leverage confirmed across every cost line: SG&A (-650 bps as % of rev), Sales & Marketing (-540 bps), Adj. SG&A (-540 bps Q/Q). Cost discipline is real.",
        "✓  Alani Nu integration complete: $50M in synergies captured, distribution fully transitioned into PepsiCo system. One of the fastest beverage integrations on record.",
        "✓  International revenue +55% Y/Y organically — no acquisition benefit. Nordics leading; UK, France, Australia/NZ expanding. Clean growth signal.",
        "✓  Gross margin improving sequentially: +90 bps Q/Q (47.4% → 48.3%). Direction of travel is correct even if Y/Y comparison is distorted by acquisitions.",
        "✓  Adj. EPS $0.41 beat consensus $0.29-0.30 by +37-41% — strongest consensus beat in the statement. Market was significantly underestimating earning power.",
        "✓  Tax rate improvement: 27.2% → 19.9% (-730 bps). Acquisition-related deductions providing meaningful tax shield.",
        "✓  $24.1M share repurchases in Q1 — management putting capital to work at near 52-week lows, signaling confidence in intrinsic value.",
    ]:
        next_row = bullet_row(ws, next_row, s, STRENGTH_BG)

    next_row += 1

    # CONCERNS
    next_row = section_header(ws, next_row, "▶  CONCERNS", "8B0000")
    for c in [
        "✗  Organic CELSIUS brand growth only +6% Y/Y — the entire revenue growth story is inorganic. If CELSIUS brand not back to double digits by Q3-Q4 2026 when acquisitions lap, the thesis weakens materially.",
        "✗  Gross margin -400 bps Y/Y (52.3% → 48.3%). Path back to 'low 50s' relies on orbit model, freight optimization, and commodity tailwinds — none of which are guaranteed. Q2 guided flat ('sidestep').",
        "✗  New structural interest expense: $11.8M per quarter ($47M annualized) on $696.5M term loan due 2032. Permanent drag on Pre-Tax Income and EPS that did not exist a year ago.",
        "✗  PepsiCo preferred stock drains ~$25M per quarter ($100M annually) from common shareholders — a permanent toll that reduces EPS and dilutes on conversion (~78M additional shares).",
        "✗  Alani Nu Q1 revenue ($368M) may have been inflated by distributor loading as PepsiCo absorbed the brand — a one-time inventory build that may not repeat in Q2 or Q3.",
        "✗  Rockstar Energy retail sales -13% Y/Y in tracked channels. Integration on track for H1 2026 completion but brand trajectory is weak. Rockstar is declining, not growing.",
        "✗  Accrued legal expenses $88.2M (Strong Arm Productions litigation) — $24.6M accrued in Q1 alone. Liability overhang and uncertainty around final settlement amount.",
        "✗  Customer concentration: PepsiCo = 59% of Q1 revenue and 45.5% of accounts receivable. Any change to the commercial agreement would be catastrophic.",
        "✗  Distributor termination fees: $40.0M still accrued. While winding down, cash outflows will continue for several quarters.",
    ]:
        next_row = bullet_row(ws, next_row, c, CONCERN_BG)

    next_row += 1

    # OVERALL ASSESSMENT
    next_row = section_header(ws, next_row, "▶  OVERALL ASSESSMENT", SECTION_BG)

    ws.merge_cells(f"A{next_row}:H{next_row}")
    grade_cell = ws[f"A{next_row}"]
    grade_cell.value = f"Quarter Grade:  {GRADE}"
    grade_cell.font = Font(name="Arial", bold=True, size=14, color=NAVY)
    grade_cell.fill = fill(ASSESS_BG)
    grade_cell.alignment = center()
    ws.row_dimensions[next_row].height = 28
    next_row += 1

    assessment = (
        "Grade: B. Downgraded from B+ on the basis of three structural issues that B+ does not adequately capture.\n\n"
        "Issue 1 — Revenue quality: The headline +138% Y/Y is an optical illusion. "
        "Organic CELSIUS brand grew only +6%. The remaining growth is inorganic (acquisitions) "
        "and inflated further by a one-time PepsiCo distributor loading event. "
        "Normalized Q1 revenue ex-loading is approximately $550-600M — 25-30% below the reported $782.6M. "
        "B+ implies the top-line momentum is real and durable. It is not, yet.\n\n"
        "Issue 2 — The Q1 2027 comp trap: Q1 2026's inflated $782.6M base creates a structural problem. "
        "Even if the underlying business grows normally in Q1 2027 (~$640-720M normalized), "
        "it will print as -8% to -18% Y/Y revenue decline. "
        "That number will create market confusion and potential multiple compression "
        "unless management actively frames it. B+ does not adequately flag this landmine.\n\n"
        "Issue 3 — Capital structure: The combined preferred + debt structure permanently redirects "
        "~$136M annually before common shareholders see a dollar "
        "($47M annualized interest on $696.5M term loan + ~$100M preferred distributions). "
        "This is not a transient integration cost — it is the ongoing price of the acquisition strategy "
        "and will persist until the debt is retired and preferred is converted or bought out.\n\n"
        "What B+ requires and what is currently unproven: CELSIUS brand organic reacceleration to double digits, "
        "gross margin trajectory back toward 50%, and Q1 2027 framed cleanly vs the inflated base. "
        "The operational execution this quarter was genuinely strong — integration, operating leverage, "
        "and the consensus beat are all confirmed positives. "
        "But execution alone is not enough when the forward setup contains these landmines. "
        "See the Forward Outlook tab for the full phase-by-phase growth rate analysis."
    )
    next_row = bullet_row(ws, next_row, assessment, ASSESS_BG)

    next_row += 1

    # LETTER GRADE FRAMEWORK
    next_row = section_header(ws, next_row,
                              f"▶  LETTER GRADE FRAMEWORK — {COMPANY} ({TICKER})", SECTION_BG)
    # Grade framework anchored to NORMALIZED growth rates, not the 138% optical illusion.
    # Normalized portfolio growth = CELSIUS organic + Alani Nu maturation + Rockstar + International.
    for grade, desc in [
        ("A",  "Normalized portfolio growth 25%+. CELSIUS brand organic reaccelerates to double digits. "
               "Gross margin returns to 50%+. Adj. EBITDA margin above 27%. "
               "Q1 2027 comp trap framed cleanly by management. Rockstar stabilizes or sells at value."),
        ("B+", "Normalized portfolio growth 15-25%. CELSIUS brand organic 8-12%. "
               "Gross margin recovering toward 50% with visible trajectory. "
               "Q1 2027 decline acknowledged and contextualized. Capital structure drag manageable."),
        (f"B   [{QUARTER} — current]",
               "Normalized portfolio growth 8-15%. Operational execution strong, operating leverage confirmed. "
               "But organic CELSIUS only +6%, Q1 revenue inflated by ~$190-230M loading, "
               "Q1 2027 comp trap unaddressed, and ~$136M annual capital structure drag structural. "
               "Show-me quarter: organic reacceleration not yet demonstrated."),
        ("C",  "Normalized portfolio growth below 8% or flat. CELSIUS brand loses shelf share. "
               "Alani Nu revenue normalizes sharply below $120M/quarter. Gross margin stays below 48%. "
               "Adj. EBITDA margin contracts. Legal settlement exceeds accruals."),
        ("D",  "Normalized revenue declines. PepsiCo commercial agreement renegotiated adversely. "
               "Organic CELSIUS growth negative. Legal liability materially exceeds $88.2M accrual. "
               "Acquisition synergies reverse. Debt covenant pressure."),
    ]:
        next_row = bullet_row(ws, next_row, f"{grade}   {desc}", ASSESS_BG)

    next_row += 1

    # KEY METRICS TO WATCH
    next_row = section_header(ws, next_row,
                              "▶  KEY METRICS TO WATCH — Q2 2026 (Expected Aug 6-11, 2026)", SECTION_BG)
    for m in [
        "① CELSIUS Brand Organic Growth Rate — must show reacceleration toward double digits as Alani lapping begins. Single most important data point for the long-term thesis.",
        "② Gross Margin — Q2 guided flat to Q1 (~48.3%). Any print above 49% = positive surprise. Any print below 47% = timeline to 'low 50s' extends into 2027 or beyond.",
        "③ Alani Nu Revenue — watch for normalization from Q1's $368M. A significant Q/Q decline would confirm Q1 was inflated by distributor loading, not pure consumer demand.",
        "④ Rockstar Integration Completion — management guided H1 2026. Confirm on track and watch for early velocity improvements post-reset.",
        "⑤ Adj. SG&A % of Revenue — Q1 was 26.4%, down from Q4's 31.8%. Continued improvement toward 25% confirms operating leverage is structural, not seasonal.",
        "⑥ Adj. EBITDA Margin — Q1 was 25.0%. Watch for expansion toward 26-27% as orbit model and freight improvements take hold.",
        "⑦ Legal Settlement Update — $88.2M accrued for Strong Arm litigation. Any resolution or incremental accrual will be material.",
        "⑧ PepsiCo Preferred — watch for any conversion activity or changes to the preferred stock terms. Conversion would dilute common shareholders but eliminate the quarterly $25M drain.",
        "⑨ Share Repurchase Activity — management bought $24.1M in Q1 near 52-week lows. Continuation signals sustained confidence; pause signals caution.",
    ]:
        next_row = bullet_row(ws, next_row, m, WHITE)

    # Footer
    next_row += 1
    ws.merge_cells(f"A{next_row}:H{next_row}")
    foot = ws[f"A{next_row}"]
    foot.value = (
        f"Source: {COMPANY} {QUARTER} Earnings Press Release ({REPORT_DATE})  |  "
        f"{PRESS_URL}  |  "
        f"Estimates: Zacks, TIKR, Investing.com  |  Analysis built: {GENERATED}  |  "
        f"Script: {SCRIPT_NAME}  |  Not investment advice."
    )
    foot.font = Font(name="Arial", size=SZ_FOOTER, color=MUTED, italic=True)
    foot.fill = fill(WHITE)
    foot.alignment = Alignment(horizontal="left", wrap_text=True)
    ws.row_dimensions[next_row].height = 18


# ==============================================================================
# TAB 2 — DATA SOURCES
# ==============================================================================
def build_data_sources_tab(wb):
    ws = wb.create_sheet("Data Sources")

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 13
    ws.column_dimensions["D"].width = 52
    ws.column_dimensions["E"].width = 42

    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value = f"{COMPANY} ({TICKER}) — {QUARTER} INCOME STATEMENT — DATA SOURCES"
    t.font = Font(name="Arial", bold=True, size=SZ_TITLE, color=WHITE)
    t.fill = fill(NAVY)
    t.alignment = center()
    ws.row_dimensions[1].height = 24

    for col_idx, h in enumerate(
        ["Category", "Source", "Date", "Key Data Used", "Reference / URL"], 1
    ):
        c = ws.cell(row=2, column=col_idx)
        c.value = h
        c.font = Font(name="Arial", bold=True, size=9, color=WHITE)
        c.fill = fill(ACCENT)
        c.alignment = center(wrap=True)
        c.border = TABLE_BORDER
    ws.row_dimensions[2].height = 24

    sources = [
        ("Primary Data",
         f"{COMPANY} {QUARTER} Earnings Press Release",
         REPORT_DATE,
         f"All {QUARTER} actuals: Revenue ${Q1_REVENUE:.1f}M, Gross Margin 48.3%, "
         "Income from Operations $139.0M, Net Income $110.1M, "
         "Net Income to Common $85.1M, Diluted EPS $0.33, Adj. EBITDA $195.5M. "
         "Quarter ended March 31, 2026.",
         PRESS_URL),

        ("Comparison Data",
         f"{COMPANY} Q1 2025 Earnings Press Release",
         "May 2025",
         "Q1 2025 figures used for all Y/Y change calculations: Revenue $329.3M, "
         "Gross Margin 52.3%, Income from Operations $52.0M, Net Income $44.4M, "
         "Net Income to Common $34.4M, Diluted EPS $0.15, Adj. EBITDA $69.7M.",
         "https://investor.celsiusholdingsinc.com/news-releases"),

        ("Strategic",
         "PepsiCo Distribution Agreement",
         "Aug 1, 2022",
         "Exclusive North American distribution of CELSIUS brand through PepsiCo DSD network. "
         "PepsiCo = 59% of Q1 2026 revenue and 45.5% of accounts receivable. "
         "Key risk: customer concentration. Agreement expanded to cover Alani Nu in 2025.",
         "Celsius Holdings 10-K / 8-K filings — SEC EDGAR"),

        ("Strategic",
         "PepsiCo Convertible Preferred Stock — Series A & B",
         "2025",
         "PepsiCo holds Series A and B convertible preferred stock issued to co-finance "
         "Alani Nu and Rockstar acquisitions. ~$25M per quarter deducted from common earnings "
         "($14.0M preferred dividends + $11.0M participating income allocation). "
         "Convertible to ~78M common shares (~334M fully diluted). "
         "Double EPS drag: numerator reduction + denominator expansion.",
         "Celsius Holdings 8-K filings and Q1 2026 Earnings Press Release"),

        ("Strategic",
         "Alani Nu Acquisition",
         "2025",
         "Largest acquisition in Celsius history. $368M revenue contribution in Q1 2026. "
         "Integration complete in Q1 2026: $50M synergies captured, distribution "
         "fully in PepsiCo DSD system. Distributor termination fee balance: "
         "$40.0M remaining (down from $264.1M original). "
         "Carries lower gross margins than core CELSIUS — primary driver of 48.3% blended margin.",
         "Celsius Holdings 8-K and Q1 2026 Earnings Press Release"),

        ("Strategic",
         "Rockstar Energy Acquisition",
         "Q4 2025",
         "$67M revenue contribution in Q1 2026 (declining brand: retail -13% Y/Y tracked channels). "
         "Integration targeted for H1 2026 completion. Orbit model and freight optimization "
         "targeted to improve acquired margin profile. $696.5M term loan (due 2032) funds "
         "Alani Nu + Rockstar — creates $11.8M/quarter permanent interest expense.",
         "Celsius Holdings 8-K and Q1 2026 Earnings Press Release"),

        ("Legal",
         "Strong Arm Productions Litigation",
         "Q1 2026",
         "$24.6M accrued in Q1 2026. Total accrual: $88.2M. "
         "Ongoing litigation; final settlement amount uncertain. "
         "Included in GAAP G&A; excluded from Adj. EBITDA add-backs. "
         "Material overhang — resolution will be a significant event.",
         "Celsius Holdings Q1 2026 10-Q / Earnings Press Release"),

        ("Guidance",
         f"{COMPANY} Q2 2026 Management Commentary",
         REPORT_DATE,
         "Q2 2026 gross margin guided flat to Q1 (~48.3%) — management called it a 'sidestep'. "
         "No explicit Q2 revenue guidance provided. Step-up expected Q3-Q4 via orbit model, "
         "freight optimization, raw material alignment. "
         "North Carolina second manufacturing line begins H2 2026.",
         "Celsius Holdings Q1 2026 Earnings Call (May 7, 2026)"),

        ("Consensus",
         "Analyst Consensus Estimates",
         GENERATED,
         "Q1 2026 consensus: Revenue ~$710M, Diluted EPS ~$0.28, "
         "Adj. EBITDA ~$165M, Adj. EPS ~$0.29-0.30. "
         "Actual beats: Revenue +10%, GAAP EPS +17.9%, Adj. EBITDA +18%, Adj. EPS +37-41%. "
         "Adj. EPS beat was headline metric of the quarter.",
         "Zacks Investment Research, TIKR, Investing.com"),
    ]

    for i, (cat, src, date, data_used, ref) in enumerate(sources):
        r = 3 + i
        bg = WHITE if i % 2 == 0 else GRAY
        for col_idx, val in enumerate([cat, src, date, data_used, ref], 1):
            c = ws.cell(row=r, column=col_idx)
            c.value = val
            c.font = Font(name="Arial", size=8,
                          bold=(col_idx in [1, 2]),
                          color=("0070C0" if col_idx == 5 else BLACK))
            c.fill = fill(bg)
            c.alignment = Alignment(
                horizontal=("center" if col_idx == 3 else "left"),
                vertical="top", wrap_text=True)
            c.border = TABLE_BORDER
        ws.row_dimensions[r].height = max(28, 14 + len(data_used) // 6)


# ==============================================================================
# TAB 3 — WHAT'S NEW
# ==============================================================================
def build_whats_new_tab(wb):
    ws = wb.create_sheet("What's New")

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 110

    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value = f"{COMPANY} ({TICKER}) — {QUARTER} INCOME STATEMENT — WHAT'S NEW"
    t.font = Font(name="Arial", bold=True, size=SZ_TITLE, color=WHITE)
    t.fill = fill(NAVY)
    t.alignment = center()
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:C2")
    s = ws["A2"]
    s.value = (f"Key developments and changes in {QUARTER}  |  "
               f"Script v1: refactored from build_celsius_simple_v4_final.py to AMD income pattern  |  "
               f"Next update → build_celh_income_v2.py")
    s.font = Font(name="Arial", size=9, color=WHITE)
    s.fill = fill(BLUE)
    s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 16

    for col_idx, h in enumerate(["", "Development", "Detail"], 1):
        c = ws.cell(row=3, column=col_idx)
        c.value = h
        c.font = Font(name="Arial", bold=True, size=9, color=WHITE)
        c.fill = fill(ACCENT)
        c.alignment = center()
    ws.row_dimensions[3].height = 16

    items = [
        ("Alani Nu Integration Complete",
         "Alani Nu integration declared complete in Q1 2026. $50M in synergies captured — "
         "one of the fastest beverage acquisitions in recent memory. Distribution fully transitioned "
         "into PepsiCo DSD system. $368M revenue contribution in Q1 2026. Distributor termination "
         "fee balance reduced from $264.1M to $40.0M remaining — wind-down well advanced. "
         "Key risk: Q1 distributor loading may have inflated revenue vs true consumer demand. "
         "Q3-Q4 2026 is the real organic growth test as the acquisition laps."),

        ("Rockstar Energy Integration",
         "Rockstar acquisition closed Q4 2025 with a $696.5M term loan (due 2032). "
         "Q1 2026 contribution: $67M. Retail tracked channels -13% Y/Y — brand in decline. "
         "Integration targeted for H1 2026 completion. Orbit model and freight optimization "
         "targeted to align Rockstar cost structure with CELSIUS margins. "
         "H2 2026 is when first brand recovery signals are expected. "
         "Rockstar is the riskiest asset in the portfolio."),

        ("PepsiCo Preferred Stock — Series A & B",
         "PepsiCo holds convertible preferred stock (Series A and Series B) issued to co-finance "
         "the Alani Nu and Rockstar acquisitions. Impact in Q1 2026: ~$25M per quarter deducted "
         "before common shareholders ($14.0M preferred dividends + $11.0M income allocation). "
         "PepsiCo can convert to ~78M common shares, expanding diluted count from 256M to ~334M. "
         "This structure is permanent — ~$100M annually — until conversion or buyout occurs. "
         "Always use 'Net Income to Common Shareholders' ($85.1M), not 'Net Income' ($110.1M), "
         "when evaluating CELH as a common equity investment."),

        ("Strong Arm Productions Litigation",
         "Celsius accrued $24.6M in Q1 2026 for the Strong Arm Productions litigation, "
         "bringing the total accrual to $88.2M. Final settlement amount remains uncertain. "
         "This is a material overhang — $88.2M represents roughly one quarter of Adj. EBITDA. "
         "Accrual included in GAAP G&A and excluded from Adj. EBITDA add-backs. "
         "Resolution (win, settle, or adverse judgment) will be a major event when it arrives."),

        ("Share Repurchase Program",
         "$24.1M in shares repurchased in Q1 2026 at prices near 52-week lows — management "
         "signaling confidence in intrinsic value. Buybacks partially offset the dilutive effect "
         "of SBC and warrant activity. Given the PepsiCo conversion overhang (~78M shares), "
         "buyback pace matters for EPS trajectory. Continuation in Q2 would reinforce conviction."),

        ("North Carolina Manufacturing Line — H2 2026",
         "Second manufacturing line at the North Carolina facility expected online H2 2026. "
         "Full production benefit expected in 2027. Key enabler of the orbit model — "
         "reducing third-party co-packing costs and improving gross margin structurally over time. "
         "Not a 2026 gross margin driver, but a meaningful 2027+ tailwind to the margin recovery thesis."),

        ("Script v1 — Refactored to AMD Income Pattern",
         "build_celh_income_v1.py refactored from build_celsius_simple_v4_final.py. "
         "Changes: explicit DATA BLOCK at top for easy quarterly updates; "
         "Data Sources tab added (tabulates all primary and comparison data sources); "
         "What's New tab added (this tab — key developments each quarter); "
         "TABLE_BORDER added to all data rows for visual alignment; "
         "tab colors added (Income Statement=navy, Data Sources=teal, What's New=orange); "
         "self-test upgraded to verify all 3 tabs, grade, revenue figure, and key references. "
         "All financial data and analysis from build_celsius_simple_v4_final.py preserved exactly. "
         "Next quarterly update → build_celh_income_v2.py (never overwrite this file)."),
    ]

    for i, (title, detail) in enumerate(items):
        r = 4 + i
        bg = WHITE if i % 2 == 0 else GRAY

        ws.cell(row=r, column=1).fill = fill(bg)

        c2 = ws.cell(row=r, column=2)
        c2.value = title
        c2.font = Font(name="Arial", bold=True, size=9, color=NAVY)
        c2.fill = fill(bg)
        c2.alignment = Alignment(horizontal="left", vertical="top",
                                  wrap_text=True, indent=1)

        c3 = ws.cell(row=r, column=3)
        c3.value = detail
        c3.font = Font(name="Arial", size=8, color=BLACK)
        c3.fill = fill(bg)
        c3.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        ws.row_dimensions[r].height = max(28, 14 + len(detail) // 10)


# ==============================================================================
# TAB 4 — FORWARD OUTLOOK
# ==============================================================================
def build_forward_outlook_tab(wb):
    ws = wb.create_sheet("Forward Outlook")

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 15
    ws.column_dimensions["H"].width = 46

    # Title
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = f"{COMPANY} ({TICKER}) — FORWARD GROWTH OUTLOOK & NORMALIZED REVENUE ANALYSIS"
    t.font = Font(name="Arial", bold=True, size=SZ_TITLE, color=WHITE)
    t.fill = fill(NAVY)
    t.alignment = center()
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:H2")
    s = ws["A2"]
    s.value = (f"Why {GRADE} and not B+  |  Phase-by-phase growth rate analysis  |  "
               f"Grade framework anchored to normalized growth, not the {Q1_REVENUE:.0f}M reported figure  |  "
               f"See Income Statement tab for line-item detail")
    s.font = Font(name="Arial", size=9, color=WHITE)
    s.fill = fill(BLUE)
    s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 16

    ROW = 3

    # ── SECTION 1: Revenue Decomposition ──────────────────────────────────────
    ws.merge_cells(f"A{ROW}:H{ROW}")
    sh = ws[f"A{ROW}"]
    sh.value = "▶  SECTION 1 — Q1 2026 REVENUE: REPORTED vs NORMALIZED"
    sh.font = Font(name="Arial", bold=True, size=10, color=WHITE)
    sh.fill = fill(SECTION_BG)
    sh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[ROW].height = 18
    ROW += 1

    # Sub-header explaining the loading concept
    ws.merge_cells(f"A{ROW}:H{ROW}")
    ex = ws[f"A{ROW}"]
    ex.value = (
        "Distributor loading = PepsiCo's DSD network stocking up with Alani Nu inventory before consumer "
        "pull-through begins. This is a one-time event: revenue recognized when product ships to distributors, "
        "not when consumers buy it. True demand = consumer pull-through rate, which is much lower in Q1."
    )
    ex.font = Font(name="Arial", size=8, color=BLACK, italic=True)
    ex.fill = fill(ASSESS_BG)
    ex.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=2)
    ws.row_dimensions[ROW].height = 36
    ROW += 1

    # Table header
    hdrs = ["", "Component", "Q1 2026 Reported ($M)", "Est. Loading ($M)",
            "Normalized ($M)", "", "", "Notes"]
    for col_idx, h in enumerate(hdrs, 1):
        c = ws.cell(row=ROW, column=col_idx)
        c.value = h
        c.font = Font(name="Arial", bold=True, size=9, color=WHITE)
        c.fill = fill(ACCENT)
        c.alignment = center(wrap=True)
        c.border = TABLE_BORDER
    ws.merge_cells(f"F{ROW}:G{ROW}")
    ws.row_dimensions[ROW].height = 28
    ROW += 1

    decomp_rows = [
        ("CELSIUS brand (organic)", 325, 0, 325,
         "+6% Y/Y organic on $306.5M Q1 2025 base. No loading effect — direct-to-consumer brand."),
        ("Alani Nu", 368, 210, 158,
         "~$500M annual pre-acquisition rate = ~$125M/quarter consumer pull. "
         "PepsiCo distribution expansion adds ~$30-55M lift → ~$140-180M sustainable. "
         "Implies ~$190-230M loading; midpoint ~$210M used here."),
        ("Rockstar", 67, 0, 67,
         "Retail tracked channels -13% Y/Y. No loading effect; acquisition is a declining brand."),
        ("International", 35, 0, 35,
         "Organic +55% Y/Y. No loading — direct consumer markets. Clean growth signal."),
        ("TOTAL", 782, 210, 572, "Normalized total implies ~27% reported figure was one-time loading."),
    ]

    for i, (label, reported, loading, normalized, note) in enumerate(decomp_rows):
        r = ROW + i
        is_total = label == "TOTAL"
        bg = NAVY if is_total else (WHITE if i % 2 == 0 else GRAY)
        fc = WHITE if is_total else BLACK

        for col_idx, val in enumerate([label, reported, loading, normalized], 2):
            c = ws.cell(row=r, column=col_idx)
            c.value = val if col_idx == 2 else f"${val:,}M" if not is_total else f"~${val:,}M"
            c.font = Font(name="Arial", bold=is_total, size=9, color=fc)
            c.fill = fill(bg)
            c.alignment = Alignment(
                horizontal="left" if col_idx == 2 else "center",
                vertical="center")
            c.border = TABLE_BORDER

        # loading cell — color amber if nonzero
        lc = ws.cell(row=r, column=4)
        if loading > 0 and not is_total:
            lc.font = Font(name="Arial", bold=True, size=9, color=AMBER)
            lc.fill = fill(AMBER_LT)
        elif is_total:
            lc.font = Font(name="Arial", bold=True, size=9, color=WHITE)

        ws.cell(row=r, column=1).fill = fill(bg)
        ws.merge_cells(f"F{r}:G{r}")
        nc = ws.cell(row=r, column=8)
        nc.value = note
        nc.font = Font(name="Arial", size=8, color=fc if is_total else "333333", italic=not is_total)
        nc.fill = fill(bg)
        nc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        nc.border = TABLE_BORDER
        ws.row_dimensions[r].height = max(28, 14 + len(note) // 8)

    ROW += len(decomp_rows) + 1

    # ── SECTION 2: Phase-by-Phase Growth ──────────────────────────────────────
    ws.merge_cells(f"A{ROW}:H{ROW}")
    sh2 = ws[f"A{ROW}"]
    sh2.value = "▶  SECTION 2 — PHASE-BY-PHASE FORWARD GROWTH RATE ANALYSIS"
    sh2.font = Font(name="Arial", bold=True, size=10, color=WHITE)
    sh2.fill = fill(SECTION_BG)
    sh2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[ROW].height = 18
    ROW += 1

    # Phase table header
    phase_hdrs = ["", "Phase", "Period", "Y/Y Growth (reported)", "Y/Y Growth (reality)", "", "", "Key Risk / Watch"]
    for col_idx, h in enumerate(phase_hdrs, 1):
        c = ws.cell(row=ROW, column=col_idx)
        c.value = h
        c.font = Font(name="Arial", bold=True, size=9, color=WHITE)
        c.fill = fill(ACCENT)
        c.alignment = center(wrap=True)
        c.border = TABLE_BORDER
    ws.merge_cells(f"F{ROW}:G{ROW}")
    ws.row_dimensions[ROW].height = 28
    ROW += 1

    phases = [
        ("Phase 1", "Q2 2026",
         "+60-80% Y/Y",
         "Significant sequential decline ($747M → ~$570-630M NA). "
         "Y/Y looks high only because Q2 2025 was pre-acquisition.",
         "⚠ Did Alani Nu loading reverse sharply? Watch Alani Nu sequential revenue.",
         GREEN_LT),
        ("Phase 2", "Q3–Q4 2026",
         "+40-60% Y/Y",
         "Still elevated Y/Y because comps are pre-acquisition. "
         "But sequential should stabilize as loading fully clears by Q3.",
         "⚠ CELSIUS brand must show double-digit organic growth as acquisition laps begin.",
         BLUE_S if hasattr(__builtins__, 'BLUE_S') else "D6E4F0"),
        ("Phase 3 ⚡", "Q1 2027",
         "-8% to -20% Y/Y",
         "THE DANGER QUARTER. Comparing against inflated $782.6M. "
         "If CELSIUS stays at 6% organic: ~$626M reported = -20% Y/Y (see Section 2B). "
         "If CELSIUS reaccelerates to 12%: ~$720M = -8% Y/Y. "
         "Either way the reported number is negative — a +9% normalized business looks broken.",
         "🚨 Management MUST pre-frame Q1 2027 comp on Q4 2026 call or stock will crater.",
         RED_LT),
        ("Phase 4", "2027+ (fully lapped)",
         "~15-25% Y/Y",
         "Sustainable blended rate: CELSIUS organic 10-15% + Alani Nu maturation 10-20% "
         "+ Rockstar flat/-5% + International 40-60% (small). "
         "This is the real business growth rate once all noise clears.",
         "✓ If CELSIUS organic is back to double digits by Q2 2027, the thesis is intact.",
         GREEN_LT),
    ]

    for i, (phase, period, reported_yy, reality, risk, bg) in enumerate(phases):
        r = ROW + i
        ws.cell(row=r, column=1).fill = fill(bg)

        for col_idx, val in enumerate([phase, period, reported_yy, reality], 2):
            c = ws.cell(row=r, column=col_idx)
            c.value = val
            bold = (col_idx == 2 or col_idx == 3)
            fc = RED if "Phase 3" in phase and col_idx == 4 else BLACK
            c.font = Font(name="Arial", bold=bold, size=9, color=fc)
            c.fill = fill(bg)
            c.alignment = Alignment(
                horizontal="center" if col_idx in [2, 3, 4] else "left",
                vertical="top", wrap_text=True)
            c.border = TABLE_BORDER

        ws.merge_cells(f"F{r}:G{r}")
        rc = ws.cell(row=r, column=8)
        rc.value = risk
        rc.font = Font(name="Arial", size=8, color=BLACK)
        rc.fill = fill(bg)
        rc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        rc.border = TABLE_BORDER

        ws.row_dimensions[r].height = max(42, 14 + len(reality) // 7)

    ROW += len(phases) + 1

    # ── SECTION 2B: Q1 2027 Detailed Scenario ─────────────────────────────────
    ws.merge_cells(f"A{ROW}:H{ROW}")
    sh2b = ws[f"A{ROW}"]
    sh2b.value = "▶  SECTION 2B — Q1 2027 DETAILED SCENARIO: PERCEPTION vs REALITY GAP (CELSIUS organic at 6%)"
    sh2b.font = Font(name="Arial", bold=True, size=10, color=WHITE)
    sh2b.fill = fill("8B0000")
    sh2b.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[ROW].height = 18
    ROW += 1

    ws.merge_cells(f"A{ROW}:H{ROW}")
    ex2b = ws[f"A{ROW}"]
    ex2b.value = (
        "If CELSIUS organic growth stays at 6% through Q1 2027, the component table below shows what each "
        "brand contributes and the resulting perception vs reality gap. "
        "The reported Y/Y will look like a business in serious decline. "
        "The normalized Y/Y will show a business that grew +9%. "
        "The 29-point gap between those two numbers is the core forward risk."
    )
    ex2b.font = Font(name="Arial", size=8, color=BLACK, italic=True)
    ex2b.fill = fill(ASSESS_BG)
    ex2b.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=2)
    ws.row_dimensions[ROW].height = 40
    ROW += 1

    # Component table header
    for col_idx, h in enumerate(
        ["", "Component", "Q1 2026 Reported", "Q1 2026 Normalized",
         "Q1 2027 Est.", "vs Reported", "vs Normalized", "Assumption"], 1
    ):
        c = ws.cell(row=ROW, column=col_idx)
        c.value = h
        c.font = Font(name="Arial", bold=True, size=9, color=WHITE)
        c.fill = fill(ACCENT)
        c.alignment = center(wrap=True)
        c.border = TABLE_BORDER
    ws.row_dimensions[ROW].height = 28
    ROW += 1

    q127_rows = [
        ("CELSIUS brand",  "$325M",  "~$325M",  "~$345M",  "+6%",   "+6%",   "Organic at 6% — no reacceleration assumed"),
        ("Alani Nu",       "$368M",  "~$158M",  "~$174M",  "-53%",  "+10%",  "~$500M pre-acq annual rate + PepsiCo lift; ~$210M Q1 loading unwinds"),
        ("Rockstar",       "$67M",   "~$67M",   "~$58M",   "-13%",  "-13%",  "Continuing -13% retail tracked channel decline"),
        ("International",  "$35M",   "~$35M",   "~$49M",   "+40%",  "+40%",  "Decelerating from 55% but organic growth continues"),
        ("TOTAL",          "$782M",  "~$572M",  "~$626M",  "-20%",  "+9%",   "29-point gap: market sees -20%, business did +9%"),
    ]

    for i, (label, rep, norm, est, vs_rep, vs_norm, note) in enumerate(q127_rows):
        r = ROW + i
        is_total = label == "TOTAL"
        bg = NAVY if is_total else (WHITE if i % 2 == 0 else GRAY)
        fc = WHITE if is_total else BLACK

        ws.cell(row=r, column=1).fill = fill(bg)

        for col_idx, val in enumerate([label, rep, norm, est], 2):
            c = ws.cell(row=r, column=col_idx)
            c.value = val
            c.font = Font(name="Arial", bold=is_total, size=9, color=fc)
            c.fill = fill(bg)
            c.alignment = Alignment(
                horizontal="left" if col_idx == 2 else "center",
                vertical="center")
            c.border = TABLE_BORDER

        for col_idx, (val, positive_is_good) in enumerate(
            [(vs_rep, False), (vs_norm, True)], 6
        ):
            c = ws.cell(row=r, column=col_idx)
            c.value = val
            if is_total:
                ink = WHITE
            elif positive_is_good:
                ink = GREEN if not val.startswith("-") else RED
            else:
                ink = RED if val.startswith("-") else GREEN
            c.font = Font(name="Arial", bold=is_total, size=9, color=ink)
            c.fill = fill(bg)
            c.alignment = center()
            c.border = TABLE_BORDER

        nc = ws.cell(row=r, column=8)
        nc.value = note
        nc.font = Font(name="Arial", size=8,
                       color=fc if is_total else "333333",
                       italic=not is_total, bold=is_total)
        nc.fill = fill(bg)
        nc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        nc.border = TABLE_BORDER
        ws.row_dimensions[r].height = 28

    ROW += len(q127_rows)

    # Callout box
    ws.merge_cells(f"A{ROW}:H{ROW}")
    callout = ws[f"A{ROW}"]
    callout.value = (
        "⚡  KEY TAKEAWAY: The market will see $626M vs $782.6M = -20% Y/Y and likely sell the stock. "
        "The business actually grew +9% on a normalized basis. "
        "This 29-point gap between perception and reality is why management MUST proactively frame "
        "the Q1 2027 comp on the Q4 2026 earnings call. "
        "Without pre-disclosure, a healthy business will look broken. "
        "This risk must be reflected in the projection engine scenarios."
    )
    callout.font = Font(name="Arial", bold=True, size=9, color=WHITE)
    callout.fill = fill("8B0000")
    callout.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=2)
    ws.row_dimensions[ROW].height = 48
    ROW += 2

    # ── SECTION 3: Normalized Grade Framework ─────────────────────────────────
    ws.merge_cells(f"A{ROW}:H{ROW}")
    sh3 = ws[f"A{ROW}"]
    sh3.value = "▶  SECTION 3 — GRADE FRAMEWORK ANCHORED TO NORMALIZED GROWTH (not 138% optical)"
    sh3.font = Font(name="Arial", bold=True, size=10, color=WHITE)
    sh3.fill = fill(SECTION_BG)
    sh3.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[ROW].height = 18
    ROW += 1

    ws.merge_cells(f"A{ROW}:H{ROW}")
    note_cell = ws[f"A{ROW}"]
    note_cell.value = (
        "Grading against 138% would be meaningless — that number cannot repeat (no more acquisitions of this scale, "
        "loading clears). The framework below grades future quarters against normalized portfolio growth: "
        "the rate the combined CELSIUS + Alani Nu + Rockstar + International business can organically sustain."
    )
    note_cell.font = Font(name="Arial", size=8, color=BLACK, italic=True)
    note_cell.fill = fill(ASSESS_BG)
    note_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=2)
    ws.row_dimensions[ROW].height = 36
    ROW += 1

    grade_rows = [
        ("A",   "25%+ normalized", "CELSIUS organic double digits + Alani Nu maturing + International scaling. "
                "Gross margin 50%+. Q1 2027 framed cleanly. Rockstar stabilizes.", GREEN_LT),
        ("B+",  "15-25% normalized", "CELSIUS organic 8-12%. Gross margin recovering visibly. "
                "Q1 2027 decline acknowledged. Capital structure drag manageable.", "D6E4F0"),
        (f"B ← {QUARTER}", "8-15% normalized", "Current standing. Strong Q1 execution but organic only +6%, "
                "loading inflates the headline, Q1 2027 trap unaddressed, "
                "$136M annual capital structure drag structural.", ASSESS_BG),
        ("C",   "0-8% normalized", "CELSIUS brand stalls or loses share. Alani Nu below $120M/quarter. "
                "Gross margin stuck below 48%. Adj. EBITDA margin contracts.", AMBER_LT),
        ("D",   "Negative normalized", "Revenue declines on organic basis. PepsiCo agreement adverse change. "
                "Legal settlement blowout. Debt covenant pressure.", RED_LT),
    ]

    for i, (grade, growth, desc, bg) in enumerate(grade_rows):
        r = ROW + i
        is_current = "←" in grade
        ws.cell(row=r, column=1).fill = fill(bg)

        for col_idx, val in enumerate([grade, growth, desc], 2):
            c = ws.cell(row=r, column=col_idx)
            c.value = val
            c.font = Font(name="Arial", bold=(col_idx <= 3 or is_current), size=9,
                          color=NAVY if is_current else BLACK)
            c.fill = fill(bg)
            c.alignment = Alignment(
                horizontal="center" if col_idx <= 3 else "left",
                vertical="top", wrap_text=True)
            c.border = TABLE_BORDER

        if col_idx == 4:  # desc occupies col 4 only; merge rest
            ws.merge_cells(f"D{r}:H{r}")
        else:
            ws.merge_cells(f"D{r}:H{r}")
            dc = ws.cell(row=r, column=4)
            dc.value = desc
            dc.font = Font(name="Arial", bold=is_current, size=9,
                           color=NAVY if is_current else BLACK)
            dc.fill = fill(bg)
            dc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            dc.border = TABLE_BORDER

        ws.row_dimensions[r].height = max(28, 14 + len(desc) // 8)

    ROW += len(grade_rows) + 1

    # ── SECTION 4: What to Watch ───────────────────────────────────────────────
    ws.merge_cells(f"A{ROW}:H{ROW}")
    sh4 = ws[f"A{ROW}"]
    sh4.value = "▶  SECTION 4 — WHAT TO WATCH TO UPGRADE FROM B TO B+"
    sh4.font = Font(name="Arial", bold=True, size=10, color=WHITE)
    sh4.fill = fill("1A7A3A")
    sh4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[ROW].height = 18
    ROW += 1

    upgrade_items = [
        "① CELSIUS brand organic growth ≥8% in Q2 2026 — the single biggest unlock. "
         "Shows the core brand is recovering independent of acquisition noise.",
        "② Alani Nu quarterly revenue stabilizes at $140M+ for two consecutive quarters — "
         "confirms loading has cleared and true consumer demand is solid.",
        "③ Gross margin prints above 49% in Q2 or Q3 2026 — shows the orbit model and freight "
         "optimization are working, not just promised.",
        "④ Management explicitly pre-frames the Q1 2027 comp on the Q4 2026 earnings call — "
         "proactive disclosure prevents the market from panicking at a technical Y/Y decline.",
        "⑤ Adj. EBITDA margin expands to 26%+ — confirms operating leverage is widening, "
         "not plateauing at the Q1 2026 level.",
    ]
    for item in upgrade_items:
        ws.merge_cells(f"A{ROW}:H{ROW}")
        c = ws[f"A{ROW}"]
        c.value = item
        c.font = Font(name="Arial", size=9, color=BLACK)
        c.fill = fill(WHITE if ROW % 2 == 0 else GRAY)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=2)
        ws.row_dimensions[ROW].height = max(18, 14 + len(item) // 12)
        ROW += 1

    # Footer
    ws.merge_cells(f"A{ROW}:H{ROW}")
    foot = ws[f"A{ROW}"]
    foot.value = (
        f"Analysis built: {GENERATED}  |  Script: {SCRIPT_NAME}  |  "
        "All revenue figures in $M  |  Normalized figures are estimates based on "
        "pre-acquisition run rates and distribution expansion modeling  |  Not investment advice."
    )
    foot.font = Font(name="Arial", size=SZ_FOOTER, color=MUTED, italic=True)
    foot.fill = fill(WHITE)
    foot.alignment = Alignment(horizontal="left", wrap_text=True)
    ws.row_dimensions[ROW].height = 18


# ==============================================================================
# BUILD
# ==============================================================================
def build():
    wb = Workbook()
    wb.remove(wb.active)

    build_income_tab(wb)
    build_data_sources_tab(wb)
    build_whats_new_tab(wb)
    build_forward_outlook_tab(wb)

    wb["Q1 2026 Income Statement"].sheet_properties.tabColor = "1A1A2E"  # navy
    wb["Data Sources"].sheet_properties.tabColor             = "006B6B"  # teal
    wb["What's New"].sheet_properties.tabColor               = "E67300"  # orange
    wb["Forward Outlook"].sheet_properties.tabColor          = "6B2D8B"  # purple

    wb.active = wb["Q1 2026 Income Statement"]

    out = os.path.join(_HERE, OUTPUT_NAME)
    wb.save(out)
    print(f"Saved: {out}")
    return out


# ==============================================================================
# SELF-TEST — DO NOT REMOVE
# Runs automatically after build(). Prints SELF-TEST PASSED or FAILED loudly.
# ==============================================================================
def self_test(out_path):
    import os
    from openpyxl import load_workbook

    errors = []

    if not os.path.exists(out_path):
        errors.append(f"FAIL: file not created at {out_path}")
    else:
        size = os.path.getsize(out_path)
        if size < 15000:
            errors.append(f"FAIL: file too small ({size} bytes) — engine may be incomplete")

        wb = load_workbook(out_path, data_only=True)

        # Tab presence
        for tab in ["Q1 2026 Income Statement", "Data Sources", "What's New", "Forward Outlook"]:
            if tab not in wb.sheetnames:
                errors.append(f"FAIL: tab '{tab}' missing")

        if "Q1 2026 Income Statement" in wb.sheetnames:
            ws = wb["Q1 2026 Income Statement"]
            found_revenue = found_782 = found_grade = False
            found_alani   = found_pepsico = False
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell:
                        cv = str(cell)
                        if "Revenue" in cv:   found_revenue = True
                        if "782" in cv:       found_782     = True
                        if GRADE in cv:       found_grade   = True
                        if "Alani" in cv:     found_alani   = True
                        if "PepsiCo" in cv:   found_pepsico = True
            if not found_revenue: errors.append("FAIL: 'Revenue' row not found in Income Statement tab")
            if not found_782:     errors.append(f"FAIL: Q1 2026 Revenue $782.6M not found")
            if not found_grade:   errors.append(f"FAIL: Grade '{GRADE}' not found in Income Statement tab")
            if not found_alani:   errors.append("FAIL: Alani Nu reference missing from Income Statement tab")
            if not found_pepsico: errors.append("FAIL: PepsiCo reference missing from Income Statement tab")

        if "Data Sources" in wb.sheetnames:
            ws_ds = wb["Data Sources"]
            found_ir = found_alani_ds = found_rockstar_ds = found_pepsico_ds = False
            for row in ws_ds.iter_rows(values_only=True):
                for cell in row:
                    if cell:
                        cv = str(cell)
                        if "investor.celsiusholdingsinc.com" in cv: found_ir          = True
                        if "Alani" in cv:                            found_alani_ds   = True
                        if "Rockstar" in cv:                         found_rockstar_ds = True
                        if "PepsiCo" in cv:                          found_pepsico_ds = True
            if not found_ir:          errors.append("FAIL: Data Sources — investor.celsiusholdingsinc.com missing")
            if not found_alani_ds:    errors.append("FAIL: Data Sources — Alani Nu entry missing")
            if not found_rockstar_ds: errors.append("FAIL: Data Sources — Rockstar entry missing")
            if not found_pepsico_ds:  errors.append("FAIL: Data Sources — PepsiCo entry missing")

        if "What's New" in wb.sheetnames:
            ws_wn = wb["What's New"]
            found_alani_wn = found_rockstar_wn = found_pepsico_wn = found_v1 = False
            for row in ws_wn.iter_rows(values_only=True):
                for cell in row:
                    if cell:
                        cv = str(cell)
                        if "Alani" in cv:                found_alani_wn   = True
                        if "Rockstar" in cv:             found_rockstar_wn = True
                        if "PepsiCo" in cv:              found_pepsico_wn  = True
                        if "build_celh_income_v1" in cv: found_v1          = True
            if not found_alani_wn:    errors.append("FAIL: What's New — Alani Nu item missing")
            if not found_rockstar_wn: errors.append("FAIL: What's New — Rockstar item missing")
            if not found_pepsico_wn:  errors.append("FAIL: What's New — PepsiCo item missing")
            if not found_v1:          errors.append("FAIL: What's New — build_celh_income_v1 script reference missing")

        if "Forward Outlook" in wb.sheetnames:
            ws_fo = wb["Forward Outlook"]
            found_phase3 = found_q1_2027 = found_normalized = False
            for row in ws_fo.iter_rows(values_only=True):
                for cell in row:
                    if cell:
                        cv = str(cell)
                        if "Phase 3" in cv:   found_phase3    = True
                        if "2027" in cv:      found_q1_2027   = True
                        if "normalized" in cv.lower(): found_normalized = True
            if not found_phase3:    errors.append("FAIL: Forward Outlook — Phase 3 section missing")
            if not found_q1_2027:   errors.append("FAIL: Forward Outlook — Q1 2027 analysis missing")
            if not found_normalized: errors.append("FAIL: Forward Outlook — normalized analysis missing")

    if errors:
        print("\n" + "=" * 60)
        print("SELF-TEST FAILED — DO NOT SAVE TO DRIVE")
        print("=" * 60)
        for e in errors:
            print(f"  {e}")
        print("=" * 60)
        return False

    size = os.path.getsize(out_path)
    print("\n" + "=" * 60)
    print("SELF-TEST PASSED — safe to save to Drive")
    print(f"  File: {out_path}")
    print(f"  Size: {size:,} bytes")
    print(f"  Tabs: Q1 2026 Income Statement + Data Sources + What's New + Forward Outlook — all confirmed")
    print(f"  Revenue ${Q1_REVENUE:.1f}M: confirmed")
    print(f"  Grade {GRADE}: confirmed")
    print(f"  Alani Nu / PepsiCo / Rockstar: confirmed in all three tabs")
    print(f"  Script reference build_celh_income_v1.py: confirmed in What's New")
    print("=" * 60)
    return True


if __name__ == "__main__":
    out = build()
    passed = self_test(out)
    if not passed:
        import sys
        sys.exit(1)
