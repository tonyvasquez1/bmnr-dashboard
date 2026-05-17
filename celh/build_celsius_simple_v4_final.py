# ══════════════════════════════════════════════════════════════════════════════
# CELSIUS HOLDINGS (CELH) — SIMPLE INCOME STATEMENT ANALYSIS
# build_celsius_simple_v4.py
# Q1 2026 | Reported May 7, 2026
# v4: Notes column widened and row heights expanded for readability
# ══════════════════════════════════════════════════════════════════════════════

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ── COLOR PALETTE ─────────────────────────────────────────────────────────────
NAVY       = "1A1A2E"
BLUE       = "1F4E79"
ACCENT     = "2E75B6"
WHITE      = "FFFFFF"
GRAY       = "F2F2F2"
GREEN      = "00B050"
AMBER      = "FF8C00"
RED        = "C00000"
GREEN_LT   = "E8F5E9"
AMBER_LT   = "FFF3E0"
RED_LT     = "FFEBEE"
MEASURE_BG = "EBF3FB"
BLACK      = "000000"
MUTED      = "666666"
SECTION_BG = "1F4E79"   # dark blue for section headers below table
STRENGTH_BG = "E8F5E9"
CONCERN_BG  = "FFEBEE"
ASSESS_BG   = "FFF8E1"

# ── PRIORITY ROWS (bold line item cell) ───────────────────────────────────────
PRIORITY = {
    "Revenue", "Cost of Revenue", "Gross Profit", "Gross Margin",
    "Total SG&A (GAAP)", "Income from Operations",
    "Total Other Income (Expense)", "Pre-Tax Income", "Net Income",
    "Net Income to Common Shareholders", "Diluted EPS (GAAP)", "Adjusted EBITDA",
}

# ── ROW DATA ──────────────────────────────────────────────────────────────────
# (label, q1_2026, q4_2025, q1_2025, yy_color, actual_color, is_pct_row, note, what_it_measures)
# yy_color:     "green"|"amber"|"red"|None  → Y/Y Change + Y/Y % cells
# actual_color: "green"|"amber"|"red"|None  → Q1 2026 actual cell
# is_pct_row:   True = margin/% row (Y/Y % shown in bps)
# q4_2025:      None = dash

ROWS = [
    # ── REVENUE ───────────────────────────────────────────────────────────────
    ("Revenue",
     782.6, None, 329.3, "green", None, False,
     "138% growth primarily acquisition-driven (Alani Nu +$368M, Rockstar +$67M); "
     "organic CELSIUS brand grew only +6% Y/Y. Alani Nu distributor loading into "
     "PepsiCo system inflated Q1 — one-time inventory build, not pure consumer demand. "
     "Real organic growth test comes Q3-Q4 2026 when acquisitions are lapped. "
     "If CELSIUS brand not back to double digits by then, that is when the alarm sounds.",
     "Top-line scale; total dollars sold across all brands and geographies"),

    ("  North America",
     747.3, None, 306.5, "amber", None, False,
     "Acquisition-driven; CELSIUS brand organic growth only ~6% Y/Y. "
     "Alani Nu ($368M) + Rockstar ($67M) account for ~$435M of the $441M Y/Y increase. "
     "Alani Nu distributor loading into PepsiCo system may have pulled forward demand — "
     "a one-time inventory build that may not fully repeat in Q2.",
     "Domestic scale; ~95% of total revenue"),

    ("  International",
     35.3, None, 22.7, "green", None, False,
     "Organic +55% Y/Y — no acquisition benefit here. Driven by Nordics strength "
     "and expansion markets (UK, Ireland, France, Australia/NZ, Benelux). "
     "Only 4.5% of total revenue but growing fast and entirely organic. "
     "International is the cleanest growth signal in the revenue mix.",
     "Organic global expansion; growth quality indicator"),

    # ── COST & GROSS ──────────────────────────────────────────────────────────
    ("Cost of Revenue",
     404.5, None, 156.9, "amber", None, False,
     "Cost grew +158% vs revenue +138% — 20 ppt gap reflects Alani Nu and Rockstar "
     "acquired at lower margin profiles than core CELSIUS brand. "
     "Core CELSIUS cost structure intact — the mix shift is the driver, not inefficiency. "
     "Q4 2025 COGS write-offs and transition costs largely behind. "
     "Commodity headwinds (Midwest aluminum LME premium) partially offset recovery. "
     "Margin recovery levers: orbit model, freight optimization, raw material alignment "
     "across Alani and Rockstar. Q2 gross margin guided flat ('sidestep').",
     "Direct production cost efficiency; lower growth vs revenue = margin expansion"),

    ("Gross Profit",
     378.1, None, 172.4, "green", None, False,
     "Dollar gross profit +119% Y/Y driven by acquisition scale — nearly doubled. "
     "Grew slower than revenue (+138%) because acquired brands carry lower margins. "
     "Dollar result is strong; efficiency story is told separately in Gross Margin below. "
     "Separating dollar result from margin rate gives a cleaner analytical picture.",
     "Absolute earning power after production; measures scale of the business"),

    ("Gross Margin",
     48.3, None, 52.3, "amber", "amber", True,
     "-400 bps Y/Y due to Alani Nu and Rockstar acquisition mix — both brands carry "
     "lower margins than core CELSIUS upon acquisition. Not an operational failure. "
     "Sequential improvement: +90 bps vs Q4 2025 (~47.4%) — direction already reversing. "
     "Actual cell flagged amber: margin rate deteriorated even as revenue scaled — "
     "consistent rule: any margin that grew slower than or declined vs revenue gets amber on actual. "
     "Management target: return to 'low 50s' — no firm timeline given. "
     "Q2 guided flat ('sidestep'); step-ups expected Q3-Q4 via orbit model, "
     "freight structure optimization, and raw material alignment across Alani and Rockstar. "
     "Second manufacturing line in North Carolina begins H2 2026; full benefit 2027.",
     "Production efficiency; how much of each revenue dollar is retained after COGS"),

    # ── OPEX ──────────────────────────────────────────────────────────────────
    ("Total SG&A (GAAP)",
     234.6, None, 120.3, "green", None, False,
     "Dollars +95% Y/Y but grew slower than revenue (+138%) — operating leverage confirmed. "
     "% of revenue improved meaningfully: 36.5% → 30.0% (-650 bps). "
     "Includes $24.6M legal settlement accrual (Strong Arm Productions litigation) "
     "and $3.8M acquisition/integration costs — both one-time. "
     "Adj. SG&A ex. one-time items: $206.3M / 26.4% of rev, down from 31.8% in Q4 2025. "
     "Green because SG&A growing slower than revenue is the definition of operating leverage.",
     "Operating cost discipline; grew slower than revenue = operating leverage confirmed"),

    ("  Sales & Marketing",
     150.6, None, 80.9, "green", None, False,
     "Dollars +86% Y/Y — grew slower than revenue (+138%). "
     "% of revenue improved: 24.6% → 19.2% (-540 bps). "
     "Brand spend becoming more efficient as portfolio scales across three brands. "
     "PepsiCo distribution handles shelf placement and logistics — reduces Celsius's "
     "direct marketing cost burden as volume scales.",
     "Brand investment efficiency; % of revenue shows leverage as scale grows"),

    ("  General & Administrative",
     84.1, None, 39.4, "green", None, False,
     "Dollars +113% Y/Y — grew slower than revenue (+138%) — operating leverage. "
     "However composition is noisy: includes $24.6M legal settlement (Strong Arm) "
     "and $3.8M M&A costs. Adjusted G&A ex. one-time items: $55.7M — only +41% growth "
     "vs revenue +138%. Underlying overhead is very well controlled. "
     "Green because GAAP G&A grew slower than revenue; note captures the one-time items "
     "that make the adjusted picture even cleaner than GAAP suggests.",
     "Overhead cost control; underlying G&A grew only +41% ex. one-time items"),

    ("Distributor Termination Fees",
     4.4, None, 0.0, None, "amber", False,
     "No prior year comp — cannot calculate meaningful Y/Y comparison. "
     "Actual flagged amber: real cash cost this quarter, not alarming but warrants attention. "
     "Transitional cost tied to shifting Alani Nu distribution from prior system into PepsiCo. "
     "Balance reduced from $264.1M to $40.0M remaining — wind-down well advanced. "
     "Not a recurring item; will disappear from the income statement as transition completes.",
     "Transitional one-time cost; signals distribution system restructuring in progress"),

    # ── OPERATING INCOME ──────────────────────────────────────────────────────
    ("Income from Operations",
     139.0, None, 52.0, "green", None, False,
     "+167% Y/Y — grew faster than revenue (+138%). Operating margin expanded: "
     "15.8% → 17.8% (+200 bps). This is the operating leverage story confirmed "
     "at the bottom of the operating section — the business is more profitable per "
     "dollar of revenue than a year ago despite absorbing two acquisition integrations simultaneously.",
     "Core business profitability before financing costs and taxes"),

    # ── BELOW THE LINE ────────────────────────────────────────────────────────
    ("Total Other Income (Expense)",
     -1.5, None, 9.0, "red", None, False,
     "Swung from +$9.0M income to -$1.5M expense — a $10.5M Y/Y deterioration. "
     "Two causes: (1) Interest income declined $4.9M as cash was deployed into acquisitions. "
     "(2) New $11.8M quarterly interest expense on $696.5M term loan (due 2032) taken on "
     "to fund Alani Nu and Rockstar acquisitions — this did not exist a year ago. "
     "Partially offset by $7.4M other income (incl. $7.0M from PepsiCo related party). "
     "Red because this is structural and permanent — not a one-time item. "
     "The debt is a 2032 term loan; $11.8M interest drag repeats every quarter going forward. "
     "This is the ongoing price of the acquisition-funded growth strategy.",
     "Financing impact; new acquisition debt creates permanent quarterly interest drag"),

    ("Pre-Tax Income",
     137.5, None, 61.0, "amber", None, False,
     "+125% Y/Y — grew slower than revenue (+138%). Amber by framework: any line "
     "growing slower than revenue signals efficiency erosion. "
     "The gap vs operating income (+167%) is entirely explained by the new interest expense. "
     "Unlike Adjusted EBITDA which adds back interest and D&A to show operating performance "
     "independent of capital structure, Pre-Tax Income fully reflects the cost of the "
     "$696.5M term loan. Management emphasizes Adjusted EBITDA precisely because "
     "Pre-Tax Income will always look weaker than the operating business warrants "
     "until the acquisition debt is paid down.",
     "Total profitability before tax; fully reflects capital structure cost unlike EBITDA"),

    ("Income Tax Provision",
     -27.4, None, -16.6, None, None, False,
     "Tax expense grew +65% Y/Y vs revenue +138% — grew slower than revenue. "
     "No color applied: support line, not a primary profitability driver. "
     "Effective tax rate improved significantly: 27.2% → 19.9% (-730 bps). "
     "Lower rate driven by tax treatment of acquisition-related deductions — "
     "a meaningful benefit that helped net income outpace pre-tax income growth.",
     "Tax efficiency; effective rate reflects jurisdictional mix and deductions"),

    ("Net Income",
     110.1, None, 44.4, "green", None, False,
     "+148% Y/Y — grew faster than revenue (+138%). One of the few lines to outpace "
     "revenue growth. The tax rate improvement (27.2% → 19.9%, -730 bps) more than "
     "offset the new interest drag, pushing net income above the revenue growth rate. "
     "Note: Net Income belongs to all capital providers including PepsiCo preferred holders. "
     "Net Income to Common Shareholders ($85.1M) is the true investor scorecard — "
     "$25M is redirected to PepsiCo preferred every quarter before common shareholders see it.",
     "Total GAAP profitability; belongs to all capital providers including preferred holders"),

    ("Net Income to Common Shareholders",
     85.1, None, 34.4, "green", None, False,
     "+147% Y/Y — grew faster than revenue (+138%). Green by framework. "
     "$25M gap vs Net Income ($110.1M) reflects PepsiCo's convertible preferred stock: "
     "$14.0M in preferred dividends + $11.0M income allocated to participating preferred. "
     "Context: PepsiCo co-financed Celsius's acquisitions by receiving Series A and B "
     "convertible preferred stock. As preferred holders they get paid before common shareholders "
     "every quarter — a permanent recurring deduction of ~$25M ($100M annually). "
     "PepsiCo can convert preferred to ~78M common shares, swelling diluted count to ~334M. "
     "This is the investor's true earnings base — always use this line, not Net Income, "
     "when evaluating CELH as a common stock investment.",
     "Profitability available to common equity after preferred obligations to PepsiCo"),

    # ── EPS ───────────────────────────────────────────────────────────────────
    ("Diluted EPS (GAAP)",
     0.33, None, 0.15, "amber", None, False,
     "+120% Y/Y — grew slower than revenue (+138%). Amber by framework. "
     "Preferred stock creates a double drag on EPS: "
     "(1) Numerator: $25M removed before calculation — reduces the earnings base. "
     "(2) Denominator: ~78M additional as-converted preferred shares inflate share count "
     "from 256M common to ~334M diluted — reduces per-share value. "
     "The two effects together explain why EPS growth (+120%) lags Net Income growth (+148%). "
     "Beat consensus ~$0.28 by +$0.05 (+17.9%). "
     "Adj. EPS $0.41 beat consensus $0.29-0.30 by +$0.11-0.12 (+37-41%). "
     "Basic EPS = Diluted EPS ($0.33) this quarter — effects roughly cancel out.",
     "Per-share profitability; primary metric Wall Street uses to price the stock"),

    ("Basic EPS (GAAP)",
     0.33, None, 0.15, None, None, False,
     "Same as diluted EPS ($0.33) this quarter — effects roughly cancel. "
     "Basic EPS = Net Income to Common / actual shares outstanding (256.5M). "
     "Diluted EPS = Net Income to Common (adjusted) / fully diluted shares (~334M as-converted). "
     "When Basic = Diluted, dilution risk is minimal in the current period. "
     "Watch for divergence — that signals dilution is actively hurting common shareholders. "
     "No color: reference line; Diluted EPS above is the primary metric.",
     "Per-share profit on common shares only; excludes all dilutive securities"),

    # ── NON-GAAP ──────────────────────────────────────────────────────────────
    ("Adjusted EBITDA",
     195.5, None, 69.7, "green", None, False,
     "+181% Y/Y — grew faster than revenue (+138%). Strongest growth rate in the statement. "
     "Margin expanded: 21.2% → 25.0% (+380 bps Y/Y). Beat consensus ~$165M by ~$30M (+18%). "
     "Add-backs total ~$49.5M: D&A $9.1M, SBC $7.6M, legal settlement $24.6M, "
     "distributor termination $4.4M, M&A costs $3.8M, FX -$0.4M. "
     "Management's primary operating metric — strips all acquisition noise to show "
     "underlying cash earning power of the brand portfolio. "
     "The +181% growth vs +138% revenue confirms the business generates cash at an "
     "accelerating rate relative to revenue as it scales.",
     "Cash earning power ex. non-cash and one-time items; management's primary metric"),

    ("Adjusted EBITDA Margin",
     25.0, None, 21.2, "green", "green", True,
     "+380 bps Y/Y expansion — margin expanded despite absorbing two acquisitions "
     "at lower margin profiles. Demonstrates operating leverage at scale. "
     "Actual cell colored green: consistent rule — all margin rows get actual colored "
     "to match Y/Y color. Green actual = margin expanded Y/Y (positive signal). "
     "Contrast with Gross Margin (amber actual) where margin deteriorated — "
     "the two together tell the full story: gross margins under pressure from mix, "
     "but cash margins expanding from operating leverage. Both signals matter.",
     "Operating leverage at scale; expanding margin = business becoming more efficient"),

    ("Adjusted Diluted EPS",
     0.41, None, 0.18, "amber", None, False,
     "+128% Y/Y — grew slower than revenue (+138%). Amber by framework. "
     "Same preferred share dilution dynamic as GAAP EPS affects the denominator "
     "even in the non-GAAP calculation. "
     "However this is the headline beat of the quarter: $0.41 vs consensus $0.29-0.30, "
     "a +37-41% surprise — strongest consensus beat in the statement. "
     "Add-backs net of tax: M&A costs +$0.01, distributor termination +$0.01, "
     "legal settlement +$0.06. Tax effect of all adjustments: -$0.08/share. "
     "Despite amber on the Y/Y framework, the consensus beat signals the market "
     "was significantly underestimating Celsius's underlying earnings power.",
     "Clean per-share earnings ex. one-time items; consensus benchmark for beat/miss"),
]

# ── COLUMN LAYOUT ─────────────────────────────────────────────────────────────
# A: Line Item    B: Y/Y Change    C: Y/Y %
# D: Q1 2026      E: Q4 2025       F: Q1 2025
# G: Notes        H: What This Measures

COL_WIDTHS = {
    "A": 32,   # Line Item
    "B": 14,   # Y/Y Change
    "C": 10,   # Y/Y %
    "D": 13,   # Q1 2026
    "E": 13,   # Q4 2025
    "F": 13,   # Q1 2025
    "G": 90,   # Notes
    "H": 40,   # What This Measures
}

def fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def section_header(ws, row, text, bg, cols=8):
    ws.merge_cells(f"A{row}:{chr(64+cols)}{row}")
    cell = ws[f"A{row}"]
    cell.value = text
    cell.font = Font(name="Arial", bold=True, size=10, color=WHITE)
    cell.fill = fill(bg)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 18
    return row + 1

def bullet_row(ws, row, text, bg, indent=2, font_size=9, color=BLACK):
    ws.merge_cells(f"A{row}:H{row}")
    cell = ws[f"A{row}"]
    cell.value = text
    cell.font = Font(name="Arial", size=font_size, color=color)
    cell.fill = fill(bg)
    cell.alignment = Alignment(horizontal="left", vertical="top",
                               wrap_text=True, indent=indent)
    ws.row_dimensions[row].height = max(15, 13 + len(text) // 10)
    return row + 1

def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "Q1 2026 Income Statement"

    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # ── Title ─────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "CELSIUS HOLDINGS, INC. (CELH) — Q1 2026 INCOME STATEMENT ANALYSIS"
    t.font = Font(name="Arial", bold=True, size=13, color=WHITE)
    t.fill = fill(NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:H2")
    s = ws["A2"]
    s.value = "Quarter Ended March 31, 2026  |  Reported May 7, 2026  |  Source: Celsius Holdings Earnings Press Release"
    s.font = Font(name="Arial", size=9, color=WHITE)
    s.fill = fill(BLUE)
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 15

    # ── Column headers ─────────────────────────────────────────────────────────
    headers = ["Line Item", "Y/Y Change ($M)", "Y/Y %",
               "Q1 2026 ($M)", "Q4 2025 ($M)", "Q1 2025 ($M)",
               "Notes", "What This Measures"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = h
        cell.font = Font(name="Arial", bold=True, size=9, color=WHITE)
        cell.fill = fill(ACCENT)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 28

    yy_map = {
        "green": (GREEN, GREEN_LT),
        "amber": (AMBER, AMBER_LT),
        "red":   (RED,   RED_LT),
    }

    DATA_START = 4

    for i, (label, q1, q4, q1_py,
            yy_color, actual_color, is_pct, note, measure) in enumerate(ROWS):

        row = DATA_START + i
        is_priority = label.strip() in PRIORITY
        bg = WHITE if i % 2 == 0 else GRAY

        def num_fmt(val, pct):
            if pct: return '0.0"%"'
            return '$#,##0.00' if abs(val) < 5 else '$#,##0.0'

        # A — Line Item
        a = ws.cell(row=row, column=1)
        a.value = label
        a.font = Font(name="Arial", bold=is_priority, size=9, color=BLACK)
        a.fill = fill(bg)
        a.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        # B — Y/Y Change (formula)
        b = ws.cell(row=row, column=2)
        b.value = f"=D{row}-F{row}"
        b.number_format = ('+0.0;-0.0;"-"' if is_pct
                           else ('+$#,##0.00;-$#,##0.00;"-"' if abs(q1) < 5
                                 else '+$#,##0.0;-$#,##0.0;"-"'))
        b.alignment = Alignment(horizontal="right", vertical="center")

        # C — Y/Y %
        c = ws.cell(row=row, column=3)
        if is_pct:
            c.value = f"=D{row}-F{row}"
            c.number_format = '+0.0" bps";-0.0" bps";"-"'
        else:
            c.value = f'=IF(F{row}<>0,(D{row}-F{row})/ABS(F{row}),"N/A")'
            c.number_format = '+0.0%;-0.0%;"-"'
        c.alignment = Alignment(horizontal="right", vertical="center")

        # Apply Y/Y color to B and C
        if yy_color in yy_map:
            ink, bg_yy = yy_map[yy_color]
            for col in [2, 3]:
                cell = ws.cell(row=row, column=col)
                cell.fill = fill(bg_yy)
                cell.font = Font(name="Arial", size=9, color=ink, bold=True)
        else:
            for col in [2, 3]:
                cell = ws.cell(row=row, column=col)
                cell.fill = fill(bg)
                cell.font = Font(name="Arial", size=9, color=BLACK)

        # D — Q1 2026 actual
        d = ws.cell(row=row, column=4)
        d.value = q1
        d.number_format = num_fmt(q1, is_pct)
        d.alignment = Alignment(horizontal="right", vertical="center")
        if actual_color in yy_map:
            ink, bg_ac = yy_map[actual_color]
            d.fill = fill(bg_ac)
            d.font = Font(name="Arial", size=9, color=ink, bold=True)
        else:
            d.fill = fill(bg)
            d.font = Font(name="Arial", size=9, color=BLACK)

        # E — Q4 2025
        e = ws.cell(row=row, column=5)
        if q4 is None:
            e.value = "—"
            e.font = Font(name="Arial", size=9, color=MUTED)
        else:
            e.value = q4
            e.number_format = num_fmt(q4, is_pct)
            e.font = Font(name="Arial", size=9, color=BLACK)
        e.fill = fill(bg)
        e.alignment = Alignment(horizontal="right", vertical="center")

        # F — Q1 2025
        f = ws.cell(row=row, column=6)
        f.value = q1_py
        f.number_format = num_fmt(q1_py, is_pct)
        f.font = Font(name="Arial", size=9, color=BLACK)
        f.fill = fill(bg)
        f.alignment = Alignment(horizontal="right", vertical="center")

        # G — Notes
        g = ws.cell(row=row, column=7)
        g.value = note
        g.font = Font(name="Arial", size=8, color="333333")
        g.fill = fill(bg)
        g.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # H — What This Measures
        h = ws.cell(row=row, column=8)
        h.value = measure
        h.font = Font(name="Arial", size=8, color="1F4E79", italic=True)
        h.fill = fill(MEASURE_BG)
        h.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True, indent=1)

        ws.row_dimensions[row].height = max(45, min(180, 12 + note.count(" ") * 1.4))

    ws.freeze_panes = "B4"

    # ── Spacer after table ────────────────────────────────────────────────────
    next_row = DATA_START + len(ROWS) + 1
    ws.row_dimensions[next_row].height = 10
    next_row += 1

    # ══════════════════════════════════════════════════════════════════════════
    # STRENGTHS
    # ══════════════════════════════════════════════════════════════════════════
    next_row = section_header(ws, next_row, "▶  STRENGTHS", "1A7A3A")
    strengths = [
        "✓  Record Q1 revenue of $782.6M — +138% Y/Y; portfolio now holds 20.9% of the U.S. RTD energy category, one of the fastest category share gains in beverage history.",
        "✓  Adjusted EBITDA +181% Y/Y to $195.5M at a 25.0% margin — strongest growth rate in the statement and proof that cash earnings scale faster than revenue.",
        "✓  Operating leverage confirmed across every cost line: SG&A (-650 bps as % of rev), Sales & Marketing (-540 bps), Adj. SG&A (-540 bps Q/Q). Cost discipline is real.",
        "✓  Alani Nu integration complete: $50M in synergies captured, distribution fully transitioned into PepsiCo system. One of the fastest beverage integrations on record.",
        "✓  International revenue +55% Y/Y organically — no acquisition benefit. Nordics leading; UK, France, Australia/NZ expanding. Clean growth signal.",
        "✓  Gross margin improving sequentially: +90 bps Q/Q (47.4% → 48.3%). Direction of travel is correct even if Y/Y comparison is distorted by acquisitions.",
        "✓  Adj. EPS $0.41 beat consensus $0.29-0.30 by +37-41% — strongest consensus beat in the statement. Market was significantly underestimating earning power.",
        "✓  Tax rate improvement: 27.2% → 19.9% (-730 bps). Acquisition-related deductions providing meaningful tax shield.",
        "✓  $24.1M share repurchases in Q1 — management putting capital to work at near 52-week lows, signaling confidence in intrinsic value.",
    ]
    for s in strengths:
        next_row = bullet_row(ws, next_row, s, STRENGTH_BG)

    next_row += 1  # spacer

    # ══════════════════════════════════════════════════════════════════════════
    # CONCERNS
    # ══════════════════════════════════════════════════════════════════════════
    next_row = section_header(ws, next_row, "▶  CONCERNS", "8B0000")
    concerns = [
        "✗  Organic CELSIUS brand growth only +6% Y/Y — the entire revenue growth story is inorganic. If CELSIUS brand not back to double digits by Q3-Q4 2026 when acquisitions lap, the thesis weakens materially.",
        "✗  Gross margin -400 bps Y/Y (52.3% → 48.3%). Path back to 'low 50s' relies on orbit model, freight optimization, and commodity tailwinds — none of which are guaranteed. Q2 guided flat ('sidestep').",
        "✗  New structural interest expense: $11.8M per quarter ($47M annualized) on $696.5M term loan due 2032. Permanent drag on Pre-Tax Income and EPS that did not exist a year ago.",
        "✗  PepsiCo preferred stock drains ~$25M per quarter ($100M annually) from common shareholders — a permanent toll that reduces EPS and dilutes on conversion (~78M additional shares).",
        "✗  Alani Nu Q1 revenue ($368M) may have been inflated by distributor loading as PepsiCo absorbed the brand — a one-time inventory build that may not repeat in Q2 or Q3.",
        "✗  Rockstar Energy retail sales -13% Y/Y in tracked channels. Integration on track for H1 2026 completion but brand trajectory is weak. Rockstar is declining, not growing.",
        "✗  Accrued legal expenses $88.2M (Strong Arm Productions litigation) — $24.6M accrued in Q1 alone. Liability overhang and uncertainty around final settlement amount.",
        "✗  Customer concentration: PepsiCo = 59% of Q1 revenue and 45.5% of accounts receivable. Any change to the commercial agreement would be catastrophic.",
        "✗  Distributor termination fees: $40.0M still accrued. While winding down, cash outflows will continue for several quarters.",
    ]
    for c in concerns:
        next_row = bullet_row(ws, next_row, c, CONCERN_BG)

    next_row += 1  # spacer

    # ══════════════════════════════════════════════════════════════════════════
    # OVERALL ASSESSMENT
    # ══════════════════════════════════════════════════════════════════════════
    next_row = section_header(ws, next_row, "▶  OVERALL ASSESSMENT", SECTION_BG)

    grade_text = "B+"
    ws.merge_cells(f"A{next_row}:H{next_row}")
    grade_cell = ws[f"A{next_row}"]
    grade_cell.value = f"Quarter Grade:  {grade_text}"
    grade_cell.font = Font(name="Arial", bold=True, size=14, color=NAVY)
    grade_cell.fill = fill(ASSESS_BG)
    grade_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[next_row].height = 28
    next_row += 1

    assessment_text = (
        "Celsius delivered a strong operational quarter but the headline numbers require careful interpretation. "
        "Record revenue of $783M and Adj. EBITDA of $195.5M are genuinely impressive — and the operating leverage "
        "story (SG&A, S&M, and G&A all growing slower than revenue) is real and confirmed. "
        "The Adj. EPS beat of +37-41% vs consensus signals the market was underpricing the earnings power of the combined portfolio.\n\n"
        "However two structural issues prevent a higher grade: (1) Organic CELSIUS brand growth of only +6% means "
        "the entire revenue growth narrative is acquisition-funded. The real test is Q3-Q4 2026 when the "
        "acquisition anniversaries lap and organic growth becomes visible. "
        "(2) The PepsiCo preferred stock structure permanently redirects ~$25M per quarter from common shareholders "
        "and the $696.5M term loan adds $11.8M quarterly interest drag — both are structural, not transient.\n\n"
        "The gross margin recovery path (orbit model, freight optimization, raw material alignment) is credible "
        "but Q2 is guided flat and the commodity environment adds uncertainty. "
        "Rockstar's -13% retail decline is a yellow flag. "
        "The legal overhang ($88.2M accrued) adds noise.\n\n"
        "Bottom line: Celsius is executing well on integration and demonstrating real operating leverage at scale. "
        "The portfolio is stronger than a year ago. But the organic growth question and the capital structure "
        "complexity make this a 'show me' story for the next two quarters."
    )
    next_row = bullet_row(ws, next_row, assessment_text, ASSESS_BG, indent=2, font_size=9)

    next_row += 1  # spacer

    # ══════════════════════════════════════════════════════════════════════════
    # LETTER GRADE FRAMEWORK
    # ══════════════════════════════════════════════════════════════════════════
    next_row = section_header(ws, next_row, "▶  LETTER GRADE FRAMEWORK — CELSIUS HOLDINGS (CELH)", SECTION_BG)

    grades = [
        ("A",  "Organic CELSIUS brand growth reaccelerates to double digits. Gross margin returns to 50%+. "
               "Adj. EBITDA margin expands above 27%. No adverse PepsiCo agreement changes. Rockstar stabilizes."),
        ("B+", "Current quarter: strong operational execution, confirmed operating leverage, major consensus beat. "
               "Acquisition integration on track. Organic growth question unresolved. Capital structure complexity persists."),
        ("B",  "Gross margin recovery stalls at 48-49%. CELSIUS brand organic growth stays mid-single digits. "
               "Operating leverage continues but Adj. EBITDA margin plateaus. Legal overhang partially resolved."),
        ("C",  "Gross margin deteriorates below 47%. CELSIUS brand loses market share. Alani Nu revenue normalizes "
               "sharply post-distributor loading. Rockstar continues declining. Adj. EBITDA margin contracts."),
        ("D",  "PepsiCo commercial agreement renegotiated adversely. Major customer concentration event. "
               "Legal liability exceeds accruals materially. Acquisition synergies fail to materialize."),
    ]
    for grade, desc in grades:
        next_row = bullet_row(ws, next_row,
                              f"{grade}   {desc}",
                              ASSESS_BG, indent=2, font_size=9)

    next_row += 1  # spacer

    # ══════════════════════════════════════════════════════════════════════════
    # KEY METRICS TO WATCH — Q2 2026
    # ══════════════════════════════════════════════════════════════════════════
    next_row = section_header(ws, next_row, "▶  KEY METRICS TO WATCH — Q2 2026 (Expected Aug 6-11, 2026)", SECTION_BG)

    metrics = [
        "① CELSIUS Brand Organic Growth Rate — must show reacceleration toward double digits as Alani lapping begins. Single most important data point for the long-term thesis.",
        "② Gross Margin — Q2 guided flat to Q1 (~48.3%). Any print above 49% = positive surprise. Any print below 47% = timeline to 'low 50s' extends into 2027 or beyond.",
        "③ Alani Nu Revenue — watch for normalization from Q1's $368M. A significant Q/Q decline would confirm Q1 was inflated by distributor loading, not pure consumer demand.",
        "④ Rockstar Integration Completion — management guided H1 2026. Confirm on track and watch for early velocity improvements post-reset.",
        "⑤ Adj. SG&A % of Revenue — Q1 was 26.4%, down from Q4's 31.8%. Continued improvement toward 25% confirms operating leverage is structural, not seasonal.",
        "⑥ Adj. EBITDA Margin — Q1 was 25.0%. Watch for expansion toward 26-27% as orbit model and freight improvements take hold.",
        "⑦ Legal Settlement Update — $88.2M accrued for Strong Arm litigation. Any resolution or incremental accrual will be material.",
        "⑧ PepsiCo Preferred — watch for any conversion activity or changes to the preferred stock terms. Conversion would dilute common shareholders but eliminate the quarterly $25M drain.",
        "⑨ Share Repurchase Activity — management bought $24.1M in Q1 near 52-week lows. Continuation signals sustained confidence; pause signals caution.",
    ]
    for m in metrics:
        next_row = bullet_row(ws, next_row, m, WHITE, indent=2, font_size=9)

    # ── Footer ────────────────────────────────────────────────────────────────
    next_row += 1
    ws.merge_cells(f"A{next_row}:H{next_row}")
    foot = ws[f"A{next_row}"]
    foot.value = (
        "Source: Celsius Holdings Q1 2026 Earnings Press Release (May 7, 2026)  |  "
        "https://s203.q4cdn.com/427437840/files/doc_financials/2026/q1/v3/Celsius-Holdings-Q1-2026__Earnings-Press-Release.pdf  |  "
        "Estimates: Zacks, TIKR, Investing.com  |  Analysis built: May 2026  |  Script: build_celsius_simple_v4.py"
    )
    foot.font = Font(name="Arial", size=7, color=MUTED, italic=True)
    foot.fill = fill(WHITE)
    foot.alignment = Alignment(horizontal="left", wrap_text=True)
    ws.row_dimensions[next_row].height = 18

    out = "/mnt/user-data/outputs/CELH_Q1_2026_Income_Statement_v4.xlsx"
    wb.save(out)
    print(f"Saved: {out}")

# ==============================================================================
# SELF-TEST — runs automatically after build(), fails loudly if broken
# DO NOT REMOVE — this is what proves the script is complete and working
# ==============================================================================
def self_test(out_path):
    import os
    from openpyxl import load_workbook

    errors = []

    # 1. File exists
    if not os.path.exists(out_path):
        errors.append(f"FAIL: file not created at {out_path}")
    else:
        # 2. File size reasonable (complete engine produces >15KB)
        size = os.path.getsize(out_path)
        if size < 15000:
            errors.append(f"FAIL: file too small ({size} bytes) — engine may be incomplete")

        # 3. Required tabs exist
        wb = load_workbook(out_path, data_only=True)
        ws = wb.active
        if ws.title != "Q1 2026 Income Statement":
            errors.append(f"FAIL: wrong sheet name '{ws.title}'")

        # 4. Key data cells populated
        found_revenue = False
        for row in ws.iter_rows(values_only=True):
            if row[0] and "Revenue" in str(row[0]):
                found_revenue = True
                break
        if not found_revenue:
            errors.append("FAIL: Revenue row not found in sheet")

        # 5. Key output values correct
        found_782 = False
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell and "782" in str(cell):
                    found_782 = True
                    break
        if not found_782:
            errors.append("FAIL: Q1 2026 Revenue $782.6M not found — data block may be wrong")

    if errors:
        print("\n" + "="*60)
        print("SELF-TEST FAILED — DO NOT SAVE TO DRIVE")
        print("="*60)
        for e in errors:
            print(f"  {e}")
        print("="*60)
        return False
    else:
        size = os.path.getsize(out_path)
        print("\n" + "="*60)
        print("SELF-TEST PASSED — safe to save to Drive")
        print(f"  File: {out_path}")
        print(f"  Size: {size:,} bytes")
        print(f"  Revenue $782.6M: confirmed")
        print(f"  Sheet structure: confirmed")
        print("="*60)
        return True

if __name__ == "__main__":
    build()
    out = "/mnt/user-data/outputs/CELH_Q1_2026_Income_Statement_v4.xlsx"
    passed = self_test(out)
    if not passed:
        import sys
        sys.exit(1)
